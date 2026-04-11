#!/usr/bin/env python3
"""
PR Extraction Script

Extracts PR information from GitHub issue comments and updates Excel file.
- Only extracts PRs from comments (not issue body)
- Only extracts PRs mentioned with fix-related keywords (fix, closes, resolved, PR #, etc.)
- For pytorch/pytorch PRs: checks "Merged" label for status
- For intel/torch-xpu-ops PRs: checks merged/merged_at field for status
- Multiple PRs per issue are separated by ';'
"""

import openpyxl
import requests
import re
import json
import sys
import os
import time
import argparse

GITHUB_API = "https://api.github.com"
REPO = "intel/torch-xpu-ops"
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")

FIX_KEYWORDS = ['fix', 'fixes', 'fixed', 'close', 'closes', 'closed', 'resolve', 'resolved', 'pr #', 'pr:', 'pull request']

HEADERS = {
    "Accept": "application/vnd.github.v3+json",
    "User-Agent": "PyTorch-Issue-Triage-Bot"
}
if GITHUB_TOKEN:
    HEADERS["Authorization"] = f"token {GITHUB_TOKEN}"


def make_request_with_retry(url, max_retries=3):
    """Make request with retry on 403/429 errors"""
    for attempt in range(max_retries):
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        elif resp.status_code in (403, 429):
            # Rate limited - wait and retry
            wait_time = 60  # 1 minute for unauthenticated
            if 'X-RateLimit-Remaining' in resp.headers:
                remaining = int(resp.headers.get('X-RateLimit-Remaining', 0))
                if remaining < 10:
                    wait_time = 120
            print(f"  [Rate Limit] Waiting {wait_time}s before retry (attempt {attempt+1}/{max_retries})...")
            time.sleep(wait_time)
        else:
            print(f"  [API Error] {resp.status_code}")
            break
    return None


def get_issue_comments(issue_num):
    """Fetch comments for a specific issue"""
    url = f"{GITHUB_API}/repos/{REPO}/issues/{issue_num}/comments"
    try:
        result = make_request_with_retry(url)
        return result if result else []
    except Exception as e:
        print(f"Error fetching comments for issue {issue_num}: {e}")
        return []


def extract_pr_refs_from_comments(issue_num):
    """Extract PR references from issue comments with fix keywords"""
    comments = get_issue_comments(issue_num)
    pr_refs = set()
    
    for comment in comments:
        body = comment.get('body', '') or ''
        lower_body = body.lower()
        
        # Find all PR URLs
        for url_match in re.finditer(r'https://github\.com/(intel/torch-xpu-ops|pytorch/pytorch)/pull/(\d+)', body):
            # Get context around the PR URL
            start = max(0, url_match.start() - 100)
            end = min(len(body), url_match.end() + 100)
            context = lower_body[start:end]
            
            # Check if fix keyword is in context
            has_fix_keyword = any(kw in context for kw in FIX_KEYWORDS)
            
            if has_fix_keyword:
                pr_key = f'{url_match.group(1)}/{url_match.group(2)}'
                pr_refs.add(pr_key)
    
    return list(pr_refs)


def get_pr_info(pr_key):
    """Get PR information with repo-specific status logic"""
    parts = pr_key.split('/')
    repo = '/'.join(parts[:-1])
    pr_num = parts[-1]
    
    url = f"https://api.github.com/repos/{repo}/pulls/{pr_num}"
    try:
        result = make_request_with_retry(url)
        if not result:
            return {'error': 'API failed'}
        
        pr = result
        # Different status logic based on repo
        if repo == 'pytorch/pytorch':
            labels = [l.get('name', '') for l in pr.get('labels', [])]
            if 'Merged' in labels:
                status = 'merged'
            else:
                status = pr.get('state', 'unknown')
        else:
            # intel/torch-xpu-ops: use merged/merged_at field
            merged = pr.get('merged', False)
            merged_at = pr.get('merged_at')
            if merged or merged_at:
                status = 'merged'
            else:
                status = pr.get('state', 'unknown')
        
        return {
            'url': pr.get('html_url'),
            'owner': pr.get('user', {}).get('login'),
            'state': status
        }
    except Exception as e:
        print(f"Error fetching PR {pr_key}: {e}")
        return {'error': str(e)}


def extract_prs_from_excel(excel_file, target_issue_ids=None):
    """Extract PRs from all issues in Excel"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Issues']
    
    issues = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is not None:
            issues.append(row[0])
    
    # Filter to target issues if specified
    if target_issue_ids:
        issues = [i for i in issues if i in target_issue_ids]
        print(f"Filtered to {len(issues)} target issues: {target_issue_ids}")
    else:
        print(f"Total issues to process: {len(issues)}")
    
    # Collect PR refs from comments
    all_pr_refs = {}
    for i, issue_num in enumerate(issues):
        pr_refs = extract_pr_refs_from_comments(issue_num)
        if pr_refs:
            all_pr_refs[issue_num] = pr_refs
            print(f"Issue {issue_num}: found {len(pr_refs)} PRs")
        
        if (i + 1) % 50 == 0:
            print(f"Processed {i + 1}/{len(issues)} issues...")
    
    print(f"\nFound {len(all_pr_refs)} issues with fixing PRs")
    return all_pr_refs


def fetch_pr_info(all_pr_refs):
    """Fetch PR information for all unique PRs"""
    unique_prs = set()
    for pr_list in all_pr_refs.values():
        for pr in pr_list:
            unique_prs.add(pr)
    
    print(f"Fetching info for {len(unique_prs)} unique PRs...")
    
    pr_info = {}
    for i, pr_key in enumerate(unique_prs):
        info = get_pr_info(pr_key)
        pr_info[pr_key] = info
        print(f"[{i + 1}/{len(unique_prs)}] {pr_key}: {info.get('owner')}, {info.get('state')}")
    
    return pr_info


def update_excel(excel_file, all_pr_refs, pr_info):
    """Update Excel with PR information"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Issues']
    
    # Build issue to row mapping
    row_by_issue = {}
    for row_idx in range(2, ws.max_row + 1):
        issue_id = ws.cell(row=row_idx, column=1).value
        if issue_id is not None:
            row_by_issue[issue_id] = row_idx
    
    # Update PR columns
    for issue_num, pr_list in all_pr_refs.items():
        if issue_num in row_by_issue:
            row = row_by_issue[issue_num]
            
            pr_urls = []
            pr_owners = []
            pr_states = []
            
            for pr_key in pr_list:
                if pr_key in pr_info:
                    info = pr_info[pr_key]
                    if 'error' not in info:
                        pr_urls.append(info.get('url', ''))
                        pr_owners.append(info.get('owner', ''))
                        pr_states.append(info.get('state', ''))
            
            # Write semicolon-separated values
            ws.cell(row=row, column=15).value = ';'.join(pr_urls)
            ws.cell(row=row, column=16).value = ';'.join(pr_owners)
            ws.cell(row=row, column=17).value = ';'.join(pr_states)
    
    wb.save(excel_file)
    print(f"Updated {len(all_pr_refs)} issues in {excel_file}")


def main():
    parser = argparse.ArgumentParser(description="Extract PRs from GitHub issue comments")
    parser.add_argument("excel_file", help="Path to Excel file with issues")
    parser.add_argument("--issues", type=str, default="", help="Comma-separated list of issue IDs to process (default: all)")
    args = parser.parse_args()
    
    excel_file = args.excel_file
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} does not exist")
        sys.exit(1)
    
    # Parse target issue IDs
    target_issue_ids = None
    if args.issues:
        target_issue_ids = set()
        for part in args.issues.split(','):
            part = part.strip()
            if part:
                try:
                    target_issue_ids.add(int(part))
                except ValueError:
                    pass
        print(f"Target issues: {sorted(target_issue_ids)}")
    
    # Step 1: Extract PR refs from comments
    print("Step 1: Extracting PRs from GitHub comments...")
    all_pr_refs = extract_prs_from_excel(excel_file, target_issue_ids)
    
    if not all_pr_refs:
        print("No PRs found in comments")
        return
    
    # Save intermediate data
    with open('/tmp/pr_refs.json', 'w') as f:
        json.dump(all_pr_refs, f)
    
    # Step 2: Fetch PR info
    print("\nStep 2: Fetching PR information...")
    pr_info = fetch_pr_info(all_pr_refs)
    
    # Save PR info
    with open('/tmp/pr_info.json', 'w') as f:
        json.dump(pr_info, f, indent=2)
    
    # Step 3: Update Excel
    print("\nStep 3: Updating Excel...")
    update_excel(excel_file, all_pr_refs, pr_info)
    
    print("\nDone!")


if __name__ == "__main__":
    main()