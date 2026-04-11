#!/usr/bin/env python3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import os
import requests

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = "/home/daisydeng"
RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")
DATA_DIR = os.path.join(ROOT_DIR, "issue_traige", "data")
DOC_DIR = os.path.join(ROOT_DIR, "issue_traige", "doc")

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_HEADERS = {
    "Accept": "application/vnd.github.v3+json",
    "Authorization": f"token {GITHUB_TOKEN}"
} if GITHUB_TOKEN else {}

# Load data - try to load from JSON, or fetch from GitHub if not exists
issues_json_path = os.path.join(DATA_DIR, "torch_xpu_ops_issues.json")
comments_json_path = os.path.join(DATA_DIR, "torch_xpu_ops_comments.json")

if os.path.exists(issues_json_path):
    with open(issues_json_path) as f:
        issues = json.load(f)
else:
    print("Fetching issues from GitHub...")
    issues = []
    page = 1
    while len(issues) < 500:
        url = f"https://api.github.com/repos/intel/torch-xpu-ops/issues?state=open&per_page=100&page={page}"
        response = requests.get(url, headers=GITHUB_HEADERS, timeout=30)
        if response.status_code != 200:
            break
        batch = response.json()
        if not batch:
            break
        # Filter out pull requests
        issues.extend([i for i in batch if 'pull_request' not in i])
        print(f"Fetched {len(issues)} issues...")
        page += 1
    with open(issues_json_path, 'w') as f:
        json.dump(issues, f)

if os.path.exists(comments_json_path):
    with open(comments_json_path) as f:
        comments = json.load(f)
else:
    print("Fetching comments from GitHub...")
    comments = {}
    for issue in issues:
        issue_num = issue.get('number')
        url = f"https://api.github.com/repos/intel/torch-xpu-ops/issues/{issue_num}/comments"
        response = requests.get(url, headers=GITHUB_HEADERS, timeout=30)
        if response.status_code == 200:
            comments[str(issue_num)] = response.json()
    with open(comments_json_path, 'w') as f:
        json.dump(comments, f)

# Load ops dependency
ops_dep = {}
with open(os.path.join(DOC_DIR, "ops_dependency.csv")) as f:
    next(f)
    for line in f:
        parts = line.strip().split(',')
        if len(parts) >= 2:
            op_name = parts[0].strip().lower()
            ops_dep[op_name] = parts[1].strip()

# Known test types
KNOWN_TEST_TYPES = ['op_ut', 'op_extend', 'e2e', 'benchmark', 'ut']

# Model lists from benchmarks
HUGGINGFACE_MODELS = [
    'AlbertForMaskedLM', 'AlbertForQuestionAnswering', 'AllenaiLongformerBase',
    'BartForCausalLM', 'BartForConditionalGeneration', 'BertForMaskedLM',
    'BertForQuestionAnswering', 'BlenderbotForCausalLM', 'BlenderbotForConditionalGeneration',
    'BlenderbotSmallForCausalLM', 'BlenderbotSmallForConditionalGeneration', 'CamemBert',
    'DebertaV2ForMaskedLM', 'DebertaV2ForQuestionAnswering', 'DistilBertForMaskedLM',
    'DistilBertForQuestionAnswering', 'DistillGPT2', 'ElectraForCausalLM',
    'ElectraForQuestionAnswering', 'GoogleFnet', 'google/gemma', 'GPT2ForSequenceClassification',
    'GPTJForCausalLM', 'GPTJForQuestionAnswering', 'GPTNeoForCausalLM',
    'GPTNeoForSequenceClassification', 'LayoutLMForMaskedLM', 'LayoutLMForSequenceClassification',
    'M2M100ForConditionalGeneration', 'MBartForCausalLM', 'MBartForConditionalGeneration',
    'MegatronBertForCausalLM', 'MegatronBertForQuestionAnswering', 'meta-llama', 'mistralai',
    'MobileBertForMaskedLM', 'MobileBertForQuestionAnswering', 'MT5ForConditionalGeneration',
    'openai/gpt', 'openai/whisper', 'OPTForCausalLM', 'PegasusForCausalLM',
    'PegasusForConditionalGeneration', 'PLBartForCausalLM', 'PLBartForConditionalGeneration',
    'Qwen', 'RobertaForCausalLM', 'RobertaForQuestionAnswering', 'T5ForConditionalGeneration',
    'T5Small', 'TrOCRForCausalLM', 'XGLMForCausalLM', 'XLNetLMHeadModel', 'YituTechConvBert',
    'hf_Albert', 'hf_Bert', 'hf_Bert_large', 'hf_DistilBert', 'hf_Roberta_base'
]

TIMM_MODELS = [
    'adv_inception_v3', 'beit_base_patch16_224', 'botnet26t_256', 'cait_m36_384',
    'coat_lite_mini', 'convit_base', 'convmixer_768_32', 'convnext_base',
    'convnextv2_nano', 'crossvit_9_240', 'cspdarknet53', 'deit_base_distilled_patch16_224',
    'deit_tiny_patch16_224', 'dla102', 'dm_nfnet_f0', 'dpn107', 'eca_botnext26ts_256',
    'eca_halonext26ts', 'ese_vovnet19b_dw', 'fbnetc_100', 'fbnetv3_b', 'gernet_l',
    'ghostnet_100', 'gluon_inception_v3', 'gmixer_24_224', 'gmlp_s16_224', 'hrnet_w18',
    'inception_v3', 'jx_nest_base', 'lcnet_050', 'levit_128', 'mixer_b16_224',
    'mixnet_l', 'mnasnet_100', 'mobilenetv2_100', 'mobilenetv3_large_100', 'mobilevit_s',
    'nfnet_l0', 'pit_b_224', 'pnasnet5large', 'poolformer_m36', 'regnety_002',
    'repvgg_a2', 'res2net101_26w_4s', 'res2net50_14w_8s', 'res2next50', 'resmlp_12_224',
    'resnest101e', 'rexnet_100', 'sebotnet33ts_256', 'selecsls42b', 'spnasnet_100',
    'swin_base_patch4_window7_224', 'swsl_resnext101_32x16d', 'tf_efficientnet_b0',
    'tf_mixnet_l', 'tinynet_a', 'tnt_s_patch16_224', 'twins_pcpvt_base', 'visformer_small',
    'vit_base_patch14_dinov2', 'vit_base_patch16_224', 'vit_base_patch16_siglip_256',
    'volo_d1_224', 'xcit_large_24_p8_224', 'timm_vision_transformer', 'timm_vision_transformer_large'
]

TORCHBENCH_MODELS = [
    'BERT_pytorch', 'Background_Matting', 'LearningToPaint', 'alexnet', 'dcgan',
    'densenet121', 'mnasnet1_0', 'mobilenet_v2', 'mobilenet_v3_large',
    'nvidia_deeprecommender', 'pytorch_unet', 'resnet18', 'resnet50',
    'resnext50_32x4d', 'shufflenet_v2_x1_0', 'squeezenet1_1', 'vgg16'
]

def identify_benchmark(model_name):
    """Identify benchmark from model name"""
    model_lower = model_name.lower()
    
    for m in HUGGINGFACE_MODELS:
        if m.lower() in model_lower or model_lower in m.lower():
            return 'huggingface'
    
    for m in TIMM_MODELS:
        if m.lower() in model_lower or model_lower in m.lower():
            return 'timm'
    
    for m in TORCHBENCH_MODELS:
        if m.lower() in model_lower or model_lower in m.lower():
            return 'torchbench'
    
    return 'unknown'

def extract_e2e_reproducer(body, title):
    """Extract reproducer command from issue body"""
    text = f"{title} {body}"
    
    reproducer_lines = []
    
    # Look for code blocks with commands (between ``` and ```)
    if '```' in text:
        parts = text.split('```')
        for i, part in enumerate(parts):
            # Code blocks are odd-indexed (1, 3, 5, ...)
            if i % 2 == 1:  # This is a code block content
                part_stripped = part.strip()
                if part_stripped:
                    lines = part_stripped.split('\n')
                    for line in lines:
                        line_stripped = line.strip()
                        # Look for actual commands (python, pytest, etc.)
                        if line_stripped and (line_stripped.startswith(('python', 'pytest', 'XPU_', './')) or 'python' in line_stripped.lower()):
                            if not line_stripped.startswith('#'):
                                reproducer_lines.append(line_stripped)
                    # If we found a command, use it
                    if reproducer_lines:
                        break
    
    # Also look for command patterns without code blocks
    if not reproducer_lines:
        # Look for python or pytest command patterns
        cmd_patterns = [
            r'(pytest\s+[^\n]+)',
            r'(python\s+test/[^\n]+)',
            r'(python\s+-m\s+pytest[^\n]+)',
            r'(XPU_QUANT_CONFIG=[^\n]+python[^\n]+)',
            r'(python\s+benchmarks/dynamo/[^\n]+)',
            r'(python\s+[^\n]+run_benchmark[^\n]+)',
        ]
        
        for pattern in cmd_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                reproducer_lines.append(match.strip())
    
    if not reproducer_lines:
        # Generic reproducer from title
        return title[:200]
    
    # Join and limit to 3 lines
    return '\n'.join(reproducer_lines[:3])


def parse_e2e_info(body, title):
    """Parse e2e benchmark information from issue body"""
    e2e_info = []
    
    text = f"{title} {body}"
    
    # Get reproducer
    reproducer = extract_e2e_reproducer(body, title)
    
    # Check for model names in title or body
    all_model_names = HUGGINGFACE_MODELS + TIMM_MODELS + TORCHBENCH_MODELS
    
    # Extract phase (training/inference)
    phase = 'inference'
    if 'training' in text.lower():
        phase = 'training'
    elif 'train' in text.lower():
        phase = 'training'
    
    # Extract dtype
    dtype = 'float32'
    dtype_patterns = [
        (r'bfloat16|bf16', 'bfloat16'),
        (r'float16|fp16', 'float16'),
        (r'float32|fp32', 'float32'),
        (r'int8|int\s*8', 'int8'),
    ]
    for pattern, dt in dtype_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            dtype = dt
            break
    
    # Extract AMP (automatic mixed precision)
    amp = False
    if '--amp' in text.lower() or 'amp' in text.lower():
        amp = True
    
    # Extract test type
    test_type = 'accuracy'
    if 'throughputs' in text.lower() or 'performance' in text.lower() or 'latency' in text.lower():
        test_type = 'performance'
    
    # Extract backend
    backend = 'inductor'
    if '--backend=(\w+)' in text:
        match = re.search(r'--backend=(\w+)', text)
        if match:
            backend = match.group(1)
    elif 'eager' in text.lower():
        backend = 'eager'
    elif 'inductor' in text.lower():
        backend = 'inductor'
    
    # Extract disable-cudagraphs
    disable_cudagraphs = 'no'
    if 'disable-cudagraphs' in text.lower() or 'disable_cudagraphs' in text.lower():
        disable_cudagraphs = 'yes'
    
    # Find model in body - need exact model name, not partial match
    found_models = set()
    for model in all_model_names:
        # Use word boundary to avoid partial matches
        if re.search(r'\b' + re.escape(model.lower()) + r'\b', text.lower()):
            benchmark = identify_benchmark(model)
            if benchmark != 'unknown' and model not in found_models:
                found_models.add(model)
                e2e_info.append({
                    'reproducer': reproducer,
                    'benchmark': benchmark,
                    'model': model,
                    'phase': phase,
                    'dtype': dtype,
                    'amp': amp,
                    'test_type': test_type,
                    'backend': backend,
                    'disable_cudagraphs': disable_cudagraphs,
                })
    
    # If no specific model found but looks like e2e issue
    if not e2e_info:
        if 'benchmark' in text.lower() or 'huggingface' in text.lower() or 'timm' in text.lower() or 'torchbench' in text.lower():
            # Try to identify benchmark from context
            if 'hf_' in text.lower() or 'huggingface' in text.lower():
                benchmark = 'huggingface'
            elif 'timm_' in text.lower() or 'timm.' in text.lower():
                benchmark = 'timm'
            elif 'torchbench' in text.lower():
                benchmark = 'torchbench'
            else:
                benchmark = 'unknown'
            
            e2e_info.append({
                'reproducer': reproducer,
                'benchmark': benchmark,
                'model': 'unknown',
                'phase': phase,
                'dtype': dtype,
                'test_type': test_type,
                'backend': backend,
                'disable_cudagraphs': disable_cudagraphs,
            })
    
    return e2e_info

# Test case patterns and torch_ops (from previous version)
TEST_CASE_OPS = {
    'test_alias': 'aten.alias',
    'test_retain_autograd': 'aten.retain_autograd',
    'test_cross_entropy': 'aten.cross_entropy',
    'test_nll_loss': 'aten.nll_loss',
    'test_mse_loss': 'aten.mse_loss',
    'test_l1_loss': 'aten.l1_loss',
    'test_layer_norm': 'aten.layer_norm',
    'test_rms_norm': 'aten.rms_norm',
    'test_group_norm': 'aten.group_norm',
    'test_batch_norm': 'aten.batch_norm',
    'test_linear': 'aten.linear',
    'test_conv': 'aten.conv',
    'test_conv2d': 'aten.conv2d',
    'test_conv3d': 'aten.conv3d',
    'test_matmul': 'aten.matmul',
    'test_mm': 'aten.mm',
    'test_bmm': 'aten.bmm',
    'test_addmm': 'aten.addmm',
    'test_dot': 'aten.dot',
    'test_vdot': 'aten.vdot',
    'test_embedding': 'aten.embedding',
    'test_softmax': 'aten.softmax',
    'test_gelu': 'aten.gelu',
    'test_relu': 'aten.relu',
    'test_scaled_dot_product': 'aten.scaled_dot_product_attention',
    'test_sdpa': 'aten.scaled_dot_product_attention',
    'test_flash_attention': 'aten._flash_attention_forward',
    'test_index_add': 'aten.index_add',
    'test_index_copy': 'aten.index_copy',
    'test_gather': 'aten.gather',
    'test_scatter': 'aten.scatter',
    'test_topk': 'aten.topk',
    'test_sort': 'aten.sort',
    'test_argsort': 'aten.argsort',
    'test_chunk': 'aten.chunk',
    'test_split': 'aten.split',
    'test_stack': 'aten.stack',
    'test_cat': 'aten.cat',
    'test_flatten': 'aten.flatten',
    'test_reshape': 'aten.reshape',
    'test_view': 'aten.view',
    'test_transpose': 'aten.transpose',
    'test_permute': 'aten.permute',
    'test_contiguous': 'aten.contiguous',
    'test_clone': 'aten.clone',
    'test_copy': 'aten.copy',
    'test_bernoulli': 'aten.bernoulli',
    'test_normal': 'aten.normal',
    'test_randn': 'aten.randn',
    'test_randint': 'aten.randint',
    'test_rand': 'aten.rand',
    'test_zeros': 'aten.zeros',
    'test_ones': 'aten.ones',
    'test_full': 'aten.full',
    'test_arange': 'aten.arange',
    'test_linspace': 'aten.linspace',
    'test_logspace': 'aten.logspace',
    'test_empty': 'aten.empty',
    'test_eye': 'aten.eye',
    'test_scalar_tensor': 'aten.scalar_tensor',
    'test_tensor': 'aten.tensor',
    'test_as_tensor': 'aten.as_tensor',
    'test_from_numpy': 'aten.from_numpy',
    'test_sparse': 'sparse',
    'test_coalesce': 'aten.coalesce',
    'test_pdist': 'aten.pdist',
    'test_cdist': 'aten.cdist',
    'test_triangular_solve': 'aten.triangular_solve',
    'test_cholesky': 'aten.cholesky',
    'test_qr': 'aten.linalg_qr',
    'test_solve': 'aten.linalg_solve',
    'test_inv': 'aten.linalg_inv',
    'test_det': 'aten.linalg_det',
    'test_slogdet': 'aten.linalg_slogdet',
    'test_eig': 'aten.linalg_eig',
    'test_eigh': 'aten.linalg_eigh',
    'test_svd': 'aten.linalg_svd',
    'test_lu': 'aten.linalg_lu',
    'test_block_addmv': 'torch.addmv',
    'test_block_addmm': 'torch.addmm',
    'test_block_triangular_solve': 'aten.triangular_solve.X',
    'test_scaled_dot_product': 'scaled_dot_product_attention',
    'test_sdpa': 'scaled_dot_product_attention',
    'test_flash_attention': '_flash_attention_forward',
    'test_triton_bsr': 'triton_bsr',
    'test_triton_scaled': 'triton_scaled_dot_product_attention',
    'test_triton': 'triton',
    'test_block': 'torch.addmm',
    'test_baddbmm': 'aten.baddbmm',
    'test_bmm': 'aten.bmm',
    'test_mm': 'aten.mm',
    'test_addmm': 'aten.addmm',
    'test_addmv': 'torch.addmv',
    'test_matmul': 'aten.matmul',
    'test_sparse': 'sparse',
    'test_triton_bsr_dense': 'triton_bsr_dense_addmm',
    'test_sampled_addmm': 'sampled_addmm',
    'test_triton_sampled': 'sampled_addmm',
    'test_conv': 'aten.conv',
    'test_conv2d': 'aten.conv2d',
    'test_conv_transpose': 'aten.conv_transpose',
    'test_norm': 'aten.layer_norm',
    'index_add': 'aten.index_add',
    'index_copy': 'aten.index_copy',
    'layer_norm': 'aten.layer_norm',
    'rms_norm': 'aten.rms_norm',
    'cross_entropy': 'aten.cross_entropy',
    'embedding': 'aten.embedding',
    'linear': 'aten.linear',
    'conv2d': 'aten.conv2d',
    'matmul': 'aten.matmul',
    'bmm': 'aten.bmm',
    'mm': 'aten.mm',
    'addmm': 'aten.addmm',
    'dot': 'aten.dot',
    'vdot': 'aten.vdot',
    'gelu': 'aten.gelu',
    'relu': 'aten.relu',
    'softmax': 'aten.softmax',
    'batch_norm': 'aten.batch_norm',
    'group_norm': 'aten.group_norm',
    'gather': 'aten.gather',
    'scatter': 'aten.scatter',
    'topk': 'aten.topk',
    'sort': 'aten.sort',
    'cholesky': 'aten.cholesky',
    'qr': 'aten.linalg_qr',
    'svd': 'aten.linalg_svd',
    'solve': 'aten.linalg_solve',
    'inverse': 'aten.linalg_inv',
    'det': 'aten.linalg_det',
}

DIRECT_OPS = [
    'scaled_dot_product_attention',
    '_flash_attention_forward',
    '_efficient_attention_forward',
    'index_add',
    'index_copy',
    'layer_norm',
    'rms_norm',
    'cross_entropy',
    'embedding',
    'linear',
    'matmul',
    'bmm',
    'mm',
    'addmm',
    'dot',
    'vdot',
    'gelu',
    'relu',
    'softmax',
    'batch_norm',
    'group_norm',
    'gather',
    'scatter',
    'topk',
    'sort',
    'cholesky',
    'qr',
    'svd',
    'solve',
]

def infer_torch_ops_from_test_case(test_case, body, error_msg="", traceback=""):
    ops = []
    all_text = f"{test_case} {body} {error_msg} {traceback}".lower()
    full_text = f"{test_case} {body} {error_msg} {traceback}"
    
    if not test_case:
        for direct_op in DIRECT_OPS:
            if direct_op.lower() in all_text:
                ops.append(direct_op)
                break
        return ops
    
    test_case_lower = test_case.lower()
    for pattern, op in TEST_CASE_OPS.items():
        if pattern in test_case_lower:
            ops.append(op)
            break
    
    if not ops:
        for pattern, op in TEST_CASE_OPS.items():
            if pattern in all_text:
                ops.append(op)
                break
    
    if not ops:
        for direct_op in DIRECT_OPS:
            if direct_op.lower() in all_text:
                ops.append(direct_op)
                break
    
    if not ops:
        aten_pattern = re.findall(r'aten::([^\s\'"]+)', full_text)
        for a_op in aten_pattern:
            clean_op = a_op.strip()
            if clean_op.startswith('_'):
                full_op = f'aten::{clean_op}'
            elif '.' in clean_op:
                full_op = f'aten.{clean_op}'
            else:
                full_op = f'aten.{clean_op}'
            if full_op not in ops:
                ops.append(full_op)
    
    if not ops:
        aten_pattern2 = re.findall(r'aten\.(\w+)', full_text)
        for a_op in aten_pattern2:
            full_op = f'aten.{a_op}'
            if full_op not in ops:
                ops.append(full_op)
    
    return ops

def map_origin_test_file(test_file):
    if not test_file:
        return ""
    match = re.search(r'test/xpu/(.+?)(?:_xpu)?\.py$', test_file)
    if match:
        return f"test/{match.group(1)}.py"
    if 'benchmarks/' in test_file:
        return test_file
    return test_file

def parse_test_cases_from_body(body):
    cases = []
    
    if 'Cases:' in body:
        cases_section = body.split('Cases:')[1]
        
        end_markers = ['\n###', '\nVersions', '\n```', '\n\n']
        min_end = len(cases_section)
        for marker in end_markers:
            idx = cases_section.find(marker)
            if idx > 0 and idx < min_end:
                min_end = idx
        cases_section = cases_section[:min_end]
        
        lines = cases_section.split('\n')
        
        for line in lines:
            line = line.strip()
            
            if not line:
                continue
            
            if line.startswith('###') or line.startswith('...'):
                continue
            
            if line.startswith('~~') and line.endswith('~~'):
                continue
            
            parts = line.split(',')
            if len(parts) < 3:
                continue
            
            test_type = parts[0].strip()
            if test_type not in KNOWN_TEST_TYPES:
                continue
            
            test_path = parts[1].strip()
            test_case = parts[2].strip()
            
            if not test_path or not test_case or len(test_case) < 3:
                continue
            
            if ' ' in test_case:
                continue
            
            if 'torch-xpu-ops' in test_path:
                path_parts = test_path.split('.')
                try:
                    txpo_idx = path_parts.index('torch-xpu-ops')
                    rel_parts = path_parts[txpo_idx+1:]
                    
                    test_class = ""
                    test_file_parts = []
                    for part in rel_parts:
                        if part.startswith('Test'):
                            test_class = part
                            break
                        test_file_parts.append(part)
                    
                    if test_file_parts:
                        test_file = 'torch-xpu-ops/' + '/'.join(test_file_parts) + '.py'
                        if not test_file.endswith('_xpu.py'):
                            test_file = test_file.replace('.py', '_xpu.py')
                    else:
                        test_file = ""
                    
                    origin_file = map_origin_test_file(test_file)
                except:
                    test_file = ""
                    test_class = ""
                    origin_file = ""
            else:
                test_file = test_path
                origin_file = ""
                test_class = ""
            
            cases.append({
                'test_type': test_type,
                'test_file': test_file,
                'origin_test_file': origin_file,
                'test_class': test_class,
                'test_case': test_case
            })
    
    if 'benchmarks/dynamo/' in body:
        matches = re.findall(r'(python\s+benchmarks/dynamo/[^\s]+)', body)
        for match in matches:
            test_file = match.replace('python ', '').strip()
            cases.append({
                'test_type': 'e2e',
                'test_file': test_file,
                'origin_test_file': test_file,
                'test_class': '',
                'test_case': match.strip()
            })
    
    if 'pytest' in body:
        k_match = re.search(r'pytest[^-]*(-k\s+[^\s]+)?', body)
        if k_match and k_match.group(1):
            cases.append({
                'test_type': 'ut',
                'test_file': '',
                'origin_test_file': '',
                'test_class': '',
                'test_case': k_match.group(1).strip()
            })
    
    return cases

def extract_error_and_traceback(body):
    error_msg = ""
    traceback = ""
    
    error_patterns = [
        r'(AssertionError[^\n]*)',
        r'(RuntimeError[^\n]*)',
        r'(ValueError[^\n]*)',
        r'(TypeError[^\n]*)',
        r'(IndexError[^\n]*)',
        r'(KeyError[^\n]*)',
        r'(ImportError[^\n]*)',
        r'(NotImplementedError[^\n]*)',
        r'(AttributeError[^\n]*)',
        r'(InductorError[^\n]*)',
    ]
    
    for pattern in error_patterns:
        match = re.search(pattern, body, re.IGNORECASE)
        if match:
            error_msg = match.group(1).strip()[:200]
            break
    
    if 'Traceback (most recent call last):' in body:
        start = body.find('Traceback (most recent call last):')
        end = body.find('\n###', start)
        if end == -1:
            end = body.find('\n\n', start)
        if end == -1:
            end = body.find('\nVersions', start)
        if end == -1:
            end = min(start + 2000, len(body))
        traceback = body[start:end].strip()
    
    elif re.search(r'_{5,}\s+Test\w+', body):
        match = re.search(r'_{5,}\s+Test\w+.*?(?=_{5,}|\n###|\nVersions|\Z)', body, re.DOTALL)
        if match:
            traceback = match.group(0).strip()[:2000]
    
    if not traceback:
        lines = body.split('\n')
        in_trace = False
        trace_lines = []
        for i, line in enumerate(lines):
            if 'File "' in line and '.py"' in line:
                in_trace = True
            if in_trace:
                trace_lines.append(line)
                if any(e in line for e in ['Error:', 'Exception:', 'raise ']) and i > 3:
                    break
        if trace_lines:
            traceback = '\n'.join(trace_lines)[:2000]
    
    return error_msg, traceback

def generate_summary(body, title, error_msg):
    if 'Error' in title or 'Exception' in title:
        match = re.search(r'(Error|Exception):\s*(.+)', title)
        if match:
            return match.group(2).strip()[:150]
        else:
            return title[:150]
    else:
        return title[:150]

def classify_issue_type(body, title, labels):
    text = f"{title} {body}".lower()
    
    for label in labels:
        ln = label.get('name', '').lower()
        if 'task' == ln or 'internal task' in ln:
            return 'internal task'
    
    performance_keywords = [
        'performance regression', 'performance dropped', 'performance issue',
        'latency', 'throughput', 'slow performance', 'performance slow',
        'execution time', 'runtime performance', 'performance fail'
    ]
    
    has_performance_keyword = any(k in text for k in performance_keywords)
    
    bug_keywords = [
        'assertionerror', 'runtimeerror', 'valueerror', 'typeerror', 'indexerror',
        'keyerror', 'importerror', 'notimplementederror', 'attributeerror',
        'inductorerror', 'crash', 'fail', 'bug', 'error', 'not implemented',
        'not supported', 'missing', 'incorrect', 'wrong', 'unexpected'
    ]
    
    has_bug_keyword = any(k in text for k in bug_keywords)
    
    feature_keywords = ['feature request', 'support for', 'implement', 'add support', 'need feature']
    has_feature_keyword = any(k in text for k in feature_keywords)
    
    if has_feature_keyword:
        return 'feature request'
    
    if has_performance_keyword:
        return 'performance issue'
    
    if has_bug_keyword:
        return 'functionality bug'
    
    return 'unknown'

def is_e2e_issue(body, title, labels):
    """Check if issue is related to E2E benchmark"""
    text = f"{title} {body}".lower()
    
    # Check labels first - only exact 'e2e' label
    for label in labels:
        ln = label.get('name', '').lower()
        if ln == 'e2e':
            return True
    
    # Check for specific E2E benchmark paths (not just the word 'benchmark')
    e2e_patterns = [
        r'benchmarks/dynamo/',           # torch-xpu-ops benchmark scripts
        r'benchmarks/timm/',             # timm benchmark
        r'benchmarks/huggingface/',     # huggingface benchmark
        r'benchmarks/torchbench/',      # torchbench benchmark
        r'run_benchmark\.py',            # torchbenchmark runner
    ]
    
    for pattern in e2e_patterns:
        if re.search(pattern, text):
            return True
    
    # Check for model names from benchmark model lists with explicit benchmark framework mention
    # Only for specific benchmark prefixes
    benchmark_model_prefixes = ['hf_', 'timm_']  # e.g., hf_Albert, timm_resnet50
    
    has_model = False
    has_benchmark_context = False
    
    for prefix in benchmark_model_prefixes:
        if prefix in text:
            has_model = True
            break
    
    # Must have explicit benchmark framework mention (as test framework)
    if has_model:
        benchmark_paths = ['benchmarks/dynamo', 'run_benchmark', 'torchbenchmark', 'benchmark.py']
        for kw in benchmark_paths:
            if kw in text:
                has_benchmark_context = True
                break
    
    if has_model and has_benchmark_context:
        return True
    
    return False


def classify_test_module(body, title, labels):
    text = f"{title} {body}".lower()
    
    # Check if it's an E2E issue first
    if is_e2e_issue(body, title, labels):
        return 'e2e'
    
    pytest_patterns = [
        r'pytest\s+.*test[/._]',
        r'python\s+.*test[/._]',
        r'test/test_',
        r'test/xpu/test_',
    ]
    
    has_test_pattern = False
    for pattern in pytest_patterns:
        if re.search(pattern, text):
            has_test_pattern = True
            break
    
    build_patterns = [
        r'\[win\]\[build\]',
        r'build from source',
        r'compile from source', 
        r'source build',
        r'build script',
        r'BUILD_SEPARATE',
        r'BUILD_SHARED',
        r'cmake build',
        r'cmake error',
        r'cmake fail',
        r'setup\.py install',
        r'pip install -e \.',
        r'python setup\.py develop',
    ]
    
    has_build = any(re.search(p, text, re.IGNORECASE) for p in build_patterns)
    
    infra_patterns = [
        r'workflow\s+(error|fail|issue|problem)',
        r'github\s+action\s+(error|fail|issue)',
        r'azure\s+pipeline\s+(error|fail)',
        r'ci\s+(runner|config|setup)\s+(error|fail)',
        r'runner\s+(error|fail|timeout)\s+in\s+ci',
        r'checkout\s+(error|fail)\s+in\s+(workflow|ci)',
        r'githubaction',
    ]
    
    has_infra = any(re.search(p, text) for p in infra_patterns)
    
    for label in labels:
        ln = label.get('name', '').lower()
        if 'infrastructure' in ln and ('ci' in ln or 'workflow' in ln or 'action' in ln):
            has_infra = True
            break
    
    if has_build:
        return 'build'
    
    if has_infra:
        return 'infrastructure'
    
    if has_test_pattern:
        if 'benchmarks/dynamo/' in text or 'benchmark' in text:
            return 'e2e'
        return 'ut'
    
    return 'ut'

def classify_module(body, title, labels):
    text = f"{title} {body}".lower()
    labels_str = ', '.join([l.get('name', '') for l in labels]).lower()
    
    # Check labels first
    for label in labels:
        ln = label.get('name', '').lower()
        if 'module: distributed' in ln:
            return 'distributed'
        if 'module: inductor' in ln:
            return 'inductor'
        if 'module: ao' in ln:
            return 'AO'
        if 'module: ut' in ln:
            return 'aten_ops'
        if 'module: quant' in ln:
            return 'low_precision'
        if 'module: profiler' in ln:
            return 'profiling'
        if 'module: dynamo' in ln:
            return 'dynamo'
        if 'module: op impl' in ln:
            return 'aten_ops'
    
    # Special case - "Torch not compiled with CUDA enabled" means test configuration issue, not inductor
    if 'torch not compiled with cuda enabled' in text:
        return 'unknown'
    
    # Random failures are not module-specific
    if 'random failure' in text or 'random failures' in text:
        return 'unknown'
    
    # Torch operations (from PyTorch docs)
    torch_ops = [
        'add', 'sub', 'mul', 'div', 'matmul', 'mm', 'dot', 'vdot', 'bmm',
        'addmm', 'addmv', 'addbmm', 'smm', 'spmm', 'mm', 'mv', 'vecdot',
        'conv', 'conv1d', 'conv2d', 'conv3d', 'conv_transpose',
        'batch_norm', 'layer_norm', 'group_norm', 'instance_norm',
        'dropout', 'embedding', 'linear', 'lstm', 'gru', 'rnn',
        'softmax', 'log_softmax', 'sigmoid', 'tanh', 'relu', 'leaky_relu',
        'pool', 'avg_pool', 'max_pool', 'adaptive_pool',
        'fft', 'ifft', 'fft2', 'ifft2',
        'chunk', 'split', 'view', 'reshape', 'transpose', 'permute',
        'cat', 'stack', 'gather', 'scatter', 'index', 'where',
        'sum', 'mean', 'std', 'var', 'min', 'max', 'argmin', 'argmax',
        'norm', 'linalg.norm', 'linalg.matrix_norm', 'linalg.vector_norm',
        'eig', 'svd', 'qr', 'cholesky', 'solve', 'inverse',
        'det', 'logdet', 'slogdet', 'trace',
        'clone', 'copy_', 'to', 'cuda', 'cpu', 'xpu', 'device',
        'zeros', 'ones', 'empty', 'full', 'arange', 'linspace', 'logspace',
        'tensor', 'scalar_tensor', 'tensor.tensor',
        'getitem', 'setitem', 'call', 'forward', 'backward',
        'relu', 'gelu', 'silu', 'mish', 'softplus', 'elu', 'selu', 'celu',
        'flash_attention', 'scaled_dot_product_attention', 'sdpa',
        'interpolate', 'grid_sample', 'affine_grid',
        'grid_sampler', 'grid_sampler_2d',
        'bernoulli', 'normal', 'uniform', 'randn', 'rand', 'randint',
        'multinomial', ' poisson', 'exponential', 'geometric',
        'lerp', 'lerp_', 'fmod', 'remainder', 'nextafter',
        'linspace', 'logspace', 'geomspace',
        'complex', 'real', 'imag', 'angle',
        'conj', 'view_as_real', 'view_as_complex',
    ]
    
    module_keywords = [
        ('distributed', ['distributed', 'device_mesh', 'ProcessGroup', 'FSDP', 'DDP', 'c10d', 'tensor parallel']),
        ('inductor', ['inductor', 'inductor error', 'compile error', 'lower', 'kernel code']),
        ('dynamo', ['dynamo', 'torch.compile', '_dynamo', 'dynamo']),
        ('autograd', ['autograd', 'backward', 'grad', 'gradient']),
        ('aten_ops', ['aten::', 'torch.ops.aten', 'test_ops']),
        ('low_precision', ['quantization', 'int8', 'fp8', 'int4', 'amp', 'bf16', 'fp16', 'float8']),
        ('optimizer', ['optimizer', 'lr_scheduler', 'adam', 'sgd']),
        ('profiling', ['profiling', 'profile', 'benchmark']),
        ('fx', ['torch.fx', 'fx.', 'symbolic']),
        ('export', ['torch.export', 'exported']),
    ]
    
    # Check torch ops first
    for op in torch_ops:
        if re.search(rf'\b{re.escape(op)}\b', text):
            return 'aten_ops'
    
    for m, kw in module_keywords:
        if any(k in text for k in kw):
            return m
    
    return 'unknown'

def get_dependency_from_body(body, labels=None):
    if labels is None:
        labels = []
    
    labels_str = ', '.join([l.get('name', '') for l in labels]).lower()
    
    # Check labels first for 'dependency component:'
    if 'dependency component: onednn' in labels_str or 'dependency component: mkl-dnn' in labels_str or 'dependency component: dnnl' in labels_str:
        return 'oneDNN'
    if 'dependency component: onemkl' in labels_str or 'dependency component: mkl' in labels_str:
        return 'oneMKL'
    if 'dependency component: triton' in labels_str:
        return 'Triton'
    if 'dependency component: torchao' in labels_str:
        return 'AO'
    if 'dependency component: transformers' in labels_str or 'dependency component: huggingface' in labels_str:
        return 'transformers'
    if 'dependency component: oneapi' in labels_str or 'dependency component: sycl' in labels_str:
        return 'oneAPI'
    if 'dependency component: driver' in labels_str:
        return 'driver'
    if 'dependency component: oneccl' in labels_str or 'dependency component: ccl' in labels_str or 'dependency component: xccl' in labels_str:
        return 'oneCCL'
    
    # Filter out version/environment sections
    if not body:
        return 'None'
    
    text = body.lower()
    
    # Remove version/environment sections
    version_headers = [
        r'###\s*version',
        r'###\s*versions',
        r'###\s*environment',
        r'###\s*reproduction',
        r'###\s*steps?\s+to\s+reproduce',
        r'###\s*additional\s*context',
    ]
    
    for header in version_headers:
        match = re.search(header, text, re.IGNORECASE)
        if match:
            text = text[:match.start()]
            break
    
    # Check for actual dependency in body (require context like "caused by", "issue", "depend on")
    dep_keywords = [
        ('transformers', [
            'caused by transformers', 'transformers issue', 'transformers bug',
            'depends on transformers', 'need transformers fix', 'waiting for transformers',
            'huggingface issue', 'huggingface bug', 'depends on huggingface'
        ]),
        ('AO', [
            'caused by torchao', 'torchao issue', 'torchao bug',
            'depends on torchao', 'need torchao fix', 'waiting for torchao'
        ]),
        ('oneDNN', [
            'caused by onednn', 'onednn issue', 'onednn bug', 'oneDNN issue',
            'depends on onednn', 'need onednn fix', 'waiting for onednn',
            'mkl-dnn issue', 'dnnl issue'
        ]),
        ('oneCCL', [
            'caused by oneccl', 'oneccl issue', 'oneccl bug',
            'depends on oneccl', 'need oneccl fix', 'waiting for oneccl',
            'xccl issue', 'ccl issue', 'depends on ccl'
        ]),
        ('oneMKL', [
            'caused by onemkl', 'onemkl issue', 'onemkl bug',
            'depends on onemkl', 'need onemkl fix', 'waiting for onemkl',
            'caused by mkl', 'mkl issue'
        ]),
        ('driver', [
            'caused by driver', 'driver issue', 'driver bug',
            'depends on driver', 'need driver fix', 'waiting for driver'
        ]),
        ('Triton', [
            'caused by triton', 'triton issue', 'triton bug',
            'depends on triton', 'need triton fix', 'waiting for triton',
            'triton-xpu issue', 'tl\\. issue'
        ]),
        ('oneAPI', [
            'caused by oneapi', 'oneapi issue', 'oneapi bug', 'sycl issue',
            'depends on oneapi', 'need oneapi fix', 'waiting for oneapi',
            'icpx issue', 'dpcpp issue', 'sycl compiler issue'
        ]),
    ]
    
    for d, kw in dep_keywords:
        if any(k in text for k in kw):
            return d
    
    return 'None'

def extract_pr_from_text(text, exclude_version_section=True):
    """Extract PR URLs and PR numbers from text, only those that could fix the issue"""
    prs = []
    
    if not text:
        return prs
    
    if exclude_version_section:
        version_headers = [
            r'###\s*Version',
            r'###\s*versions',
            r'###\s*environment',
            r'###\s*reproduction',
            r'###\s*steps?\s+to\s+reproduce',
            r'###\s*additional\s+context',
        ]
        
        for header in version_headers:
            match = re.search(header, text, re.IGNORECASE)
            if match:
                text = text[:match.start()]
                break
    
    fix_context_patterns = [
        r'fixed\s+(?:in|by|with)?\s*PR',
        r'fixed\s+in\s+commit',
        r'resolved\s+(?:in|by)?\s*PR',
        r'resolution:.*PR',
        r'merged\s+in\s+PR',
        r'PR\s*:\s*https?://',  # PR: https://...
        r'PR\s*#\d+',  # PR #1234
        r'pull\s+request\s*#\d+',
        r'landed\s+(?:in|via)?\s*PR',
    ]
    
    fix_context = any(re.search(p, text, re.IGNORECASE) for p in fix_context_patterns)
    
    # Always extract PR URLs from the text (not just fix context)
    pr_url_patterns = [
        (r'https://github\.com/([\w-]+)/([\w-]+)/pull/(\d+)', None),
        (r'https://github\.com/([\w-]+)/([\w-]+)/issues/(\d+)', 'issue'),
    ]
    
    seen = set()
    
    for pattern, url_type in pr_url_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            org, repo, num = match[0], match[1], match[-1]
            pr_num = str(num).strip()
            if pr_num.isdigit() and pr_num not in seen:
                seen.add(pr_num)
                html_url = f"https://github.com/{org}/{repo}/pull/{pr_num}"
                prs.append({
                    'number': pr_num,
                    'html_url': html_url,
                    'source_repo': f"{org}/{repo}"
                })
    
    if fix_context:
        bare_pr_patterns = [
            r'PR\s*#(\d+)',
            r'pull\s+request\s*#(\d+)',
            r'#(\d+)\s+(?:to|fix|resolve|address|close)',
        ]
        
        for pattern in bare_pr_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                pr_num = str(match).strip()
                if pr_num.isdigit() and pr_num not in seen:
                    seen.add(pr_num)
                    prs.append({
                        'number': pr_num,
                        'html_url': None,
                        'source_repo': None
                    })
    
    return prs

def validate_pr_and_get_info(pr_list, issue_title="", issue_body=""):
    """Validate PR numbers against repos and get PR info - only include PRs that fix the issue"""
    pr_info_list = []
    
    if not pr_list:
        return pr_info_list
    
    repos_to_try = ["pytorch/pytorch", "intel/torch-xpu-ops"]
    
    issue_text = f"{issue_title} {issue_body}".lower()
    
    keywords_for_fixing = [
        'fix', 'resolve', 'address', 'close', 'solved', 'resolved', 
        'bug', 'error', 'patch', 'implement', 'support', 'add', 'enable',
    ]
    
    for pr_item in pr_list[:5]:
        pr_num = pr_item['number']
        
        source_repo = pr_item.get('source_repo')
        if source_repo:
            repos_to_check = [source_repo]
        else:
            repos_to_check = repos_to_try.copy()
        
        pr_found = False
        pr_data = None
        repo = None
        
        for repo in repos_to_check:
            try:
                url = f"https://api.github.com/repos/{repo}/pulls/{pr_num}"
                response = requests.get(url, headers=GITHUB_HEADERS, timeout=10)
                
                if response.status_code == 200:
                    pr_data = response.json()
                    pr_found = True
                    break
            except:
                continue
        
        if not pr_found or not pr_data:
            continue
        
        pr_title = pr_data.get('title', '').lower()
        pr_body = pr_data.get('body', '') or ''
        pr_body = pr_body.lower() if pr_body else ""
        
        pr_text = f"{pr_title} {pr_body}"
        
        # Check if PR is relevant to the issue
        is_fix_pr = any(kw in pr_text for kw in keywords_for_fixing)
        
        title_match = False
        title_keywords = []
        for word in issue_text.split():
            if len(word) > 3 and word in pr_text:
                title_keywords.append(word)
        
        if len(title_keywords) >= 2:
            title_match = True
        
        if not is_fix_pr and not title_match:
            continue
        
        # Check for intel/torch-xpu-ops repo: Skip if "Closed with unmerged commits"
        if repo == "intel/torch-xpu-ops":
            if pr_data.get('state') == 'closed':
                # Check if merged (not unmerged)
                if not pr_data.get('merged'):
                    # Check for unmerged commits
                    commits_url = pr_data.get('commits_url')
                    if commits_url:
                        try:
                            commits_response = requests.get(commits_url, headers=GITHUB_HEADERS, timeout=10)
                            if commits_response.status_code == 200:
                                commits_data = commits_response.json()
                                if commits_data.get('total_count', 0) > 0:
                                    # Has commits, likely unmerged - skip
                                    continue
                        except:
                            pass
        
        # Check for pytorch/pytorch repo: Only include if has "Merged" label
        if repo == "pytorch/pytorch":
            if pr_data.get('state') == 'closed':
                # Check if merged
                if not pr_data.get('merged'):
                    # Check labels
                    labels = pr_data.get('labels', [])
                    label_names = [l.get('name', '') for l in labels]
                    if 'merged' not in [l.lower() for l in label_names]:
                        # Not merged with "Merged" label - skip
                        continue
        
        pr_info_list.append({
            'number': pr_num,
            'state': pr_data.get('state', 'unknown'),
            'user': pr_data.get('user', {}).get('login', 'unknown'),
            'html_url': pr_data.get('html_url', ''),
            'title': pr_data.get('title', ''),
            'source_repo': repo if pr_found else None
        })
    
    return pr_info_list

def get_issue_comments_with_pr(issue_num, repo="intel/torch-xpu-ops"):
    """Get PR references from issue comments - only PRs that could fix the issue"""
    all_prs = []
    
    issue_num_str = str(issue_num)
    if issue_num_str not in comments:
        return all_prs
    
    issue_comments = comments[issue_num_str]
    
    fix_context_patterns = [
        r'fixed\s+(?:in|by|with)?\s*PR',
        r'fixed\s+in\s+commit',
        r'resolved\s+(?:in|by)?\s*PR',
        r'resolution:.*PR',
        r'merged\s+in\s+PR',
        r'PR\s*:\s*https?://',  # PR: https://...
        r'PR\s*#\d+',  # PR #1234
        r'pull\s+request\s*#\d+',
        r'landed\s+(?:in|via)?\s*PR',
    ]
    
    for comment in issue_comments:
        body = comment.get('body', '') or ''
        if not body:
            continue
        
        fix_context = any(re.search(p, body, re.IGNORECASE) for p in fix_context_patterns)
        
        if not fix_context:
            continue
        
        seen = set()
        pr_url_patterns = [
            r'https://github\.com/([\w-]+)/([\w-]+)/pull/(\d+)',
        ]
        
        for pattern in pr_url_patterns:
            matches = re.findall(pattern, body, re.IGNORECASE)
            for match in matches:
                org, repo, num = match[0], match[1], match[-1]
                pr_num = str(num).strip()
                if pr_num.isdigit() and pr_num not in seen:
                    seen.add(pr_num)
                    all_prs.append({
                        'number': pr_num,
                        'html_url': f"https://github.com/{org}/{repo}/pull/{pr_num}",
                        'source_repo': f"{org}/{repo}"
                    })
        
        bare_pr_patterns = [
            r'PR\s*#(\d+)',
            r'pull\s+request\s*#(\d+)',
            r'#(\d+)\s+(?:to|fix|resolve|address|close)',
        ]
        
        for pattern in bare_pr_patterns:
            matches = re.findall(pattern, body, re.IGNORECASE)
            for match in matches:
                pr_num = str(match).strip()
                if pr_num.isdigit() and pr_num not in seen:
                    seen.add(pr_num)
                    all_prs.append({
                        'number': pr_num,
                        'html_url': None,
                        'source_repo': None
                    })
    
    return all_prs

# Create Excel
wb = openpyxl.Workbook()

# Sheet 1: Issues
ws_issues = wb.active
ws_issues.title = "Issues"

headers = ["Issue ID", "Title", "Status", "Assignee", "Reporter", "Labels",
           "Created Time", "Updated Time", "Milestone", "Summary", "Type",
           "Module", "Test Module", "Dependency", "PR", "PR Owner", "PR Status", "PR Description",
           "owner_transfer", "action_TBD", "action_TBD_reason", "duplicated_issue", 
           "priority", "priority_reason", "Category", "category_reason", "Root Cause"]

for col, header in enumerate(headers, 1):
    cell = ws_issues.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 2: Test Cases (ut)
ws_cases = wb.create_sheet("Test Cases")

case_headers = ["Issue ID", "Test Reproducer", "Test Type", "Test File", 
                  "Origin Test File", "Test Class", "Test Case", 
                  "Error Message", "Traceback", "torch-ops", "dependency",
                  "XPU Status", "Stock Status", "Is SKIP", "Is CUDA Skip",
                  "CUDA Case Exist", "XPU Case Exist", "case_existence_comments",
                  "can_enable_on_xpu", "duplicated_issue"]

for col, header in enumerate(case_headers, 1):
    cell = ws_cases.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 3: E2E Test Cases
ws_e2e = wb.create_sheet("E2E Test Cases")

e2e_headers = ["Issue ID", "Test Reproducer", "Benchmark", "Model", "Phase", "Dtype", "AMP",
               "Backend", "Test Type", "Cudagraph", "Error Message", "Traceback"]

for col, header in enumerate(e2e_headers, 1):
    cell = ws_e2e.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

issue_row = 2
case_row = 2
e2e_row = 2

MAX_ISSUES_FOR_TEST = 10
TESTING_MODE = True

for issue in issues[:MAX_ISSUES_FOR_TEST] if TESTING_MODE else issues:
    num = issue['number']
    title = issue['title']
    body = issue.get('body', '') or ''
    
    status = issue['state']
    assignee = ", ".join([a['login'] for a in issue.get('assignees', [])]) or "None"
    reporter = issue['user']['login']
    labels = issue.get('labels', [])
    label_str = ", ".join([l['name'] for l in labels])
    created = issue['created_at']
    updated = issue['updated_at']
    milestone = issue.get('milestone', {})
    milestone_name = milestone.get('title', 'None') if milestone else 'None'
    
    issue_type = classify_issue_type(body, title, labels)
    module = classify_module(body, title, labels)
    test_module = classify_test_module(body, title, labels)
    dependency = get_dependency_from_body(body, labels)
    
    error_msg, traceback = extract_error_and_traceback(body)
    summary = generate_summary(body, title, error_msg)
    
    ws_issues.cell(row=issue_row, column=1, value=num)
    ws_issues.cell(row=issue_row, column=2, value=title)
    ws_issues.cell(row=issue_row, column=3, value=status)
    ws_issues.cell(row=issue_row, column=4, value=assignee)
    ws_issues.cell(row=issue_row, column=5, value=reporter)
    ws_issues.cell(row=issue_row, column=6, value=label_str)
    ws_issues.cell(row=issue_row, column=7, value=created)
    ws_issues.cell(row=issue_row, column=8, value=updated)
    ws_issues.cell(row=issue_row, column=9, value=milestone_name)
    ws_issues.cell(row=issue_row, column=10, value=summary)
    ws_issues.cell(row=issue_row, column=11, value=issue_type)
    ws_issues.cell(row=issue_row, column=12, value=module)
    ws_issues.cell(row=issue_row, column=13, value=test_module)
    ws_issues.cell(row=issue_row, column=14, value=dependency)
    
    pr_numbers = get_issue_comments_with_pr(num)
    
    pr_info_list = validate_pr_and_get_info(pr_numbers[:3], title, body)
    
    pr_str = ", ".join([p['html_url'] for p in pr_info_list if p.get('html_url')])
    pr_owner = ", ".join([p['user'] for p in pr_info_list])
    pr_status = ", ".join([p['state'] for p in pr_info_list])
    pr_desc = ", ".join([p.get('title', '')[:80] for p in pr_info_list if p.get('title')])
    
    ws_issues.cell(row=issue_row, column=15, value=pr_str)
    ws_issues.cell(row=issue_row, column=16, value=pr_owner)
    ws_issues.cell(row=issue_row, column=17, value=pr_status)
    ws_issues.cell(row=issue_row, column=18, value=pr_desc)
    ws_issues.cell(row=issue_row, column=19, value="")
    
    # Parse test cases and e2e info
    test_cases = parse_test_cases_from_body(body)
    
    # Only parse e2e info if it's actually an e2e issue
    e2e_info = []
    if test_module == 'e2e':
        e2e_info = parse_e2e_info(body, title)
    
    # Add to test cases sheet (non-e2e)
    if test_cases:
        for tc in test_cases:
            # Skip e2e cases - they go to e2e sheet
            if tc.get('test_type') == 'e2e':
                continue
                
            test_case = tc.get('test_case', '')
            torch_ops = infer_torch_ops_from_test_case(test_case, body, error_msg, traceback)
            
            if not torch_ops:
                torch_ops = ['unknown']
            
            case_dep = 'sycl'
            if torch_ops[0] != 'unknown':
                for op in torch_ops:
                    op_lower = op.lower().replace('aten.', '')
                    if op_lower in ops_dep:
                        case_dep = ops_dep[op_lower]
                        break
            
            ws_cases.cell(row=case_row, column=1, value=num)
            ws_cases.cell(row=case_row, column=2, value=title[:150])
            ws_cases.cell(row=case_row, column=3, value=tc.get('test_type', ''))
            ws_cases.cell(row=case_row, column=4, value=tc.get('test_file', ''))
            ws_cases.cell(row=case_row, column=5, value=tc.get('origin_test_file', ''))
            ws_cases.cell(row=case_row, column=6, value=tc.get('test_class', ''))
            ws_cases.cell(row=case_row, column=7, value=tc.get('test_case', ''))
            ws_cases.cell(row=case_row, column=8, value=error_msg)
            ws_cases.cell(row=case_row, column=9, value=traceback)
            ws_cases.cell(row=case_row, column=10, value=", ".join(torch_ops))
            ws_cases.cell(row=case_row, column=11, value=case_dep)
            # Leave columns 12-17 empty for CI status (to be filled by update_test_results.py)
            # Column 18: cuda_case_exist
            # Column 19: xpu_case_exist  
            # Column 20: case_existence_comments
            case_row += 1
    
    # Add to e2e sheet
    if e2e_info:
        for info in e2e_info:
            reproducer = info.get('reproducer', title[:150])
            ws_e2e.cell(row=e2e_row, column=1, value=num)
            ws_e2e.cell(row=e2e_row, column=2, value=reproducer[:200] if reproducer else title[:150])
            ws_e2e.cell(row=e2e_row, column=3, value=info.get('benchmark', ''))
            ws_e2e.cell(row=e2e_row, column=4, value=info.get('model', ''))
            ws_e2e.cell(row=e2e_row, column=5, value=info.get('phase', ''))
            ws_e2e.cell(row=e2e_row, column=6, value=info.get('dtype', ''))
            ws_e2e.cell(row=e2e_row, column=7, value=info.get('amp', False))
            ws_e2e.cell(row=e2e_row, column=8, value=info.get('backend', ''))
            ws_e2e.cell(row=e2e_row, column=9, value=info.get('test_type', ''))
            ws_e2e.cell(row=e2e_row, column=10, value=info.get('disable_cudagraphs', ''))
            ws_e2e.cell(row=e2e_row, column=11, value=error_msg)
            ws_e2e.cell(row=e2e_row, column=12, value=traceback)
            e2e_row += 1
    elif test_module == 'e2e':
        # Add e2e issues without specific model info
        ws_e2e.cell(row=e2e_row, column=1, value=num)
        ws_e2e.cell(row=e2e_row, column=2, value=title[:150])
        ws_e2e.cell(row=e2e_row, column=3, value='unknown')
        ws_e2e.cell(row=e2e_row, column=11, value=error_msg)
        ws_e2e.cell(row=e2e_row, column=12, value=traceback)
        e2e_row += 1

    issue_row += 1

    if issue_row % 50 == 0:
        print(f"Processed {issue_row-1} issues...")

print(f"\nTotal issues: {issue_row-2}")
print(f"Total test case rows: {case_row-2}")
print(f"Total e2e case rows: {e2e_row-2}")

for ws in [ws_issues, ws_cases, ws_e2e]:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

# Ensure result directory exists
os.makedirs(RESULT_DIR, exist_ok=True)

output_path = os.path.join(RESULT_DIR, "torch_xpu_ops_issues.xlsx")
wb.save(output_path)
print(f"\nSaved to {output_path}")
