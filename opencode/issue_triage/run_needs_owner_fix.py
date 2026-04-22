"""Fix NEEDS_OWNER mis-classification for issues that already have an owner.

Rules:
  owner ∈ {Triage, unassigned}           → keep NEEDS_OWNER (real owner still needed)
  real owner + pure NEEDS_OWNER          → reclassify to ROOT_CAUSE
  real owner + IMPLEMENT+NEEDS_OWNER     → drop NEEDS_OWNER (keep IMPLEMENT)
"""
import json

import openpyxl

EXCEL = "opencode/issue_triage/result/torch_xpu_ops_issues.xlsx"
STUB_OWNERS = {"triage", "unassigned", "none"}


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "none" else s


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Issues"]
    hdr = [c.value for c in ws[1]]
    I = {k: hdr.index(k) for k in
         ("Issue ID", "Assignee", "owner_transferred",
          "action_Type", "action_TBD", "action_reason")}

    n_root = n_impl = n_keep = 0
    for row in ws.iter_rows(min_row=2):
        at = clean(row[I["action_Type"]].value)
        parts = at.split("+") if at else []
        if "NEEDS_OWNER" not in parts:
            continue
        owner = clean(row[I["Assignee"]].value) or clean(row[I["owner_transferred"]].value)
        if not owner or owner.lower() in STUB_OWNERS:
            n_keep += 1
            continue

        if parts == ["NEEDS_OWNER"]:
            # pure NEEDS_OWNER with real owner → ROOT_CAUSE
            row[I["action_Type"]].value = "ROOT_CAUSE"
            row[I["action_TBD"]].value = json.dumps(
                [f"Assignee @{owner} to investigate"])
            row[I["action_reason"]].value = json.dumps(
                [f"Issue already assigned to @{owner}; owner to lead root-cause."])
            n_root += 1
        elif "IMPLEMENT" in parts:
            # drop NEEDS_OWNER, keep IMPLEMENT
            new_parts = [p for p in parts if p != "NEEDS_OWNER"]
            row[I["action_Type"]].value = "+".join(new_parts)
            row[I["action_TBD"]].value = json.dumps(
                [f"Owner @{owner} to file fix PR"])
            row[I["action_reason"]].value = json.dumps(
                [f"Issue assigned to @{owner}; owner to implement fix."])
            n_impl += 1
        else:
            n_keep += 1

    wb.save(EXCEL)
    print(f"ROOT_CAUSE reassigned:          {n_root}")
    print(f"NEEDS_OWNER dropped (IMPLEMENT): {n_impl}")
    print(f"Kept (stub owner):              {n_keep}")


if __name__ == "__main__":
    main()
