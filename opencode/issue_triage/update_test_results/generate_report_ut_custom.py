#!/usr/bin/env python3
"""
Generate UT-specific custom markdown report with only selected issues
Same structure as generate_report.py but only for filtered UT issues
"""

import openpyxl
from datetime import datetime, timedelta
import os

ROOT_DIR = "/home/daisydeng"
RESULT_DIR = "/home/daisydeng/ai_for_validation/opencode/issue_triage/result"

# List of issue IDs to include (as provided by user)
INCLUDED_ISSUES = {
    1893, 1962, 1972, 2015, 2024, 2164, 2169, 2214, 2229, 2244, 2245, 2253, 2255, 2270,
    2283, 2285, 2287, 2295, 2301, 2309, 2329, 2376, 2436, 2442, 2482, 2512, 2531, 2532,
    2536, 2541, 2554, 2578, 2609, 2611, 2613, 2615, 2620, 2694, 2697, 2698, 2715, 2720,
    2783, 2800, 2802, 2806, 2810, 2853, 2888, 2891, 2958, 2997, 2999, 3004, 3006, 3007,
    3033, 3126, 3127, 3128, 3129, 3131, 3132, 3133, 3136, 3137, 3140, 3141, 3142, 3143,
    3163, 3166, 3170, 3177, 3187, 3238
}


def generate_report():
    wb = openpyxl.load_workbook(os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx'))
    ws_issues = wb['Issues']
    
    # Collect issue data - only included IDs
    issues = []
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        if issue_id not in INCLUDED_ISSUES:
            continue
            
        issue = {
            'id': issue_id,
            'title': ws_issues.cell(row, 2).value,
            'status': ws_issues.cell(row, 3).value,
            'assignee': ws_issues.cell(row, 4).value,
            'reporter': ws_issues.cell(row, 5).value,
            'labels': ws_issues.cell(row, 6).value,
            'created_time': ws_issues.cell(row, 7).value,
            'type': ws_issues.cell(row, 11).value,
            'module': ws_issues.cell(row, 12).value,
            'test_module': ws_issues.cell(row, 13).value,
            'dependency': ws_issues.cell(row, 14).value,
            'pr': ws_issues.cell(row, 15).value,
            'owner_transfer': ws_issues.cell(row, 19).value,
            'action_TBD': ws_issues.cell(row, 20).value,
            'duplicated_issue': ws_issues.cell(row, 21).value,
            'priority': ws_issues.cell(row, 22).value,
            'priority_reason': ws_issues.cell(row, 23).value,
            'category': ws_issues.cell(row, 24).value,
            'root_cause': ws_issues.cell(row, 25).value,
        }
        issues.append(issue)
    
    # Categorize issues
    action_required = []
    no_assignee = []
    duplicated = []
    with_dependency = []
    others = []
    
    for issue in issues:
        dep = issue['dependency']
        is_valid_dep = dep and dep not in ['None', None, '']
        
        if issue['duplicated_issue']:
            duplicated.append(issue)
        elif issue['action_TBD']:
            action_required.append(issue)
        elif not issue['assignee'] or issue['assignee'] == 'None':
            no_assignee.append(issue)
        elif is_valid_dep:
            with_dependency.append(issue)
        else:
            others.append(issue)
    
    # Sort action_required by action_TBD
    action_required.sort(key=lambda x: (x['action_TBD'] or '', x['id']))
    
    # Sort duplicated by ID
    duplicated.sort(key=lambda x: x['id'])
    
    # Sort with_dependency by ID
    with_dependency.sort(key=lambda x: x['id'])
    
    # Group action_required by action_TBD
    action_groups = {}
    for issue in action_required:
        action = issue['action_TBD'] or 'Unknown'
        if action not in action_groups:
            action_groups[action] = []
        action_groups[action].append(issue)
    
    # Group by category
    category_groups = {}
    for issue in issues:
        cat = issue['category'] or 'unknown'
        if cat not in category_groups:
            category_groups[cat] = []
        category_groups[cat].append(issue)
    
    # Build statistics
    test_module_stats = {}
    module_stats = {}
    dep_stats = {}
    action_tbd_stats = {}
    category_stats = {}
    no_assignee_count = len(no_assignee)
    duplicated_count = len(duplicated)
    with_dependency_count = len(with_dependency)
    action_required_count = len(action_required)
    
    for issue in issues:
        tm = issue['test_module'] or 'unknown'
        test_module_stats[tm] = test_module_stats.get(tm, 0) + 1
        
        m = issue['module'] or 'unknown'
        module_stats[m] = module_stats.get(m, 0) + 1
        
        dep = issue['dependency']
        if dep and dep != 'None':
            dep_stats[dep] = dep_stats.get(dep, 0) + 1
        
        action = issue['action_TBD']
        if action:
            action_tbd_stats[action] = action_tbd_stats.get(action, 0) + 1
        
        cat = issue['category'] or 'unknown'
        category_stats[cat] = category_stats.get(cat, 0) + 1
    
    priority_stats = {}
    for issue in issues:
        p = issue['priority'] or 'unknown'
        priority_stats[p] = priority_stats.get(p, 0) + 1
    
    # Recent issues (last 7 days)
    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_issues = []
    for issue in issues:
        created_str = issue.get('created_time')
        if created_str:
            try:
                created_date = datetime.strptime(created_str.replace('Z', '+00:00'), '%Y-%m-%dT%H:%M:%S%z')
                created_date = created_date.replace(tzinfo=None)
                if created_date >= seven_days_ago:
                    recent_issues.append(issue)
            except:
                pass
    recent_issues.sort(key=lambda x: x['id'], reverse=True)
    
    # Build markdown
    md = ""
    md += "# Torch XPU Ops UT Issue Report (Custom Filtered)\n\n"
    md += "**Repository:** [intel/torch-xpu-ops](https://github.com/intel/torch-xpu-ops)\n\n"
    md += "**Report Type:** UT (Unit Test) Issues - Custom Filtered List\n\n"
    md += f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    md += "---\n\n"
    
    # Index
    md += "## Index\n\n"
    md += "1. [Summary](#1-summary)\n"
    md += "2. [Statistics](#2-statistics)\n"
    md += "   - [By Test Module](#by-test-module)\n"
    md += "   - [By Module](#by-module)\n"
    md += "   - [By Dependency](#by-dependency)\n"
    md += "   - [By Action TBD](#by-action-tbd)\n"
    md += "   - [By Category](#by-category-stats)\n"
    md += "   - [By Priority](#by-priority)\n"
    md += "3. [New Submitted Issues (Past Week)](#3-new-submitted-issues-past-week)\n"
    md += "4. [Action Required](#4-action-required)\n"
    md += "   - [Reporter Actions](#reporter-actions)\n"
    md += "     - [Information Required](#information-required)\n"
    md += "     - [Close Fixed Issue](#close-fixed-issue)\n"
    md += "     - [Enable Test](#enable-test)\n"
    md += "     - [Add to Skiplist](#add-to-skiplist)\n"
    md += "     - [Verify the Issue](#verify-the-issue)\n"
    md += "     - [Need Reproduce Steps](#need-reproduce-steps)\n"
    md += "   - [Engineer Actions](#engineer-actions)\n"
    md += "     - [Needs PyTorch Repo Changes (upstream)](#needs-pytorch-repo-changes-upstream)\n"
    md += "     - [Revisit the PR as Case Failed](#revisit-the-pr-as-case-failed)\n"
    md += "5. [By Category](#5-by-category)\n"
    md += "6. [Duplicated Issues](#6-duplicated-issues)\n"
    md += "7. [Issues with Dependency](#7-issues-with-dependency)\n\n"
    md += "---\n\n"
    
    # 1. Summary
    md += "## 1. Summary {#1-summary}\n\n"
    md += "| Category | Count |\n"
    md += "|----------|-------|\n"
    md += f"| Action Required | {action_required_count} |\n"
    md += f"| No Assignee | {no_assignee_count} |\n"
    md += f"| Duplicated Issues | {duplicated_count} |\n"
    md += f"| With Dependency | {with_dependency_count} |\n"
    md += f"| Others | {len(others)} |\n"
    md += f"| **Total** | {len(issues)} |\n\n"
    md += "---\n\n"
    
    # 2. Statistics
    md += "## 2. Statistics {#2-statistics}\n\n"
    
    md += "### By Test Module {#by-test-module}\n\n"
    md += "| Test Module | Count |\n"
    md += "|-------------|-------|\n"
    for tm, count in sorted(test_module_stats.items(), key=lambda x: -x[1]):
        md += f"| {tm} | {count} |\n"
    md += "\n"
    
    md += "### By Module {#by-module}\n\n"
    md += "| Module | Count |\n"
    md += "|--------|-------|\n"
    for m, count in sorted(module_stats.items(), key=lambda x: -x[1]):
        md += f"| {m} | {count} |\n"
    md += "\n"
    
    md += "### By Dependency {#by-dependency}\n\n"
    md += "| Dependency | Count |\n"
    md += "|------------|-------|\n"
    for dep, count in sorted(dep_stats.items(), key=lambda x: -x[1]):
        md += f"| {dep} | {count} |\n"
    md += "\n"
    
    md += "### By Action TBD {#by-action-tbd}\n\n"
    md += "| Action TBD | Count |\n"
    md += "|------------|-------|\n"
    for action, count in sorted(action_tbd_stats.items(), key=lambda x: -x[1]):
        md += f"| {action} | {count} |\n"
    md += "\n"
    
    md += "### By Category (Statistics) {#by-category-stats}\n\n"
    md += "| Category | Count |\n"
    md += "|----------|-------|\n"
    for cat, count in sorted(category_stats.items(), key=lambda x: -x[1]):
        md += f"| {cat} | {count} |\n"
    md += "\n"
    
    md += "### By Priority {#by-priority}\n\n"
    md += "| Priority | Count |\n"
    md += "|----------|-------|\n"
    for p, count in sorted(priority_stats.items()):
        md += f"| {p} | {count} |\n"
    md += "\n"
    
    # 3. New submitted issues in past week
    md += f"---\n\n"
    md += f"## 3. New Submitted Issues (Past Week) {{#3-new-submitted-issues-past-week}}\n\n"
    md += f"Issues created in the past 7 days (as of {datetime.now().strftime('%Y-%m-%d')}).\n\n"
    md += "| ID | Title | Status | Owner | Priority | Reason | Category | Root Cause | Labels | Module | Test Module |\n"
    md += "|---|-------|--------|-------|---------|--------|----------|-----------|--------|--------|-------------|\n"
    
    for issue in recent_issues:
        title = (issue['title'] or '')[:40]
        status = issue['status'] or ''
        owner = issue['assignee'] or ''
        labels = issue['labels'] or ''
        module = issue['module'] or ''
        test_module = issue['test_module'] or ''
        priority = issue['priority'] or ''
        reason = issue['priority_reason'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {status} | {owner} | {priority} | {reason} | {category} | {root_cause} | {labels} | {module} | {test_module} |\n"
    
    # 4. Action Required
    reporter_actions = {
        'information_required': [],
        'close_fixed': [],
        'enable_test': [],
        'add_skiplist': [],
        'verify_issue': [],
        'need_reproduce': [],
    }
    
    engineer_actions = {
        'upstream': [],
        'revisit_pr': [],
    }
    
    info_keywords = ['reproduce', 'information', 'missing', 'awaiting']
    
    for action, issues_list in action_groups.items():
        action_lower = action.lower()
        
        if 'upstream' in action_lower or 'pytorch' in action_lower:
            for iss in issues_list:
                engineer_actions['upstream'].append(iss)
        elif 'revisit' in action_lower:
            for iss in issues_list:
                engineer_actions['revisit_pr'].append(iss)
        elif 'information' in action_lower or any(kw in action_lower for kw in info_keywords):
            for iss in issues_list:
                reporter_actions['information_required'].append(iss)
        elif 'close' in action_lower or 'fix' in action_lower:
            for iss in issues_list:
                reporter_actions['close_fixed'].append(iss)
        elif 'enable' in action_lower:
            for iss in issues_list:
                reporter_actions['enable_test'].append(iss)
        elif 'skiplist' in action_lower or 'not target' in action_lower:
            for iss in issues_list:
                reporter_actions['add_skiplist'].append(iss)
        elif 'verify' in action_lower:
            for iss in issues_list:
                reporter_actions['verify_issue'].append(iss)
        elif 'reproduce' in action_lower:
            for iss in issues_list:
                reporter_actions['need_reproduce'].append(iss)
    
    md += "\n---\n\n"
    md += "## 4. Action Required {#4-action-required}\n\n"
    
    # Reporter Actions
    md += "### Reporter Actions {#reporter-actions}\n\n"
    
    md += "#### Information Required {#information-required}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in reporter_actions['information_required']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    md += "\n#### Close Fixed Issue {#close-fixed-issue}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in reporter_actions['close_fixed']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    md += "\n#### Enable Test {#enable-test}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in reporter_actions['enable_test']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    md += "\n#### Add to Skiplist {#add-to-skiplist}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in reporter_actions['add_skiplist']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    md += "\n#### Verify the Issue {#verify-the-issue}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in reporter_actions['verify_issue']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    md += "\n#### Need Reproduce Steps {#need-reproduce-steps}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in reporter_actions['need_reproduce']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    # Engineer Actions
    md += "\n### Engineer Actions {#engineer-actions}\n\n"
    
    md += "#### Needs PyTorch Repo Changes (upstream) {#needs-pytorch-repo-changes-upstream}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in engineer_actions['upstream']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    md += "\n#### Revisit the PR as Case Failed {#revisit-the-pr-as-case-failed}\n\n"
    md += "| ID | Title | Owner | Owner Transferred | Priority | Category | Root Cause | PR | Test Module |\n"
    md += "|---|-------|-------|-------------------|---------|----------|-----------|-----|-------------|\n"
    
    for issue in engineer_actions['revisit_pr']:
        title = (issue['title'] or '')[:35]
        owner = issue['assignee'] or ''
        owner_transfer = issue['owner_transfer'] or ''
        priority = issue['priority'] or ''
        category = issue['category'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {owner_transfer} | {priority} | {category} | {root_cause} | {pr_link} | {test_module} |\n"
    
    # 5. By Category
    md += "\n---\n\n"
    md += "## 5. By Category {#5-by-category}\n\n"
    
    for cat in sorted(category_groups.keys()):
        issues_list = category_groups[cat]
        issues_list.sort(key=lambda x: x['id'])
        cat_count = len(issues_list)
        cat_anchor = "#" + cat.lower().replace(" ", "-").replace("/", "-")
        md += "#### " + cat + " (" + cat_anchor + ")\n\n"
        md += "| ID | Title | Status | Owner | Priority | Root Cause | PR | Labels | Module | Test Module |\n"
        md += "|---|-------|--------|-------|---------|-----------|-----|--------|--------|-------------|\n"
        for issue in issues_list:
            title = (issue['title'] or '')[:30]
            status = issue['status'] or ''
            owner = issue['assignee'] or ''
            priority = issue['priority'] or ''
            root_cause = issue['root_cause'] or ''
            pr = issue['pr'] or ''
            pr_link = f"[PR]({pr})" if pr else ''
            labels = issue['labels'] or ''
            module = issue['module'] or ''
            test_module = issue['test_module'] or ''
            md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {status} | {owner} | {priority} | {root_cause} | {pr_link} | {labels} | {module} | {test_module} |\n"
        md += "\n"
    
    # 6. Duplicated Issues
    md += "\n---\n\n"
    md += "## 6. Duplicated Issues {#6-duplicated-issues}\n\n"
    md += "Issues that share test cases with other issues.\n\n"
    md += "| ID | Title | Owner | Reporter | Duplicated With | Priority | Root Cause | PR | Labels | Module | Test Module |\n"
    md += "|---|-------|-------|----------|-----------------|---------|-----------|-----|--------|--------|-------------|\n"
    
    for issue in duplicated:
        title = (issue['title'] or '')[:30]
        owner = issue['assignee'] or ''
        reporter = issue['reporter'] or ''
        dup = issue['duplicated_issue'] or ''
        priority = issue['priority'] or ''
        root_cause = issue['root_cause'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        labels = issue['labels'] or ''
        module = issue['module'] or ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {reporter} | {dup} | {priority} | {root_cause} | {pr_link} | {labels} | {module} | {test_module} |\n"
    
    # 7. Issues with Dependency
    md += "\n---\n\n"
    md += "## 7. Issues with Dependency {#7-issues-with-dependency}\n\n"
    md += "Issues that have dependencies on other components.\n\n"
    md += "| ID | Title | Owner | Priority | Root Cause | Dependency | Category | PR | Labels | Test Module |\n"
    md += "|---|-------|-------|---------|-----------|------------|----------|-----|--------|-------------|\n"
    
    with_dep_sorted = [i for i in with_dependency if i['dependency']]
    with_dep_sorted.sort(key=lambda x: x['id'])
    for issue in with_dep_sorted:
        title = (issue['title'] or '')[:25]
        owner = issue['assignee'] or ''
        priority = issue['priority'] or ''
        root_cause = issue['root_cause'] or ''
        dep = issue['dependency'] or ''
        category = issue['category'] or ''
        pr = issue['pr'] or ''
        pr_link = f"[PR]({pr})" if pr else ''
        labels = issue['labels'] or ''
        test_module = issue['test_module'] or ''
        md += f"| [{issue['id']}](https://github.com/intel/torch-xpu-ops/issues/{issue['id']}) | {title} | {owner} | {priority} | {root_cause} | {dep} | {category} | {pr_link} | {labels} | {test_module} |\n"
    
    # Save to file
    output_path = os.path.join(RESULT_DIR, 'issue_report_ut.md')
    with open(output_path, 'w') as f:
        f.write(md)
    
    print(f"UT Custom Report saved to: {output_path}")
    print(f"Total issues included: {len(issues)}")
    print(f"  Action Required: {action_required_count}")
    print(f"  No Assignee: {no_assignee_count}")
    print(f"  Duplicated: {duplicated_count}")
    print(f"  With Dependency: {with_dependency_count}")
    print(f"  Others: {len(others)}")


if __name__ == '__main__':
    generate_report()