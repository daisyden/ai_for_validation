#!/usr/bin/env python3
"""
Script to update torch_xpu_ops_issues.xlsx with test case results from:
1. torch-xpu-ops nightly CI (XML files in Inductor-XPU-UT-Data-*-op_ut-*)
2. Stock PyTorch XPU CI (test-reports from pytorch xpu CI artifacts)
3. Case existence analysis (checking if tests exist in pytorch/test and torch-xpu-ops/test/xpu)
4. Duplicated issue detection (by Test Class + Test Case or similar Traceback)
5. E2E test case status from torch-xpu-ops nightly
6. Generate markdown report

Usage:
    python update_test_results.py

Input:
    - /home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx
    - /home/daisydeng/issue_traige/ci_results/torch-xpu-ops/
    - /home/daisydeng/issue_traige/ci_results/stock/
    - ~/pytorch/test/
    - ~/pytorch/third_party/torch-xpu-ops/test/xpu/

Output:
    - Updated /home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx
    - /home/daisydeng/ai_for_validation/opencode/issue_triage/result/issue_report.md
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import xml.etree.ElementTree as ET
import os
import re
import glob
import time
import sys

ROOT_DIR = "/home/daisydeng"
RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")
LOG_FILE = os.path.join(RESULT_DIR, "pipeline.log")

def log(msg, print_also=True):
    """Log message to file and optionally print to console."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {msg}"
    try:
        with open(LOG_FILE, "a") as f:
            f.write(log_msg + "\n")
    except Exception:
        pass
    if print_also:
        print(log_msg)


def get_torch_xpu_ops_xml_files():
    """Get all XML files from torch-xpu-ops nightly artifacts"""
    base_dir = '/home/daisydeng/issue_traige/ci_results/torch-xpu-ops'
    
    ut_folders = []
    for d in os.listdir(base_dir):
        if d.startswith('Inductor-XPU-UT-Data-'):
            match = re.match(r'Inductor-XPU-UT-Data-([a-f0-9]+)-.*-(\d+)-1$', d)
            if match:
                ut_folders.append((d, match.group(1), match.group(2)))
    
    xml_files = {}
    for folder_name, commit, run_id in ut_folders:
        folder_path = os.path.join(base_dir, folder_name, folder_name)
        if not os.path.exists(folder_path):
            continue
        for f in os.listdir(folder_path):
            if f.endswith('.xml') and (f.startswith('op_ut_with_all') or f.startswith('op_ut_with_skip')):
                xml_path = os.path.join(folder_path, f)
                try:
                    tree = ET.parse(xml_path)
                    root = tree.getroot()
                    count = len(root.findall('.//testcase'))
                    if count > 0:
                        prefix = f.replace('.xml', '')
                        xml_files[prefix] = (xml_path, commit, run_id, count)
                except:
                    pass
    
    return xml_files


def get_stock_xml_files():
    """Get all XML files from stock PyTorch XPU CI"""
    stock_base = '/home/daisydeng/issue_traige/ci_results/stock'
    stock_xml_files = {}
    
    for zip_file in glob.glob(f'{stock_base}/test-reports-*.zip'):
        pytest_dir = os.path.join(zip_file, 'test-reports', 'python-pytest')
        if not os.path.exists(pytest_dir):
            continue
        for root, dirs, files in os.walk(pytest_dir):
            for f in files:
                if f.endswith('.xml'):
                    xml_path = os.path.join(root, f)
                    test_module = os.path.basename(root)
                    stock_xml_files[test_module] = xml_path
    
    return stock_xml_files


def convert_test_file_to_xml_prefix(test_file):
    """Convert test file to XML prefix for torch-xpu-ops"""
    if not test_file:
        return None, 'No test file'
    
    test_file = str(test_file)
    
    if '/' in test_file:
        test_file = test_file.replace('torch-xpu-ops/test/xpu/', '')
        test_file = test_file.replace('.py', '')
        return 'op_ut_with_all.' + test_file.replace('/', '.'), None
    
    parts = test_file.split('.')
    parts_len = len(parts)
    
    if parts_len == 2:
        module = parts[0]
        if '_xpu' in module:
            return 'op_ut_with_skip.' + module, None
        return 'op_ut_with_all.' + module, None
    
    if parts_len == 3:
        module = parts[0]
        test_module = parts[1]
        if module == 'test':
            return 'op_ut_with_all.test_' + test_module + '_xpu', None
        if module == 'inductor':
            return None, 'inductor not in torch-xpu-ops'
        return 'op_ut_with_all.' + module + '.' + test_module, None
    
    if parts_len == 4:
        module = parts[0]
        subdir = parts[1]
        test_module = parts[2]
        if module == 'test':
            if subdir == 'functorch':
                return f'op_ut_with_all.{subdir}.{test_module}_xpu', None
            return None, f'{subdir} not in torch-xpu-ops'
        if module == 'inductor':
            return None, 'inductor not in torch-xpu-ops'
    
    return None, 'unknown pattern'


def convert_to_stock_prefix(test_file):
    """Convert test file to stock test module name"""
    if not test_file:
        return None
    
    test_file = str(test_file)
    
    if '/' in test_file:
        test_file = test_file.replace('torch-xpu-ops/test/xpu/', '')
        test_file = test_file.replace('.py', '')
        return test_file.replace('/', '.')
    
    parts = test_file.split('.')
    parts_len = len(parts)
    
    if parts_len == 2:
        return parts[0]
    if parts_len == 3:
        module = parts[0]
        test_module = parts[1]
        if module == 'test':
            return test_module
        return f'{module}.{test_module}'
    if parts_len == 4:
        module = parts[0]
        subdir = parts[1]
        test_module = parts[2]
        if module == 'test':
            return f'{subdir}.{test_module}'
        if module == 'inductor':
            return f'{module}.{test_module}'
    
    return None


def find_best_xml_match(xml_prefix, xml_files):
    """Find XML file matching prefix"""
    if not xml_prefix:
        return None
    if xml_prefix in xml_files:
        return xml_files[xml_prefix]
    return None


def parse_failure_content(content):
    """Parse failure message to extract error_msg and traceback until error type."""
    error_msg = ""
    traceback = ""

    if not content:
        return error_msg, traceback

    lines = content.split('\n')

    # Find error types - look for exception raising patterns
    # Pattern 1: Error type at start of line (RuntimeError: message)
    # Pattern 2: raise ErrorType("message") within traceback
    error_patterns = [
        (r'^RuntimeError', 'RuntimeError'),
        (r'^AssertionError', 'AssertionError'),
        (r'^ValueError', 'ValueError'),
        (r'^TypeError', 'TypeError'),
        (r'^IndexError', 'IndexError'),
        (r'^KeyError', 'KeyError'),
        (r'^ImportError', 'ImportError'),
        (r'^NotImplementedError', 'NotImplementedError'),
        (r'^AttributeError', 'AttributeError'),
        (r'^InductorError', 'InductorError'),
    ]

    exception_raise_patterns = [
        r'\braise\s+(RuntimeError|AssertionError|ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|AttributeError|InductorError)\s*[\(\'"]',
    ]

    error_line_idx = -1
    error_type = None
    last_error_msg = ""

    for idx, line in enumerate(lines):
        stripped = line.strip()
        # Pattern 1: Error type at start of line
        for pattern, etype in error_patterns:
            if re.match(pattern, stripped):
                error_line_idx = idx
                error_type = etype
                # Clean up the line - remove quotes and extra info
                clean_line = re.sub(r'^' + etype + r'[:\s]*', '', stripped)
                error_msg = clean_line[:200]
                break
        if error_line_idx >= 0:
            break
        # Pattern 2: raise ErrorType("message") - look for the last exception raise
        for ep in exception_raise_patterns:
            if re.search(ep, stripped):
                error_line_idx = idx
                # Extract the message part after the error type
                match = re.search(r'raise\s+\w+\s*[\(\'"](.+?)[\'\"]?', stripped)
                if match:
                    last_error_msg = match.group(1).strip()[:200]

    # Check if we have Traceback header
    traceback = ""
    if 'Traceback (most recent call last):' in content:
        # Build traceback: from Traceback line until error line (inclusive)
        tb_lines = []
        end_idx = error_line_idx if error_line_idx >= 0 else len(lines)
        for idx, line in enumerate(lines):
            if 'Traceback (most recent call last):' in line:
                # Include all lines from here until error line (inclusive)
                for j in range(idx, end_idx + 1):
                    tb_lines.append(lines[j])
                break

        if tb_lines:
            traceback = '\n'.join(tb_lines)
        # If no traceback but we found exception raise, extract from that point
        elif last_error_msg:
            for idx, line in enumerate(lines):
                stripped = line.strip()
                for ep in exception_raise_patterns:
                    if re.search(ep, stripped):
                        traceback = '\n'.join(lines[idx:])
                        break
                if traceback:
                    break
    else:
        # No traceback, just take the error message
        traceback = ""
        # Use content as error message
        error_msg = content[:200]

    # If we found an exception raise but no clear error line, use the extracted message
    if last_error_msg and not error_msg:
        error_msg = last_error_msg

    # Ensure traceback includes the actual error line if we have an error message
    if error_msg and traceback and error_msg not in traceback:
        traceback += f"\n{error_msg}"

    return error_msg, traceback[:3000] if traceback else traceback


def get_test_result(xml_path, test_case):
    """Get test case result from XML file. Returns status, error_msg, and traceback."""
    if not xml_path:
        return None, None, None

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        for testcase in root.findall('.//testcase'):
            if testcase.get('name') == test_case:
                failure = testcase.find('failure')
                if failure is not None:
                    msg = failure.text or failure.get('message', '')
                    error_msg, traceback = parse_failure_content(msg)
                    return 'failed', error_msg, traceback
                skipped = testcase.find('skipped')
                if skipped is not None:
                    msg = skipped.text or skipped.get('message', '')
                    return 'skipped', msg[:500] if msg else 'skipped', None
                return 'passed', '', None

        return 'not found', 'Test case not found', None
    except Exception as e:
        return 'error', str(e), None


def analyze_test_case_with_llm(test_file, test_class, test_case, origin_test_file):
    """
    Use LLM skills to check CUDA and XPU test case existence.
    Calls check-cuda-test-existence and check-xpu-test-existence skills via opencode.
    
    Note: This requires opencode to be available and pytorch skills to be loaded.
    """
    import subprocess
    import json
    
    if not test_file or not test_case:
        return {
            'cuda_exists': 'No',
            'xpu_exists': 'No',
            'explanation': 'No test file or test case',
            'cuda_decorators': [],
            'xpu_decorators': []
        }
    
    # Try to find pytorch root
    pytorch_root = os.path.expanduser('~/pytorch')
    if not os.path.exists(pytorch_root):
        pytorch_root = os.path.expanduser('~/issue_traige/pytorch')
    
    prompt = f"""You are in the pytorch directory: {pytorch_root}

Use the check-cuda-test-existence skill to check if the CUDA test exists in the original PyTorch test file.
Use the check-xpu-test-existence skill to check if the XPU test exists in torch-xpu-ops repo.

Paths:
- PyTorch test files: {pytorch_root}/test/
- torch-xpu-ops test files: {pytorch_root}/third_party/torch-xpu-ops/test/xpu/

Test File: {test_file}
Origin Test File: {origin_test_file if origin_test_file else 'Not provided'}
Test Class: {test_class}
Test Case: {test_case}

IMPORTANT: The base test name is NOT just removing '_xpu' suffix. The base test is the actual test function in the test file that can be parameterized to generate the XPU test case. For example:
- test_compare_cpu_add_float16_xpu has base test test_compare_cpu in test_ops.py
- test_conv2d_hipdnn_backend_selection_xpu has base test test_conv2d_hipdnn_backend_selection

The base test is often a function that gets parameterized with decorators like @dtypes, @parametrize_test, etc. to generate multiple test cases with different parameters (dtypes, devices, arguments).

IMPORTANT: In the explanation, you MUST explain WHY the XPU test does not exist if cuda_exists is "No" or xpu_exists is "No". The reasons can be:
1. SKIP DECORATORS: Test has decorators like @onlyCUDA, @skipCUDAIfNoHipdnn, @skipIfXpu, @requires_xccl that prevent it from running on XPU
2. PARAMETERIZATION: Test is generated from a parameterized base test (e.g., @dtypes, @parametrize_test) - explain what parameters are used
3. REMOVED/RENAMED: Test was removed or renamed in newer PyTorch versions
4. NOT APPLICABLE: Test is specific to CUDA/ROCm hardware (e.g., hipdnn backend)
5. OTHER: Other reasons

For each check, provide:
1. Whether CUDA test exists (Yes/No)
2. Whether XPU test exists (Yes/N/A)
3. Key decorators found (e.g., @onlyCUDA, @onlyXPU, @skipCUDAIfNoHipdnn, @skipIfXpu, @dtypes, @parametrize_test, @requires_xccl)
4. Base test name (the actual function in the test file that this test case derives from)
5. CUDA test file path if found
6. XPU test file path if found
7. CUDA test name found (full or base name)
8. XPU test name found (full or base name)
9. Detailed explanation of why XPU test exists or does not exist, including:
   - Which decorators affect XPU compatibility
   - What parameterization generates this test case
   - Whether test was removed/renamed
   - Whether test is CUDA/ROCm specific

Return ONLY valid JSON format (no additional text):
{{
    "cuda_exists": "Yes/No",
    "xpu_exists": "Yes/No/N/A", 
    "cuda_decorators": ["decorator1", "decorator2"],
    "xpu_decorators": ["decorator1"],
    "base_test_name": "original_test_function_name",
    "cuda_test_file": "path/to/test_file.py",
    "xpu_test_file": "path/to/test_xpu.py", 
    "cuda_test_name": "test_name_found",
    "xpu_test_name": "test_name_found",
    "explanation": "detailed explanation including why XPU test does or does not exist"
}}
"""
    
    try:
        # Use 'opencode run' with a faster free model
        result = subprocess.run(
            ['opencode', 'run', '-m', 'opencode/gpt-5-nano', prompt],
            capture_output=True,
            text=True,
            timeout=180,
            cwd=pytorch_root
        )
        
        if result.returncode == 0:
            output = result.stdout.strip()
            # Try to extract JSON from output
            try:
                # Look for JSON object in output
                import re
                json_match = re.search(r'\{[^{}]*\}', output, re.DOTALL)
                if json_match:
                    data = json.loads(json_match.group())
                    return data
                # If full output is JSON
                if output.startswith('{'):
                    data = json.loads(output)
                    return data
            except (json.JSONDecodeError, AttributeError) as e:
                pass
            # If no valid JSON found, return explanation
            return {
                'cuda_exists': 'Unknown',
                'xpu_exists': 'Unknown',
                'cuda_decorators': [],
                'xpu_decorators': [],
                'explanation': f"Output: {output[:500]}"
            }
        else:
            return {
                'cuda_exists': 'Error',
                'xpu_exists': 'Error',
                'cuda_decorators': [],
                'xpu_decorators': [],
                'explanation': f"LLM error (rc={result.returncode}): {result.stderr[:300]}"
            }
    except subprocess.TimeoutExpired:
        return {
            'cuda_exists': 'Timeout',
            'xpu_exists': 'Timeout',
            'cuda_decorators': [],
            'xpu_decorators': [],
            'explanation': 'LLM analysis timed out after 240s'
        }
    except FileNotFoundError:
        return {
            'cuda_exists': 'Error',
            'xpu_exists': 'Error',
            'cuda_decorators': [],
            'xpu_decorators': [],
            'explanation': 'opencode command not found'
        }
    except Exception as e:
        return {
            'cuda_exists': 'Error',
            'xpu_exists': 'Error',
            'cuda_decorators': [],
            'xpu_decorators': [],
            'explanation': f"Failed to run LLM: {str(e)}"
        }


def analyze_test_case(test_file, test_case):
    """
    Analyze why test case exists or not.
    Uses basic file existence check - for detailed analysis use check-cuda-test-existence 
    and check-xpu-test-existence skills via opencode.
    """
    result = {
        'cuda_file': None, 'cuda_exists': 'No',
        'xpu_file': None, 'xpu_exists': 'No',
        'explanation': '', 'expected_name': test_case
    }
    
    if not test_file:
        result['explanation'] = 'No test file'
        return result
    
    test_file = str(test_file)
    
    if '/' in test_file:
        cuda_rel = test_file.replace('torch-xpu-ops/test/xpu/', '').replace('_xpu', '')
        xpu_rel = test_file.replace('torch-xpu-ops/test/xpu/', '')
    elif '.' in test_file:
        parts = test_file.split('.')
        parts_len = len(parts)
        
        if parts_len == 4:
            cuda_rel = f"{parts[1]}/{parts[2]}.py"
            xpu_rel = f"{parts[1]}/{parts[2]}_xpu.py"
            result['expected_name'] = parts[3]
        elif parts_len == 3:
            if parts[0] == 'test':
                cuda_rel = f"{parts[1]}.py"
                xpu_rel = f"{parts[1]}_xpu.py"
            elif parts[0] == 'inductor':
                cuda_rel = f"inductor/{parts[1]}.py"
                xpu_rel = None
        elif parts_len == 2:
            cuda_rel = f"{parts[0].replace('_xpu', '')}.py"
            xpu_rel = f"{parts[0]}.py"
        else:
            return result
    else:
        return result
    
    pytorch_test_dir = os.path.expanduser('~/pytorch/test')
    torch_xpu_ops_dir = os.path.expanduser('~/pytorch/third_party/torch-xpu-ops/test/xpu')
    
    xpu_path = os.path.join(torch_xpu_ops_dir, xpu_rel) if xpu_rel else None
    xpu_content = None
    uses_xpu_patch = False
    
    if xpu_path and os.path.exists(xpu_path):
        with open(xpu_path, 'r') as f:
            xpu_content = f.read()
        uses_xpu_patch = 'XPUPatchForImport' in xpu_content
    
    cuda_path = os.path.join(pytorch_test_dir, cuda_rel) if cuda_rel else None
    cuda_content = None
    
    if cuda_path and os.path.exists(cuda_path):
        with open(cuda_path, 'r') as f:
            cuda_content = f.read()
    
    base_name = test_case.replace('_xpu', '') if test_case else None
    cuda_name = test_case.replace('_xpu', '_cuda') if test_case else None
    
    if cuda_content:
        result['cuda_file'] = cuda_rel
        result['cuda_exists'] = 'Yes'
        
        if test_case and test_case in cuda_content:
            result['explanation'] += 'CUDA: Found. '
        elif base_name and base_name in cuda_content:
            result['explanation'] += f"CUDA: Found (base: {base_name}). "
        elif cuda_name and cuda_name in cuda_content:
            result['explanation'] += f"CUDA: Found (cuda: {cuda_name}). "
        else:
            result['explanation'] += 'CUDA: Test not found in pytorch/test. '
            result['cuda_exists'] = 'No'
    else:
        result['cuda_file'] = f'Not found: {cuda_rel}'
        result['explanation'] += 'CUDA file missing. '
    
    uses_xpu_patch = xpu_content and 'XPUPatchForImport' in xpu_content
    
    if uses_xpu_patch:
        if cuda_content:
            if test_case in cuda_content:
                result['explanation'] += 'XPU: Found (imported from CUDA). '
            elif base_name and base_name in cuda_content:
                result['explanation'] += 'XPU: Found (imported from CUDA, base name). '
            elif cuda_name and cuda_name in cuda_content:
                result['explanation'] += 'XPU: Found (imported from CUDA, cuda name). '
            else:
                result['explanation'] += 'XPU: Imported but test not found in CUDA. '
        else:
            result['explanation'] += 'XPU: Uses XPUPatchForImport but CUDA file missing. '
    elif xpu_path and os.path.exists(xpu_path):
        result['xpu_file'] = xpu_rel
        result['xpu_exists'] = 'Yes'
        
        if test_case in xpu_content:
            result['explanation'] += 'XPU: Found in XPU file. '
        elif base_name and base_name in xpu_content:
            result['explanation'] += 'XPU: Found without _xpu. '
        else:
            result['explanation'] += 'XPU: Test not in file. '
    elif xpu_rel is None:
        result['xpu_file'] = 'N/A'
        result['explanation'] += 'XPU: Inductor tests use stock CI. '
    else:
        result['xpu_file'] = f'Not found: {xpu_rel}'
        result['explanation'] += 'XPU file missing. '
    
    return result


def analyze_test_case_with_llm_qwen(test_file, test_class, test_case, origin_test_file=None):
    """
    Use Qwen3-32B via internal API to check CUDA and XPU test case existence.
    Returns: dict with test existence info and measures elapsed time.
    Used for double "not found" cases analysis.
    """
    import requests
    import json
    import time
    
    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"
    
    pytorch_root = os.path.expanduser('~/pytorch')
    if not os.path.exists(pytorch_root):
        pytorch_root = os.path.expanduser('~/issue_traige/pytorch')
    
    prompt = f"""You are in the pytorch directory: {pytorch_root}

Check if the CUDA and XPU test exists in respective test files.

Paths:
- PyTorch test files: {pytorch_root}/test/
- torch-xpu-ops test files: {pytorch_root}/third_party/torch-xpu-ops/test/xpu/

Test File: {test_file}
Origin Test File: {origin_test_file if origin_test_file else 'Not provided'}
Test Class: {test_class}
Test Case: {test_case}

Base test is the actual test function in the test file,NOT just removing '_xpu' suffix.

IMPORTANT: Explain WHY XPU test does not exist if cuda_exists is "No" or xpu_exists is "No":
1. SKIP DECORators: @onlyCUDA, @skipCUDAIfNoHipdnn, @skipIfXpu, @requires_xccl
2. PARAMETERIZATION: @dtypes, @parametrize_test generating tests
3. REMOVED/RENAMED: Test removed/renamed in newer versions
4. NOT APPLICABLE: CUDA/ROCm specific (hipdnn backend)
5. OTHER: Other reasons

Return ONLY valid JSON:
{{
    "explanation": "detailed explanation why XPU test exists or not"
    "cuda_exists": "Yes/No",
    "xpu_exists": "Yes/No/N/A",
    "cuda_decorators": ["decorator1"],
    "xpu_decorators": ["decorator1"],
    "base_test_name": "original_test_function_name",
    "cuda_test_file": "path/to/test_file.py",
    "xpu_test_file": "path/to/test_xpu.py",
    "cuda_test_name": "test_name_found",
    "xpu_test_name": "test_name_found",
}}
"""
    
    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "You are a PyTorch test analysis assistant. Return ONLY valid JSON."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 4096
    }
    
    start_time = time.time()
    
    try:
        response = requests.post(
            LLM_ENDPOINT,
            headers=headers,
            json=payload,
            timeout=180
        )
        response.raise_for_status()
        result = response.json()
        elapsed = time.time() - start_time
        
        if result.get('choices') and len(result['choices']) > 0:
            content = result['choices'][0].get('message', {}).get('content', '')
            import re
            json_match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
            if json_match:
                data = json.loads(json_match.group())
                data['elapsed_time'] = elapsed
                return data
        
        return {
            'explanation': f'API response parsing failed: {str(result)[:200]}',
            'cuda_exists': 'Unknown',
            'xpu_exists': 'Unknown',
            'elapsed_time': elapsed
        }
        
    except requests.exceptions.Timeout:
        elapsed = time.time() - start_time
        return {
            'explanation': 'Request timed out after 180s',
            'cuda_exists': 'Timeout',
            'xpu_exists': 'Timeout',
            'elapsed_time': elapsed
        }
    except Exception as e:
        elapsed = time.time() - start_time
        return {
            'explanation': f'API error: {str(e)}',
            'cuda_exists': 'Error',
            'xpu_exists': 'Error',
            'elapsed_time': elapsed
        }


def determine_category_llm(title, summary, test_cases_info, test_module, labels):
    """
    Use Qwen3-32B via internal API to determine the category of an issue.
    Categories: Distributed, TorchAO, PT2E, Flash Attention/Transformer, Sparse, Inductor/Compilation, Dtype/Precision, Backend/Device, Others

    Returns: tuple of (category string, reason string)
    """
    import requests
    import json
    import time
    import re

    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"

    if not title and not summary:
        return "Others", ''

    tc_info_str = ""
    if test_cases_info:
        for tc in test_cases_info:
            error_val = tc.get('error_msg') or ''
            tc_info_str += f"- Test: {tc.get('test_case', '')}, Error: {str(error_val)[:100]}\n"

    test_module_str = test_module or "Unknown"
    labels_str = labels or "None"

    prompt = f"""You are analyzing PyTorch XPU issue to determine its category.

Issue Title: {title}
Issue Summary: {summary}
Test Module: {test_module_str}
Labels: {labels_str}

Test Cases Info:
{tc_info_str}

Categorize this issue into ONE of:

Analyze the provided PyTorch runtime error and classify it into exactly ONE category from the list below.

Categories:

    1. Distributed - Keywords: distributed, XCCL, NCCL, Gloo, ProcessGroup, DDP, FSDP, torch.distributed, collective communication, multi-node, timeout, rendezvous, init_method, reduce_scatter, all_gather

    2. TorchAO - Keywords: torchao, quantize_, int4_weight_only, int8_dynamic_activation, fp8, nf4, autoquant, quantization_config, Adam8bit, AdamW4bit, Lion8bit, PagedAdam, OptimizerWithQuantization, quantized_optimizer, ao/sparsity, apply_dynamic_quant, change_linear_weights_to_int8_packed, QuantizedLinear, from_float (AO context), packed weight, dequantization, int4, int8 (when paired with torchao)

    3. PT2E - Keywords: torch.export(), Dynamo, fake_tensor, ExportedProgram, AOT, torch._export, graph break, tracing, exported_program.run_decompositions

    4. Flash Attention/Transformer - Keywords: flash_attention, scaled_dot_product_attention, SDPA, attention mask, transformer layer, F.scaled_dot_product_attention, memory-efficient attention, FlexAttention

    5. Sparse - Keywords: sparse tensor, CSR, CSC, COO, torch.sparse, sparse matrix multiplication, sparse_mask, to_sparse(), sparse_coo_tensor

    6. Inductor/Compilation - Keywords: torch.compile(), Inductor, Triton, codegen, AOT Autograd, FX graph, torch._inductor, compilation cache, torch._dynamo

    7. Torch Runtime - Keywords: CUDA runtime, cudaMalloc, cudaMemcpy, out of memory (OOM), device kernel launch, stream synchronization, cudaStreamSynchronize, device-side assert, cudaError, illegal memory access, context management, cudaSetDevice, device initialization, cudaGetDevice, driver error, CUDA_VISIBLE_DEVICES, memory leak, allocation failure, device reset

    8. Torch Operations - Keywords: operator implementation, aten::, native::, custom op, register_operator, operator overloading, tensor operation dispatch, kernel selection, op signature mismatch, unsupported op on device, op not implemented for device, device-specific op behavior, backward pass operation, autograd op

    9. Dtype/Precision - Keywords: dtype mismatch, float16, bfloat16, float32, mixed precision, autocast, GradScaler, precision loss, NaN/inf due to dtype, to(dtype=...), torch.int8 (without torchao), legacy torch.quantization

    10. Feature Not Supported - Keywords: unimplemented operator, missing kernel, feature not available in this build, unsupported combination, "not implemented for"

    11. Skip/No Test Exists - Keywords: test skipped, @unittest.skip, missing test decorator, CI test gap, skipIfNoTorchAO

    12. Others - None of the above (only if truly uncategorizable)

    Classification Rules:
        - Select exactly ONE category
        - For int4/int8/fp8 errors: TorchAO takes precedence over Dtype/Precision
        - For quantized optimizers (Adam8bit, Lion8bit): TorchAO takes precedence
        - For CUDA runtime errors (memory, sync, context): Choose Torch Runtime
        - For operator dispatch/kernel errors: Choose Torch Operations
        - For device-specific op implementation issues: Choose Torch Operations
        - Use Others only as a last resort

        Distinction between Torch Runtime and Torch Operations:
            - Torch Runtime = Errors related to device management, memory allocation, synchronization, driver/runtime API calls
            - Torch Operations = Errors related to specific operator execution, kernel selection, op dispatch, custom op registration

Return the category AND a detailed reason for your classification:
- The reason is REQUIRED, not optional
- Make it detailed and specific (full sentences, 150-300 characters)
- Include: specific ops/functions mentioned, dtypes (float16, bf16, int8, Long, etc.), arguments, or patterns that led to this categorization
- Explain clearly WHY you chose this category based on the issue details

Format: "Category Name | detailed_reason"
Example: "Dtype/Precision Issue | The aten.memory_efficient_attention kernel encounters dtype mismatch when processing fp32 input tensors with bfloat16 scaling factor on XPU device. The dot_xpu_mkl operation is not implemented for Long dtype, causing NotImplementedError."

YOUR ANSWER (must include detailed reason after the pipe symbol):"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "Output ONLY 'X - Category Name | brief_reason'. No markdown. No JSON. No thinking tags."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 100
    }

    start_time = time.time()

    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time

        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            
            # Remove all thinking/markdown tags patterns
            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            
            # Remove all bracket patterns and their content (but not cause issues with escape)
            # Using a more careful approach
            content = content.replace('[', ' <').replace(']', '> ')
            
            # Clean up markers
            content = re.sub(r'<[^>]*>', '', content)
            
            # Remove common thinking tag leftovers
            content = re.sub(r'has ATTR\b', '', content)
            
            # Clean up whitespace
            content = re.sub(r'\s+', ' ', content).strip()

            # Parse category and reason - format: "X - Category Name | reason"
            if '|' in content:
                parts = content.split('|', 1)
                category_part = parts[0].strip()
                reason = parts[1].strip() if len(parts) > 1 else ''
                return category_part, reason
            else:
                match = re.search(r'(\d+)\s*[-–]\s*[\w\s/]+', content)
                if match:
                    category = f"{match.group(1)} - " + content[match.start():match.end()].split('-')[-1].strip()
                    return category, ''
                else:
                    return content.strip()[:50], ''

        return f"API Error: {response.status_code}", ''

    except Exception as e:
        return f"Error: {str(e)[:30]}", ''


def check_info_requested_to_reporter_llm(issue_title, issue_summary, error_msg, traceback):
    """
    Use Qwen3-32B via internal API to check if more info needs to be requested from reporter.
    Returns: action string (e.g., "Need reproduce steps", "Ready to analyze", "Need more information")
    """
    import requests
    import json
    import time
    import re
    
    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"
    
    issue_content = f"{issue_title} {issue_summary} {error_msg} {traceback}".strip()
    
    if not issue_content or len(issue_content) < 20:
        return "Need more information"
    
    prompt = f"""You are analyzing a PyTorch XPU GitHub issue to determine if more information is needed from the reporter.

Issue Content:
{issue_content[:1500]}

Determine:
1. Is this a feature request or enhancement?
2. Does it have sufficient error information?
3. Does it have reproduction steps?
4. Is it a clear bug with complete information?

Respond with ONE of:
- "Ready to analyze" - with sufficient info to start debugging
- "Need reproduce steps" - if a bug without clear repro steps
- "Need more information - [specific missing info]" - if the comments requesting key details
- "Feature Request - needs triage" - if it's a feature request

YOUR ANSWER:"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "Output ONLY what action to take. Be brief. No markdown. No JSON."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 60
    }
    
    start_time = time.time()
    
    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time
        
        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            content = content.replace("[TO]", "").replace("[/TO]", "").replace("<think>", "").replace("]", "").strip()
            
            if "Ready to analyze" in content:
                return "Ready to analyze"
            elif "reproduce" in content.lower():
                return "Need reproduce steps"
            elif "Feature" in content:
                return "Feature Request"
            else:
                return content.strip()[:60]
        
        return "Need more information"
        
    except Exception as e:
        return "Need more information"


def determine_priority_llm(title, summary, error_msg, test_module, labels_str, test_cases_info):
    """
    Use Qwen3-32B via internal API to determine the priority of an issue.
    Returns: (priority, reason) tuple
    """
    import requests
    import json
    import time
    import re
    
    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"
    
    tc_info_str = ""
    if test_cases_info:
        for tc in test_cases_info:
            tc_info_str += f"- {tc.get('test_case', '')}: {str(tc.get('error_msg', ''))[:80]}\n"
    
    prompt = f"""You are analyzing PyTorch XPU issue priority.

Title: {title}
Summary: {summary[:500]}
Test Module: {test_module or 'Unknown'}
Labels: {labels_str}

Error Info:
{error_msg[:300] if error_msg else 'N/A'}

Test Cases:
{tc_info_str}

Determine priority (P0=critical, P1=high, P2=medium, P3=low):
- P0: Build crash, regression (was passing), real model failure, security
- P1: Many test failures, e2e accuracy issue, performance regression
- P2: Few UT failures, feature gaps, minor issues
- P3: Minor, cosmetic, documentation

Return ONLY format: "P# - detailed_reason" (full sentences, 150-300 characters)

Example: "P0 - Build crash during aten.neg kernel compilation for XPU backend due to undefined reference to device-specific Triton template implementation"
Example: "P1 - E2E regression in HuggingFace models involving torch.nn.functional.scaled_dot_product_attention failing with precision mismatch on fp16 input for XPU device"
Example: "P2 - aten.dot_xpu_mkl kernel NotImplementedError when called with Long tensors, indicating dtype support gap for aten.matmul operation on XPU"

Include specific details about: ops/functions involved, dtype transitions, arguments/parameters, failure patterns, severity indicators.

YOUR ANSWER:"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "Output ONLY 'Priority - reason'. No markdown. No JSON. No thinking tags."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 50
    }
    
    start_time = time.time()
    
    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time
        
        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            content = content.replace("<think>", "").replace("]", "").strip()
            
            match = re.search(r'(P[0-3])\s*[-–]\s*[^\n]+', content, re.IGNORECASE)
            if match:
                priority = match.group(1).upper()
                reason = content[match.start():].split('-')[-1].strip()[:50]
                return priority, reason, elapsed
            
            return "P2", "Default priority", elapsed
        
        return "P2", f"API Error: {response.status_code}", 0
        
    except Exception as e:
        return "P2", f"Error: {str(e)[:30]}", 0


def find_duplicated_issues(ws):
    """Find duplicated issues based on Test Class + Test Case or similar Traceback"""
    from collections import defaultdict
    
    # Build index: (Test Class, Test Case) -> [(row, issue_id)]
    class_case_index = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        issue_id = ws.cell(row, 1).value
        if test_class and test_case:
            key = (test_class, test_case)
            class_case_index[key].append((row, issue_id))
    
    # Build index for traceback similarity (excluding AssertionError about tensor-likes)
    traceback_index = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        traceback = ws.cell(row, 9).value or ''
        issue_id = ws.cell(row, 1).value
        
        # Skip if it's about "AssertionError: Tensor-likes are not close!"
        if 'AssertionError: Tensor-likes are not close!' in traceback:
            continue
        
        # Normalize traceback for comparison
        norm = traceback.strip()
        if norm:
            traceback_index[norm].append((row, issue_id))
    
    # Find duplicates: Test Class + Test Case
    class_case_duplicates = {}
    for key, rows in class_case_index.items():
        if len(rows) > 1:
            issue_ids = [rid for _, rid in rows]
            for row, rid in rows:
                other_issues = [i for i in issue_ids if i != rid]
                if other_issues:
                    class_case_duplicates[row] = other_issues
    
    # Find duplicates: similar Traceback
    traceback_duplicates = {}
    for key, rows in traceback_index.items():
        if len(rows) > 1:
            issue_ids = [rid for _, rid in rows]
            for row, rid in rows:
                other_issues = [i for i in issue_ids if i != rid]
                if other_issues:
                    if row in traceback_duplicates:
                        traceback_duplicates[row].extend(other_issues)
                    else:
                        traceback_duplicates[row] = other_issues
    
    # Merge duplicates
    merged_duplicates = {}
    for row in range(2, ws.max_row + 1):
        dup_set = set()
        if row in class_case_duplicates:
            dup_set.update(class_case_duplicates[row])
        if row in traceback_duplicates:
            dup_set.update(traceback_duplicates[row])
        if dup_set:
            merged_duplicates[row] = sorted(list(dup_set))
    
    return merged_duplicates


def determine_category(title, summary, test_cases_str, traceback, test_module, labels):
    """
    Determine the category of an issue based on its content.
    Categories:
    1. Dtype / Precision Related
    2. Sparse Operations Related
    3. Inductor / Compilation Related
    4. Flash Attention / Transformer Related
    5. PT2E
    6. Distributed
    7. TorchAO
    8. Others
    """
    text = f"{title} {summary} {test_cases_str} {traceback}".lower()
    labels_lower = str(labels).lower() if labels else ""
    
    # 1. Distributed - check first as distributed is a clear module
    distributed_keywords = [
        'distributed', 'device_mesh', 'ProcessGroup', 'FSDP', 'DDP', 'c10d',
        'tensor parallel', 'all_reduce', 'all_gather', 'reduce_scatter',
        'comm', 'rank', 'world_size', 'process group'
    ]
    if any(k in text for k in distributed_keywords):
        return "Distributed"
    
    # 2. TorchAO (quantization, optimizer, etc.)
    torchao_keywords = [
        'torchao', 'quantization', 'quantize', 'int8', 'int4', 'fp8',
        'optimizer', 'Adam', 'SGD', 'adamw', 'qat', 'lora', 'adapter'
    ]
    if any(k in text for k in torchao_keywords):
        return "TorchAO"
    
    # 3. PT2E (torch.export, ExportedProgram, fake tensors)
    pt2e_keywords = [
        'torch.export', 'export', 'exported', 'dynamo', 'fake_tensor',
        'graph_code', 'graph_submodule', 'capture', 'aot', 'aotautograd',
        'forward_from_graph', '_export', 'exported_program'
    ]
    if any(k in text for k in pt2e_keywords):
        return "PT2E"
    
    # 4. Flash Attention / Transformer Related
    flash_attention_keywords = [
        'flash', 'flash_attention', 'flashattention', 'sdpa', 'scaled_dot_product',
        'scaled_dot_product_attention', 'mem_eff', 'memory efficient',
        'transformer', 'attention', 'qwen', 'llama', 'bert', 'gpt',
        'mha', 'mqa', 'gqa', 'rope', 'rms_norm', 'layernorm',
        'linear', 'mlp', 'feed forward', 'feedforward'
    ]
    if any(k in text for k in flash_attention_keywords):
        return "Flash Attention / Transformer Related"
    
    # 5. Sparse Operations Related
    sparse_keywords = [
        'sparse', 'csr', 'csc', 'coo', 'sampled_addmm', 'sampled_addmm',
        'spmm', 'sparse_ops', 'sparse_matmul', 'torch.sparse'
    ]
    if any(k in text for k in sparse_keywords):
        return "Sparse Operations Related"
    
    # 6. Inductor / Compilation Related
    inductor_keywords = [
        'inductor', 'compile', 'compilation', 'codegen', 'triton',
        'kernel', 'loop', 'schedule', 'fx', 'graph', 'lower',
        'tile', 'vectorize', 'scheduler', 'abs_float'
    ]
    if any(k in text for k in inductor_keywords):
        return "Inductor / Compilation Related"
    
    # 7. Dtype / Precision Related
    dtype_precision_keywords = [
        'dtype', 'precision', 'accuracy', 'type promotion', 'typepromotion',
        'bf16', 'fp16', 'float16', 'float32', 'int8', 'int4', 'amp',
        'atomic', 'nan', 'inf', 'numerical', 'round', 'ceil', 'floor',
        'small', 'close', 'assertionerror'
    ]
    if any(k in text for k in dtype_precision_keywords):
        return "Dtype / Precision Related"
    
    # Default: Others
    return "Others"


def extract_version_info(issue_content):
    """
    Extract version info from issue content (title, summary, comments).
    Returns version string or None.
    """
    if not issue_content:
        return None

    version_patterns = [
        r'2\.\d+(\.\d+)?[a-z0-9]*',
        r'v2\.\d+(\.\d+)?',
        r'git[a-f0-9]+',
        r' nightly',
        r'a\d+',
    ]

    for pattern in version_patterns:
        match = re.search(pattern, issue_content, re.IGNORECASE)
        if match:
            return match.group(0)

    return None


def check_reproduce_step(issue_content):
    """
    Check if reproduce step is available in issue content.
    Returns True if reproduce step exists, False otherwise.
    """
    if not issue_content:
        return False

    content_lower = issue_content.lower()

    strong_indicators = [
        'pip install',
        'git clone',
        'git checkout',
        'python3',
        'benchmark',
        'reproducer',
        '```',
        'bash',
        'sh ',
        './',
        'conda',
        'pip3 install',
    ]

    medium_indicators = [
        'reproduce',
        'steps to reproduce',
        'how to reproduce',
        'minimal example',
        'test case',
        'run the',
        'execute',
        'command',
        'script',
        'code snippet',
    ]

    soft_indicators = [
        'git ',
        'python',
        'run ',
    ]

    strong_count = sum(1 for kw in strong_indicators if kw in content_lower)
    medium_count = sum(1 for kw in medium_indicators if kw in content_lower)
    soft_count = sum(1 for kw in soft_indicators if kw in content_lower)

    if strong_count >= 1:
        return True
    if strong_count >= 1 and medium_count >= 1:
        return True
    if medium_count >= 2:
        return True
    if medium_count >= 1 and soft_count >= 2:
        return True
    if soft_count >= 3:
        return True

    return False


def check_info_requested_to_reporter(issue_content):
    """
    Check if maintainer has requested more information from reporter.
    Returns True if info was requested from reporter.
    """
    if not issue_content:
        return False

    request_keywords = [
        'could you please provide',
        'please provide more',
        'can you provide additional',
        'need more information',
        'needs more info',
        'please add',
        'please share',
        'need the reproduce',
        'we need',
        'please attach',
        'please run',
        'please check',
        'please verify',
    ]

    content_lower = issue_content.lower()
    return any(kw in content_lower for kw in request_keywords)


def is_public_branch(version_str):
    """
    Check if version indicates a public branch (main, release) vs private branch/PR.
    Returns True if public branch, False if private.
    """
    if not version_str:
        return False

    version_lower = version_str.lower()

    if 'pr' in version_lower and 'http' in version_lower:
        return False

    if version_lower in ['main', 'master']:
        return True

    if re.match(r'^v?2\.\d+(\.\d+)?$', version_lower):
        return True

    if re.match(r'^\d+\.\d+\.\d+a\d+\+git[0-9a-f]+$', version_lower):
        return True

    if '+git' in version_lower and not 'pr' in version_lower:
        return True

    return False


def analyze_root_cause_llm(issue_id, issue_title, issue_summary, test_file, test_class, test_case, error_msg, traceback, test_module=None):
    """
    Use internal Qwen3-32B LLM to analyze root cause of an issue.
    Returns brief but specific root cause description.
    Logs to ~/ai_for_validation/opencode/issue_triage/result/root_cause.txt
    """
    import time
    import requests
    import os
    import re
    
    ROOT_CAUSE_LOG = os.path.expanduser('~/ai_for_validation/opencode/issue_triage/result/root_cause.txt')
    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxx")
    LLM_MODEL = "Qwen3-32B"
    
    if not issue_title and not issue_summary and not error_msg and not traceback:
        return ""
    
    def log_result(issue_id, root_cause, elapsed):
        msg = f"Issue {issue_id}: {root_cause} ({elapsed:.2f}s)"
        print(f"  {msg}")
        with open(ROOT_CAUSE_LOG, 'a') as log:
            log.write(msg + "\n")
    
    prompt = f"""You are analyzing PyTorch XPU issue root cause.

Issue ID: {issue_id}
Title: {issue_title}
Summary: {issue_summary}
Test: {test_class}.{test_case}
Error: {error_msg}
Traceback: {traceback[:500] if traceback else 'N/A'}

Classify into ONE category (internal classification):
1. Memory/Shared Memory Issue
2. Dtype/Precision Issue
3. Inductor/Compilation Issue
4. DNNL/OneDNN Issue
5. Flash Attention/Specific Ops Issue
6. Distributed/Gloo Issue
7. Skip/No Test Exists
8. Backend/Device Issue
9. API/Template Mismatch
10. Feature Not Supported
11. Timeout/Performance Issue
12. Runtime Error
13. Assertion Failure
14. Type/Value Error
15. Others

Answer format: "CATEGORY - detailed_reason" (full sentences, 150-300 characters)

Include in your reason:
- Specific PyTorch ops (aten.xxx), functions (torch.nn.functional.xxx), or APIs involved
- Dtypes that caused issues (float32, bf16, fp16, int8, Long, etc.)
- Arguments or parameters that triggered the failure
- Error patterns or signature mismatches
- Device-specific context (XPU, Inductor, Triton)

Example: "Dtype/Precision Issue - aten.memory_efficient_attention kernel fails when invoked with fp32 input tensors combined with bfloat16 scaling factor on XPU device. The torch.ops.scaled_dot_product_attention call encounters dtype promotion mismatch between query (float32), key (float32), and value (bfloat16 tensors) during kernel selection phase"
Example: "Backend/Device Issue - XPU device initialization fails during kernel compilation stage with sycl::queue creation error. The test case involves torch.nn.functional.conv2d with input tensor of ndarry type requiring device placement from CPU to XPU"

YOUR ANSWER (no JSON, no thinking tags, just the answer):"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    data = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "Output ONLY 'Category - reason'. No markdown. No JSON. No thinking tags."},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 400,
        "temperature": 0.0
    }
    
    start = time.time()
    response = None
    try:
        max_retries = 3
        for attempt in range(max_retries):
            response = requests.post(LLM_ENDPOINT, headers=headers, json=data, timeout=180)
            elapsed = time.time() - start
            
            if response.status_code == 200:
                break
            elif response.status_code in (403, 429):
                wait_time = 60
                print(f"    [Rate Limit] Waiting {wait_time}s (attempt {attempt+1}/{max_retries})...")
                time.sleep(wait_time)
            else:
                elapsed = time.time() - start
                log_result(issue_id, f"API Error: {response.status_code}", elapsed)
                return ""
        
        if response is None or response.status_code != 200:
            elapsed = time.time() - start
            log_result(issue_id, f"API Error: {response.status_code if response else 'No response'}", elapsed)
            return ""
        
        resp_data = response.json()
        content = resp_data.get("choices", [{}])[0].get("message", {}).get("content", "")
        
        content = content.replace("<think>", "").replace("]", "")
        content = content.replace("</think>", "").strip()
        
        match = re.search(r'(Memory|Dtype|Precision|Inductor|Compilation|DNNL|OneDNN|Flash|Attention|Distributed|Gloo|Backend|Device|API|Template|Mismatch|Feature|Not|Supported|Timeout|Performance|Runtime|Error|Assertion|Failure|Type|Value|Others)\s*[-_\-\u2013]\s*.{50,}', content, re.IGNORECASE)
        
        if match:
            root_cause = match.group(0).strip()
        else:
            lines = [l.strip() for l in content.split("\n") if l.strip() and len(l.strip()) > 20]
            root_cause = lines[-1][:300] if lines else content.strip()[:300]
        
        log_result(issue_id, root_cause, elapsed)
        return root_cause
            
    except Exception as e:
        elapsed = time.time() - start
        log_result(issue_id, f"Exception: {str(e)[:50]}", elapsed)
        return ""


def analyze_root_cause(issue_title, issue_summary, test_file, test_class, test_case, error_msg, traceback, test_module=None):
    """
    Determine root cause category based on issue information using keyword matching.
    Returns root cause description.
    """
    import re
    
    text = f"{issue_title} {issue_summary} {error_msg or ''} {traceback or ''}".lower()
    
    if any(k in text for k in ['out of memory', 'oom', 'alloc', 'memory', 'cuda out of memory']):
        if 'shared' in text:
            return "Memory/Shared Memory Issue - shared memory allocation failed"
        return "Memory/Shared Memory Issue"
    
    if any(k in text for k in ['precision', 'dtype', 'accuracy', 'numerical', 'fp16', 'bf16', 'float16']):
        return "Dtype/Precision Issue"
    
    if any(k in text for k in ['graph break', 'inductor', 'compile', 'symbolic', 'fx pass']):
        return "Inductor/Compilation Issue"
    
    if 'dnnl' in text or 'onednn' in text:
        return "DNNL/OneDNN Issue"
    
    if 'flash attention' in text or 'flash_attn' in text or 'fused_attention' in text:
        return "Flash Attention/Specific Ops Issue"
    
    if any(k in text for k in ['distributed', 'gloo', 'nccl', 'all_reduce', 'all_gather']):
        return "Distributed/Gloo Issue"
    
    if any(k in text for k in ['skip', 'decorator', 'xfail', 'expected failure']):
        return "Skip/No Test Exists"
    
    if any(k in text for k in ['device', 'xpu', 'placement', 'cuda']) and ('init' in text or 'set' in text or 'empty_cache' in text):
        return "Backend/Device Issue"
    
    if any(k in text for k in ['api', 'template', 'signature', 'missing parameter']):
        return "API/Template Mismatch"
    
    if any(k in text for k in ['not supported', 'unimplemented', 'not implement']):
        return "Feature Not Supported"
    
    if any(k in text for k in ['timeout', 'slow', 'performance', 'hang']):
        return "Timeout/Performance Issue"
    
    if 'runtime error' in text or 'runtimewarning' in text:
        return "Runtime Error"
    
    if 'assertion' in text:
        return "Assertion Failure"
    
    if any(k in text for k in ['typeerror', 'valueerror', 'type error', 'value error']):
        return "Type/Value Error"
    
    return ""


def analyze_root_cause(issue_title, issue_summary, test_file, test_class, test_case, error_msg, traceback, test_module=None):
    """
    Analyze root cause of an issue based on available information.
    Returns brief but specific root cause description.
    """
    if not any([error_msg, traceback, issue_summary]):
        return ""
    
    text = ' '.join([
        str(issue_title or ''),
        str(issue_summary or ''),
        str(error_msg or ''),
        str(traceback or ''),
        str(test_file or ''),
        str(test_class or ''),
        str(test_case or '')
    ]).lower()
    
    if 'permanent kill' in text or 'xgboost' in text:
        return "XGBoost/External Dependency"
    
    if 'out of memory' in text or 'cannot allocate memory' in text or 'allocation fails' in text:
        return "Memory/Shared Memory Issue"
    
    if any(k in text for k in ['requires xccl', 'no xccl', 'xccl not found']):
        return "XCCL/Dependency Issue"
    
    if any(k in text for k in ['inductor', 'compile', 'graph break', 'symbolic', 'fx']) and 'test/inductor' in str(test_file or '').lower():
        return "Inductor / Compilation Issue"
    
    if any(k in text for k in ['dynamo']) and ('test/dynamo' in str(test_file or '').lower() or 'testinductor' in str(test_file or '').lower()):
        return "Inductor / Compilation Issue"
    
    if 'dnnl' in text or 'onednn' in text or 'mkldnn' in text:
        return "DNNL/OneDNN Issue"
    
    if any(k in text for k in ['flash attention', 'flash_attention', 'flashattn']) and 'attention' in text:
        return "Flash Attention / Specific Ops Issue"
    
    if 'distributed' in text or 'gloo' in text or ' nccl' in text or 'nccl' in text:
        if 'test/distributed' in str(test_file or '').lower():
            return "Distributed / Gloo Issue"
    
    if 'dtype' in text or 'precision' in text or 'accuracy' in text or 'typepromotion' in text:
        return "Dtype / Precision Issue"
    
    if 'float' in text and ('16' in text or '32' in text or 'bf' in text):
        return "Dtype / Precision Issue"
    
    if 'skip' in text or 'decorator' in text or ('test' in text and 'not found' in text):
        return "Skip / No Test Exists"
    
    if 'device' in text or 'xpu' in text and 'init' in text:
        return "Backend / Device Issue"
    
    if 'api' in text or 'template' in text or 'signature' in text:
        return "API / Template Mismatch"
    
    if 'not implemented' in text or 'not support' in text:
        return "Feature Not Supported"
    
    if 'import error' in text or 'no module' in text:
        return "Import / Dependency Issue"
    
    if 'assertionerror' in text or 'assert' in text:
        if 'dtype' in text or 'precision' in text or 'float' in text:
            return "Dtype / Precision Issue"
        return "Assertion Failure"
    
    if 'runtimeerror' in text:
        if 'memory' in text or 'allocate' in text:
            return "Memory/Shared Memory Issue"
        if 'dtype' in text or 'cast' in text or 'convert' in text:
            return "Dtype / Precision Issue"
        if 'inductor' in text or 'compile' in text:
            return "Inductor / Compilation Issue"
        return "Runtime Error"
    
    if 'typeerror' in text or 'valueerror' in text:
        if 'dtype' in text or 'cast' in text:
            return "Dtype / Precision Issue"
        return "Type / Value Error"
    
    if 'timeout' in text:
        return "Timeout / Performance Issue"
    
    if 'test/nn' in str(test_file or '').lower():
        if 'conv' in str(test_case or '').lower() or 'linear' in str(test_case or '').lower():
            return "DNNL / Specific Ops Issue"
    
    return "Others"


def process_issues_sheet(wb):
    """Process Issues sheet to add owner_transfer, action_TBD, and duplicated_issue columns"""
    ws_issues = wb['Issues']
    ws_test = wb['Test Cases']

    ws_issues.cell(1, 19, 'owner_transfer')
    ws_issues.cell(1, 20, 'action_TBD')
    ws_issues.cell(1, 21, 'duplicated_issue')
    ws_issues.cell(1, 22, 'priority')
    ws_issues.cell(1, 23, 'priority_reason')
    ws_issues.cell(1, 24, 'Category')
    ws_issues.cell(1, 25, 'category_reason')
    ws_issues.cell(1, 26, 'Root Cause')

    MAX_LLM_ROOT_CAUSE = 500
    MAX_LLM_CATEGORY = 500
    MAX_LLM_PRIORITY = 500
    llm_root_cause_count = 0
    llm_category_count = 0
    llm_priority_count = 0
    root_cause_cache = {}
    
    issue_test_results = {}
    for row in range(2, ws_test.max_row + 1):
        issue_id = ws_test.cell(row, 1).value
        xpu_status = ws_test.cell(row, 11).value
        stock_status = ws_test.cell(row, 16).value
        cuda_case_exist = ws_test.cell(row, 18).value
        dup_issue = ws_test.cell(row, 21).value
        
        if issue_id not in issue_test_results:
            issue_test_results[issue_id] = {
                'xpu_statuses': set(),
                'stock_statuses': set(),
                'cuda_case_not_exist': False,
                'duplicated_issues': set()
            }
        
        if xpu_status:
            issue_test_results[issue_id]['xpu_statuses'].add(xpu_status)
        if stock_status:
            issue_test_results[issue_id]['stock_statuses'].add(stock_status)
        if cuda_case_exist == 'No':
            issue_test_results[issue_id]['cuda_case_not_exist'] = True
        if dup_issue:
            for dup_id in str(dup_issue).split(','):
                issue_test_results[issue_id]['duplicated_issues'].add(dup_id.strip())
    
    ws_e2e = wb['E2E Test Cases']
    for row in range(2, ws_e2e.max_row + 1):
        issue_id = ws_e2e.cell(row, 1).value
        e2e_status = ws_e2e.cell(row, 13).value
        
        if issue_id not in issue_test_results:
            issue_test_results[issue_id] = {
                'xpu_statuses': set(),
                'stock_statuses': set(),
                'cuda_case_not_exist': False,
                'duplicated_issues': set(),
                'e2e_statuses': set()
            }
        
        if 'e2e_statuses' not in issue_test_results[issue_id]:
            issue_test_results[issue_id]['e2e_statuses'] = set()
        
        if e2e_status:
            issue_test_results[issue_id]['e2e_statuses'].add(e2e_status)
    
    issue_prs = {}
    issue_reporters = {}
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        pr = ws_issues.cell(row, 15).value
        reporter = ws_issues.cell(row, 5).value
        assignee = ws_issues.cell(row, 4).value
        
        issue_prs[issue_id] = pr
        issue_reporters[issue_id] = reporter
    
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        pr = ws_issues.cell(row, 15).value
        reporter = ws_issues.cell(row, 5).value
        assignee = ws_issues.cell(row, 4).value
        labels = ws_issues.cell(row, 6).value
        title_raw = str(ws_issues.cell(row, 2).value) if ws_issues.cell(row, 2).value else ''
        module = str(ws_issues.cell(row, 12).value) if ws_issues.cell(row, 12).value else ''
        summary_raw = str(ws_issues.cell(row, 10).value) if ws_issues.cell(row, 10).value else ''
        test_module_raw = ws_issues.cell(row, 13).value
        test_module = str(test_module_raw) if test_module_raw else ''
        traceback = ''
        for tr in range(2, ws_test.max_row + 1):
            if ws_test.cell(tr, 1).value == issue_id:
                traceback = str(ws_test.cell(tr, 9).value) if ws_test.cell(tr, 9).value else ''
                break
        
        test_cases_str = ''
        for tr in range(2, ws_test.max_row + 1):
            if ws_test.cell(tr, 1).value == issue_id:
                tc = str(ws_test.cell(tr, 7).value) if ws_test.cell(tr, 7).value else ''
                test_cases_str += ' ' + tc
        
        test_info = issue_test_results.get(issue_id, {
            'xpu_statuses': set(),
            'stock_statuses': set(),
            'cuda_case_not_exist': False,
            'duplicated_issues': set()
        })
        
        xpu_statuses = test_info['xpu_statuses']
        stock_statuses = test_info['stock_statuses']
        cuda_case_not_exist = test_info['cuda_case_not_exist']
        duplicated_issues = test_info['duplicated_issues']
        e2e_statuses = test_info.get('e2e_statuses', set())
        
        owner_transfer = ''
        action_tbd = ''
        
        labels_str = str(labels).lower() if labels else ''
        title_lower = title_raw.lower()
        summary_lower = summary_raw.lower() if summary_raw else ''
        
        is_not_target = ('not target' in labels_str or 'wont' in labels_str or "won't" in labels_str)
        
        if is_not_target:
            owner_transfer = reporter
            action_tbd = 'add to skiplist'
        
        is_random = 'random' in labels_str
        ut_passed = False
        if not is_random and xpu_statuses and stock_statuses:
            xpu_all_passed = (xpu_statuses == {'passed'})
            stock_all_passed = (stock_statuses == {'passed'})
            
            if xpu_all_passed or stock_all_passed:
                owner_transfer = reporter
                action_tbd = 'Close fixed issue'
                ut_passed = True
        
        e2e_all_passed = all(s == 'pass' for s in e2e_statuses) if e2e_statuses else False
        if not ut_passed and e2e_all_passed:
            owner_transfer = reporter
            action_tbd = 'Close fixed issue'
        
        pr_status = ws_issues.cell(row, 17).value
        pr_closed = pr_status in ['closed', 'merged']
        
        if not owner_transfer and pr_closed:
            has_failed = ('failed' in xpu_statuses) or ('failed' in stock_statuses)
            
            if not has_failed:
                owner_transfer = reporter
                action_tbd = 'Verify the issue'
            else:
                owner_transfer = assignee
                action_tbd = 'Revisit the PR as case failed'
        
        is_upstream = 'ut_upstream' in labels_str or 'inductor' in labels_str
        is_wontfix = 'wont ' in labels_str or ' wont ' in labels_str or 'wontfix' in labels_str or 'not target' in labels_str.replace('nottarget', '')
        is_not_target_upstream = is_not_target and is_upstream
        
        if not action_tbd:
            if is_not_target_upstream:
                action_tbd = 'Needs Upstream Skip PR (not_target + ut_upstream)'
                owner_transfer = assignee
            elif is_wontfix:
                action_tbd = 'Needs Skip PR (wontfix / not_target)'
                owner_transfer = assignee
            elif is_upstream:
                action_tbd = ''
                owner_transfer = assignee
        
        priority = 'P2'
        priority_reason = ''
        
        is_model_issue = ('model' in title_raw.lower() or 'model' in summary_raw.lower() or 
                        'application' in title_raw.lower() or 'application' in summary_raw.lower() or
                        'huggingface' in title_raw.lower() or 'timm' in title_raw.lower() or 'torchbench' in title_raw.lower())
        is_e2e = test_module == 'e2e'
        is_ut = test_module == 'ut'
        
        is_regression = ('regression' in labels_str or 'regression' in title_raw.lower() or 
                        'was pass' in summary_raw.lower() or 'previously pass' in summary_raw.lower() or
                        ('before' in summary_raw.lower() and 'now' in summary_raw.lower()))
        
        is_build_crash = ('build' in test_module.lower() or 'build' in title_raw.lower() or 
                         'crash' in title_raw.lower() or 'segmentation' in title_raw.lower() or
                         'segfault' in title_raw.lower() or 'signal' in summary_raw.lower())
        
        failed_count = 0
        for tr in range(2, ws_test.max_row + 1):
            if ws_test.cell(tr, 1).value == issue_id:
                tc_status = ws_test.cell(tr, 11).value
                if tc_status in ['failed', 'error']:
                    failed_count += 1
        
        if is_build_crash:
            priority = 'P0'
            priority_reason = 'Build crash - critical blocking issue'
        elif is_model_issue and not ('test' in title_raw.lower() and 'case' in title_raw.lower()):
            priority = 'P0'
            priority_reason = 'Impacts real model/application'
        elif is_regression:
            priority = 'P0'
            priority_reason = 'Regression - passed before but failed now'
        elif is_e2e and ('accuracy' in title_raw.lower() or 'accuracy' in summary_raw.lower() or 
                        'fail' in title_raw.lower() or 'fail' in summary_raw.lower()):
            priority = 'P1'
            priority_reason = 'E2E benchmark accuracy/functionality issue'
        elif is_e2e and ('performance' in title_raw.lower() or 'slow' in title_raw.lower() or 
                        'latency' in title_raw.lower()):
            priority = 'P2'
            priority_reason = 'E2E benchmark performance issue'
        elif is_ut and failed_count > 20:
            priority = 'P1'
            priority_reason = f'UT with {failed_count} failed test cases'
        else:
            priority = 'P2'
            priority_reason = 'UT issue with few failures'
        
        if llm_priority_count < MAX_LLM_PRIORITY:
            test_cases_for_llm = []
            for tr in range(2, ws_test.max_row + 1):
                if ws_test.cell(tr, 1).value == issue_id:
                    tc_info = {
                        'test_case': ws_test.cell(tr, 7).value,
                        'error_msg': ws_test.cell(tr, 8).value,
                        'traceback': ws_test.cell(tr, 9).value
                    }
                    test_cases_for_llm.append(tc_info)
            llm_priority, llm_reason, _ = determine_priority_llm(
                title_raw, summary_raw, error_msg if 'error_msg' in dir() else '',
                test_module, str(labels) if labels else '',
                test_cases_for_llm
            )
            if llm_priority.startswith('P') and not llm_reason.startswith('API'):
                priority = llm_priority
                priority_reason = llm_reason
                llm_priority_count += 1
                print(f"  [LLM PRIORITY #{llm_priority_count}] Issue {issue_id}: {priority} - {priority_reason}")
        
        if duplicated_issues:
            ws_issues.cell(row, 21, ','.join(sorted(duplicated_issues)))
        
        if not action_tbd:
            version_info = extract_version_info(title_raw + ' ' + summary_raw)
            is_public = is_public_branch(version_info)
            
            has_xpu_status = bool(xpu_statuses)
            has_stock_status = bool(stock_statuses)
            has_e2e_status = bool(e2e_statuses)
            
            is_e2e_issue = test_module_raw == 'e2e'
            
            all_statuses_empty = not has_xpu_status and not has_stock_status and not has_e2e_status
            
            is_feature_request = (
                ('feature' in title_lower or 'feature' in summary_lower) or
                ('request' in title_lower or 'request' in summary_lower) or
                ('implement' in title_lower or 'implement' in summary_lower) or
                ('add support' in title_lower or 'add support' in summary_lower) or
                ('enable' in title_lower and 'test' not in title_lower and 'feature' in summary_lower) or
                (labels and 'enhancement' in str(labels).lower())
            )
            
            # E2E issues are always tracked for upstream changes - they have their own test framework
            if is_e2e_issue:
                # For all E2E issues: if they have pass status and are resolved, close them
                # Regardless of whether they have status data, E2E issues don't need case availability check
                if is_e2e_issue and has_e2e_status and e2e_all_passed:
                    owner_transfer = reporter
                    action_tbd = 'Close fixed issue'
                else:
                    # E2E issue needs upstream investigation
                    action_tbd = ''
                    owner_transfer = assignee
            elif is_public and all_statuses_empty and not is_feature_request:
                if e2e_all_passed:
                    owner_transfer = reporter
                    action_tbd = 'Close fixed issue'
                else:
                    action_tbd = ''
                    owner_transfer = assignee
            elif is_public and all_statuses_empty and not is_feature_request and not is_e2e_issue:
                action_tbd = 'Check case availability'
                owner_transfer = reporter
            
            if not action_tbd and not is_feature_request and not is_e2e_issue:
                issue_content = title_raw + ' ' + summary_raw
                
                is_bug_or_perf = any(kw in title_lower or kw in summary_lower for kw in [
                    'bug', 'fail', 'error', 'crash', 'assertion', 'exception',
                    'performance', 'slow', 'latency', 'timeout', 'regression',
                    'accuracy', 'wrong result', 'precision', 'dtype'
                ])
                
                has_already_requested = check_info_requested_to_reporter(issue_content)
                llm_error_msg = ''
                llm_traceback = ''
                test_file = ''
                test_case = ''
                for tr in range(2, ws_test.max_row + 1):
                    if ws_test.cell(tr, 1).value == issue_id:
                        llm_error_msg = str(ws_test.cell(tr, 8).value) if ws_test.cell(tr, 8).value else ''
                        llm_traceback = str(ws_test.cell(tr, 9).value) if ws_test.cell(tr, 9).value else ''
                        test_file = str(ws_test.cell(tr, 4).value) if ws_test.cell(tr, 4).value else ''
                        test_case = str(ws_test.cell(tr, 7).value) if ws_test.cell(tr, 7).value else ''
                        break
                llm_info_action = check_info_requested_to_reporter_llm(
                    title_raw, summary_raw, llm_error_msg, llm_traceback
                )
                if llm_info_action and llm_info_action not in ['Ready to analyze']:
                    has_already_requested = True
                    if 'reproduce' in llm_info_action.lower():
                        action_tbd = 'Need reproduce steps'
                    else:
                        action_tbd = f'LLM Suggestion: {llm_info_action}'
                    owner_transfer = reporter

                has_test_info = bool(test_file and test_case)
                
                if is_bug_or_perf:
                    
                    if has_already_requested and 'reproduce' in llm_info_action.lower():
                        action_tbd = 'Need reproduce steps'
                        owner_transfer = reporter
                    elif has_test_info:
                        # Test info provided, can proceed with upstream changes
                        action_tbd = ''
                    elif has_already_requested:
                        action_tbd = 'Awaiting response from reporter'
                        owner_transfer = reporter
                    else:
                        action_tbd = ''
                else:
                    if has_already_requested:
                        action_tbd = 'Awaiting response from reporter'
                        owner_transfer = reporter
                    else:
                        info_needed = []
                        if 'accuracy' in title_lower or 'accuracy' in summary_lower:
                            info_needed.append('accuracy comparison data')
                        if 'performance' in title_lower or 'performance' in summary_lower:
                            info_needed.append('performance numbers/baseline')
                        if 'regression' in title_lower or 'regression' in summary_lower:
                            info_needed.append('previous good version info')
                        
                        if info_needed:
                            info_str = ', '.join(info_needed)
                            action_tbd = f'Need more information - {info_str}'
                            owner_transfer = reporter
        
        if owner_transfer:
            ws_issues.cell(row, 19, owner_transfer)
        if action_tbd:
            ws_issues.cell(row, 20, action_tbd)
        # Remove old format entries that should be replaced
        old_format = 'Need reproduce step and more information'
        current_cell_value = ws_issues.cell(row, 20).value
        if current_cell_value == old_format:
            ws_issues.cell(row, 20, action_tbd if action_tbd else '')
        
        ws_issues.cell(row, 22, priority)
        ws_issues.cell(row, 23, priority_reason)
        
        category = determine_category(title_raw, summary_raw, test_cases_str, traceback, test_module, labels)
        category_reason = ''
        if llm_category_count < MAX_LLM_CATEGORY:
            test_cases_for_llm = []
            for tr in range(2, ws_test.max_row + 1):
                if ws_test.cell(tr, 1).value == issue_id:
                    tc_info = {
                        'test_case': ws_test.cell(tr, 7).value,
                        'test_file': ws_test.cell(tr, 4).value,
                        'error_msg': ws_test.cell(tr, 8).value,
                        'torch_ops': ws_test.cell(tr, 10).value
                    }
                    test_cases_for_llm.append(tc_info)
            category_llm, category_reason = determine_category_llm(
                title_raw, summary_raw, test_cases_for_llm, test_module, labels
            )
            if category_llm and not category_llm.startswith('API') and not category_llm.startswith('Error'):
                category = category_llm
                llm_category_count += 1
                print(f"  [LLM CATEGORY #{llm_category_count}] Issue {issue_id}: {category}")
        ws_issues.cell(row, 24, category)
        ws_issues.cell(row, 25, category_reason or '')
        
        current_action_tbd = ws_issues.cell(row, 20).value
        if not current_action_tbd:
            issue_test_file = ''
            issue_test_class = ''
            issue_test_case = ''
            issue_error_msg = ''
            issue_traceback = ''
            
            for tr in range(2, ws_test.max_row + 1):
                if ws_test.cell(tr, 1).value == issue_id:
                    if not issue_test_file:
                        issue_test_file = str(ws_test.cell(tr, 4).value) if ws_test.cell(tr, 4).value else ''
                    if not issue_test_class:
                        issue_test_class = str(ws_test.cell(tr, 6).value) if ws_test.cell(tr, 6).value else ''
                    issue_test_case = str(ws_test.cell(tr, 7).value) if ws_test.cell(tr, 7).value else ''
                    if not issue_error_msg:
                        issue_error_msg = str(ws_test.cell(tr, 8).value) if ws_test.cell(tr, 8).value else ''
                    if not issue_traceback:
                        issue_traceback = str(ws_test.cell(tr, 9).value) if ws_test.cell(tr, 9).value else ''
                    if issue_error_msg and issue_traceback:
                        break
            
            root_cause = analyze_root_cause(
                title_raw,
                summary_raw,
                issue_test_file,
                issue_test_class,
                issue_test_case,
                issue_error_msg,
                issue_traceback,
                test_module
            )
            
            if llm_root_cause_count < MAX_LLM_ROOT_CAUSE:
                print(f"  [LLM ROOT CAUSE #{llm_root_cause_count+1}] Issue {issue_id}: Calling LLM for root cause analysis...")
                llm_start = time.time()
                llm_root_cause = analyze_root_cause_llm(
                    issue_id,
                    title_raw,
                    summary_raw,
                    issue_test_file,
                    issue_test_class,
                    issue_test_case,
                    issue_error_msg,
                    issue_traceback,
                    test_module
                )
                llm_elapsed = time.time() - llm_start
                if llm_root_cause:
                    root_cause = llm_root_cause
                    llm_root_cause_count += 1
                    print(f"  [LLM ROOT CAUSE #{llm_root_cause_count}] Issue {issue_id} -> '{llm_root_cause}' ({llm_elapsed:.2f}s)")
                else:
                    print(f"  [LLM ROOT CAUSE #{llm_root_cause_count+1}] Issue {issue_id}: LLM returned empty")
            
            if root_cause:
                ws_issues.cell(row, 26, root_cause)
    
    print(f"Processed {ws_issues.max_row - 1} issues")
    return wb


def process_e2e_cases(wb):
    """Process E2E Test Cases sheet to add accuracy status from torch-xpu-ops nightly"""
    ws_e2e = wb['E2E Test Cases']
    
    # Add new column for accuracy status
    ws_e2e.cell(1, 13, 'torch-xpu-ops nightly status - accuracy')
    
    import glob
    import os
    base_dir = '/home/daisydeng/issue_traige/ci_results/torch-xpu-ops'
    
    # Build mapping: (benchmark, dtype, amp, phase, model) -> status
    # Phase: inf (inference), tra (training)
    e2e_model_status = {}
    
    for report_path in glob.glob(f'{base_dir}/*E2E*/Inductor_E2E_Test_Report.xlsx'):
        try:
            # Extract benchmark from path
            dirname = os.path.basename(os.path.dirname(report_path))
            if 'huggingface' in dirname:
                benchmark = 'huggingface'
            elif 'timm' in dirname:
                benchmark = 'timm_models'
            elif 'torchbench' in dirname:
                benchmark = 'torchbench'
            else:
                continue
            
            report_wb = openpyxl.load_workbook(report_path)
            
            # Parse all accuracy sheets (sheets ending with _acc)
            for sheet_name in report_wb.sheetnames:
                if not sheet_name.endswith('_acc'):
                    continue
                
                # Parse sheet name: huggingface_float32_inf_acc -> benchmark=huggingface, dtype=float32, phase=inf
                # Also: huggingface_amp_bf16_inf_acc -> benchmark=huggingface, dtype=bf16, amp=True, phase=inf
                parts = sheet_name.replace(f'{benchmark}_', '').replace('_acc', '').split('_')
                
                dtype = 'float32'
                amp = False
                phase = ''
                
                if 'amp' in parts:
                    amp = True
                    idx = parts.index('amp')
                    if idx + 1 < len(parts):
                        dtype = parts[idx + 1]
                else:
                    if len(parts) >= 1:
                        dtype = parts[0]
                
                if 'inf' in parts:
                    phase = 'inf'
                elif 'tra' in parts:
                    phase = 'tra'
                
                # Parse model status
                ws = report_wb[sheet_name]
                for row in range(3, ws.max_row + 1):  # Skip header rows
                    model_name = ws.cell(row, 2).value  # Column B: name
                    status = ws.cell(row, 4).value  # Column D: target/accuracy
                    
                    if model_name and status:
                        key = (benchmark, dtype, amp, phase, model_name.lower())
                        e2e_model_status[key] = status
        except Exception as e:
            print(f"  Warning: Failed to parse {report_path}: {e}")
    
    print(f"  Found {len(e2e_model_status)} E2E model status entries")
    
    # Map E2E test case row to status
    matched = 0
    for row in range(2, ws_e2e.max_row + 1):
        benchmark = ws_e2e.cell(row, 3).value  # Benchmark column
        model = ws_e2e.cell(row, 4).value  # Model column
        dtype = ws_e2e.cell(row, 6).value  # Dtype column
        amp = ws_e2e.cell(row, 7).value  # AMP column
        phase = ws_e2e.cell(row, 5).value  # Phase column
        
        if benchmark and model and dtype and phase:
            # Normalize
            dtype_key = dtype.lower().replace(' ', '_')
            phase_key = phase.lower().replace(' ', '_').replace('inference', 'inf').replace('training', 'tra')
            model_key = model.lower()
            
            # Try exact match first
            key = (benchmark, dtype_key, bool(amp), phase_key, model_key)
            if key in e2e_model_status:
                ws_e2e.cell(row, 13, e2e_model_status[key])
                matched += 1
            else:
                # Try without amp
                key = (benchmark, dtype_key, False, phase_key, model_key)
                if key in e2e_model_status:
                    ws_e2e.cell(row, 13, e2e_model_status[key])
                    matched += 1
    
def main():
    import time as time_module
    pipeline_start = time_module.time()
    
    # Load workbook
    wb = openpyxl.load_workbook(os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx'))
    ws = wb['Test Cases']
    
    # Add new columns
    ws.cell(1, 11, 'status in torch-xpu-ops nightly')
    ws.cell(1, 12, 'comments in torch-xpu-ops nightly')
    ws.cell(1, 13, 'Commit')
    ws.cell(1, 14, 'Run_id')
    ws.cell(1, 15, 'XML')
    ws.cell(1, 16, 'status in stock CI')
    ws.cell(1, 17, 'comments in stock CI')
    ws.cell(1, 18, 'cuda_case_exist')
    ws.cell(1, 19, 'xpu_case_exist')
    ws.cell(1, 20, 'case_existence_comments')
    ws.cell(1, 21, 'duplicated_issue')
    
    # Get XML files
    log("\n" + "=" * 60)
    log("[STEP 1/5] Loading XML files...")
    xpu_start = time_module.time()
    xpu_xml_files = get_torch_xpu_ops_xml_files()
    xpu_elapsed = time_module.time() - xpu_start
    log(f"  Found {len(xpu_xml_files)} torch-xpu-ops XML files ({xpu_elapsed:.1f}s)")
    
    stock_xml_files = get_stock_xml_files()
    log(f"  Found {len(stock_xml_files)} stock XML files")
    
    # TWO-PASS APPROACH for case existence
    # Pass 1: Process CI results and collect unique issues needing LLM
    # Pass 2: Run LLM ONCE per unique issue, then apply to all test cases
    
    total = ws.max_row - 1
    log("\n" + "=" * 60)
    log(f"[STEP 2/5] Processing {total} test cases (two-pass approach)...")
    log(f"  [START TIME] {time_module.strftime('%Y-%m-%d %H:%M:%S')}")
    
    pass1_start = time_module.time()
    issues_needing_llm = {}
    
    log("  [PASS 1/2] Processing CI results and collecting unique issues...")
    for i, row in enumerate(range(2, ws.max_row + 1), 1):
        test_file = ws.cell(row, 4).value
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        issue_id = ws.cell(row, 1).value
        origin_test_file = ws.cell(row, 5).value
        
        xml_prefix, reason = convert_test_file_to_xml_prefix(test_file)
        if xml_prefix:
            matched = find_best_xml_match(xml_prefix, xpu_xml_files)
            if matched:
                xml_path, commit, run_id, _ = matched
                status, error_msg, traceback = get_test_result(xml_path, test_case)
                ws.cell(row, 11, status)
                ws.cell(row, 12, error_msg if error_msg else '')
                if traceback and not ws.cell(row, 9).value:
                    ws.cell(row, 9, traceback[:3000])
                ws.cell(row, 13, commit)
                ws.cell(row, 14, run_id)
                ws.cell(row, 15, os.path.basename(xml_path))
            else:
                ws.cell(row, 11, 'not found')
                ws.cell(row, 12, f'No XML: {xml_prefix}')
        else:
            ws.cell(row, 11, 'not found')
            ws.cell(row, 12, reason)

        stock_prefix = convert_to_stock_prefix(test_file)
        if stock_prefix and stock_prefix in stock_xml_files:
            stock_xml = stock_xml_files[stock_prefix]
            stock_status, stock_error_msg, stock_traceback = get_test_result(stock_xml, test_case)
            ws.cell(row, 16, stock_status)
            ws.cell(row, 17, stock_error_msg if stock_error_msg else '')
            if stock_traceback and not ws.cell(row, 9).value:
                ws.cell(row, 9, stock_traceback[:3000])
        else:
            ws.cell(row, 16, 'not found')
            ws.cell(row, 17, 'Not in stock CI')
        
        xpu_status = ws.cell(row, 11).value
        stock_status = ws.cell(row, 16).value
        ci_not_found = (xpu_status == 'not found' or not xpu_status) and (stock_status == 'not found' or not stock_status)
        
        if ci_not_found and issue_id not in issues_needing_llm:
            issues_needing_llm[issue_id] = {
                'test_file': test_file,
                'test_class': test_class,
                'test_case': test_case,
                'origin_test_file': origin_test_file
            }
        
        if i % 500 == 0:
            log(f"    Progress: {i}/{total} cases, {len(issues_needing_llm)} unique issues needing LLM")
    
    pass1_elapsed = time_module.time() - pass1_start
    issue_count = len(issues_needing_llm)
    log(f"  PASS 1 complete: {issue_count} unique issues ({pass1_elapsed:.1f}s)")
    
    pass2_start = time_module.time()
    log(f"  [PASS 2/2] Running LLM for {issue_count} unique issues...")
    llm_results = {}
    
    for idx, (issue_id, info) in enumerate(issues_needing_llm.items(), 1):
        try:
            test_file = info['test_file']
            test_class = info['test_class']
            test_case = info['test_case']
            origin_test_file = info['origin_test_file']
            
            result = analyze_test_case_with_llm_qwen(test_file, test_class, test_case, origin_test_file)
            elapsed = result.get('elapsed_time', 0)
            llm_results[issue_id] = result
            log(f"    [LLM {idx}/{issue_count}] Issue {issue_id}: {elapsed:.1f}s")
        except Exception as e:
            log(f"    [LLM ERROR] Issue {issue_id}: {e}")
            llm_results[issue_id] = {'error': str(e)}
    
    pass2_elapsed = time_module.time() - pass2_start
    log(f"  PASS 2 complete: {len(llm_results)} LLM results ({pass2_elapsed:.1f}s)")
    
    log("  [APPLY] Writing LLM results to all test cases...")
    applied = 0
    skipped = 0
    for row in range(2, ws.max_row + 1):
        issue_id = ws.cell(row, 1).value
        xpu_status = ws.cell(row, 11).value
        stock_status = ws.cell(row, 16).value
        ci_not_found = (xpu_status == 'not found' or not xpu_status) and (stock_status == 'not found' or not stock_status)
        
        # Only fill case_existence when LLM was actually called for this issue
        if ci_not_found and issue_id in llm_results:
            result = llm_results[issue_id]
            cuda_exists = result.get('cuda_exists', 'Unknown')
            xpu_exists = result.get('xpu_exists', 'Unknown')
            
            parts = []
            explanation = result.get('explanation', '')
            if explanation:
                parts.append('explanation: ' + explanation + '\n')
 
            for key in ['base_test_name', 'cuda_test_file', 'xpu_test_file', 'cuda_test_name', 'xpu_test_name']:
                val = result.get(key)
                if val:
                    parts.append(key + ':' + val)
            comment = '\n'.join(parts) if parts else 'Double not found - LLM analysis'
            
            # Only write if we have meaningful LLM results
            if result.get('cuda_exists') or result.get('xpu_exists'):
                ws.cell(row, 18, cuda_exists)
                ws.cell(row, 19, xpu_exists)
                ws.cell(row, 20, comment)
                applied += 1
            else:
                skipped += 1
        else:
            skipped += 1
    
    log(f"  Applied LLM results to {applied} test cases")
    log(f"  Skipped (no LLM analysis): {skipped} test cases")
    
    wb.save(os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx'))
    log("Saved Test Cases sheet!")
    
    # Process E2E and Issues sheets
    log("\n" + "=" * 60)
    log("[STEP 3/5] Processing E2E Test Cases sheet...")
    e2e_start = time_module.time()
    process_e2e_cases(wb)
    e2e_elapsed = time_module.time() - e2e_start
    log(f"  E2E processing completed ({e2e_elapsed:.1f}s)")
    
    log("\n" + "=" * 60)
    log("[STEP 4/5] Processing Issues sheet...")
    issues_start = time_module.time()
    process_issues_sheet(wb)
    issues_elapsed = time_module.time() - issues_start
    log(f"  Issues processing completed ({issues_elapsed:.1f}s)")
    wb.save(os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx'))
    log("Saved Issues sheet!")

    # Generate markdown report
    log("\n" + "=" * 60)
    log("[STEP 5/5] Generating markdown report...")
    from generate_report import generate_report
    generate_report()
    
    total_elapsed = time_module.time() - pipeline_start
    log(f"\n{'=' * 60}")
    log(f"[COMPLETE] Total time: {total_elapsed:.0f}s")
    log(f"  [END TIME] {time_module.strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == '__main__':
    main()
