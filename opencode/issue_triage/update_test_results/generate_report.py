#!/usr/bin/env python3
"""
Generate markdown report from torch_xpu_ops_issues.xlsx
Reads Excel columns by header names, not hardcoded positions.
Filters out 'enhancement' labeled issues.

Generates two reports:
1. issue_report.md - all issues
2. issue_report_sh_ut.md - only upstream_ut issues from specific list
"""

import openpyxl
from datetime import datetime, timedelta
import os
import re

ROOT_DIR = "/home/daisydeng"
RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")

UPSTREAM_UT_ISSUE_IDS = [
    1893, 1962, 1972, 2015, 2024, 2164, 2169, 2214, 2229, 2244, 2245, 2253, 2255, 2270,
    2283, 2285, 2287, 2295, 2301, 2309, 2329, 2376, 2436, 2442, 2482, 2512, 2531, 2532,
    2536, 2541, 2554, 2578, 2609, 2611, 2613, 2615, 2620, 2694, 2697, 2698, 2715, 2720,
    2783, 2800, 2802, 2806, 2810, 2853, 2888, 2891, 2958, 2997, 2999, 3004, 3006, 3007,
    3033, 3126, 3127, 3128, 3129, 3131, 3132, 3133, 3136, 3137, 3140, 3141, 3142, 3143,
    3163, 3166, 3170, 3177, 3187, 3238
]


def get_all_columns_by_header(ws):
    """Get dict of all columns by header name."""
    col_map = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h:
            col_map[h] = col
    return col_map


def clean_category_name(cat):
    """Remove number prefix like '9 - ', '3. ' from category names."""
    if not cat:
        return 'unknown'
    cat = re.sub(r'^\d+\s*-\s*', '', cat)
    cat = re.sub(r'^\d+\.\s*', '', cat)
    return cat.strip()


def slugify(text):
    """Create URL-friendly anchor from text."""
    text = str(text).lower().replace('/', '-').replace('&', 'and').replace(' ', '-')
    text = re.sub(r'[^a-z0-9\s\-]', '', text)
    text = re.sub(r'\s+', '-', text)
    return text


def extract_action_type(action_reason):
    """Extract action type prefix from action reason (before first colon)."""
    if not action_reason:
        return 'Unknown Action'
    reason = str(action_reason).strip()
    if ':' in reason:
        return reason.split(':')[0].strip()
    return 'Unknown Action'


def get_priority_key(issue):
    """Get sort key for priority. P0=0, P1=1, P2=2, etc."""
    priority = issue.get('priority') or ''
    if priority.startswith('P'):
        try:
            return int(priority[1:])
        except ValueError:
            pass
    return 99


def sort_by_priority(issues_list):
    """Sort issues by priority (P0 first), then by ID ascending."""
    return sorted(issues_list, key=lambda x: (get_priority_key(x), -int(x.get('id', 0))))


def clean_cell(text):
    """Remove newlines and excess whitespace from cell content for markdown tables."""
    if text is None:
        return ''
    return ' '.join(str(text).split())


def is_last_week(created_time):
    """Check if issue was created in the last 7 days."""
    if not created_time:
        return False
    seven_days_ago = datetime.now() - timedelta(days=7)
    if isinstance(created_time, datetime):
        return created_time >= seven_days_ago
    if isinstance(created_time, str) and len(created_time) >= 10:
        try:
            created_dt = datetime.strptime(created_time[:10], '%Y-%m-%d')
            return created_dt >= seven_days_ago
        except ValueError:
            return False
    return False


def is_stale(updated_time):
    """Check if issue has not been updated for 2+ weeks."""
    if not updated_time:
        return False
    fourteen_days_ago = datetime.now() - timedelta(days=14)
    if isinstance(updated_time, datetime):
        return updated_time < fourteen_days_ago
    if isinstance(updated_time, str) and len(updated_time) >= 10:
        try:
            updated_dt = datetime.strptime(updated_time[:10], '%Y-%m-%d')
            return updated_dt < fourteen_days_ago
        except ValueError:
            return False
    return False


def days_since_update(updated_time):
    """Calculate days since last update."""
    if not updated_time:
        return None
    try:
        if isinstance(updated_time, datetime):
            return (datetime.now() - updated_time).days
        if isinstance(updated_time, str) and len(updated_time) >= 10:
            updated_dt = datetime.strptime(updated_time[:10], '%Y-%m-%d')
            return (datetime.now() - updated_dt).days
    except:
        return None
    return None


def generate_report_content(issues, report_title, upstream_only=False):
    """Generate report content for given issues."""
    enhancement_count = 0
    
    action_groups = {}
    for issue in issues:
        action = str(issue.get('action_TBD') or 'Unknown')
        if action not in action_groups:
            action_groups[action] = []
        action_groups[action].append(issue)
    
    need_inv_issues = action_groups.get('Need Investigation', [])
    need_inv_by_category = {}
    need_inv_by_action_type = {}
    for issue in need_inv_issues:
        cat = clean_category_name(issue.get('category') or 'unknown')
        if cat not in need_inv_by_category:
            need_inv_by_category[cat] = []
        need_inv_by_category[cat].append(issue)
        
        if cat == 'unknown':
            action_type = issue.get('action_type', 'Unknown Action')
            if action_type not in need_inv_by_action_type:
                need_inv_by_action_type[action_type] = []
            need_inv_by_action_type[action_type].append(issue)
    
    other_action_groups = {}
    for action, issues_list in action_groups.items():
        if action != 'Need Investigation':
            other_action_groups[action] = issues_list
    
    with_dependency = [i for i in issues if i.get('dependency') and str(i.get('dependency', '')).strip() not in ['None', '', None]]
    duplicated = [i for i in issues if i.get('duplicated_issue') is not None and str(i.get('duplicated_issue', '')).strip()]
    
    last_week = [i for i in issues if is_last_week(i.get('created_time'))]
    stale_issues = [i for i in issues if is_stale(i.get('updated_time'))]
    
    dep_stats = {}
    for issue in issues:
        dep = issue.get('dependency')
        if dep and dep != 'None':
            dep_stats[dep] = dep_stats.get(dep, 0) + 1
    
    md = ""
    md += f"# {report_title}\n\n"
    md += "**Repository:** [intel/torch-xpu-ops](https://github.com/intel/torch-xpu-ops)\n\n"
    md += "**CI Data Sources:**\n"
    md += "- Torch-XPU-OPS Nightly CI: XML files from `Inductor-XPU-UT-Data-*` + `Inductor-wheel-nightly-LTS2-XPU-E2E-Data-*`\n"
    md += "- Stock PyTorch XPU CI: XML files from `test-default-*-linux.idc.xpu_*.zip`\n\n"
    md += f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    md += f"**Total Issues:** {len(issues)}\n\n"
    md += "---\n\n"
    
    total = len(issues)
    need_inv_count = len(need_inv_issues)
    dup_count = len(duplicated)
    dep_count = len(with_dependency)
    last_week_count = len(last_week)
    stale_count = len(stale_issues)
    other_count = total - need_inv_count
    
    md += "## <span id='toc'>Index</span>\n\n"
    md += f"- [1. Summary (#1-summary)](#1-summary) - {total} issues |\n"
    md += f"- [2. Need Investigation by Category (#2-need-investigation-by-category)](#2-need-investigation-by-category) - {need_inv_count} issues |\n"
    for cat in sorted(need_inv_by_category.keys(), key=lambda x: -len(need_inv_by_category[x])):
        anchor = slugify(cat)
        md += f"   - [{cat}](#{anchor}) - {len(need_inv_by_category[cat])} issues |\n"
        if cat == 'unknown':
            for action_type in sorted(need_inv_by_action_type.keys(), key=lambda x: -len(need_inv_by_action_type[x])):
                type_anchor = f"{anchor}-{slugify(action_type)}"
                md += f"      - [{action_type}](#{type_anchor}) - {len(need_inv_by_action_type[action_type])} issues |\n"
    md += f"- [3. Other Actions by Type (#3-other-actions-by-type)](#3-other-actions-by-type) - {other_count} issues |\n"
    for action in other_action_groups:
        anchor = slugify(action)
        md += f"   - [{action}](#{anchor}) - {len(other_action_groups[action])} issues |\n"
    md += f"- [4. Last Week Issues (#4-last-week-issues)](#4-last-week-issues) - {last_week_count} issues |\n"
    md += f"- [5. Stale Issues (#5-stale-issues)](#5-stale-issues) - {stale_count} issues |\n"
    md += f"- [6. Duplicated Issues (#6-duplicated-issues)](#6-duplicated-issues) - {dup_count} issues |\n"
    md += f"- [7. Issues with Dependency (#7-issues-with-dependency)](#7-issues-with-dependency) - {dep_count} issues |\n"
    md += f"- [8. Statistics (#8-statistics)](#8-statistics) - Dependency stats |\n\n"
    
    md += "---\n\n"
    
    md += "## <span id='1-summary'>1. Summary</span>\n\n"
    md += f"**Total: {total} issues**\n\n"
    md += "| # | Action Type | Count | Link |\n"
    md += "|--:|-------------|-------|------|\n"
    idx = 1
    for action, issues_list in sorted(action_groups.items(), key=lambda x: -len(x[1])):
        anchor = slugify(action)
        md += f"| {idx} | [{action}](#{anchor}) | {len(issues_list)} | [View Issues](#{anchor}) |\n"
        idx += 1
    md += f"| | **Total** | **{total}** | |\n\n"
    
    md += "## <span id='2-need-investigation-by-category'>2. Need Investigation by Category</span>\n\n"
    md += f"**Total: {need_inv_count} issues** - Issues requiring further investigation\n\n"
    
    thread_idx = 1
    for cat in sorted(need_inv_by_category.keys(), key=lambda x: -len(need_inv_by_category[x])):
        cat_issues = need_inv_by_category[cat]
        cat_issues.sort(key=lambda x: x.get('id', 0))
        anchor = slugify(cat)
        
        if cat == 'unknown':
            md += f"### <span id='{anchor}'>{cat}</span> ({len(cat_issues)} issues)\n\n"
            md += f"**Grouped by Action Type:**\n\n"
            
            for action_type, type_issues in sorted(need_inv_by_action_type.items(), key=lambda x: -len(x[1])):
                type_anchor = f"{anchor}-{slugify(action_type)}"
                md += f"#### <span id='{type_anchor}'>{action_type}</span> ({len(type_issues)} issues)\n"
                md += "| # | ID | Title | Priority | Priority Reason | Action Reason | Summary | Assignee | Test Module | Related PR |\n"
                md += "|--:|----|-------|----------|---------------|---------------|---------|----------|-------------|-------------|\n"
                
                sub_idx = 1
                for issue in sort_by_priority(type_issues):
                    issue_id = issue.get('id') or ''
                    title = clean_cell(issue.get('title'))
                    priority = clean_cell(issue.get('priority'))
                    priority_reason = clean_cell(issue.get('priority_reason'))
                    action_reason = clean_cell(issue.get('action_reason'))
                    summary = clean_cell(issue.get('summary'))
                    assignee = clean_cell(issue.get('assignee'))
                    test_module = clean_cell(issue.get('test_module'))
                    pr = clean_cell(issue.get('pr'))
                    
                    md += f"| {sub_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {action_reason} | {summary} | {assignee} | {test_module} | {pr} |\n"
                    sub_idx += 1
                md += f"| | | **Subtotal: {len(type_issues)} issues** | | | | | | | |\n\n"
            md += f"| | | **Category Total: {len(cat_issues)} issues** | | | | | |\n\n"
        
        else:
            md += f"### <span id='{anchor}'>{cat}</span> ({len(cat_issues)} issues)\n\n"
            md += "| # | ID | Title | Priority | Priority Reason | Action Reason | Summary | Assignee | Test Module | Related PR |\n"
            md += "|--:|----|-------|----------|---------------|---------------|---------|----------|-------------|-------------|\n"
            
            for issue in sort_by_priority(cat_issues):
                issue_id = issue.get('id') or ''
                title = clean_cell(issue.get('title'))
                priority = clean_cell(issue.get('priority'))
                priority_reason = clean_cell(issue.get('priority_reason'))
                action_reason = clean_cell(issue.get('action_reason'))
                summary = clean_cell(issue.get('summary'))
                assignee = clean_cell(issue.get('assignee'))
                test_module = clean_cell(issue.get('test_module'))
                pr = clean_cell(issue.get('pr'))
                
                md += f"| {thread_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {action_reason} | {summary} | {assignee} | {test_module} | {pr} |\n"
                thread_idx += 1
            md += f"| | | **Subtotal: {len(cat_issues)} issues** | | | | | | | |\n\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "## <span id='3-other-actions-by-type'>3. Other Actions by Type</span>\n\n"
    md += f"**Total: {other_count} issues** - Actions other than Need Investigation\n\n"
    
    action_order = ['Close fixed issue', 'Verify the issue', 'Enable test', 'add to skiplist', 'Revisit the PR as case failed']
    ordered_actions = []
    for a in action_order:
        if a in other_action_groups:
            ordered_actions.append(a)
    for a in sorted(other_action_groups.keys()):
        if a not in ordered_actions:
            ordered_actions.append(a)
    
    thread_idx = 1
    for action in ordered_actions:
        issues_list = other_action_groups.get(action, [])
        if not issues_list:
            continue
        
        anchor = slugify(action)
        md += f"### <span id='{anchor}'>{action}</span> ({len(issues_list)} issues)\n\n"
        md += "| # | ID | Title | Priority | Priority Reason | Action Reason | Summary | Assignee | Test Module | Related PR |\n"
        md += "|--:|----|-------|----------|---------------|---------------|---------|----------|-------------|-------------|\n"
        
        for issue in sort_by_priority(issues_list):
            issue_id = issue.get('id') or ''
            title = clean_cell(issue.get('title'))
            priority = clean_cell(issue.get('priority'))
            priority_reason = clean_cell(issue.get('priority_reason'))
            action_reason = clean_cell(issue.get('action_reason'))
            summary = clean_cell(issue.get('summary'))
            assignee = clean_cell(issue.get('assignee'))
            test_module = clean_cell(issue.get('test_module'))
            pr = clean_cell(issue.get('pr'))
            
            md += f"| {thread_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {action_reason} | {summary} | {assignee} | {test_module} | {pr} |\n"
            thread_idx += 1
        md += f"| | | **Subtotal: {len(issues_list)} issues** | | | | | | | |\n\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "## <span id='4-last-week-issues'>4. Last Week Issues</span>\n\n"
    md += f"**Total: {last_week_count} issues** - Issues created in the last 7 days\n\n"
    
    if last_week:
        md += "| # | ID | Title | Priority | Priority Reason | Action Reason | Summary | Category | Created Time | Related PR |\n"
        md += "|--:|----|-------|----------|---------------|---------------|---------|----------|--------------|-------------|\n"
        
        for issue in sort_by_priority(last_week):
            issue_id = issue.get('id') or ''
            title = clean_cell(issue.get('title'))
            priority = clean_cell(issue.get('priority'))
            priority_reason = clean_cell(issue.get('priority_reason'))
            action_reason = clean_cell(issue.get('action_reason'))
            summary = clean_cell(issue.get('summary'))
            category = clean_cell(issue.get('category'))
            created_time = issue.get('created_time') or ''
            if isinstance(created_time, datetime):
                created_time = created_time.strftime('%Y-%m-%d')
            pr = clean_cell(issue.get('pr'))
            
            md += f"| {thread_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {action_reason} | {summary} | {category} | {created_time} | {pr} |\n"
            thread_idx += 1
        md += f"| | | **Subtotal: {thread_idx-1} issues** | | | | | | | |\n\n"
    else:
        md += "No issues created in the last 7 days.\n\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "## <span id='5-stale-issues'>5. Stale Issues</span>\n\n"
    md += f"**Total: {stale_count} issues** - Issues not updated in 2+ weeks\n\n"
    
    if stale_issues:
        md += "| # | ID | Title | Priority | Priority Reason | Action Reason | Summary | Category | Updated Time | Days Since Update | Related PR |\n"
        md += "|--:|----|-------|----------|---------------|---------------|---------|----------|---------------|-------------------|-------------|\n"
        
        for issue in sort_by_priority(stale_issues):
            issue_id = issue.get('id') or ''
            title = clean_cell(issue.get('title'))
            priority = clean_cell(issue.get('priority'))
            priority_reason = clean_cell(issue.get('priority_reason'))
            action_reason = clean_cell(issue.get('action_reason'))
            summary = clean_cell(issue.get('summary'))
            category = clean_cell(issue.get('category'))
            updated_time = issue.get('updated_time') or ''
            if isinstance(updated_time, datetime):
                updated_time = updated_time.strftime('%Y-%m-%d')
            days = days_since_update(issue.get('updated_time')) or 'N/A'
            pr = clean_cell(issue.get('pr'))
            
            md += f"| {thread_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {action_reason} | {summary} | {category} | {updated_time} | {days} | {pr} |\n"
            thread_idx += 1
        md += f"| | | **Subtotal: {thread_idx-1} issues** | | | | | | | | |\n\n"
    else:
        md += "No stale issues.\n\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "## <span id='6-duplicated-issues'>6. Duplicated Issues</span>\n\n"
    md += f"**Total: {dup_count} issues** - Issues sharing test cases with other issues\n\n"
    md += "| # | ID | Title | Priority | Priority Reason | Summary | Assignee | Root Cause | Dependency | Duplicated With | Test Module | Related PR |\n"
    md += "|--:|----|-------|----------|---------------|---------|----------|---------|-----------|----------------|-------------|-------------|\n"
    
    for issue in sort_by_priority(duplicated):
        issue_id = issue.get('id') or ''
        title = clean_cell(issue.get('title'))
        priority = clean_cell(issue.get('priority'))
        priority_reason = clean_cell(issue.get('priority_reason'))
        summary = clean_cell(issue.get('summary'))
        assignee = clean_cell(issue.get('assignee'))
        root_cause = clean_cell(issue.get('root_cause'))
        dependency = issue.get('dependency') or ''
        dup = issue.get('duplicated_issue') or ''
        test_module = clean_cell(issue.get('test_module'))
        pr = clean_cell(issue.get('pr'))
        
        md += f"| {thread_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {summary} | {assignee} | {root_cause} | {dependency} | {dup} | {test_module} | {pr} |\n"
        thread_idx += 1
    md += f"| | | **Subtotal: {thread_idx-1} issues** | | | | | | | | |\n\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "## <span id='7-issues-with-dependency'>7. Issues with Dependency</span>\n\n"
    md += f"**Total: {dep_count} issues** - Issues with external dependencies\n\n"
    md += "| # | ID | Title | Priority | Priority Reason | Summary | Assignee | Root Cause | Category | Dependency | Test Module | Related PR |\n"
    md += "|--:|----|-------|----------|---------------|---------|----------|---------|----------|------------|-------------|-------------|\n"
    
    for issue in sort_by_priority(with_dependency):
        issue_id = issue.get('id') or ''
        title = clean_cell(issue.get('title'))
        priority = clean_cell(issue.get('priority'))
        priority_reason = clean_cell(issue.get('priority_reason'))
        summary = clean_cell(issue.get('summary'))
        assignee = clean_cell(issue.get('assignee'))
        root_cause = clean_cell(issue.get('root_cause'))
        category = clean_cell(issue.get('category'))
        dependency = issue.get('dependency') or ''
        test_module = clean_cell(issue.get('test_module'))
        pr = clean_cell(issue.get('pr'))
        
        md += f"| {thread_idx} | {issue_id} | {title} | {priority} | {priority_reason} | {summary} | {assignee} | {root_cause} | {category} | {dependency} | {test_module} | {pr} |\n"
        thread_idx += 1
    md += f"| | | **Subtotal: {thread_idx-1} issues** | | | | | | | | |\n\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "## <span id='8-statistics'>8. Statistics</span>\n\n"
    
    md += "### <span id='stats-dependency'>By Dependency</span>\n\n"
    md += "| Dependency | Count |\n"
    md += "|------------|-------|\n"
    for dep, count in sorted(dep_stats.items(), key=lambda x: -x[1]):
        md += f"| {dep} | {count} |\n"
    md += "\n"
    
    md += "[Back to Index](#toc) |\n\n"
    
    md += "---\n"
    md += f"*Report generated with {total} issues*\n"
    
    return md, enhancement_count, total, need_inv_count, other_count, dup_count, dep_count


def generate_report():
    wb = openpyxl.load_workbook(os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx'))
    ws_issues = wb['Issues']
    
    issue_cols = get_all_columns_by_header(ws_issues)
    print(f"Available columns: {list(issue_cols.keys())}")
    
    all_issues = []
    upstream_ut_issues = []
    
    for row in range(2, ws_issues.max_row + 1):
        def get_val(header, default=None):
            col = issue_cols.get(header)
            return ws_issues.cell(row, col).value if col else default
        
        issue_id = get_val('Issue ID')
        
        labels = get_val('Labels') or ''
        if 'enhancement' in str(labels).lower():
            continue
        
        issue = {
            'id': issue_id,
            'title': get_val('Title'),
            'summary': get_val('Summary'),
            'status': get_val('Status'),
            'assignee': get_val('Assignee'),
            'reporter': get_val('Reporter'),
            'labels': labels,
            'created_time': get_val('Created Time'),
            'updated_time': get_val('Updated Time'),
            'type': get_val('Type'),
            'module': get_val('Module'),
            'test_module': get_val('Test Module'),
            'dependency': get_val('Dependency'),
            'pr': get_val('PR'),
            'pr_owner': get_val('PR Owner'),
            'pr_status': get_val('PR Status'),
            'pr_desc': get_val('PR Description'),
            'category': clean_category_name(get_val('Category')),
            'category_reason': get_val('Category Reason'),
            'priority': get_val('Priority'),
            'priority_reason': get_val('Priority Reason'),
            'root_cause': get_val('Root Cause'),
            'root_cause_reason': get_val('Root Cause Reason'),
            'owner_transfer': get_val('Owner Transfer'),
            'action_TBD': get_val('Action TBD'),
            'action_reason': get_val('Action Reason'),
            'action_type': extract_action_type(get_val('Action Reason')),
            'duplicated_issue': get_val('duplicated_issue'),
        }
        
        all_issues.append(issue)
        
        if issue_id in UPSTREAM_UT_ISSUE_IDS:
            upstream_ut_issues.append(issue)
    
    print(f"Total loaded: {len(all_issues)}")
    print(f"Upstream UT issues: {len(upstream_ut_issues)}")
    
    md1, enh1, tot1, ni1, oth1, dup1, dep1 = generate_report_content(
        all_issues, "Torch XPU Ops Issue Report"
    )
    
    output_path1 = os.path.join(RESULT_DIR, 'issue_report.md')
    with open(output_path1, 'w') as f:
        f.write(md1)
    
    print(f"\nReport 1 saved to: {output_path1}")
    print(f"Total issues: {tot1}")
    print(f"  Need Investigation: {ni1}")
    print(f"  Other Actions: {oth1}")
    print(f"  Duplicated: {dup1}")
    print(f"  With Dependency: {dep1}")
    
    md2, enh2, tot2, ni2, oth2, dup2, dep2 = generate_report_content(
        upstream_ut_issues, "Torch XPU Ops Issue Report (upstream_ut only)"
    )
    
    output_path2 = os.path.join(RESULT_DIR, 'issue_report_sh_ut.md')
    with open(output_path2, 'w') as f:
        f.write(md2)
    
    print(f"\nReport 2 saved to: {output_path2}")
    print(f"Upstream UT issues: {tot2}")
    print(f"  Need Investigation: {ni2}")
    print(f"  Other Actions: {oth2}")
    print(f"  Duplicated: {dup2}")
    print(f"  With Dependency: {dep2}")


if __name__ == '__main__':
    generate_report()