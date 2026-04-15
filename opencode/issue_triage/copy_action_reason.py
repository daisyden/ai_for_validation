#!/usr/bin/env python3
"""
Copy Action Reason from source Excel to target Excel.

Only updates the 'Action Reason' column in the Issues sheet.
Does not modify any other columns or sheets.
Issues are matched by exact Issue ID.
"""
import openpyxl
import sys
import os

def copy_action_reason(
    source_path: str = '~/torch_xpu_ops_issues_action_reason2.xlsx',
    target_path: str = '~/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx',
    sheet_name: str = 'Issues'
):
    """
    Copy Action Reason from source to target Excel file.
    
    Args:
        source_path: Source Excel file with Action Reason column
        target_path: Target Excel file to update
        sheet_name: Name of sheet to update
    """
    # Expand paths
    source_path = os.path.expanduser(source_path)
    target_path = os.path.expanduser(target_path)
    
    # Validate paths exist
    if not os.path.exists(source_path):
        print(f"Error: Source file not found: {source_path}")
        return
    
    if not os.path.exists(target_path):
        print(f"Error: Target file not found: {target_path}")
        return
    
    # Load source workbook
    print(f"Loading source: {source_path}")
    try:
        wb_source = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading source file: {e}")
        return
    
    if sheet_name not in wb_source.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in source file")
        wb_source.close()
        return
    
    ws_source = wb_source[sheet_name]
    
    # Find Issue ID and Action Reason columns in source
    source_issue_col = None
    source_action_col = None
    
    for i, cell in enumerate(ws_source[1], 1):
        if cell.value == 'Issue ID':
            source_issue_col = i
        elif cell.value == 'Action Reason':
            source_action_col = i
    
    if source_issue_col is None:
        print("Error: 'Issue ID' column not found in source sheet header")
        wb_source.close()
        return
    
    if source_action_col is None:
        print("Error: 'Action Reason' column not found in source sheet header")
        wb_source.close()
        return
    
    print(f"Source: Issue ID at col {source_issue_col}, Action Reason at col {source_action_col}")
    
    # Build mapping: Issue ID -> Action Reason (EXACT match only)
    source_data = {}
    source_issue_count = 0
    
    for row in ws_source.iter_rows(min_row=2, values_only=True):
        issue_id = row[source_issue_col - 1]
        action_reason = row[source_action_col - 1] if source_action_col <= len(row) else None
        
        # Requirement: Issue ID must EXACTLY match
        # Skip if issue_id is None, empty, or not a valid identifier
        if issue_id is None:
            continue
        if isinstance(issue_id, str) and issue_id.strip() == '':
            continue
        if isinstance(issue_id, (int, float)) and issue_id == 0:
            continue
            
        source_issue_count += 1
        
        # Store only non-None Action Reasons
        if action_reason is not None and str(action_reason).strip() != '':
            # Use exact issue_id as key - ensure type matches
            source_data[issue_id] = action_reason
    
    print(f"Source: Found {source_issue_count} total issues, {len(source_data)} with Action Reason")
    print(f"Source: Issue IDs sample: {list(source_data.keys())[:10]}")
    
    wb_source.close()
    
    # Load target workbook (read-write)
    print(f"\nLoading target: {target_path}")
    try:
        wb_target = openpyxl.load_workbook(target_path)
    except Exception as e:
        print(f"Error loading target file: {e}")
        return
    
    if sheet_name not in wb_target.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in target file")
        wb_target.close()
        return
    
    ws_target = wb_target[sheet_name]
    
    # Find Issue ID and Action Reason columns in target
    target_issue_col = None
    target_action_col = None
    
    for i, cell in enumerate(ws_target[1], 1):
        if cell.value == 'Issue ID':
            target_issue_col = i
        elif cell.value == 'Action Reason':
            target_action_col = i
    
    if target_issue_col is None:
        print("Error: 'Issue ID' column not found in target sheet header")
        wb_target.close()
        return
    
    if target_action_col is None:
        print("Error: 'Action Reason' column not found in target sheet header")
        wb_target.close()
        return
    
    print(f"Target: Issue ID at col {target_issue_col}, Action Reason at col {target_action_col}")
    
    # Update Action Reason in target for EXACT Issue ID matches only
    updated_count = 0
    skipped_count = 0
    not_found_count = 0
    empty_source_count = 0
    
    for row_idx, row in enumerate(ws_target.iter_rows(min_row=2), start=2):
        target_issue_id = row[target_issue_col - 1].value
        
        # Skip if Issue ID is None or invalid
        if target_issue_id is None:
            continue
        if isinstance(target_issue_id, str) and target_issue_id.strip() == '':
            continue
            
        # ONLY copy if Issue ID EXACTLY matches AND source has Action Reason
        if target_issue_id in source_data:
            new_action_reason = source_data[target_issue_id]
            
            if new_action_reason is None or str(new_action_reason).strip() == '':
                # Source has no Action Reason for this issue
                empty_source_count += 1
                continue
            
            # Get current Action Reason value
            current_action_reason = row[target_action_col - 1].value
            
            # Only update if different
            if current_action_reason != new_action_reason:
                row[target_action_col - 1].value = new_action_reason
                updated_count += 1
            else:
                # Same value - skip
                skipped_count += 1
        else:
            # Issue ID not found in source (no match)
            not_found_count += 1
    
    print(f"\n{'='*60}")
    print("Update Summary:")
    print(f"{'='*60}")
    print(f"  Action Reasons copied/updated: {updated_count}")
    print(f"  Skipped (same value): {skipped_count}")
    print(f"  Not found in source: {not_found_count}")  
    print(f"  Empty in source: {empty_source_count}")
    print(f"{'='*60}")
    
    # Save target workbook (ONLY Action Reason column modified)
    print(f"\nSaving target: {target_path}")
    try:
        wb_target.save(target_path)
        print("Save successful!")
    except Exception as e:
        print(f"Error saving target file: {e}")
        wb_target.close()
        return
    
    wb_target.close()

if __name__ == '__main__':
    source = sys.argv[1] if len(sys.argv) > 1 else '~/torch_xpu_ops_issues_action_reason2.xlsx'
    target = sys.argv[2] if len(sys.argv) > 2 else '~/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx'
    
    print("=" * 70)
    print("Copy Action Reason Script")
    print("=" * 70)
    print(f"Source: {source}")
    print(f"Target: {target}")
    print("=" * 70)
    
    copy_action_reason(
        source_path=source,
        target_path=target,
        sheet_name='Issues'
    )