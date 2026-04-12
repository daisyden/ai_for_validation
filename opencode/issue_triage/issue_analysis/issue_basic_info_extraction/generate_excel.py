#!/usr/bin/env python3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import os
import requests
import argparse

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = "/home/daisydeng"
RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")
DATA_DIR = os.path.join(ROOT_DIR, "issue_traige", "data")
DOC_DIR = os.path.join(ROOT_DIR, "issue_traige", "doc")

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_HEADERS = {
    "Accept": "application/vnd.github.v3+json",
    "Authorization": f"token {GITHUB_TOKEN}"
} if GITHUB_TOKEN else {}

# Parse command line arguments
parser = argparse.ArgumentParser(description="Generate Excel report for torch-xpu-ops issues")
parser.add_argument("--issues", type=str, default="", help="Comma-separated list of issue IDs to process (default: all)")
args = parser.parse_args()

# Parse target issue IDs
TARGET_ISSUE_IDS = None
if args.issues:
    TARGET_ISSUE_IDS = set()
    for part in args.issues.split(','):
        part = part.strip()
        if part:
            try:
                TARGET_ISSUE_IDS.add(int(part))
            except ValueError:
                pass

# Load data - try to load from JSON, or fetch from GitHub if not exists
issues_json_path = os.path.join(DATA_DIR, "torch_xpu_ops_issues.json")
comments_json_path = os.path.join(DATA_DIR, "torch_xpu_ops_comments.json")

if os.path.exists(issues_json_path):
    with open(issues_json_path) as f:
        issues = json.load(f)
else:
    print("Fetching issues from GitHub...")
    issues = []
    page = 1
    while len(issues) < 500:
        url = f"https://api.github.com/repos/intel/torch-xpu-ops/issues?state=open&per_page=100&page={page}"
        response = requests.get(url, headers=GITHUB_HEADERS, timeout=30)
        if response.status_code != 200:
            break
        batch = response.json()
        if not batch:
            break
        issues.extend([i for i in batch if 'pull_request' not in i])
        print(f"Fetched {len(issues)} issues...")
        page += 1
    with open(issues_json_path, 'w') as f:
        json.dump(issues, f)

# Filter to target issues if specified
if TARGET_ISSUE_IDS:
    issues = [i for i in issues if i['number'] in TARGET_ISSUE_IDS]
    print(f"Filtered to {len(issues)} target issues: {sorted(TARGET_ISSUE_IDS)}")



# Load ops dependency
ops_dep = {}
with open(os.path.join(DOC_DIR, "ops_dependency.csv")) as f:
    next(f)
    for line in f:
        parts = line.strip().split(',')
        if len(parts) >= 2:
            op_name = parts[0].strip().lower()
            ops_dep[op_name] = parts[1].strip()

# Known test types
KNOWN_TEST_TYPES = ['op_ut', 'op_extend', 'op_extended', 'e2e', 'benchmark', 'ut', 'test_xpu']

# Model lists from benchmarks
HUGGINGFACE_MODELS = [
    'AlbertForMaskedLM', 'AlbertForQuestionAnswering', 'AllenaiLongformerBase',
    'BartForCausalLM', 'BartForConditionalGeneration', 'BertForMaskedLM',
    'BertForQuestionAnswering', 'BlenderbotForCausalLM', 'BlenderbotForConditionalGeneration',
    'BlenderbotSmallForCausalLM', 'BlenderbotSmallForConditionalGeneration', 'CamemBert',
    'DebertaV2ForMaskedLM', 'DebertaV2ForQuestionAnswering', 'DistilBertForMaskedLM',
    'DistilBertForQuestionAnswering', 'DistillGPT2', 'ElectraForCausalLM',
    'ElectraForQuestionAnswering', 'GoogleFnet', 'google/gemma', 'GPT2ForSequenceClassification',
    'GPTJForCausalLM', 'GPTJForQuestionAnswering', 'GPTNeoForCausalLM',
    'GPTNeoForSequenceClassification', 'LayoutLMForMaskedLM', 'LayoutLMForSequenceClassification',
    'M2M100ForConditionalGeneration', 'MBartForCausalLM', 'MBartForConditionalGeneration',
    'MegatronBertForCausalLM', 'MegatronBertForQuestionAnswering', 'meta-llama', 'mistralai',
    'MobileBertForMaskedLM', 'MobileBertForQuestionAnswering', 'MT5ForConditionalGeneration',
    'openai/gpt', 'openai/whisper', 'OPTForCausalLM', 'PegasusForCausalLM',
    'PegasusForConditionalGeneration', 'PLBartForCausalLM', 'PLBartForConditionalGeneration',
    'Qwen', 'RobertaForCausalLM', 'RobertaForQuestionAnswering', 'T5ForConditionalGeneration',
    'T5Small', 'TrOCRForCausalLM', 'XGLMForCausalLM', 'XLNetLMHeadModel', 'YituTechConvBert',
    'hf_Albert', 'hf_Bert', 'hf_Bert_large', 'hf_DistilBert', 'hf_Roberta_base'
]

TIMM_MODELS = [
    'adv_inception_v3', 'beit_base_patch16_224', 'botnet26t_256', 'cait_m36_384',
    'coat_lite_mini', 'convit_base', 'convmixer_768_32', 'convnext_base',
    'convnextv2_nano', 'crossvit_9_240', 'cspdarknet53', 'deit_base_distilled_patch16_224',
    'deit_tiny_patch16_224', 'dla102', 'dm_nfnet_f0', 'dpn107', 'eca_botnext26ts_256',
    'eca_halonext26ts', 'ese_vovnet19b_dw', 'fbnetc_100', 'fbnetv3_b', 'gernet_l',
    'ghostnet_100', 'gluon_inception_v3', 'gmixer_24_224', 'gmlp_s16_224', 'hrnet_w18',
    'inception_v3', 'jx_nest_base', 'lcnet_050', 'levit_128', 'mixer_b16_224',
    'mixnet_l', 'mnasnet_100', 'mobilenetv2_100', 'mobilenetv3_large_100', 'mobilevit_s',
    'nfnet_l0', 'pit_b_224', 'pnasnet5large', 'poolformer_m36', 'regnety_002',
    'repvgg_a2', 'res2net101_26w_4s', 'res2net50_14w_8s', 'res2next50', 'resmlp_12_224',
    'resnest101e', 'rexnet_100', 'sebotnet33ts_256', 'selecsls42b', 'spnasnet_100',
    'swin_base_patch4_window7_224', 'swsl_resnext101_32x16d', 'tf_efficientnet_b0',
    'tf_mixnet_l', 'tinynet_a', 'tnt_s_patch16_224', 'twins_pcpvt_base', 'visformer_small',
    'vit_base_patch14_dinov2', 'vit_base_patch16_224', 'vit_base_patch16_siglip_256',
    'volo_d1_224', 'xcit_large_24_p8_224', 'timm_vision_transformer', 'timm_vision_transformer_large'
]

TORCHBENCH_MODELS = [
    'BERT_pytorch', 'Background_Matting', 'LearningToPaint', 'alexnet', 'dcgan',
    'densenet121', 'mnasnet1_0', 'mobilenet_v2', 'mobilenet_v3_large',
    'nvidia_deeprecommender', 'pytorch_unet', 'resnet18', 'resnet50',
    'resnext50_32x4d', 'shufflenet_v2_x1_0', 'squeezenet1_1', 'vgg16'
]

def identify_benchmark(model_name):
    """Identify benchmark from model name"""
    model_lower = model_name.lower()
    
    for m in HUGGINGFACE_MODELS:
        if m.lower() in model_lower or model_lower in m.lower():
            return 'huggingface'
    
    for m in TIMM_MODELS:
        if m.lower() in model_lower or model_lower in m.lower():
            return 'timm'
    
    for m in TORCHBENCH_MODELS:
        if m.lower() in model_lower or model_lower in m.lower():
            return 'torchbench'
    
    return 'unknown'

def extract_e2e_reproducer(body, title):
    """Extract reproducer command from issue body"""
    text = f"{title} {body}"
    
    reproducer_lines = []
    
    # Look for code blocks with commands (between ``` and ```)
    if '```' in text:
        parts = text.split('```')
        for i, part in enumerate(parts):
            # Code blocks are odd-indexed (1, 3, 5, ...)
            if i % 2 == 1:  # This is a code block content
                part_stripped = part.strip()
                if part_stripped:
                    lines = part_stripped.split('\n')
                    for line in lines:
                        line_stripped = line.strip()
                        # Look for actual commands (python, pytest, etc.)
                        if line_stripped and (line_stripped.startswith(('python', 'pytest', 'XPU_', './')) or 'python' in line_stripped.lower()):
                            if not line_stripped.startswith('#'):
                                reproducer_lines.append(line_stripped)
                    # If we found a command, use it
                    if reproducer_lines:
                        break
    
    # Also look for command patterns without code blocks
    if not reproducer_lines:
        # Look for python or pytest command patterns
        cmd_patterns = [
            r'(pytest\s+[^\n]+)',
            r'(python\s+test/[^\n]+)',
            r'(python\s+-m\s+pytest[^\n]+)',
            r'(XPU_QUANT_CONFIG=[^\n]+python[^\n]+)',
            r'(python\s+benchmarks/dynamo/[^\n]+)',
            r'(python\s+[^\n]+run_benchmark[^\n]+)',
        ]
        
        for pattern in cmd_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                reproducer_lines.append(match.strip())
    
    if not reproducer_lines:
        # Generic reproducer from title
        return title[:200]
    
    # Join and limit to 3 lines
    return '\n'.join(reproducer_lines[:3])


def parse_e2e_info(body, title):
    """Parse e2e benchmark information from issue body"""
    e2e_info = []
    
    text = f"{title} {body}"
    
    # Get reproducer
    reproducer = extract_e2e_reproducer(body, title)
    
    # Check for model names in title or body
    all_model_names = HUGGINGFACE_MODELS + TIMM_MODELS + TORCHBENCH_MODELS
    
    # Extract phase (training/inference)
    phase = 'inference'
    if 'training' in text.lower():
        phase = 'training'
    elif 'train' in text.lower():
        phase = 'training'
    
    # Extract dtype
    dtype = 'float32'
    dtype_patterns = [
        (r'bfloat16|bf16', 'bfloat16'),
        (r'float16|fp16', 'float16'),
        (r'float32|fp32', 'float32'),
        (r'int8|int\s*8', 'int8'),
    ]
    for pattern, dt in dtype_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            dtype = dt
            break
    
    # Extract AMP (automatic mixed precision)
    amp = False
    if '--amp' in text.lower() or 'amp' in text.lower():
        amp = True
    
    # Extract test type
    test_type = 'accuracy'
    if 'throughputs' in text.lower() or 'performance' in text.lower() or 'latency' in text.lower():
        test_type = 'performance'
    
    # Extract backend
    backend = 'inductor'
    if '--backend=(\w+)' in text:
        match = re.search(r'--backend=(\w+)', text)
        if match:
            backend = match.group(1)
    elif 'eager' in text.lower():
        backend = 'eager'
    elif 'inductor' in text.lower():
        backend = 'inductor'
    
    # Extract disable-cudagraphs
    disable_cudagraphs = 'no'
    if 'disable-cudagraphs' in text.lower() or 'disable_cudagraphs' in text.lower():
        disable_cudagraphs = 'yes'
    
    # Find model in body - need exact model name, not partial match
    found_models = set()
    for model in all_model_names:
        # Use word boundary to avoid partial matches
        if re.search(r'\b' + re.escape(model.lower()) + r'\b', text.lower()):
            benchmark = identify_benchmark(model)
            if benchmark != 'unknown' and model not in found_models:
                found_models.add(model)
                e2e_info.append({
                    'reproducer': reproducer,
                    'benchmark': benchmark,
                    'model': model,
                    'phase': phase,
                    'dtype': dtype,
                    'amp': amp,
                    'test_type': test_type,
                    'backend': backend,
                    'disable_cudagraphs': disable_cudagraphs,
                })
    
    # If no specific model found but looks like e2e issue
    if not e2e_info:
        if 'benchmark' in text.lower() or 'huggingface' in text.lower() or 'timm' in text.lower() or 'torchbench' in text.lower():
            # Try to identify benchmark from context
            if 'hf_' in text.lower() or 'huggingface' in text.lower():
                benchmark = 'huggingface'
            elif 'timm_' in text.lower() or 'timm.' in text.lower():
                benchmark = 'timm'
            elif 'torchbench' in text.lower():
                benchmark = 'torchbench'
            else:
                benchmark = 'unknown'
            
            e2e_info.append({
                'reproducer': reproducer,
                'benchmark': benchmark,
                'model': 'unknown',
                'phase': phase,
                'dtype': dtype,
                'test_type': test_type,
                'backend': backend,
                'disable_cudagraphs': disable_cudagraphs,
            })
    
    return e2e_info



def map_origin_test_file(test_file):
    if not test_file:
        return ""
    match = re.search(r'test/xpu/(.+?)(?:_xpu)?\.py$', test_file)
    if match:
        return f"test/{match.group(1)}.py"
    if 'benchmarks/' in test_file:
        return test_file
    return test_file

def parse_test_cases_from_body(body):
    cases = []

    if 'Cases:' in body:
        cases_section = body.split('Cases:')[1]

        end_markers = ['\n###', '\nVersions', '\n```', '\n\n']
        min_end = len(cases_section)
        for marker in end_markers:
            idx = cases_section.find(marker)
            if idx > 0 and idx < min_end:
                min_end = idx
        cases_section = cases_section[:min_end]

        lines = cases_section.split('\n')

        for line in lines:
            line = line.strip()

            if not line:
                continue

            if line.startswith('###') or line.startswith('...'):
                continue

            if line.startswith('~~') and line.endswith('~~'):
                continue

            parts = line.split(',')
            if len(parts) < 3:
                continue

            test_type = parts[0].strip()
            if test_type not in KNOWN_TEST_TYPES:
                continue

            test_path = parts[1].strip()
            test_case = parts[2].strip()

            if not test_path or not test_case or len(test_case) < 3:
                continue

            if ' ' in test_case:
                continue

            if 'torch-xpu-ops' in test_path:
                path_parts = test_path.split('.')
                try:
                    txpo_idx = path_parts.index('torch-xpu-ops')
                    rel_parts = path_parts[txpo_idx+1:]

                    test_class = ""
                    test_file_parts = []
                    for part in rel_parts:
                        if part.startswith('Test'):
                            test_class = part
                            break
                        test_file_parts.append(part)

                    if test_file_parts:
                        test_file = 'torch-xpu-ops/' + '/'.join(test_file_parts) + '.py'
                        if not test_file.endswith('_xpu.py'):
                            test_file = test_file.replace('.py', '_xpu.py')
                    else:
                        test_file = ""

                    origin_file = map_origin_test_file(test_file)
                except:
                    test_file = ""
                    test_class = ""
                    origin_file = ""
            else:
                test_file = test_path
                origin_file = ""
                test_class = ""

            cases.append({
                'test_type': test_type,
                'test_file': test_file,
                'origin_test_file': origin_file,
                'test_class': test_class,
                'test_case': test_case
            })

    # Extract from pytest code blocks (format: pytest -v test/test_ops.py -k test_name)
    if '```' in body:
        code_blocks = body.split('```')
        for block in code_blocks:
            # Look for pytest patterns with test path and test method
            # Handles formats: test/test_ops.py or test/distributed/test_c10d_xccl.py::ClassName::method
            pytest_pattern = r'pytest\s+-v\s+(test[/a-zA-Z0-9_/.]+\.py(?:::[a-zA-Z0-9_]+)*)'
            matches = re.findall(pytest_pattern, block)
            for match in matches:
                test_path = match.strip()
                if '::' in test_path:
                    parts = test_path.split('::')
                    file_path = parts[0]
                    test_class = parts[1] if len(parts) > 1 else ""
                    test_method = parts[2] if len(parts) > 2 else parts[-1]
                    cases.append({
                        'test_type': 'ut',
                        'test_file': file_path,
                        'origin_test_file': file_path,
                        'test_class': test_class,
                        'test_case': test_method
                    })
                else:
                    # No class/method, just file
                    cases.append({
                        'test_type': 'ut',
                        'test_file': test_path,
                        'origin_test_file': test_path,
                        'test_class': '',
                        'test_case': ''
                    })

            # Also look for test_xpu,...,... format in code blocks
            test_xpu_pattern = r'(test_xpu),([a-zA-Z0-9_\.]+),([a-zA-Z0-9_]+)'
            matches = re.findall(test_xpu_pattern, block)
            for match in matches:
                test_type, test_path, test_method = match[0], match[1], match[2]
                test_class = ""
                if '.test_' in test_path:
                    # e.g., test.test_xpu.TestXpuAutocast -> TestXpuAutocast
                    class_parts = test_path.split('.test_')
                    if len(class_parts) > 1:
                        class_name = class_parts[1]
                        if '.' in class_name:
                            test_class = class_name.rsplit('.', 1)[1] if '.' in class_name else class_name
                        else:
                            test_class = class_name
                cases.append({
                    'test_type': test_type,
                    'test_file': test_path.replace('.', '/') + '.py',
                    'origin_test_file': test_path.replace('.', '/') + '.py',
                    'test_class': test_class,
                    'test_case': test_method
                })

            # Also handle pytest commands with -k pattern (extract test method from -k value)
            # Look for: pytest ... -k test_python_ref__refs_logspace_tensor_overload_xpu_float64
            k_pattern_matches = re.findall(r'-k\s+([a-zA-Z0-9_]+)', block)
            for test_name in k_pattern_matches:
                # Try to find associated test file in the same block
                pytest_v_match = re.search(r'pytest\s+-v\s+(test[/a-zA-Z0-9_]+\.py)', block)
                if pytest_v_match:
                    file_path = pytest_v_match.group(1)
                    cases.append({
                        'test_type': 'ut',
                        'test_file': file_path,
                        'origin_test_file': file_path,
                        'test_class': '',
                        'test_case': test_name
                    })

    # Extract from pytest commands outside code blocks
    # Look for patterns like: pytest -v test/test_ops.py -k test_name
    re_pattern = r'pytest\s+-v\s+(test[/a-zA-Z0-9_]+\.py)\s*-k\s+([a-zA-Z0-9_]+)'
    matches = re.findall(re_pattern, body)
    for file_path, test_name in matches:
        cases.append({
            'test_type': 'ut',
            'test_file': file_path,
            'origin_test_file': file_path,
            'test_class': '',
            'test_case': test_name
        })

    if 'benchmarks/dynamo/' in body:
        matches = re.findall(r'(python\s+benchmarks/dynamo/[^\s]+)', body)
        for match in matches:
            test_file = match.replace('python ', '').strip()
            cases.append({
                'test_type': 'e2e',
                'test_file': test_file,
                'origin_test_file': test_file,
                'test_class': '',
                'test_case': match.strip()
            })

    if 'pytest' in body:
        k_match = re.search(r'pytest[^-]*(-k\s+[^\s]+)?', body)
        if k_match and k_match.group(1):
            cases.append({
                'test_type': 'ut',
                'test_file': '',
                'origin_test_file': '',
                'test_class': '',
                'test_case': k_match.group(1).strip()
            })

    return cases

def generate_summary(body, title):
    # Summary based on issue title
    return title[:150]

def classify_issue_type(body, title, labels):
    text = f"{title} {body}".lower()
    
    for label in labels:
        ln = label.get('name', '').lower()
        if 'task' == ln or 'internal task' in ln:
            return 'internal task'
    
    performance_keywords = [
        'performance regression', 'performance dropped', 'performance issue',
        'latency', 'throughput', 'slow performance', 'performance slow',
        'execution time', 'runtime performance', 'performance fail'
    ]
    
    has_performance_keyword = any(k in text for k in performance_keywords)
    
    bug_keywords = [
        'assertionerror', 'runtimeerror', 'valueerror', 'typeerror', 'indexerror',
        'keyerror', 'importerror', 'notimplementederror', 'attributeerror',
        'inductorerror', 'crash', 'fail', 'bug', 'error', 'not implemented',
        'not supported', 'missing', 'incorrect', 'wrong', 'unexpected'
    ]
    
    has_bug_keyword = any(k in text for k in bug_keywords)
    
    feature_keywords = ['feature request', 'support for', 'implement', 'add support', 'need feature']
    has_feature_keyword = any(k in text for k in feature_keywords)
    
    if has_feature_keyword:
        return 'feature request'
    
    if has_performance_keyword:
        return 'performance issue'
    
    if has_bug_keyword:
        return 'functionality bug'
    
    return 'unknown'

def is_e2e_issue(body, title, labels):
    """Check if issue is related to E2E benchmark"""
    text = f"{title} {body}".lower()
    
    # Check labels first - only exact 'e2e' label
    for label in labels:
        ln = label.get('name', '').lower()
        if ln == 'e2e':
            return True
    
    # Check for specific E2E benchmark paths (not just the word 'benchmark')
    e2e_patterns = [
        r'benchmarks/dynamo/',           # torch-xpu-ops benchmark scripts
        r'benchmarks/timm/',             # timm benchmark
        r'benchmarks/huggingface/',     # huggingface benchmark
        r'benchmarks/torchbench/',      # torchbench benchmark
        r'run_benchmark\.py',            # torchbenchmark runner
    ]
    
    for pattern in e2e_patterns:
        if re.search(pattern, text):
            return True
    
    # Check for model names from benchmark model lists with explicit benchmark framework mention
    # Only for specific benchmark prefixes
    benchmark_model_prefixes = ['hf_', 'timm_']  # e.g., hf_Albert, timm_resnet50
    
    has_model = False
    has_benchmark_context = False
    
    for prefix in benchmark_model_prefixes:
        if prefix in text:
            has_model = True
            break
    
    # Must have explicit benchmark framework mention (as test framework)
    if has_model:
        benchmark_paths = ['benchmarks/dynamo', 'run_benchmark', 'torchbenchmark', 'benchmark.py']
        for kw in benchmark_paths:
            if kw in text:
                has_benchmark_context = True
                break
    
    if has_model and has_benchmark_context:
        return True
    
    return False


def classify_test_module(body, title, labels):
    text = f"{title} {body}".lower()
    
    # Check if it's an E2E issue first
    if is_e2e_issue(body, title, labels):
        return 'e2e'
    
    pytest_patterns = [
        r'pytest\s+.*test[/._]',
        r'python\s+.*test[/._]',
        r'test/test_',
        r'test/xpu/test_',
    ]
    
    has_test_pattern = False
    for pattern in pytest_patterns:
        if re.search(pattern, text):
            has_test_pattern = True
            break
    
    build_patterns = [
        r'\[win\]\[build\]',
        r'build from source',
        r'compile from source', 
        r'source build',
        r'build script',
        r'BUILD_SEPARATE',
        r'BUILD_SHARED',
        r'cmake build',
        r'cmake error',
        r'cmake fail',
        r'setup\.py install',
        r'pip install -e \.',
        r'python setup\.py develop',
    ]
    
    has_build = any(re.search(p, text, re.IGNORECASE) for p in build_patterns)
    
    infra_patterns = [
        r'workflow\s+(error|fail|issue|problem)',
        r'github\s+action\s+(error|fail|issue)',
        r'azure\s+pipeline\s+(error|fail)',
        r'ci\s+(runner|config|setup)\s+(error|fail)',
        r'runner\s+(error|fail|timeout)\s+in\s+ci',
        r'checkout\s+(error|fail)\s+in\s+(workflow|ci)',
        r'githubaction',
    ]
    
    has_infra = any(re.search(p, text) for p in infra_patterns)
    
    for label in labels:
        ln = label.get('name', '').lower()
        if 'infrastructure' in ln and ('ci' in ln or 'workflow' in ln or 'action' in ln):
            has_infra = True
            break
    
    if has_build:
        return 'build'
    
    if has_infra:
        return 'infrastructure'
    
    if has_test_pattern:
        if 'benchmarks/dynamo/' in text or 'benchmark' in text:
            return 'e2e'
        return 'ut'
    
    return 'ut'

def classify_module(body, title, labels):
    text = f"{title} {body}".lower()
    labels_str = ', '.join([l.get('name', '') for l in labels]).lower()
    
    # Check labels first
    for label in labels:
        ln = label.get('name', '').lower()
        if 'module: distributed' in ln:
            return 'distributed'
        if 'module: inductor' in ln:
            return 'inductor'
        if 'module: ao' in ln:
            return 'AO'
        if 'module: ut' in ln:
            return 'aten_ops'
        if 'module: quant' in ln:
            return 'low_precision'
        if 'module: profiler' in ln:
            return 'profiling'
        if 'module: dynamo' in ln:
            return 'dynamo'
        if 'module: op impl' in ln:
            return 'aten_ops'
    
    # Special case - "Torch not compiled with CUDA enabled" means test configuration issue, not inductor
    if 'torch not compiled with cuda enabled' in text:
        return 'unknown'
    
    # Random failures are not module-specific
    if 'random failure' in text or 'random failures' in text:
        return 'unknown'
    
    # Torch operations (from PyTorch docs)
    torch_ops = [
        'add', 'sub', 'mul', 'div', 'matmul', 'mm', 'dot', 'vdot', 'bmm',
        'addmm', 'addmv', 'addbmm', 'smm', 'spmm', 'mm', 'mv', 'vecdot',
        'conv', 'conv1d', 'conv2d', 'conv3d', 'conv_transpose',
        'batch_norm', 'layer_norm', 'group_norm', 'instance_norm',
        'dropout', 'embedding', 'linear', 'lstm', 'gru', 'rnn',
        'softmax', 'log_softmax', 'sigmoid', 'tanh', 'relu', 'leaky_relu',
        'pool', 'avg_pool', 'max_pool', 'adaptive_pool',
        'fft', 'ifft', 'fft2', 'ifft2',
        'chunk', 'split', 'view', 'reshape', 'transpose', 'permute',
        'cat', 'stack', 'gather', 'scatter', 'index', 'where',
        'sum', 'mean', 'std', 'var', 'min', 'max', 'argmin', 'argmax',
        'norm', 'linalg.norm', 'linalg.matrix_norm', 'linalg.vector_norm',
        'eig', 'svd', 'qr', 'cholesky', 'solve', 'inverse',
        'det', 'logdet', 'slogdet', 'trace',
        'clone', 'copy_', 'to', 'cuda', 'cpu', 'xpu', 'device',
        'zeros', 'ones', 'empty', 'full', 'arange', 'linspace', 'logspace',
        'tensor', 'scalar_tensor', 'tensor.tensor',
        'getitem', 'setitem', 'call', 'forward', 'backward',
        'relu', 'gelu', 'silu', 'mish', 'softplus', 'elu', 'selu', 'celu',
        'flash_attention', 'scaled_dot_product_attention', 'sdpa',
        'interpolate', 'grid_sample', 'affine_grid',
        'grid_sampler', 'grid_sampler_2d',
        'bernoulli', 'normal', 'uniform', 'randn', 'rand', 'randint',
        'multinomial', ' poisson', 'exponential', 'geometric',
        'lerp', 'lerp_', 'fmod', 'remainder', 'nextafter',
        'linspace', 'logspace', 'geomspace',
        'complex', 'real', 'imag', 'angle',
        'conj', 'view_as_real', 'view_as_complex',
    ]
    
    module_keywords = [
        ('distributed', ['distributed', 'device_mesh', 'ProcessGroup', 'FSDP', 'DDP', 'c10d', 'tensor parallel']),
        ('inductor', ['inductor', 'inductor error', 'compile error', 'lower', 'kernel code']),
        ('dynamo', ['dynamo', 'torch.compile', '_dynamo', 'dynamo']),
        ('autograd', ['autograd', 'backward', 'grad', 'gradient']),
        ('aten_ops', ['aten::', 'torch.ops.aten', 'test_ops']),
        ('low_precision', ['quantization', 'int8', 'fp8', 'int4', 'amp', 'bf16', 'fp16', 'float8']),
        ('optimizer', ['optimizer', 'lr_scheduler', 'adam', 'sgd']),
        ('profiling', ['profiling', 'profile', 'benchmark']),
        ('fx', ['torch.fx', 'fx.', 'symbolic']),
        ('export', ['torch.export', 'exported']),
    ]
    
    # Check torch ops first
    for op in torch_ops:
        if re.search(rf'\b{re.escape(op)}\b', text):
            return 'aten_ops'
    
    for m, kw in module_keywords:
        if any(k in text for k in kw):
            return m
    
    return 'unknown'

def get_dependency_from_body(body, labels=None):
    if labels is None:
        labels = []
    
    labels_str = ', '.join([l.get('name', '') for l in labels]).lower()
    
    # Check labels first for 'dependency component:'
    if 'dependency component: onednn' in labels_str or 'dependency component: mkl-dnn' in labels_str or 'dependency component: dnnl' in labels_str:
        return 'oneDNN'
    if 'dependency component: onemkl' in labels_str or 'dependency component: mkl' in labels_str:
        return 'oneMKL'
    if 'dependency component: triton' in labels_str:
        return 'Triton'
    if 'dependency component: torchao' in labels_str:
        return 'AO'
    if 'dependency component: transformers' in labels_str or 'dependency component: huggingface' in labels_str:
        return 'transformers'
    if 'dependency component: oneapi' in labels_str or 'dependency component: sycl' in labels_str:
        return 'oneAPI'
    if 'dependency component: driver' in labels_str:
        return 'driver'
    if 'dependency component: oneccl' in labels_str or 'dependency component: ccl' in labels_str or 'dependency component: xccl' in labels_str:
        return 'oneCCL'
    
    # Filter out version/environment sections
    if not body:
        return 'None'
    
    text = body.lower()
    
    # Remove version/environment sections
    version_headers = [
        r'###\s*version',
        r'###\s*versions',
        r'###\s*environment',
        r'###\s*reproduction',
        r'###\s*steps?\s+to\s+reproduce',
        r'###\s*additional\s*context',
    ]
    
    for header in version_headers:
        match = re.search(header, text, re.IGNORECASE)
        if match:
            text = text[:match.start()]
            break
    
    # Check for actual dependency in body (require context like "caused by", "issue", "depend on")
    dep_keywords = [
        ('transformers', [
            'caused by transformers', 'transformers issue', 'transformers bug',
            'depends on transformers', 'need transformers fix', 'waiting for transformers',
            'huggingface issue', 'huggingface bug', 'depends on huggingface'
        ]),
        ('AO', [
            'caused by torchao', 'torchao issue', 'torchao bug',
            'depends on torchao', 'need torchao fix', 'waiting for torchao'
        ]),
        ('oneDNN', [
            'caused by onednn', 'onednn issue', 'onednn bug', 'oneDNN issue',
            'depends on onednn', 'need onednn fix', 'waiting for onednn',
            'mkl-dnn issue', 'dnnl issue'
        ]),
        ('oneCCL', [
            'caused by oneccl', 'oneccl issue', 'oneccl bug',
            'depends on oneccl', 'need oneccl fix', 'waiting for oneccl',
            'xccl issue', 'ccl issue', 'depends on ccl'
        ]),
        ('oneMKL', [
            'caused by onemkl', 'onemkl issue', 'onemkl bug',
            'depends on onemkl', 'need onemkl fix', 'waiting for onemkl',
            'caused by mkl', 'mkl issue'
        ]),
        ('driver', [
            'caused by driver', 'driver issue', 'driver bug',
            'depends on driver', 'need driver fix', 'waiting for driver'
        ]),
        ('Triton', [
            'caused by triton', 'triton issue', 'triton bug',
            'depends on triton', 'need triton fix', 'waiting for triton',
            'triton-xpu issue', 'tl\\. issue'
        ]),
        ('oneAPI', [
            'caused by oneapi', 'oneapi issue', 'oneapi bug', 'sycl issue',
            'depends on oneapi', 'need oneapi fix', 'waiting for oneapi',
            'icpx issue', 'dpcpp issue', 'sycl compiler issue'
        ]),
    ]
    
    for d, kw in dep_keywords:
        if any(k in text for k in kw):
            return d
    
    return 'None'


# Create Excel
wb = openpyxl.Workbook()

# Sheet 1: Issues
ws_issues = wb.active
ws_issues.title = "Issues"

# Core columns for basic issue info
# Note: PR columns (PR, PR Owner, PR Status, PR Description) populated by ../pr-extraction/
# Note: owner_transfer, action_TBD, priority, Category, Root Cause columns populated by update_test_results/
headers = ["Issue ID", "Title", "Status", "Assignee", "Reporter", "Labels",
           "Created Time", "Updated Time", "Milestone", "Summary", "Type",
           "Module", "Test Module", "Dependency"]

for col, header in enumerate(headers, 1):
    cell = ws_issues.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 2: Test Cases (ut)
ws_cases = wb.create_sheet("Test Cases")

# Core columns for test case basic info
# Note: Error Message, Traceback, torch-ops, dependency filled by test_result_analysis/Test_Cases
# Note: XPU Status, Stock Status, CUDA/XPU case exist filled by update_test_results/
case_headers = ["Issue ID", "Test Reproducer", "Test Type", "Test File",
                "Origin Test File", "Test Class", "Test Case",
                "Error Message", "Traceback", "torch-ops", "dependency",
                "XPU Status", "Stock Status", "Is SKIP", "Is CUDA Skip",
                "CUDA Case Exist", "XPU Case Exist", "case_existence_comments",
                "can_enable_on_xpu", "duplicated_issue"]

for col, header in enumerate(case_headers, 1):
    cell = ws_cases.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 3: E2E Test Cases
ws_e2e = wb.create_sheet("E2E Test Cases")

# Core columns for E2E test case basic info
# Note: Error Message, Traceback filled by test_result_analysis/E2E_Test_Cases
e2e_headers = ["Issue ID", "Test Reproducer", "Benchmark", "Model", "Phase", "Dtype", "AMP",
               "Backend", "Test Type", "Cudagraph", "Error Message", "Traceback"]

for col, header in enumerate(e2e_headers, 1):
    cell = ws_e2e.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

issue_row = 2
case_row = 2
e2e_row = 2

for issue in issues:
    num = issue['number']
    title = issue['title']
    body = issue.get('body', '') or ''
    
    status = issue['state']
    assignee = ", ".join([a['login'] for a in issue.get('assignees', [])]) or "None"
    reporter = issue['user']['login']
    labels = issue.get('labels', [])
    label_str = ", ".join([l['name'] for l in labels])
    created = issue['created_at']
    updated = issue['updated_at']
    milestone = issue.get('milestone', {})
    milestone_name = milestone.get('title', 'None') if milestone else 'None'
    
    issue_type = classify_issue_type(body, title, labels)
    module = classify_module(body, title, labels)
    test_module = classify_test_module(body, title, labels)
    dependency = get_dependency_from_body(body, labels)
    
    # Error Message, Traceback, torch-ops, dependency deferred to test_result_analysis/
    # Leave blank for now - will be populated by test_result_analysis/Test_Cases/
    summary = generate_summary(body, title)
    
    ws_issues.cell(row=issue_row, column=1, value=num)
    ws_issues.cell(row=issue_row, column=2, value=title)
    ws_issues.cell(row=issue_row, column=3, value=status)
    ws_issues.cell(row=issue_row, column=4, value=assignee)
    ws_issues.cell(row=issue_row, column=5, value=reporter)
    ws_issues.cell(row=issue_row, column=6, value=label_str)
    ws_issues.cell(row=issue_row, column=7, value=created)
    ws_issues.cell(row=issue_row, column=8, value=updated)
    ws_issues.cell(row=issue_row, column=9, value=milestone_name)
    ws_issues.cell(row=issue_row, column=10, value=summary)
    ws_issues.cell(row=issue_row, column=11, value=issue_type)
    ws_issues.cell(row=issue_row, column=12, value=module)
    ws_issues.cell(row=issue_row, column=13, value=test_module)
    ws_issues.cell(row=issue_row, column=14, value=dependency)
    
    # Parse test cases and e2e info
    test_cases = parse_test_cases_from_body(body)
    
    # Only parse e2e info if it's actually an e2e issue
    e2e_info = []
    if test_module == 'e2e':
        e2e_info = parse_e2e_info(body, title)
    
    # Add to test cases sheet (non-e2e)
    if test_cases:
        for tc in test_cases:
            # Skip e2e cases - they go to e2e sheet
            if tc.get('test_type') == 'e2e':
                continue

            ws_cases.cell(row=case_row, column=1, value=num)
            ws_cases.cell(row=case_row, column=2, value=title[:150])
            ws_cases.cell(row=case_row, column=3, value=tc.get('test_type', ''))
            ws_cases.cell(row=case_row, column=4, value=tc.get('test_file', ''))
            ws_cases.cell(row=case_row, column=5, value=tc.get('origin_test_file', ''))
            ws_cases.cell(row=case_row, column=6, value=tc.get('test_class', ''))
            ws_cases.cell(row=case_row, column=7, value=tc.get('test_case', ''))
            # Columns 8-11: Error Message, Traceback, torch-ops, dependency - left blank for test_result_analysis/
            case_row += 1

    # Add to e2e sheet
    if e2e_info:
        for info in e2e_info:
            reproducer = info.get('reproducer', title[:150])
            ws_e2e.cell(row=e2e_row, column=1, value=num)
            ws_e2e.cell(row=e2e_row, column=2, value=reproducer[:200] if reproducer else title[:150])
            ws_e2e.cell(row=e2e_row, column=3, value=info.get('benchmark', ''))
            ws_e2e.cell(row=e2e_row, column=4, value=info.get('model', ''))
            ws_e2e.cell(row=e2e_row, column=5, value=info.get('phase', ''))
            ws_e2e.cell(row=e2e_row, column=6, value=info.get('dtype', ''))
            ws_e2e.cell(row=e2e_row, column=7, value=info.get('amp', False))
            ws_e2e.cell(row=e2e_row, column=8, value=info.get('backend', ''))
            ws_e2e.cell(row=e2e_row, column=9, value=info.get('test_type', ''))
            ws_e2e.cell(row=e2e_row, column=10, value=info.get('disable_cudagraphs', ''))
            # Columns 11-12: Error Message, Traceback - left blank for test_result_analysis/
            e2e_row += 1
    elif test_module == 'e2e':
        # Add e2e issues without specific model info
        ws_e2e.cell(row=e2e_row, column=1, value=num)
        ws_e2e.cell(row=e2e_row, column=2, value=title[:150])
        ws_e2e.cell(row=e2e_row, column=3, value='unknown')
        # Columns 11-12: Error Message, Traceback - left blank for test_result_analysis/
        e2e_row += 1

    issue_row += 1

    if issue_row % 50 == 0:
        print(f"Processed {issue_row-1} issues...")

print(f"\nTotal issues: {issue_row-2}")
print(f"Total test case rows: {case_row-2}")
print(f"Total e2e case rows: {e2e_row-2}")

for ws in [ws_issues, ws_cases, ws_e2e]:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

# Ensure result directory exists
os.makedirs(RESULT_DIR, exist_ok=True)

output_path = os.path.join(RESULT_DIR, "torch_xpu_ops_issues.xlsx")
wb.save(output_path)
print(f"\nSaved to {output_path}")
