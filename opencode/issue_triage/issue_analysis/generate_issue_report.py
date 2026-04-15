#!/usr/bin/env python3
"""
Generate issue_report.md

Creates a comprehensive Markdown report with:
1. Summary
2. Action Required
   2.1 Developer AR (Need Investigation by Action Reason type)
   2.2 Reporter AR (other Action TBD values)
3. Issues by Category
4. Last Week Issues
5. Stale Issues
6. Dependency Issues
7. Duplicated Issues
8. Statistics

Each table includes: Priority, Priority Reason, Action Reason columns
"""

import os
import sys
import re
from datetime import datetime, timedelta
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ISSUE_TRIAGE_DIR = os.path.dirname(SCRIPT_DIR)
ROOT_DIR = os.path.dirname(ISSUE_TRIAGE_DIR)
RESULT_DIR = os.environ.get('RESULT_DIR', os.path.join(ISSUE_TRIAGE_DIR, 'result'))

# Priority mapping from SKILL.md
PRIORITY_MAP = {
    'add to skiplist': 1,
    'Close fixed issue': 2,
    'Verify the issue': 3,
    'No Test Status in CI': 4,
    'Needs Upstream Skip PR': 5,
    'Needs Skip PR': 6,
    'E2E accuracy issue': 7,
    'Awaiting response': 8,
    'Awaiting response from reporter': 9,
}

PRIORITY_REASON_MAP = {
    1: 'Issues marked as not_target/wontfix or cannot be enabled',
    2: 'All test cases passed on both XPU and stock',
    3: 'PR exists but no failures',
    4: 'No test status available - needs testing',
    5: 'Issue is upstream - needs skip PR upstream',
    6: 'Issue marked as wontfix/not_target - needs skip PR',
    7: 'E2E accuracy issue pending - needs upstream investigation',
    8: 'Bug/Perf issue pending reporter response',
    9: 'Maintainer requested info from reporter',
    'N': 'Fallback - no specific action identified',
    99: 'No specific action identified',
}

# Action Reason type patterns for categorizing "Need Investigation"
ACTION_REASON_PATTERNS = [
    (r'All test cases passed', 'All test cases passed on both XPU and stock'),
    (r'PR closed but tests? still failing', 'PR closed but tests still failing'),
    (r'PR closed but no failed tests', 'PR closed but no failed tests'),
    (r'No specific action identified', 'No specific action identified - needs investigation'),
    (r'Bug/Perf issue awaiting reporter response', 'Bug/Perf issue awaiting reporter response'),
    (r'Fix (distributed|Gloo|backend|initialization)', 'Backend/initialization fix needed'),
    (r'Fix (precision|dtype|accuracy)', 'Precision/dtype fix needed'),
    (r'Fix (memory|allocation|deallocation)', 'Memory management fix needed'),
    (r'Fix (flash attention|torch\._decomp|sdpa)', 'Flash attention/op fix needed'),
    (r'Fix (Inductor|lowering|decomposition)', 'Inductor compilation fix needed'),
    (r'Implement XPU-specific kernel', 'Missing XPU kernel implementation'),
    (r'Implement (distributed|backend)', 'Missing distributed backend'),
    (r'Add test case to skiplist|xfail', 'Needs skiplist/xfail entry'),
    (r'Optimize performance', 'Performance optimization needed'),
    (r'Investigate|Fix error:', 'Test failure investigation needed'),
    (r'Fix output mismatch', 'Output mismatch to fix'),
    (r'needs investigation', 'needs investigation'),
]

# Column width limits
TITLE_LEN = 65
REASON_LEN = 45
ACTION_LEN = 120
DEPENDENCY_LEN = 50
DUPLICATE_LEN = 50


def get_priority(issue_row: dict) -> int:
    """Get priority number based on Action TBD."""
    action_tbd = issue_row.get('Action TBD', '')
    if action_tbd in PRIORITY_MAP:
        return PRIORITY_MAP[action_tbd]
    return 99  # Will be mapped to 'N' later


def get_priority_reason(priority: int) -> str:
    """Get priority reason based on priority number."""
    return PRIORITY_REASON_MAP.get(priority, PRIORITY_REASON_MAP['N'])


def format_date(date_val) -> str:
    """Format date value to string."""
    if date_val is None:
        return ''
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')
    if isinstance(date_val, str):
        return date_val[:10] if len(str(date_val)) >= 10 else str(date_val)
    return str(date_val)[:10] if len(str(date_val)) >= 10 else str(date_val)


def truncate(text: str, max_len: int = 80) -> str:
    """Truncate text to fit table columns while showing key info."""
    if text is None:
        return ''
    text = str(text)
    
    # Clean up text: remove excessive whitespace
    text = ' '.join(text.split())
    
    # Apply truncation with ellipsis
    if len(text) > max_len:
        truncated = text[:max_len-3]
        last_space = truncated.rfind(' ')
        if last_space > max_len * 0.5:
            truncated = truncated[:last_space]
        return truncated + '...'
    
    return text


def load_issues(excel_file: str) -> list:
    """Load issues from Excel."""
    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Issues']
    
    # Get headers
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    
    # Build issue list - handle duplicate 'Action Reason' column
    issues = []
    for row in range(2, ws.max_row + 1):
        issue = {}
        for col, header in enumerate(headers, 1):
            # Rename duplicate headers for clarity
            if header == 'Action Reason':
                if col == 21:
                    display_header = 'Action Reason'
                elif col == 27:
                    display_header = 'Action Reason Derived'
                else:
                    display_header = header
            else:
                display_header = header
            issue[display_header] = ws.cell(row, col).value
        issues.append(issue)
    
    print(f"Loaded {len(issues)} issues")
    return issues


def extract_reason_type(action_reason: str) -> str:
    """Extract a categorization type from Action Reason text."""
    if not action_reason:
        return 'Needs investigation - general'
    ar = str(action_reason)
    for pattern, label in ACTION_REASON_PATTERNS:
        if re.search(pattern, ar, re.IGNORECASE):
            return label
    if ':' in ar:
        return ar.split(':')[0].strip()[:60]
    return ar.strip()[:60]


def build_developer_action_types(issues: list) -> list:
    """Get unique Action Reason types from Need Investigation issues.
    
    Returns sorted list of (type_name, list_of_issues) tuples, sorted by count descending.
    """
    type_map = {}
    for issue in issues:
        action_tbd = issue.get('Action TBD', '') or ''
        if action_tbd == 'Need Investigation':
            action_reason = issue.get('Action Reason', '') or ''
            reason_type = extract_reason_type(action_reason)
            if reason_type not in type_map:
                type_map[reason_type] = []
            type_map[reason_type].append(issue)
    
    # Sort by count descending
    sorted_types = sorted(type_map.items(), key=lambda x: -len(x[1]))
    return sorted_types


def build_reporter_action_types(issues: list) -> list:
    """Get Action TBD values for Reporter AR (all except Need Investigation).
    
    Returns list of (action_tbd_name, list_of_issues) tuples.
    """
    reporter_types = [
        'Needs Upstream Skip PR',
        'add to skiplist',
        'Close fixed issue',
        'Verify the issue',
        'Awaiting response',
        'Awaiting response from reporter',
        'E2E accuracy issue',
    ]
    type_map = {}
    for issue in issues:
        action_tbd = issue.get('Action TBD', '') or ''
        if action_tbd and action_tbd != 'Need Investigation':
            if action_tbd not in type_map:
                type_map[action_tbd] = []
            type_map[action_tbd].append(issue)
    
    # Return only types that have issues, in the defined order
    result = []
    for action in reporter_types:
        if action in type_map and type_map[action]:
            result.append((action, type_map[action]))
    return result


def build_categories() -> list:
    """Define all categories and their display names."""
    return [
        'Distributed',
        'TorchAO',
        'PT2E',
        'Flash Attention / Transformer Related',
        'Sparse Operations Related',
        'Inductor / Compilation Related',
        'Others',
        'Dtype / Precision Related',
    ]


def get_category_display_name(cat: str) -> str:
    """Map category codes/names to display names."""
    category_map = {
        '1 - Distributed': 'Distributed',
        '2 - TorchAO': 'TorchAO',
        '3 - PT2E': 'PT2E',
        '4 - Flash Attention/Transformer': 'Flash Attention / Transformer Related',
        '5 - Sparse': 'Sparse Operations Related',
        '6 - Inductor/Compilation': 'Inductor / Compilation Related',
        '7 - Torch Runtime': 'Others',
        '8 - Torch Operations': 'Others',
        '9 - Dtype/Precision': 'Dtype / Precision Related',
        '10 - Feature Not Supported': 'Others',
        '11 - Skip/No Test Exists': 'Others',
        '12 - Others': 'Others',
        'Distributed': 'Distributed',
        'TorchAO': 'TorchAO',
        'PT2E': 'PT2E',
        'Flash Attention / Transformer': 'Flash Attention / Transformer Related',
        'Flash Attention/Transformer': 'Flash Attention / Transformer Related',
        'Sparse': 'Sparse Operations Related',
        'Inductor/Compilation': 'Inductor / Compilation Related',
        'Dtype/Precision': 'Dtype / Precision Related',
    }
    return category_map.get(cat, cat)


def generate_table_header() -> str:
    """Generate standard table header."""
    return '| # | ID | Title | Priority | Priority Reason | Action Reason | Owner Transfer | Category | Test Module |'


def generate_table_header2() -> str:
    """Generate standard table header with Action TBD column."""
    return '| # | ID | Title | Priority | Priority Reason | Action Reason | Owner Transfer | Action TBD | Category | Test Module |'


def generate_table_separator() -> str:
    """Generate standard table separator with consistent GFM alignment."""
    return '|---|------|------|----------|--------------------------------------------|----------|----------|------------|------------|'


def generate_table_separator2() -> str:
    """Generate table separator with Action TBD column (11 columns)."""
    return '|---|------|------|----------|--------------------------------------------|----------|----------|------------|------------|-----------|'


def generate_table_row(idx: int, issue: dict, show_action: bool = False, use_display_category: bool = True, full_action_reason: bool = False) -> str:
    """Generate a table row with balanced column widths."""
    # Use actual Priority and Priority Reason from Excel if available
    priority_display = issue.get('Priority', 'N') or 'N'
    priority_reason = issue.get('Priority Reason', '') or ''
    
    # Get category with display name mapping
    cat = issue.get('Category', '')
    if use_display_category:
        cat = get_category_display_name(cat)
    
    # Get action reason - either full text or truncated
    action_reason = issue.get('Action Reason', '')
    if not full_action_reason:
        action_reason = truncate(action_reason, ACTION_LEN)
    
    row = [
        idx,
        issue.get('Issue ID', ''),
        truncate(issue.get('Title', ''), TITLE_LEN),
        priority_display,
        truncate(priority_reason, REASON_LEN),
        action_reason,
        issue.get('Owner Transfer', ''),
    ]
    
    if show_action:
        row.extend([
            issue.get('Action TBD', ''),
        ])
    
    row.extend([
        cat,
        issue.get('Test Module', ''),
    ])
    
    return '| ' + ' | '.join(str(x) for x in row) + ' |'


def generate_action_required_section(issues: list) -> str:
    """Generate section 2: Action Required with Developer and Reporter subsections."""
    lines = []
    lines.append('## <span id=\'2-action-required\'>2. Action Required</span>\n')
    
    # 2.1 Developer AR - Need Investigation by Action Reason type
    lines.append('### <span id=\'action-required-developer\'>2.1 Developer AR (Need Investigation by Action Reason)</span>\n')
    lines.append('*Issues pending investigation, grouped by type of action needed - Developer*\n')
    
    dev_types = build_developer_action_types(issues)
    
    # Separate feature requests from other issues
    feature_requests = []
    regular_dev_issues = []
    small_type_issues = []  # For tables with < 3 issues
    
    for reason_type, type_issues in dev_types:
        # Separate feature request typed issues from regular issues
        non_feature_issues = [issue for issue in type_issues if 'feature request' not in str(issue.get('Type', '')).lower()]
        feature_issues = [issue for issue in type_issues if 'feature request' in str(issue.get('Type', '')).lower()]
        
        # Add feature typed issues to feature_requests
        feature_requests.extend(feature_issues)
        
        # Use non-feature issues for regular tables
        if len(non_feature_issues) >= 3:
            regular_dev_issues.append((reason_type, non_feature_issues))
        elif len(non_feature_issues) > 0:
            small_type_issues.extend(non_feature_issues)
    
    # Generate tables for regular dev types (>= 3 issues)
    dev_idx = 1
    for reason_type, type_issues in regular_dev_issues:
        issue_list = sorted(type_issues, key=lambda x: x.get('Issue ID', 0), reverse=True)
        count = len(issue_list)
        
        # Create anchor
        anchor = f'2.1-{dev_idx}-' + reason_type.lower().replace(' ', '-').replace('/', '-').replace(',', '')[:40]
        lines.append(f'#### <span id=\'{anchor}\'>2.1.{dev_idx} {reason_type} - Developer</span> ({count} issues)\n')
        lines.append(generate_table_header2())
        lines.append(generate_table_separator2())
        
        for row_idx, issue in enumerate(issue_list, 1):
            lines.append(generate_table_row(row_idx, issue, show_action=True, full_action_reason=True))
        
        lines.append('')
        dev_idx += 1
    
    # Merge small tables (< 3 issues) into "Others"
    if small_type_issues:
        issue_list = sorted(small_type_issues, key=lambda x: x.get('Issue ID', 0), reverse=True)
        count = len(issue_list)
        
        lines.append(f'#### <span id=\'2.1-{dev_idx}-others\'>2.1.{dev_idx} Others - Developer</span> ({count} issues)\n')
        lines.append('*Tables merged due to low issue count (< 3 issues each)*\n')
        lines.append(generate_table_header2())
        lines.append(generate_table_separator2())
        
        for row_idx, issue in enumerate(issue_list, 1):
            lines.append(generate_table_row(row_idx, issue, show_action=True, full_action_reason=True))
        
        lines.append('')
        dev_idx += 1
    
    # Feature Requests Table
    if feature_requests:
        issue_list = sorted(feature_requests, key=lambda x: x.get('Issue ID', 0), reverse=True)
        count = len(issue_list)
        
        lines.append(f'#### <span id=\'2.1-{dev_idx}-feature-requests\'>2.1.{dev_idx} Feature Requests - Developer</span> ({count} issues)\n')
        lines.append(generate_table_header2())
        lines.append(generate_table_separator2())
        
        for row_idx, issue in enumerate(issue_list, 1):
            lines.append(generate_table_row(row_idx, issue, show_action=True, full_action_reason=True))
        
        lines.append('')
    
    # 2.2 Reporter AR - Other Action TBD values
    lines.append('### <span id=\'action-required-reporter\'>2.2 Reporter AR (Other Action TBD)</span>\n')
    lines.append('*Action TBD values requiring reporter/community response*\n')
    
    reporter_types = build_reporter_action_types(issues)
    reporter_idx = 1
    for action, action_issues in reporter_types:
        issue_list = sorted(action_issues, key=lambda x: x.get('Issue ID', 0), reverse=True)
        count = len(issue_list)
        
        # Create anchor
        anchor = f'2.2-{reporter_idx}-' + action.lower().replace(' ', '-').replace('/', '-')[:40]
        lines.append(f'#### <span id=\'{anchor}\'>2.2.{reporter_idx} {action} - Reporter</span> ({count} issues)\n')
        lines.append(generate_table_header2())
        lines.append(generate_table_separator2())
        
        for row_idx, issue in enumerate(issue_list, 1):
            lines.append(generate_table_row(row_idx, issue, show_action=True, full_action_reason=True))
        
        lines.append('')
        reporter_idx += 1
    
    return '\n'.join(lines)


def generate_summary_section(issues: list) -> str:
    """Generate section 1: Summary."""
    lines = []
    lines.append('## <span id=\'1-summary\'>1. Summary</span>\n')
    
    total = len(issues)
    lines.append(f'**Total Issues: {total}**\n')
    
    # Count by category using normalized display names
    cat_counts = {}
    for issue in issues:
        raw_cat = issue.get('Category', 'unknown') or 'unknown'
        cat = get_category_display_name(raw_cat)
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    
    lines.append('### <span id=\'category-summary\'>Issues by Category</span>\n')
    lines.append('| Category | Count |')
    lines.append('|----------|------:|')
    for cat in sorted(cat_counts.keys()):
        lines.append(f'| {cat} | {cat_counts[cat]} |')
    
    lines.append('')
    return '\n'.join(lines)


def generate_category_section(issues: list) -> str:
    """Generate section 3: Issues by Category."""
    lines = []
    lines.append('## <span id=\'3-issues-by-category\'>3. Issues by Category</span>\n')
    
    # Group by Category using normalized display names
    by_category = {}
    for issue in issues:
        raw_cat = issue.get('Category', 'unknown') or 'unknown'
        cat = get_category_display_name(raw_cat)
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(issue)
    
    # Build category display name list (preserve order)
    categories = build_categories()
    
    # Add any categories not in predefined list
    for cat in by_category:
        disp_cat = get_category_display_name(cat)
        if disp_cat not in categories:
            categories.append(disp_cat)
    
    # Generate tables for each category
    for cat in categories:
        if cat not in by_category or not by_category[cat]:
            continue
        
        issue_list = by_category[cat]
        issue_list.sort(key=lambda x: x.get('Issue ID', 0), reverse=True)
        count = len(issue_list)
        
        anchor = cat.lower().replace(' ', '-').replace('/', '-')
        lines.append(f'### <span id=\'{anchor}\'>{cat}</span> ({count} issues)\n')
        lines.append(generate_table_header())
        lines.append(generate_table_separator())
        
        for row_idx, issue in enumerate(issue_list, 1):
            lines.append(generate_table_row(row_idx, issue))
        
        lines.append('')
    
    return '\n'.join(lines)


def generate_last_week_section(issues: list) -> str:
    """Generate section 4: Last Week Issues."""
    lines = []
    lines.append('## <span id=\'4-last-week-issues\'>4. Last Week Issues</span>\n')
    
    # Calculate date 7 days ago
    seven_days_ago = datetime.now() - timedelta(days=7)
    
    # Filter issues from last week
    last_week = []
    for issue in issues:
        created = issue.get('Created Time')
        if created:
            if isinstance(created, datetime):
                if created >= seven_days_ago:
                    last_week.append(issue)
            elif isinstance(created, str) and len(created) >= 10:
                try:
                    created_dt = datetime.strptime(created[:10], '%Y-%m-%d')
                    if created_dt >= seven_days_ago:
                        last_week.append(issue)
                except ValueError:
                    pass
    
    # Sort by ID descending
    last_week.sort(key=lambda x: x.get('Issue ID', 0), reverse=True)
    count = len(last_week)
    
    lines.append(f'**Issues reported in last 7 days: {count}**\n')
    
    if count > 0:
        lines.append('| # | ID | Title | Priority | Action Reason | Category | Created Time |')
        lines.append('|---|------|------|----------|--------------------------------------------|----------|--------------|')
        
        for idx, issue in enumerate(last_week, 1):
            priority = get_priority(issue)
            priority_display = priority if priority < 99 else 'N'
            created = format_date(issue.get('Created Time'))
            
            lines.append(f"| {idx} | {issue.get('Issue ID', '')} | {truncate(issue.get('Title', ''), TITLE_LEN)} | {priority_display} | {truncate(issue.get('Action Reason', ''), ACTION_LEN)} | {get_category_display_name(issue.get('Category', ''))} | {created} |")
    else:
        lines.append('No issues reported in the last 7 days.\n')
    
    lines.append('')
    return '\n'.join(lines)


def generate_stale_section(issues: list) -> str:
    """Generate section 5: Stale Issues."""
    lines = []
    lines.append('## <span id=\'4-stale-issues\'>4. Stale Issues - No Update 2+ Weeks</span>\n')
    
    # Calculate date 14 days ago
    fourteen_days_ago = datetime.now() - timedelta(days=14)
    
    # Filter stale, open issues
    stale = []
    for issue in issues:
        # Skip closed issues
        status = str(issue.get('Status', '')).lower()
        if 'closed' in status:
            continue
        
        updated = issue.get('Updated Time')
        if updated:
            if isinstance(updated, datetime):
                if updated < fourteen_days_ago:
                    stale.append(issue)
            elif isinstance(updated, str) and len(updated) >= 10:
                try:
                    updated_dt = datetime.strptime(updated[:10], '%Y-%m-%d')
                    if updated_dt < fourteen_days_ago:
                        stale.append(issue)
                except ValueError:
                    pass
    
    # Sort by Updated Time ascending (oldest first)
    stale.sort(key=lambda x: x.get('Updated Time', datetime.min) if x.get('Updated Time') else datetime.min)
    count = len(stale)
    
    lines.append(f'**Issues without update for 2+ weeks (excluding closed): {count}**\n')
    
    if count > 0:
        lines.append('| # | ID | Title | Priority | Action Reason | Category | Updated Time | Days Since Update |')
        lines.append('|---|------|------|----------|--------------------------------------------|----------|---------------|----------------|')
        
        for idx, issue in enumerate(stale, 1):
            priority = get_priority(issue)
            priority_display = priority if priority < 99 else 'N'
            updated = format_date(issue.get('Updated Time'))
            
            # Calculate days since update
            try:
                if isinstance(issue.get('Updated Time'), datetime):
                    days = (datetime.now() - issue.get('Updated Time')).days
                elif isinstance(issue.get('Updated Time'), str):
                    days = (datetime.now() - datetime.strptime(issue.get('Updated Time')[:10], '%Y-%m-%d')).days
                else:
                    days = 'N/A'
            except:
                days = 'N/A'
            
            lines.append(f"| {idx} | {issue.get('Issue ID', '')} | {truncate(issue.get('Title', ''), TITLE_LEN)} | {priority_display} | {truncate(issue.get('Action Reason', ''), ACTION_LEN)} | {get_category_display_name(issue.get('Category', ''))} | {updated} | {days} |")
    else:
        lines.append('No stale issues found.\n')
    
    lines.append('')
    return '\n'.join(lines)


def generate_dependency_section(issues: list) -> str:
    """Generate section 6: Dependency Issues."""
    lines = []
    lines.append('## <span id=\'6-dependency-issues\'>6. Dependency Issues</span>\n')
    
    # Filter issues with dependencies
    deps = []
    for issue in issues:
        dep = issue.get('Dependency', '')
        if dep and str(dep).strip():
            deps.append(issue)
    
    # Sort by ID descending
    deps.sort(key=lambda x: x.get('Issue ID', 0), reverse=True)
    count = len(deps)
    
    lines.append(f'**Issues with dependencies: {count}**\n')
    
    if count > 0:
        lines.append('| # | ID | Title | Priority | Dependency | Category |')
        lines.append('|---|------|------|----------|--------------------------------------------|----------|')
        
        for idx, issue in enumerate(deps, 1):
            priority = get_priority(issue)
            priority_display = priority if priority < 99 else 'N'
            
            lines.append(f"| {idx} | {issue.get('Issue ID', '')} | {truncate(issue.get('Title', ''), TITLE_LEN)} | {priority_display} | {truncate(issue.get('Dependency', ''), DEPENDENCY_LEN)} | {get_category_display_name(issue.get('Category', ''))} |")
    else:
        lines.append('No issues with dependencies.\n')
    
    lines.append('')
    return '\n'.join(lines)


def generate_duplicated_section(issues: list) -> str:
    """Generate section 7: Duplicated Issues."""
    lines = []
    lines.append('## <span id=\'7-duplicated-issues\'>7. Duplicated Issues</span>\n')
    
    # Filter duplicated issues
    dupes = []
    for issue in issues:
        dupe = issue.get('duplicated_issue', '')
        if dupe and str(dupe).strip():
            dupes.append(issue)
    
    # Sort by ID descending
    dupes.sort(key=lambda x: x.get('Issue ID', 0), reverse=True)
    count = len(dupes)
    
    lines.append(f'**Issues marked as duplicated: {count}**\n')
    
    if count > 0:
        lines.append('| # | ID | Title | Priority | Duplicated Issue |')
        lines.append('|---|------|------|----------|--------------------------------------------|')
        
        for idx, issue in enumerate(dupes, 1):
            priority = get_priority(issue)
            priority_display = priority if priority < 99 else 'N'
            
            lines.append(f"| {idx} | {issue.get('Issue ID', '')} | {truncate(issue.get('Title', ''), TITLE_LEN)} | {priority_display} | {truncate(issue.get('duplicated_issue', ''), DUPLICATE_LEN)} |")
    else:
        lines.append('No duplicated issues.\n')
    
    lines.append('')
    return '\n'.join(lines)


def generate_statistics_section(issues: list) -> str:
    """Generate section 7: Statistics."""
    lines = []
    lines.append('## <span id=\'8-statistics\'>8. Statistics</span>\n')
    
    # Action TBD counts
    lines.append('### Action TBD Distribution\n')
    action_counts = {}
    for issue in issues:
        action = issue.get('Action TBD', '') or 'Need Investigation'
        action_counts[action] = action_counts.get(action, 0) + 1
    
    lines.append('| Action TBD | Count |')
    lines.append('|------------|------:|')
    for action in sorted(action_counts.keys()):
        lines.append(f'| {action} | {action_counts[action]} |')
    
    # Category counts (using normalized display names)
    lines.append('\n### Category Distribution\n')
    cat_counts = {}
    for issue in issues:
        raw_cat = issue.get('Category', 'unknown') or 'unknown'
        cat = get_category_display_name(raw_cat)
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    
    lines.append('| Category | Count |')
    lines.append('|----------|------:|')
    for cat in sorted(cat_counts.keys()):
        lines.append(f'| {cat} | {cat_counts[cat]} |')
    
    # Test Module counts
    lines.append('\n### Test Module Distribution\n')
    module_counts = {}
    for issue in issues:
        module = issue.get('Test Module', 'unknown') or 'unknown'
        module_counts[module] = module_counts.get(module, 0) + 1
    
    lines.append('| Test Module | Count |')
    lines.append('|-------------|------:|')
    for module in sorted(module_counts.keys()):
        lines.append(f'| {module} | {module_counts[module]} |')
    
    lines.append('')
    return '\n'.join(lines)


def generate_index(issues: list) -> str:
    """Generate index/TOC."""
    lines = []
    lines.append('## Index\n')
    
    # Action Required count
    action_counts = {}
    for issue in issues:
        action = issue.get('Action TBD', '') or 'Need Investigation'
        action_counts[action] = action_counts.get(action, 0) + 1
    ni_count = action_counts.get('Need Investigation', 0)
    
    lines.append(f'- [1. Summary](#1-summary) - {len(issues)} issues')
    lines.append('- [2. Action Required](#2-action-required)')
    
    # Developer AR - Need Investigation by type
    dev_types = build_developer_action_types(issues)
    
    # Separate feature requests and small tables
    feature_requests = []
    small_type_issues = []
    regular_dev_issues = []
    
    for reason_type, type_issues in dev_types:
        # Separate feature request typed issues from regular issues
        non_feature_issues = [issue for issue in type_issues if 'feature request' not in str(issue.get('Type', '')).lower()]
        feature_issues = [issue for issue in type_issues if 'feature request' in str(issue.get('Type', '')).lower()]
        
        # Add feature typed issues to feature_requests
        feature_requests.extend(feature_issues)
        
        # Use non-feature issues for regular tables
        if len(non_feature_issues) >= 3:
            regular_dev_issues.append((reason_type, non_feature_issues))
        elif len(non_feature_issues) > 0:
            small_type_issues.extend(non_feature_issues)
    
    # Index for regular dev types
    for idx, (reason_type, type_issues) in enumerate(regular_dev_issues, 1):
        anchor = f'2.1-{idx}-' + reason_type.lower().replace(' ', '-').replace('/', '-').replace(',', '')[:40]
        lines.append(f'    - [2.1.{idx}. {reason_type} - Developer](#{anchor}) - {len(type_issues)}')
    
    # Index for merged "Others"
    if small_type_issues:
        lines.append(f'    - [2.1.{len(regular_dev_issues) + 1}. Others - Developer](#2.1-{len(regular_dev_issues) + 1}-others) - {len(small_type_issues)}')
    
    # Index for Feature Requests - use correct numbering
    if feature_requests:
        # Feature index = regular_dev_issues count + (1 if Others exists) + 1 for Feature
        feature_idx = len(regular_dev_issues) + (1 if small_type_issues else 0) + 1
        lines.append(f'    - [2.1.{feature_idx}. Feature Requests - Developer](#2.1-{feature_idx}-feature-requests) - {len(feature_requests)}')
    
    # Reporter AR - Other actions
    reporter_types = build_reporter_action_types(issues)
    for idx, (action, action_issues) in enumerate(reporter_types, 1):
        anchor = f'2.2-{idx}-' + action.lower().replace(' ', '-').replace('/', '-')[:40]
        lines.append(f'    - [2.2.{idx}. {action} - Reporter](#{anchor}) - {len(action_issues)}')
    
    # Category count
    cat_counts = {}
    for issue in issues:
        cat = issue.get('Category', 'unknown') or 'unknown'
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    total_cat = sum(cat_counts.values())
    lines.append(f'- [3. Issues by Category](#3-issues-by-category) - {total_cat} issues')
    
    # Last week
    seven_days_ago = datetime.now() - timedelta(days=7)
    last_week_count = 0
    for issue in issues:
        created = issue.get('Created Time')
        if created and isinstance(created, datetime):
            if created >= seven_days_ago:
                last_week_count += 1
        elif isinstance(created, str) and len(created) >= 10:
            try:
                if datetime.strptime(created[:10], '%Y-%m-%d') >= seven_days_ago:
                    last_week_count += 1
            except ValueError:
                pass
    lines.append(f'- [4. Last Week Issues](#4-last-week-issues) - {last_week_count} issues')
    
    # Stale
    stale_count = 0
    fourteen_days_ago = datetime.now() - timedelta(days=14)
    for issue in issues:
        status = str(issue.get('Status', '')).lower()
        if 'closed' in status:
            continue
        updated = issue.get('Updated Time')
        if updated and isinstance(updated, datetime):
            if updated < fourteen_days_ago:
                stale_count += 1
        elif isinstance(updated, str) and len(updated) >= 10:
            try:
                if datetime.strptime(updated[:10], '%Y-%m-%d') < fourteen_days_ago:
                    stale_count += 1
            except ValueError:
                pass
    lines.append(f'- [5. Stale Issues - No Update 2+ Weeks](#5-stale-issues) - {stale_count} issues')
    
    # Dependency
    dep_count = sum(1 for issue in issues if issue.get('Dependency', '') and str(issue.get('Dependency', '')).strip())
    lines.append(f'- [6. Dependency Issues](#6-dependency-issues) - {dep_count} issues')
    
    # Duplicated
    dupe_count = sum(1 for issue in issues if issue.get('duplicated_issue', '') and str(issue.get('duplicated_issue', '')).strip())
    lines.append(f'- [7. Duplicated Issues](#7-duplicated-issues) - {dupe_count} issues')
    
    lines.append('- [8. Statistics](#8-statistics)\n')
    
    return '\n'.join(lines)


def generate_report(excel_file: str, output_file: str):
    """Generate the complete issue report."""
    print(f"\nGenerating issue report...")
    print(f"Excel: {excel_file}")
    print(f"Output: {output_file}")
    
    # Load issues
    issues = load_issues(excel_file)
    
    # Generate sections
    print("Generating sections...")
    
    lines = []
    lines.append('# Torch XPU Ops Issue Report\n')
    lines.append(f'**Generated:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append(f'**Total Issues:** {len(issues)}\n')
    lines.append('---\n')
    
    lines.append(generate_index(issues))
    lines.append('\n---\n')
    lines.append(generate_summary_section(issues))
    lines.append('\n---\n')
    lines.append(generate_action_required_section(issues))
    lines.append('\n---\n')
    lines.append(generate_category_section(issues))
    lines.append('\n---\n')
    lines.append(generate_last_week_section(issues))
    lines.append(generate_stale_section(issues))
    lines.append(generate_dependency_section(issues))
    lines.append(generate_duplicated_section(issues))
    lines.append(generate_statistics_section(issues))
    
    # Write to file
    content = '\n'.join(lines)
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f"\nReport saved to: {output_file}")
    print(f"Report size: {len(content)} characters")


def main():
    """Main entry point."""
    excel_file = os.environ.get('ISSUE_EXCEL', os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx'))
    excel_file = sys.argv[1] if len(sys.argv) > 1 else excel_file
    
    output_file = os.environ.get('ISSUE_REPORT', os.path.join(RESULT_DIR, 'issue_report.md'))
    output_file = sys.argv[2] if len(sys.argv) > 2 else output_file
    
    if not os.path.exists(excel_file):
        print(f"ERROR: Excel file not found: {excel_file}")
        return 1
    
    generate_report(excel_file, output_file)
    return 0


if __name__ == '__main__':
    sys.exit(main())