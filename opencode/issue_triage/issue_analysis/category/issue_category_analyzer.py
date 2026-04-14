#!/usr/bin/env python3
"""
Issue Category Analysis Script

Determines issue category using LLM-based analysis with Qwen3-32B model.
Populates Category and Category Reason columns in Issues sheet.

Usage:
    python3 issue_category_analyzer.py <excel_file>
    python3 issue_category_analyzer.py <excel_file> --issues "3246,3243"
    python3 issue_category_analyzer.py <excel_file> --force
"""

import os
import sys
import time
import argparse
import requests
import re
import logging
from openpyxl.styles import Font, PatternFill
import openpyxl

LLM_ENDPOINT = os.environ.get("LLM_ENDPOINT", "http://10.239.15.43/v1/chat/completions")
LLM_MODEL = "Qwen3-32B"
LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")

RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")
LOG_FILE = os.path.join(RESULT_DIR, "pipeline.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)


def determine_category_llm(title, summary, test_cases_info, test_module, labels):
    """Use Qwen3-32B to determine issue category."""
    if not title and not summary:
        return "12 - Others", ""

    tc_info_str = ""
    if test_cases_info:
        for tc in test_cases_info:
            error_val = tc.get('error_msg') or ''
            tc_info_str += f"- Test: {tc.get('test_case', '')}, Error: {str(error_val)[:100]}\n"

    test_module_str = test_module or "Unknown"

    filtered_labels = labels or ""
    if filtered_labels:
        labels_list = [l.strip() for l in str(filtered_labels).split(",") if l.strip()]
        filtered_labels = ", ".join([l for l in labels_list if l.lower() != 'skipped' and l.lower() != 'skipped_windows'])
    labels_str_for_llm = filtered_labels if filtered_labels else "None"

    prompt = f"""You are analyzing PyTorch XPU issue to determine its category.
IMPORTANT: The GitHub labels 'skipped' and 'skipped_windows' should be IGNORED.

Issue Title: {title}
Issue Summary: {summary}
Test Module: {test_module_str}
Labels: {labels_str_for_llm}

Test Cases Info:
{tc_info_str}

Categorize into ONE of these categories:

1. Distributed - distributed, XCCL, NCCL, Gloo, ProcessGroup, DDP, FSDP, torch.distributed
2. TorchAO - torchao, quantize_, int4_weight_only, int8_dynamic_activation, fp8, nf4, autoquant
3. PT2E - torch.export(), Dynamo, fake_tensor, ExportedProgram, AOT, graph break
4. Flash Attention/Transformer - flash_attention, scaled_dot_product_attention, SDPA, attention mask
5. Sparse - sparse tensor, CSR, CSC, COO, torch.sparse, sparse matrix multiplication
6. Inductor/Compilation - torch.compile(), Inductor, Triton, codegen, FX graph
7. Torch Runtime - CUDA runtime, OOM, device kernel launch, stream sync, memory allocation
8. Torch Operations - aten::, native::, custom op, operator dispatch, kernel selection
9. Dtype/Precision - float16, bfloat16, float32, mixed precision, autocast, dtype mismatch
10. Feature Not Supported - unimplemented operator, missing kernel
11. Skip/No Test Exists - missing tests, CI infrastructure problems
12. Others - None of the above

Rules:
- Select exactly ONE category
- "Skip/No Test Exists" is ONLY for MISSING tests, not SKIPPED tests
- For int4/int8/fp8: TorchAO over Dtype/Precision
- For CUDA runtime errors: Torch Runtime
- For operator errors: Torch Operations

Return: "Category Name | detailed_reason"

YOUR ANSWER:"""

    headers = {"Authorization": f"Bearer {LLM_API_KEY}", "Content-Type": "application/json"}
    payload = {"model": LLM_MODEL, "messages": [
        {"role": "system", "content": "Output ONLY 'X - Category Name | brief_reason'. No markdown. No JSON."},
        {"role": "user", "content": prompt}
    ], "temperature": 0.0, "max_tokens": 400}

    start_time = time.time()
    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time

        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            content = re.sub(r'<[^>]*>', '', content)
            content = re.sub(r'\s+', ' ', content).strip()

            if '|' in content:
                parts = content.split('|', 1)
                return parts[0].strip(), parts[1].strip() if len(parts) > 1 else '', elapsed
            else:
                match = re.search(r'(\d+)\s*[-–]\s*[\w\s/]+', content)
                if match:
                    return f"{match.group(1)} - {content[match.start():match.end()].split('-')[-1].strip()}", '', elapsed
                return content.strip()[:50], '', elapsed
        return f"API Error: {response.status_code}", '', elapsed
    except Exception as e:
        return f"Error: {str(e)[:30]}", '', elapsed


def run_category_analysis(excel_file, target_issue_ids=None, force=False):
    """Run category analysis on Excel file."""
    if not os.path.exists(excel_file):
        logger.error(f"File not found: {excel_file}")
        return 0

    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Issues']

    max_col = ws.max_column
    has_category = max_col >= 20

    if not has_category:
        logger.info("Adding Category columns (19=S, 20=T)...")
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=19).value = ''
            ws.cell(row=row, column=20).value = ''
        ws.cell(row=1, column=19, value="Category")
        ws.cell(row=1, column=20, value="Category Reason")

    issues_to_process = []
    already_processed = []

    for row_idx in range(2, ws.max_row + 1):
        issue_id = ws.cell(row=row_idx, column=1).value
        existing_category = ws.cell(row=row_idx, column=19).value
        if issue_id is not None:
            if existing_category and not target_issue_ids and not force:
                already_processed.append(issue_id)
            else:
                issues_to_process.append(row_idx)

    if target_issue_ids:
        issues_to_process = [row for row in issues_to_process if ws.cell(row=row, column=1).value in target_issue_ids]
    elif already_processed and not force:
        logger.info(f"Skipping {len(already_processed)} issues that already have categories")

    total_issues = len(issues_to_process)
    logger.info(f"Total issues to process: {total_issues}")

    ws_cases = wb['Test Cases'] if 'Test Cases' in wb.sheetnames else None
    category_count = 0
    error_count = 0

    for i, row_idx in enumerate(issues_to_process):
        issue_id = ws.cell(row=row_idx, column=1).value
        title = ws.cell(row=row_idx, column=2).value or ''
        title_short = title[:60] + '...' if len(title) > 60 else title
        logger.info(f"[Category] Issue {issue_id}: Starting... ({i+1}/{total_issues})")

        summary = ws.cell(row=row_idx, column=10).value or ''
        test_module = ws.cell(row=row_idx, column=13).value or ''
        labels = ws.cell(row=row_idx, column=6).value or ''

        test_cases_info = []
        if ws_cases:
            for case_row in range(2, ws_cases.max_row + 1):
                case_issue_id = ws_cases.cell(row=case_row, column=1).value
                if case_issue_id == issue_id:
                    test_case = ws_cases.cell(row=case_row, column=7).value or ''
                    error_msg = ws_cases.cell(row=case_row, column=8).value or ''
                    test_cases_info.append({'test_case': test_case, 'error_msg': error_msg})

        category, reason, elapsed = determine_category_llm(title, summary, test_cases_info, test_module, labels)
        ws.cell(row=row_idx, column=19).value = category
        ws.cell(row=row_idx, column=20).value = reason if reason else category

        if category and not category.startswith('API Error') and not category.startswith('Error'):
            logger.info(f"[Category] Issue {issue_id}: OK | {category} | {elapsed:.1f}s")
            category_count += 1
        else:
            logger.error(f"[Category] Issue {issue_id}: ERROR | {category}")
            error_count += 1

        if (i + 1) % 10 == 0:
            wb.save(excel_file)

    wb.save(excel_file)
    logger.info(f"\n===== Summary =====")
    logger.info(f"Processed: {total_issues}")
    logger.info(f"Success: {category_count}, Errors: {error_count}")
    logger.info(f"Output: {excel_file}")

    return category_count


def main():
    parser = argparse.ArgumentParser(description="Determine issue category using LLM")
    parser.add_argument("excel_file", help="Path to Excel file with issues")
    parser.add_argument("--issues", type=str, default="", help="Comma-separated issue IDs")
    parser.add_argument("--force", action="store_true", help="Re-process all")
    args = parser.parse_args()

    target_ids = None
    if args.issues:
        target_ids = set(int(p.strip()) for p in args.issues.split(',') if p.strip())
        logger.info(f"Target issues: {sorted(target_ids)}")

    run_category_analysis(args.excel_file, target_ids, args.force)


if __name__ == "__main__":
    main()