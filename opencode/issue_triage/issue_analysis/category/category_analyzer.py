#!/usr/bin/env python3
"""
Category Analysis Script

Determines issue category using LLM-based analysis with Qwen3-32B model.
Populates Category and Category Reason columns in Issues sheet.
Logs progress and results to result/pipeline.log.

Usage:
    python3 category_analyzer.py <excel_file>
    python3 category_analyzer.py <excel_file> --issues "3246,3243"
"""

import os
import sys
import time
import argparse
import requests
import json
import re
import logging
from datetime import datetime
from openpyxl.styles import Font, PatternFill
import openpyxl

LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
LLM_MODEL = "Qwen3-32B"

RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")

# Setup logging
LOG_FILE = os.path.join(RESULT_DIR, "pipeline.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def determine_category_llm(title, summary, test_cases_info, test_module, labels):
    """
    Use Qwen3-32B via internal API to determine the category of an issue.

    Returns:
        tuple: (category, reason)
    """
    from openpyxl.styles import Font, PatternFill
    import openpyxl

    if not title and not summary:
        return "12 - Others", ''

    tc_info_str = ""
    if test_cases_info:
        for tc in test_cases_info:
            error_val = tc.get('error_msg') or ''
            tc_info_str += f"- Test: {tc.get('test_case', '')}, Error: {str(error_val)[:100]}\n"

    test_module_str = test_module or "Unknown"

    # Filter out 'skipped' and 'skipped_windows' from labels - these should NOT influence category
    filtered_labels = labels or ""
    if filtered_labels:
        # Remove skipped-related labels from influencing classification
        labels_list = [l.strip() for l in str(filtered_labels).split(",") if l.strip()]
        filtered_labels = ", ".join([l for l in labels_list if l.lower() != 'skipped' and l.lower() != 'skipped_windows'])
    labels_str_for_llm = filtered_labels if filtered_labels else "None"

    prompt = f"""You are analyzing PyTorch XPU issue to determine its category.
IMPORTANT: The GitHub labels 'skipped' and 'skipped_windows' should be IGNORED for classification. These only indicate test status, not the actual issue category.

Issue Title: {title}
Issue Summary: {summary}
Test Module: {test_module_str}
Labels (skip-related labels filtered out): {labels_str_for_llm}

Test Cases Info:
{tc_info_str}

Categorize this issue into ONE of:

Categories:

    1. Distributed - Keywords: distributed, XCCL, NCCL, Gloo, ProcessGroup, DDP, FSDP, torch.distributed, collective communication, multi-node, timeout, rendezvous, init_method, reduce_scatter, all_gather

    2. TorchAO - Keywords: torchao, quantize_, int4_weight_only, int8_dynamic_activation, fp8, nf4, autoquant, quantization_config, Adam8bit, AdamW4bit, Lion8bit, PagedAdam, OptimizerWithQuantization, quantized_optimizer, ao/sparsity, apply_dynamic_quant, change_linear_weights_to_int8_packed, QuantizedLinear, from_float (AO context), packed weight, dequantization, int4, int8 (when paired with torchao)

    3. PT2E - Keywords: torch.export(), Dynamo, fake_tensor, ExportedProgram, AOT, torch._export, graph break, tracing, exported_program.run_decompositions

    4. Flash Attention/Transformer - Keywords: flash_attention, scaled_dot_product_attention, SDPA, attention mask, transformer layer, F.scaled_dot_product_attention, memory-efficient attention, FlexAttention

    5. Sparse - Keywords: sparse tensor, CSR, CSC, COO, torch.sparse, sparse matrix multiplication, sparse_mask, to_sparse(), sparse_coo_tensor

    6. Inductor/Compilation - Keywords: torch.compile(), Inductor, Triton, codegen, AOT Autograd, FX graph, torch._inductor, compilation cache, torch._dynamo

    7. Torch Runtime - Keywords: CUDA runtime, cudaMalloc, cudaMemcpy, out of memory (OOM), device kernel launch, stream synchronization, cudaStreamSynchronize, device-side assert, cudaError, illegal memory access, context management, cudaSetDevice, device initialization, cudaGetDevice, driver error, CUDA_VISIBLE_DEVICES, memory leak, allocation failure, device reset

    8. Torch Operations - Keywords: operator implementation, aten::, native::, custom op, register_operator, operator overloading, tensor operation dispatch, kernel selection, op signature mismatch, unsupported op on device, op not implemented for device, device-specific op behavior, backward pass operation, autograd op

    9. Dtype/Precision - Keywords: dtype mismatch, float16, bfloat16, float32, mixed precision, autocast, GradScaler, precision loss, NaN/inf due to dtype, to(dtype=...), torch.int8 (without torchao), legacy torch.quantization

    10. Feature Not Supported - Keywords: unimplemented operator, missing kernel, feature not available in this build, unsupported combination, "not implemented for"

    11. Skip/No Test Exists - ONLY use when the issue is about:
        - A test that SHOULD exist but is MISSING (test gap)
        - Test infrastructure problems (missing test decorator, CI configuration)
        - @unittest.skip decorators explicitly mentioned as problem
        - NOT for tests that happen to be labeled "skipped" due to failures!

    12. Others - None of the above (only if truly uncategorizable)

Classification Rules:
    - Select exactly ONE category
    - A test being labeled "skipped" does NOT make it a "Skip/No Test Exists" issue
    - If tests have "skipped" labels but the issue describes a REAL FAILURE (OOM, dtype mismatch,
      operator not implemented, etc.), categorize based on the ACTUAL ERROR, NOT the skip label
    - "Skip/No Test Exists" is ONLY for MISSING tests, not for SKIPPED tests
    - For int4/int8/fp8 errors: TorchAO takes precedence over Dtype/Precision
    - For quantized optimizers (Adam8bit, Lion8bit): TorchAO takes precedence
    - For CUDA runtime errors (memory, sync, context): Choose Torch Runtime
    - For operator dispatch/kernel errors: Choose Torch Operations
    - For device-specific op implementation issues: Choose Torch Operations
    - Use Others only as a last resort

    Distinction between Torch Runtime and Torch Operations:
        - Torch Runtime = Errors related to device management, memory allocation, synchronization, driver/runtime API calls
        - Torch Operations = Errors related to specific operator execution, kernel selection, op dispatch, custom op registration

Return the category AND a detailed reason for your classification:
- The reason is REQUIRED, not optional
- Make it detailed and specific (full sentences, 150-300 characters)
- Include: specific ops/functions mentioned, dtypes (float16, bf16, int8, Long, etc.), arguments, or patterns that led to this categorization
- Explain clearly WHY you chose this category based on the issue details

Format: "Category Name | detailed_reason"
Example: "Dtype/Precision Issue | The aten.memory_efficient_attention kernel encounters dtype mismatch when processing fp32 input tensors with bfloat16 scaling factor on XPU device. The dot_xpu_mkl operation is not implemented for Long dtype, causing NotImplementedError."

YOUR ANSWER (must include detailed reason after the pipe symbol):"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "Output ONLY 'X - Category Name | brief_reason'. No markdown. No JSON. No thinking tags."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 400
    }

    start_time = time.time()

    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time

        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")

            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            content = content.replace('[', ' <').replace(']', '> ')
            content = re.sub(r'<[^>]*>', '', content)
            content = re.sub(r'has ATTR\b', '', content)
            content = re.sub(r'\s+', ' ', content).strip()

            if '|' in content:
                parts = content.split('|', 1)
                category_part = parts[0].strip()
                reason = parts[1].strip() if len(parts) > 1 else ''
                return category_part, reason, elapsed
            else:
                match = re.search(r'(\d+)\s*[-–]\s*[\w\s/]+', content)
                if match:
                    category = f"{match.group(1)} - " + content[match.start():match.end()].split('-')[-1].strip()
                    return category, '', elapsed
                else:
                    return content.strip()[:50], '', elapsed

        return f"API Error: {response.status_code}", '', elapsed

    except Exception as e:
        return f"Error: {str(e)[:30]}", '', elapsed


def main():
    parser = argparse.ArgumentParser(description="Determine issue category using LLM")
    parser.add_argument("excel_file", help="Path to Excel file with issues")
    parser.add_argument("--issues", type=str, default="", help="Comma-separated list of issue IDs to process")
    parser.add_argument("--force", action="store_true", help="Re-process all issues, overwriting existing categories")
    args = parser.parse_args()

    excel_file = args.excel_file

    if not os.path.exists(excel_file):
        logger.error(f"File not found: {excel_file}")
        sys.exit(1)

    # Parse target issue IDs
    target_issue_ids = None
    if args.issues:
        target_issue_ids = set()
        for part in args.issues.split(','):
            part = part.strip()
            if part:
                try:
                    target_issue_ids.add(int(part))
                except ValueError:
                    pass
        logger.info(f"Target issues: {sorted(target_issue_ids)}")

    logger.info(f"Loading Excel: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Issues']

    # Check if Category columns exist
    max_col = ws.max_column
    has_category = max_col >= 20

    if not has_category:
        logger.info("Adding Category columns (S=19, T=20)...")
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=19).value = ''  # Category (S) - col 19
            ws.cell(row=row, column=20).value = ''  # Category Reason (T) - col 20
        # Add headers
        cat_cell = ws.cell(row=1, column=19, value="Category")
        cat_cell.font = Font(bold=True, color="FFFFFF")
        cat_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        reason_cell = ws.cell(row=1, column=20, value="Category Reason")
        reason_cell.font = Font(bold=True, color="FFFFFF")
        reason_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Get all issues to process (skip if already has category unless --issues specified)
    issues_to_process = []
    already_processed = []
    skipped_count = 0

    force_mode = getattr(args, 'force', False)

    for row_idx in range(2, ws.max_row + 1):
        issue_id = ws.cell(row=row_idx, column=1).value
        existing_category = ws.cell(row=row_idx, column=19).value
        if issue_id is not None:
            if existing_category and not target_issue_ids and not force_mode:
                # Skip if already has category (unless specific issues requested or --force)
                already_processed.append(issue_id)
            else:
                issues_to_process.append(row_idx)

    if target_issue_ids:
        issues_to_process = [row for row in issues_to_process if ws.cell(row=row, column=1).value in target_issue_ids]
    elif already_processed and not force_mode:
        skipped_count = len(already_processed)
        logger.info(f"Skipping {skipped_count} issues that already have categories")

    total_issues = len(issues_to_process)
    logger.info(f"Total issues to process: {total_issues}")

    category_count = 0
    error_count = 0

    for i, row_idx in enumerate(issues_to_process):
        issue_id = ws.cell(row=row_idx, column=1).value
        title = ws.cell(row=row_idx, column=2).value or ''

        # Truncate title for logging
        title_short = title[:60] + '...' if len(title) > 60 else title

        logger.info(f"[Category] Issue {issue_id}: Starting LLM analysis... ({i+1}/{total_issues})")

        # Get issue details
        summary = ws.cell(row=row_idx, column=10).value or ''  # J
        test_module = ws.cell(row=row_idx, column=13).value or ''  # M
        labels = ws.cell(row=row_idx, column=6).value or ''  # F

        # Get test case info from Test Cases sheet
        wb = ws.parent
        ws_cases = wb['Test Cases'] if 'Test Cases' in wb.sheetnames else None

        test_cases_info = []
        if ws_cases:
            for case_row in range(2, ws_cases.max_row + 1):
                case_issue_id = ws_cases.cell(row=case_row, column=1).value
                if case_issue_id == issue_id:
                    test_case = ws_cases.cell(row=case_row, column=7).value or ''  # G
                    error_msg = ws_cases.cell(row=case_row, column=8).value or ''  # H
                    test_cases_info.append({
                        'test_case': test_case,
                        'error_msg': error_msg
                    })

        category, reason, elapsed = determine_category_llm(
            title=title,
            summary=summary,
            test_cases_info=test_cases_info,
            test_module=test_module,
            labels=labels
        )

        # Save to Excel
        ws.cell(row=row_idx, column=19).value = category  # S (19) - Category
        ws.cell(row=row_idx, column=20).value = reason if reason else category  # T (20) - Category Reason

        # Log result
        if category and not category.startswith('API Error') and not category.startswith('Error'):
            logger.info(f"[Category] Issue {issue_id}: SUCCESS | Category: {category} | Time: {elapsed:.1f}s")
            category_count += 1
            if reason:
                # Log reason (truncated if too long)
                reason_short = reason[:150] + '...' if len(reason) > 150 else reason
                logger.info(f"[Category] Issue {issue_id}: Reason: {reason_short}")
        else:
            logger.error(f"[Category] Issue {issue_id}: ERROR | {category} | Time: {elapsed:.1f}s")
            error_count += 1

        # Progress log every 10 issues
        if (i + 1) % 10 == 0 or (i + 1) == total_issues:
            progress = (i + 1) / total_issues * 100
            logger.info(f"[Category] Progress: {i+1}/{total_issues} ({progress:.1f}%)")

        # Save every 10 issues
        if (i + 1) % 10 == 0:
            wb.save(excel_file)
            logger.info(f"[Category] Saved progress: {i+1} issues")

    # Final save
    wb.save(excel_file)

    logger.info(f"\n===== Category Analysis Summary =====")
    logger.info(f"Total issues: {total_issues}")
    logger.info(f"Category assigned: {category_count}")
    logger.info(f"Errors: {error_count}")
    logger.info(f"Output file: {excel_file}")
    logger.info(f"Category column: S (19)")
    logger.info(f"Category Reason column: T (20)")
    logger.info(f"Log file: {LOG_FILE}")


if __name__ == "__main__":
    main()