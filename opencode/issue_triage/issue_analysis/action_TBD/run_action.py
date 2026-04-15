#!/usr/bin/env python3
"""
Action TBD Analysis Runner

Adds Action TBD, Owner Transfer, and Action Reason columns to Issues sheet using ActionAnalyzer.

Usage:
    python3 run_action.py [--excel EXCEL_FILE] [--limit N] [--force]

Example:
    python3 run_action.py
    python3 run_action.py --limit 10
    python3 run_action.py --force
"""

import os
import sys
import argparse
import time

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
RESULT_DIR = os.environ.get('RESULT_DIR', os.path.join(ROOT_DIR, 'result'))

sys.path.insert(0, ROOT_DIR)

from issue_analysis.action_TBD.action_analyzer import analyze_action_all


def get_column_by_header(ws, header_name):
    """Find column index by header name (case-insensitive)."""
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header and str(header).lower() == header_name.lower():
            return col
    return None


def get_issues_columns(ws):
    """Get column indices for Issues sheet by header names."""
    return {
        'issue_id': get_column_by_header(ws, 'Issue ID'),
        'title': get_column_by_header(ws, 'Title'),
        'status': get_column_by_header(ws, 'Status'),
        'assignee': get_column_by_header(ws, 'Assignee'),
        'reporter': get_column_by_header(ws, 'Reporter'),
        'labels': get_column_by_header(ws, 'Labels'),
        'created_time': get_column_by_header(ws, 'Created Time'),
        'updated_time': get_column_by_header(ws, 'Updated Time'),
        'milestone': get_column_by_header(ws, 'Milestone'),
        'summary': get_column_by_header(ws, 'Summary'),
        'type': get_column_by_header(ws, 'Type'),
        'module': get_column_by_header(ws, 'Module'),
        'test_module': get_column_by_header(ws, 'Test Module'),
        'dependency': get_column_by_header(ws, 'Dependency'),
        'pr': get_column_by_header(ws, 'PR'),
        'pr_owner': get_column_by_header(ws, 'PR Owner'),
        'pr_status': get_column_by_header(ws, 'PR Status'),
        'pr_description': get_column_by_header(ws, 'PR Description'),
        'category': get_column_by_header(ws, 'Category'),
        'category_reason': get_column_by_header(ws, 'Category Reason'),
        'priority': get_column_by_header(ws, 'Priority'),
        'priority_reason': get_column_by_header(ws, 'Priority Reason'),
        'root_cause': get_column_by_header(ws, 'Root Cause'),
        'root_cause_reason': get_column_by_header(ws, 'Root Cause Reason'),
    }


def get_test_columns(ws):
    """Get column indices for Test Cases sheet by header names."""
    return {
        'issue_id': get_column_by_header(ws, 'Issue ID'),
        'test_reproducer': get_column_by_header(ws, 'Test Reproducer'),
        'test_type': get_column_by_header(ws, 'Test Type'),
        'test_file': get_column_by_header(ws, 'Test File'),
        'origin_test_file': get_column_by_header(ws, 'Origin Test File'),
        'test_class': get_column_by_header(ws, 'Test Class'),
        'test_case': get_column_by_header(ws, 'Test Case'),
        'error_msg': get_column_by_header(ws, 'Error Message'),
        'traceback': get_column_by_header(ws, 'Traceback'),
        'torch_ops': get_column_by_header(ws, 'torch-ops'),
        'xpu_status': get_column_by_header(ws, 'XPU Status'),
        'stock_status': get_column_by_header(ws, 'Stock Status'),
        'no_match_reason': get_column_by_header(ws, 'No Match Reason'),
        'can_enable_on_xpu': get_column_by_header(ws, 'can_enable_on_xpu'),
        'cuda_case_exist': get_column_by_header(ws, 'CUDA Case Exist'),
        'xpu_case_exist': get_column_by_header(ws, 'XPU Case Exist'),
    }


def add_action_columns(excel_file, limit=None, force=False):
    """
    Add Action TBD columns to Issues sheet using ActionAnalyzer.

    Args:
        excel_file: Path to Excel file
        limit: Limit number of issues to process (for testing)
        force: Overwrite existing action values
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    ws_issues = wb['Issues']
    ws_test = wb['Test Cases']

    total_issues = ws_issues.max_row - 1
    print(f"Total issues: {total_issues}")

    # Get column indices by header names
    issue_cols = get_issues_columns(ws_issues)
    test_cols = get_test_columns(ws_test)

    # Verify required columns exist (pr_status is optional)
    required_issue_cols = ['issue_id', 'title', 'summary', 'labels', 'reporter', 'assignee', 'test_module']
    missing = [c for c in required_issue_cols if not issue_cols.get(c)]
    if missing:
        print(f"  Error: Missing required columns in Issues sheet: {missing}")
        return
    if not issue_cols.get('pr_status'):
        print(f"  Note: Missing 'PR Status' column - PR-based actions will be skipped")

    # Find existing Action columns or use first blank
    action_tbd_col = get_column_by_header(ws_issues, 'Action TBD')
    owner_transfer_col = get_column_by_header(ws_issues, 'Owner Transfer')
    action_reason_col = get_column_by_header(ws_issues, 'Action Reason')

    # If not found, find first blank for new columns
    if action_tbd_col is None:
        first_blank = None
        for col in range(1, 51):
            if ws_issues.cell(1, col).value is None:
                first_blank = col
                break

        if first_blank is None:
            print("  Error: No blank column found!")
            return

        action_tbd_col = first_blank
        owner_transfer_col = first_blank + 1
        action_reason_col = first_blank + 2

        # Add headers
        headers = ["Action TBD", "Owner Transfer", "Action Reason"]
        colors = ["FF6B35", "4472C4", "70AD47"]  # Orange, Blue, Green
        for idx, header in enumerate(headers):
            cell = ws_issues.cell(row=1, column=action_tbd_col + idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")

        print(f"  Added columns: {action_tbd_col}=Action TBD, {owner_transfer_col}=Owner Transfer, {action_reason_col}=Action Reason")
    else:
        print(f"  Using existing columns: {action_tbd_col}=Action TBD, {owner_transfer_col}=Owner Transfer, {action_reason_col}=Action Reason")

    # Build issue row map
    issue_row_map = {}
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, issue_cols['issue_id']).value
        if issue_id:
            issue_row_map[issue_id] = row

    # Build test case info map
    test_case_map = {}
    for tr in range(2, ws_test.max_row + 1):
        issue_id = ws_test.cell(tr, test_cols['issue_id']).value
        if issue_id:
            if issue_id not in test_case_map:
                test_case_map[issue_id] = []
            test_case_map[issue_id].append({
                'test_file': ws_test.cell(tr, test_cols['test_file']).value if test_cols['test_file'] else None,
                'test_class': ws_test.cell(tr, test_cols['test_class']).value if test_cols['test_class'] else None,
                'test_case': ws_test.cell(tr, test_cols['test_case']).value if test_cols['test_case'] else None,
                'error_msg': ws_test.cell(tr, test_cols['error_msg']).value if test_cols['error_msg'] else None,
                'traceback': ws_test.cell(tr, test_cols['traceback']).value if test_cols['traceback'] else None,
                'xpu_status': ws_test.cell(tr, test_cols['xpu_status']).value if test_cols['xpu_status'] else None,
                'stock_status': ws_test.cell(tr, test_cols['stock_status']).value if test_cols['stock_status'] else None,
                'can_enable_on_xpu': ws_test.cell(tr, test_cols.get('can_enable_on_xpu', 0)).value if test_cols.get('can_enable_on_xpu') else None,
                'cuda_case_exist': ws_test.cell(tr, test_cols.get('cuda_case_exist', 0)).value if test_cols.get('cuda_case_exist') else None,
                'xpu_case_exist': ws_test.cell(tr, test_cols.get('xpu_case_exist', 0)).value if test_cols.get('xpu_case_exist') else None,
            })

    # Process issues
    processed = 0
    skipped = 0
    errors = 0
    start_time = time.time()

    # Count action distribution
    action_counts = {}

    for i, (issue_id, row_idx) in enumerate(issue_row_map.items()):
        if limit and i >= limit:
            print(f"  Reached limit of {limit} issues")
            break

        # Skip if already has action (unless force=True)
        existing_action = ws_issues.cell(row_idx, action_tbd_col).value
        if existing_action and not force:
            skipped += 1
            continue

        # Get issue data from Issues sheet using column indices
        title = ws_issues.cell(row_idx, issue_cols['title']).value or ''
        summary = ws_issues.cell(row_idx, issue_cols['summary']).value or ''
        labels = ws_issues.cell(row_idx, issue_cols['labels']).value or ''
        reporter = ws_issues.cell(row_idx, issue_cols['reporter']).value or ''
        assignee = ws_issues.cell(row_idx, issue_cols['assignee']).value or ''
        test_module = ws_issues.cell(row_idx, issue_cols['test_module']).value or ''
        pr_status = ws_issues.cell(row_idx, issue_cols['pr_status']).value if issue_cols.get('pr_status') else ''

        # Get test case info
        tc_list = test_case_map.get(issue_id, [])
        xpu_statuses = set()
        stock_statuses = set()
        e2e_statuses = set()
        can_enable_list = []
        
        # For CUDA Enable Test logic: check first test case with blank status
        has_cuda_enabled_error = False
        cuda_case_exists = False
        xpu_case_exists = False
        error_msg = None

        for tc in tc_list:
            xpu_status = tc.get('xpu_status', '')
            stock_status = tc.get('stock_status', '')
            if xpu_status:
                xpu_statuses.add(str(xpu_status).lower())
            if stock_status:
                stock_statuses.add(str(stock_status).lower())
            can_enable_val = tc.get('can_enable_on_xpu')
            if can_enable_val is not None:
                can_enable_list.append(can_enable_val)
            
            # Check for CUDA enabled error and CUDA/XPU case existence
            # Only use values from rows where BOTH XPU and Stock status are blank
            tc_xpu = tc.get('xpu_status', '')
            tc_stock = tc.get('stock_status', '')
            if (tc_xpu is None or tc_xpu == '') and (tc_stock is None or tc_stock == ''):
                # Both blank - this row qualifies for CUDA->XPU check
                tc_error_msg = tc.get('error_msg', '')
                if tc_error_msg and 'cuda enabled' in str(tc_error_msg).lower():
                    has_cuda_enabled_error = True
                    error_msg = tc_error_msg
                
                tc_cuda = tc.get('cuda_case_exist', '')
                if tc_cuda == 'True':
                    cuda_case_exists = True
                
                tc_xpu_exist = tc.get('xpu_case_exist', '')
                if tc_xpu_exist == 'True':
                    xpu_case_exists = True

        # Format issue_can_enable for analyze_action_all
        issue_can_enable = {
            issue_id: {
                'can_enable_list': can_enable_list,
                'comments_list': []
            }
        }

        # Run action analysis with new CUDA Enable Test parameters
        owner_transfer, action_tbd, action_reason = analyze_action_all(
            issue_id=str(issue_id),
            labels=labels,
            title_raw=title,
            summary_raw=summary,
            reporter=reporter,
            assignee=assignee,
            test_module=test_module,
            xpu_statuses=xpu_statuses,
            stock_statuses=stock_statuses,
            e2e_statuses=e2e_statuses,
            issue_can_enable=issue_can_enable,
            issue_duplicated_map={},
            pr_status=pr_status,
            test_cases_info=tc_list,
            llm_info_action=None,
            version_info=None,
            error_msg=error_msg,
            test_case_cuda_exists=cuda_case_exists,
            test_case_xpu_exists=xpu_case_exists,
            has_cuda_enabled_error=has_cuda_enabled_error
        )

        if not action_tbd:
            errors += 1
            continue

        # Count for stats
        if action_tbd:
            action_counts[action_tbd] = action_counts.get(action_tbd, 0) + 1

        # Write to Excel
        ws_issues.cell(row_idx, action_tbd_col).value = action_tbd[:100] if action_tbd else ''
        ws_issues.cell(row_idx, owner_transfer_col).value = owner_transfer[:50] if owner_transfer else ''
        ws_issues.cell(row_idx, action_reason_col).value = action_reason[:500] if action_reason else ''
        processed += 1

        if (processed + skipped) % 20 == 0:
            elapsed = time.time() - start_time
            print(f"  Processed {processed + skipped}/{len(issue_row_map)} issues ({elapsed:.1f}s)")

    elapsed_total = time.time() - start_time

    print(f"\nAction TBD Distribution:")
    for action, count in sorted(action_counts.items(), key=lambda x: -x[1]):
        print(f"  {action}: {count}")

    stats = f"\nComplete: {processed} processed, {skipped} skipped (had existing), {errors} no action"
    print(stats)
    print(f"  Total time: {elapsed_total:.1f}s ({elapsed_total/60:.1f} min)")

    wb.save(excel_file)
    print(f"Saved: {excel_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Add Action TBD columns to Issues sheet using ActionAnalyzer',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '--excel', '-e', type=str, default=None,
        help=f'Path to Excel file (default: {RESULT_DIR}/torch_xpu_ops_issues.xlsx)'
    )
    parser.add_argument(
        '--limit', '-l', type=int, default=None,
        help='Limit number of issues to process (for testing)'
    )
    parser.add_argument(
        '--force', '-f', action='store_true',
        help='Overwrite existing action values'
    )

    args = parser.parse_args()
    excel_file = args.excel or os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx')

    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        return 1

    print("\n" + "=" * 60)
    print("Action TBD Analysis (using ActionAnalyzer)")
    print("=" * 60)

    add_action_columns(excel_file, limit=args.limit, force=args.force)
    return 0


if __name__ == '__main__':
    sys.exit(main())