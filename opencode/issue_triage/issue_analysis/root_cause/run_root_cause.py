#!/usr/bin/env python3
"""
Root Cause Analysis Runner (LLM-based)

Adds Root Cause and Root Cause Reason columns to Issues sheet using LLM-based analysis with Qwen3-32B.

Usage:
    python3 run_root_cause.py [--excel EXCEL_FILE] [--limit N] [--force]

Example:
    python3 run_root_cause.py
    python3 run_root_cause.py --limit 10
    python3 run_root_cause.py --force
"""

import os
import sys
import argparse
import time

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
RESULT_DIR = os.environ.get('RESULT_DIR', os.path.join(ROOT_DIR, 'result'))

sys.path.insert(0, ROOT_DIR)

from issue_analysis.root_cause.root_cause_analyzer import (
    analyze_root_cause_llm,
    analyze_root_cause_keyword
)


def add_root_cause_column(excel_file, limit=None, force=False, max_llm_calls=None):
    """
    Add Root Cause and Root Cause Reason columns to Issues sheet using LLM analysis.
    
    Args:
        excel_file: Path to Excel file
        limit: Limit number of issues to process (for testing)
        force: Overwrite existing root cause values
        max_llm_calls: Limit number of LLM calls (None = unlimited but rate-limited by API)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    ws_issues = wb['Issues']
    ws_test = wb['Test Cases']

    total_issues = ws_issues.max_row - 1
    print(f"Total issues: {total_issues}")

    # Find existing Root Cause columns or use first blank
    root_col = None
    reason_col = None
    for col in range(1, ws_issues.max_column + 1):
        header = ws_issues.cell(1, col).value
        if header == "Root Cause":
            root_col = col
        elif header == "Root Cause Reason":
            reason_col = col

    # If not found, find first blank for new columns
    if root_col is None:
        first_blank = None
        for col in range(1, 51):
            if ws_issues.cell(1, col).value is None:
                first_blank = col
                break
        if first_blank is None:
            print("  Error: No blank column found!")
            return
        root_col = first_blank
        reason_col = first_blank + 1
        # Add headers
        root_headers = ["Root Cause", "Root Cause Reason"]
        for idx, header in enumerate(root_headers):
            cell = ws_issues.cell(row=1, column=root_col + idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        reason_col = root_col + 1
        print(f"  Added columns: {root_col}=Root Cause, {reason_col}=Root Cause Reason")
    else:
        print(f"  Using existing columns: {root_col}=Root Cause, {reason_col}=Root Cause Reason")

    # Build issue row map
    issue_row_map = {}
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        if issue_id:
            issue_row_map[issue_id] = row

    # Preprocess: build test case info map
    test_case_map = {}
    for tr in range(2, ws_test.max_row + 1):
        issue_id = ws_test.cell(tr, 1).value
        if issue_id:
            if issue_id not in test_case_map:
                test_case_map[issue_id] = []
            test_case_map[issue_id].append({
                'test_file': ws_test.cell(tr, 3).value,
                'test_class': ws_test.cell(tr, 6).value,
                'test_case': ws_test.cell(tr, 7).value,
                'error_msg': ws_test.cell(tr, 8).value,
                'traceback': ws_test.cell(tr, 9).value
            })

    # Process issues
    processed = 0
    skipped = 0
    skipped_llm = 0
    llm_errors = 0
    start_time = time.time()
    llm_call_count = 0

    # Count categories
    category_counts = {}

    for i, (issue_id, row_idx) in enumerate(issue_row_map.items()):
        if limit and i >= limit:
            print(f"  Reached limit of {limit} issues")
            break

        title = ws_issues.cell(row_idx, 2).value or ''
        summary = ws_issues.cell(row_idx, 10).value or ''

        # Skip if already has root cause (unless force=True)
        existing = ws_issues.cell(row_idx, root_col).value
        if existing and not force:
            skipped += 1
            continue

        # Skip if max LLM calls reached
        if max_llm_calls is not None and llm_call_count >= max_llm_calls:
            skipped_llm += 1
            continue

        # Get test case info for this issue
        tc_info = test_case_map.get(issue_id, [])
        if tc_info:
            tc = tc_info[0]
            test_file = tc.get('test_file', '')
            test_class = tc.get('test_class', '')
            test_case = tc.get('test_case', '')
            error_msg = tc.get('error_msg', '')
            traceback = tc.get('traceback', '')
        else:
            test_file = test_class = test_case = error_msg = traceback = ''

        # Check if we have valid content to analyze
        has_content = title or summary or error_msg or traceback

        if has_content:
            # Use LLM analysis
            root_cause = analyze_root_cause_llm(
                issue_id=str(issue_id),
                issue_title=title,
                issue_summary=summary,
                test_file=test_file or '',
                test_class=test_class or '',
                test_case=test_case or '',
                error_msg=error_msg or '',
                traceback=traceback or ''
            )
            llm_call_count += 1
        else:
            # Fallback to keyword-based if no content
            root_cause = analyze_root_cause_keyword(
                title, summary, test_file, test_class, test_case, error_msg, traceback
            )

        if not root_cause or root_cause.startswith('API Error') or root_cause.startswith('Exception'):
            llm_errors += 1
            # Fallback to keyword analysis
            root_cause = analyze_root_cause_keyword(
                title, summary, test_file, test_class, test_case, error_msg, traceback
            )

        # Parse category and reason
        if ' - ' in root_cause:
            parts = root_cause.split(' - ', 1)
            category = parts[0]
            reason = parts[1] if len(parts) > 1 else root_cause
        else:
            category = root_cause
            reason = root_cause

        # Count for stats
        category_counts[category] = category_counts.get(category, 0) + 1

        ws_issues.cell(row_idx, root_col).value = category[:100] if category else 'Others'
        ws_issues.cell(row_idx, reason_col).value = reason[:500] if reason else ''
        processed += 1

        if (processed + skipped + skipped_llm) % 20 == 0:
            elapsed = time.time() - start_time
            print(f"  Processed {processed + skipped + skipped_llm}/{len(issue_row_map)} issues ({elapsed:.1f}s, LLM calls: {llm_call_count})")

    elapsed_total = time.time() - start_time
    print(f"\nRoot Cause Category Distribution:")
    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")

    stats = f"\nComplete: {processed} processed, {skipped} skipped (had existing), {skipped_llm} skipped (max LLM), {llm_errors} LLM errors"
    print(stats)
    print(f"  Total LLM calls: {llm_call_count}")
    print(f"  Total time: {elapsed_total:.1f}s ({elapsed_total/60:.1f} min)")

    wb.save(excel_file)
    print(f"Saved: {excel_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Add Root Cause columns to Issues sheet using LLM analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '--excel', '-e', type=str, default=None,
        help=f'Path to Excel file (default: {RESULT_DIR}/torch_xpu_ops_issues.xlsx)'
    )
    parser.add_argument(
        '--limit', '-l', type=int, default=None,
        help='Limit number of issues to process (for testing)'
    )
    parser.add_argument(
        '--max-llm', '-m', type=int, default=None,
        help='Maximum number of LLM calls (default: unlimited)'
    )
    parser.add_argument(
        '--force', '-f', action='store_true',
        help='Overwrite existing root cause values'
    )

    args = parser.parse_args()
    excel_file = args.excel or os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx')

    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        return 1

    print("\n" + "=" * 60)
    print("Root Cause Analysis (LLM-based with Qwen3-32B)")
    print("=" * 60)

    add_root_cause_column(excel_file, limit=args.limit, force=args.force, max_llm_calls=args.max_llm)
    return 0


if __name__ == '__main__':
    sys.exit(main())