#!/usr/bin/env python3
"""
Priority Analysis Runner

Adds Priority and Priority Reason columns to Issues sheet using rule-based priority determination.

Usage:
    python3 run_priority.py [--excel EXCEL_FILE] [--limit N] [--force]

Example:
    python3 run_priority.py
    python3 run_priority.py --limit 10
    python3 run_priority.py --force
"""

import os
import sys
import argparse
import time
import requests
import re
from functools import lru_cache

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
RESULT_DIR = os.environ.get('RESULT_DIR', os.path.join(ROOT_DIR, 'result'))

sys.path.insert(0, ROOT_DIR)

GITHUB_BASE = "https://raw.githubusercontent.com/intel/torch-xpu-ops/main/.ci/benchmarks"

BENCHMARK_MODELS = {
    'huggingface': set(),
    'timm': set(),
    'torchbench': set()
}

def load_benchmark_models():
    """Load benchmark models from GitHub."""
    sources = {
        'huggingface': f'{GITHUB_BASE}/huggingface_models_list.txt',
        'timm': f'{GITHUB_BASE}/timm_models_list.txt',
        'torchbench': f'{GITHUB_BASE}/torchbench_models_list.txt'
    }
    
    for name, url in sources.items():
        try:
            resp = requests.get(url, timeout=30)
            if resp.status_code == 200:
                for line in resp.text.strip().split('\n'):
                    line = line.strip()
                    if line and not line.startswith('#'):
                        model = line.split(',')[0].strip()
                        BENCHMARK_MODELS[name].add(model.lower())
                print(f"  Loaded {len(BENCHMARK_MODELS[name])} {name} models")
        except Exception as e:
            print(f"  Warning: Failed to load {name} models: {e}")
    
    print(f"  Total benchmark models: {sum(len(v) for v in BENCHMARK_MODELS.values())}")


def is_benchmark_model_issue(title_raw, summary_raw):
    """
    Check if issue is about a known benchmark model (huggingface, timm, torchbench).
    Returns True if this is a benchmark model issue.
    """
    text_lower = (str(title_raw) + ' ' + str(summary_raw)).lower()
    
    for name, models in BENCHMARK_MODELS.items():
        for model in models:
            if model and len(model) > 2 and model in text_lower:
                return True
    return False


def is_custom_model_issue(title_raw, summary_raw):
    """
    Check if issue is about a custom model (not a known benchmark).
    Returns True only for custom/unknown models.
    """
    is_model = (
        'model' in str(title_raw).lower() or 'model' in str(summary_raw).lower() or
        'application' in str(title_raw).lower() or 'application' in str(summary_raw).lower()
    )
    
    if not is_model:
        return False
    
    return not is_benchmark_model_issue(title_raw, summary_raw)


def determine_priority_rules(title_raw, summary_raw, test_module, labels, ws_test, issue_id, MAX_LLM_PRIORITY, llm_priority_count):
    """
    Rule-based priority determination with heuristics.
    
    Priority levels:
    - P0: Build crash, regression, custom model impact (NOT benchmark models)
    - P1: E2E accuracy/functionality, many UT failures (>20)
    - P2: E2E performance, few UT failures, benchmark model issues
    - P3: Minor issues
    """
    priority = 'P3'
    priority_reason = ''

    # Custom model check (not benchmark models like huggingface, timm, torchbench)
    is_custom_model = is_custom_model_issue(title_raw, summary_raw)
    
    is_e2e = str(test_module).lower() == 'e2e'
    is_ut = str(test_module).lower() == 'ut'
    is_build = 'build' in str(test_module).lower()

    is_regression = (
        'regression' in str(labels).lower() or 'regression' in str(title_raw).lower() or
        'was pass' in str(summary_raw).lower() or 'previously pass' in str(summary_raw).lower()
    )

    is_build_crash = (
        is_build or 'build' in str(title_raw).lower() or
        'crash' in str(title_raw).lower() or 'segmentation' in str(title_raw).lower() or
        'segfault' in str(title_raw).lower() or 'signal' in str(summary_raw).lower()
    )

    failed_count = 0
    for tr in range(2, ws_test.max_row + 1):
        if ws_test.cell(tr, 1).value == issue_id:
            tc_status = ws_test.cell(tr, 11).value
            if str(tc_status).lower() in ['failed', 'error']:
                failed_count += 1

    if is_build_crash:
        priority = 'P0'
        priority_reason = 'Build crash - critical blocking issue'
    elif is_custom_model and not ('test' in str(title_raw).lower() and 'case' in str(title_raw).lower()):
        priority = 'P0'
        priority_reason = 'Impacts customer custom model/application'
    elif is_regression:
        priority = 'P0'
        priority_reason = 'Regression - passed before but failed now'
    elif is_e2e and is_bench_model_issue(title_raw, summary_raw):
        if 'accuracy' in str(title_raw).lower() or 'accuracy' in str(summary_raw).lower():
            priority = 'P1'
            priority_reason = 'E2E benchmark accuracy issue'
        elif 'performance' in str(title_raw).lower() or 'slow' in str(title_raw).lower():
            priority = 'P2'
            priority_reason = 'E2E benchmark performance issue'
        else:
            priority = 'P2'
            priority_reason = 'E2E benchmark model issue'
    elif is_e2e and is_custom_model:
        priority = 'P1'
        priority_reason = 'E2E custom model accuracy/functionality issue'
    elif is_e2e and ('accuracy' in str(title_raw).lower() or 'accuracy' in str(summary_raw).lower() or
                    'fail' in str(title_raw).lower() or 'fail' in str(summary_raw).lower()):
        priority = 'P1'
        priority_reason = 'E2E accuracy/functionality issue'
    elif is_e2e and ('performance' in str(title_raw).lower() or 'slow' in str(title_raw).lower() or
                    'latency' in str(title_raw).lower()):
        priority = 'P2'
        priority_reason = 'E2E performance issue'
    elif is_ut and failed_count > 20:
        priority = 'P1'
        priority_reason = f'UT with {failed_count} failed test cases'
    else:
        priority = 'P2'
        priority_reason = 'UT issue with few failures'

    return priority, priority_reason, llm_priority_count, failed_count

def is_bench_model_issue(title_raw, summary_raw):
    """Check if issue is about a benchmark model (huggingface, timm, torchbench)."""
    return is_benchmark_model_issue(title_raw, summary_raw)


def add_priority_column(excel_file, limit=None, force=False):
    """Add Priority and Priority Reason columns to Issues sheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    print(f"Loading: {excel_file}")
    print("Loading benchmark models...")
    load_benchmark_models()

    wb = openpyxl.load_workbook(excel_file)
    ws_issues = wb['Issues']
    ws_test = wb['Test Cases']

    total_issues = ws_issues.max_row - 1
    print(f"Total issues: {total_issues}")

    # Find first blank column for Priority headers
    first_blank = None
    for col in range(1, 51):
        if ws_issues.cell(1, col).value is None:
            first_blank = col
            break

    if first_blank is None:
        print("  Error: No blank column found!")
        return

    # Add Priority headers at first blank column
    priority_headers = ["Priority", "Priority Reason"]
    for idx, header in enumerate(priority_headers):
        cell = ws_issues.cell(row=1, column=first_blank + idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    print(f"  Added columns: {first_blank}=Priority, {first_blank+1}=Priority Reason")

    # Build issue row map
    issue_row_map = {}
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        if issue_id:
            issue_row_map[issue_id] = row

    # Process issues
    processed = 0
    skipped = 0
    updated = 0
    start_time = time.time()

    for i, (issue_id, row_idx) in enumerate(issue_row_map.items()):
        if limit and i >= limit:
            print(f"  Reached limit of {limit} issues")
            break

        title = ws_issues.cell(row_idx, 2).value or ''
        summary = ws_issues.cell(row_idx, 10).value or ''
        test_module = ws_issues.cell(row_idx, 13).value or 'ut'
        labels = ws_issues.cell(row_idx, 6).value or ''

        # Skip if already has priority (unless force=True)
        existing = ws_issues.cell(row_idx, first_blank).value
        if existing and not force:
            skipped += 1
            continue

        priority, reason, _, failed_count = determine_priority_rules(
            title, summary, test_module, labels, ws_test, issue_id, 999, 0
        )

        ws_issues.cell(row_idx, first_blank).value = priority
        ws_issues.cell(row_idx, first_blank + 1).value = reason
        processed += 1
        updated += 1

        if (processed + skipped) % 50 == 0:
            elapsed = time.time() - start_time
            print(f"  Processed {processed + skipped}/{len(issue_row_map)} issues ({elapsed:.1f}s)")

    elapsed_total = time.time() - start_time
    print(f"\nComplete: {updated} issues updated, {skipped} skipped, {elapsed_total:.1f}s elapsed")

    wb.save(excel_file)
    print(f"Saved: {excel_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Add Priority and Priority Reason columns to Issues sheet',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '--excel', '-e', type=str, default=None,
        help=f'Path to Excel file (default: {RESULT_DIR}/torch_xpu_ops_issues.xlsx)'
    )
    parser.add_argument(
        '--limit', '-l', type=int, default=None,
        help='Limit number of issues to process'
    )
    parser.add_argument(
        '--force', '-f', action='store_true',
        help='Overwrite existing priority values'
    )

    args = parser.parse_args()
    excel_file = args.excel or os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx')

    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        return 1

    add_priority_column(excel_file, limit=args.limit, force=args.force)
    return 0


if __name__ == '__main__':
    sys.exit(main())