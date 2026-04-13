#!/usr/bin/env python3
"""
Standalone script to detect duplicated issues and add duplicated_issue column to Issues sheet.

Usage:
    python3 duplicated_issue_detector.py [--excel EXCEL_FILE]

Output:
    - Adds 'duplicated_issue' column at first blank column in Issues sheet
    - Populates with comma-separated list of duplicate issue IDs
"""

import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
sys.path.insert(0, PROJECT_ROOT)

import openpyxl
from collections import defaultdict


def find_duplicated_issues(ws) -> dict:
    """Find duplicated issues based on Test Class + Test Case."""
    class_case_index = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        issue_id = ws.cell(row, 1).value
        if test_class and test_case:
            key = (test_class, test_case)
            class_case_index[key].append((row, issue_id))

    duplicates = {}
    for key, rows in class_case_index.items():
        if len(rows) > 1:
            issue_ids = [rid for _, rid in rows]
            for row, rid in rows:
                other_issues = [i for i in issue_ids if i != rid]
                if other_issues:
                    duplicates[row] = sorted([str(i) for i in other_issues])
    return duplicates


def add_duplicated_column(excel_file: str):
    """Add duplicated_issue column to Issues sheet."""
    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)

    if "Issues" not in wb.sheetnames:
        print("  Error: 'Issues' sheet not found!")
        return 0

    if "Test Cases" not in wb.sheetnames:
        print("  Error: 'Test Cases' sheet not found!")
        return 0

    ws_issues = wb["Issues"]
    ws_test = wb["Test Cases"]
    total_issues = ws_issues.max_row - 1
    total_tests = ws_test.max_row - 1

    print(f"  Issues sheet: {total_issues} rows")
    print(f"  Test Cases sheet: {total_tests} rows")

    first_blank = None
    for col in range(1, 51):
        if ws_issues.cell(1, col).value is None:
            first_blank = col
            break

    if first_blank is None:
        print("  Error: No blank column found!")
        return 0

    ws_issues.cell(1, first_blank, "duplicated_issue")
    print(f"  Added column {first_blank}: duplicated_issue")

    print("  Detecting duplicates...")
    dups = find_duplicated_issues(ws_test)
    print(f"  Found {len(dups)} duplicate entries")

    issue_row_map = {}
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        if issue_id:
            issue_row_map[issue_id] = row

    updated = 0
    for row in range(2, ws_test.max_row + 1):
        issue_id = ws_test.cell(row, 1).value
        if issue_id and issue_id in issue_row_map:
            dup_issues = dups.get(row, [])
            if dup_issues:
                issue_row = issue_row_map[issue_id]
                ws_issues.cell(issue_row, first_blank, ",".join(dup_issues))
                updated += 1

    print(f"  Updated {updated} issues with duplicate info")

    wb.save(excel_file)
    print(f"  Saved: {excel_file}")

    wb.close()
    return updated


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Detect duplicated issues and add duplicated_issue column"
    )
    parser.add_argument(
        "--excel",
        default="/home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx",
        help="Excel file with Issues and Test Cases sheets",
    )
    args = parser.parse_args()
    updated = add_duplicated_column(args.excel)
    print(f"\nComplete: Updated {updated} issues")


if __name__ == "__main__":
    main()