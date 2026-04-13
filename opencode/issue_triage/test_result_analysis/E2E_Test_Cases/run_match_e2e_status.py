#!/usr/bin/env python3
"""
Standalone script to run match_e2e_status function.

This script loads E2E test status from Inductor_E2E_Test_Report.xlsx files and matches them
to E2E test cases in the torch_xpu_ops_issues.xlsx workbook.

Usage:
    python run_match_e2e_status.py [--excel EXCEL_FILE] [--base-dir BASE_DIR] [--no-save]

Requirements:
    - torch_xpu_ops_issues.xlsx with 'E2E Test Cases' sheet
    - E2E test report files in directories like:
        */Inductor_E2E_Test_Report.xlsx
"""

import os
import sys
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(SCRIPT_DIR)))
sys.path.insert(0, PROJECT_ROOT)

import glob
import argparse
import openpyxl


def normalize_key_value(value):
    if value is None:
        return ''
    return str(value).lower().strip().replace(' ', '_').replace('-', '_')


def create_model_variants(name):
    if not name:
        return []
    name = str(name).lower().strip()
    variants = set()
    variants.add(name)
    if '/' in name:
        parts = name.split('/')
        variants.add(parts[-1])
        variants.add(name.replace('/', ''))
        variants.add(parts[0] + parts[1])
    camel_split = re.sub(r'([a-z])([A-Z])', r'\1_\2', name).split('_')
    if len(camel_split) > 1:
        variants.add('_'.join(camel_split).lower())
        variants.add(''.join(camel_split).lower())
    variants.add(name.replace('_', ''))
    return list(variants)


def parse_sheet_name(sheet_name):
    parts = sheet_name.replace('_acc', '').split('_')
    if parts[0] in ['huggingface', 'timm', 'torchbench']:
        benchmark = parts[0]
        idx = 1
    else:
        benchmark = 'unknown'
        idx = 0
    amp = False
    dtype = 'float32'
    phase = 'inf'
    remaining = parts[idx:]
    for i, part in enumerate(remaining):
        if part == 'amp':
            amp = True
        elif part in ['bf16', 'float16', 'float32', 'fp16', 'fp32']:
            dtype = part.replace('bf16', 'bfloat16').replace('fp16', 'float16').replace('fp32', 'float32')
        elif part in ['inf', 'tra', 'training']:
            phase = 'inf' if part == 'inf' else 'tra'
    return benchmark, amp, dtype, phase


def load_all_e2e_reports(base_dir):
    all_models = {}
    for xlsx_path in glob.glob(f'{base_dir}/*/Inductor_E2E_Test_Report.xlsx'):
        try:
            dirname = os.path.basename(os.path.dirname(xlsx_path))
            print(f"  Loading: {dirname}/Inductor_E2E_Test_Report.xlsx")
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            for sheet_name in wb.sheetnames:
                if not sheet_name.endswith('_acc'):
                    continue
                benchmark, amp, dtype, phase = parse_sheet_name(sheet_name)
                model_prefix = f"{benchmark}|{dtype}|{phase}|{amp}"
                sheet = wb[sheet_name]
                row_count = 0
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    if not row or not row[1]:
                        continue
                    model_name = str(row[1]).strip()
                    accuracy = row[3] if len(row) > 3 else None
                    if accuracy:
                        for v in create_model_variants(model_name):
                            all_models[f"{model_prefix}|{v}"] = accuracy
                        row_count += 1
                print(f"    Sheet {sheet_name}: {row_count} models")
            wb.close()
        except Exception as e:
            print(f"  Warning: Failed to parse {xlsx_path}: {e}")
    return all_models


def run_match_e2e_status(excel_file, base_dir, save=True):
    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    if 'E2E Test Cases' not in wb.sheetnames:
        print("  Error: 'E2E Test Cases' sheet not found!")
        return 0
    ws_e2e = wb['E2E Test Cases']
    total_rows = ws_e2e.max_row - 1
    print(f"  Total rows: {total_rows}")
    ws_e2e.cell(1, 13, 'XPU Accuracy Status')
    print(f"  Loading E2E model status from reports...")
    status_map = load_all_e2e_reports(base_dir)
    print(f"  Built status map with {len(status_map)} entries")
    matched = 0
    not_found = 0
    for row in range(2, ws_e2e.max_row + 1):
        benchmark = ws_e2e.cell(row, 3).value
        model = ws_e2e.cell(row, 4).value
        dtype = ws_e2e.cell(row, 6).value
        amp = ws_e2e.cell(row, 7).value
        phase = ws_e2e.cell(row, 5).value
        if benchmark and model and dtype and phase:
            dtype_key = normalize_key_value(dtype)
            phase_key = 'inf' if normalize_key_value(phase).startswith('inf') else 'tra'
            amp_val = bool(amp) if amp else False
            model_variants = create_model_variants(model)
            found = False
            for model_var in model_variants:
                for check_amp in [amp_val, False] if amp_val else [False]:
                    for dtype_fallback in [dtype_key, 'float32', 'float16']:
                        key = f"{benchmark}|{dtype_fallback}|{phase_key}|{check_amp}|{model_var}"
                        if key in status_map:
                            ws_e2e.cell(row, 13, status_map[key])
                            matched += 1
                            found = True
                            break
                    if found:
                        break
                if found:
                    break
            if not found:
                ws_e2e.cell(row, 13, 'Status not found')
                not_found += 1
    print(f"  Matched {matched} E2E test cases, {not_found} not found")
    if save:
        wb.save(excel_file)
        print(f"  Saved: {excel_file}")
    return matched


def main():
    parser = argparse.ArgumentParser(description='Match E2E test cases to accuracy status')
    parser.add_argument('--excel', default='/home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx',
                        help='Excel file with E2E_Test_Cases sheet')
    parser.add_argument('--base-dir', default='/home/daisydeng/issue_traige/ci_results/torch-xpu-ops/',
                        help='Base directory containing E2E report folders')
    parser.add_argument('--no-save', action='store_true', help='Do not save results')
    args = parser.parse_args()
    matched = run_match_e2e_status(args.excel, args.base_dir, save=not args.no_save)
    print(f"\nComplete: Matched {matched} E2E test cases")


if __name__ == '__main__':
    main()