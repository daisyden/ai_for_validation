#!/usr/bin/env python3
"""
E2E Test Cases processor module.

Contains all logic for filling in fields of 'E2E_Test_Cases' sheet in torch_xpu_ops_issues.xlsx:
- Accuracy status from torch-xpu-ops nightly CI E2E reports
- Maps E2E test cases to accuracy results from huggingface/timm/torchbench reports

Usage:
    from test_result.E2E_Test_Cases.e2e_cases_processor import process_e2e_cases

    wb = openpyxl.load_workbook('torch_xpu_ops_issues.xlsx')
    process_e2e_cases(wb)
    wb.save('torch_xpu_ops_issues.xlsx')
"""

import glob
import os


def process_e2e_cases(wb):
    """
    Process E2E Test Cases sheet to add accuracy status from torch-xpu-ops nightly.
    
    This function reads E2E test reports from torch-xpu-ops nightly CI and maps
    accuracy status to the E2E Test Cases sheet.
    
    Input:
        - wb: openpyxl Workbook with 'E2E Test Cases' sheet
        - CI artifacts at: /home/daisydeng/issue_traige/ci_results/torch-xpu-ops/*E2E*/Inductor_E2E_Test_Report.xlsx
    
    Output:
        - Col 13: torch-xpu-ops nightly status - accuracy
    
    Report parsing:
        - Supports huggingface, timm_models, torchbench benchmarks
        - Sheet naming: {benchmark}_{dtype}_{phase}_acc (e.g., huggingface_float32_inf_acc)
        - Phase: inf (inference), tra (training)
        - AMP support: _amp_ in sheet name indicates AMP enabled
    
    Args:
        wb: openpyxl Workbook object (loaded)
    """
    import openpyxl
    
    ws_e2e = wb['E2E Test Cases']
    
    ws_e2e.cell(1, 13, 'torch-xpu-ops nightly status - accuracy')
    
    base_dir = '/home/daisydeng/issue_traige/ci_results/torch-xpu-ops'
    
    e2e_model_status = {}
    
    for report_path in glob.glob(f'{base_dir}/*E2E*/Inductor_E2E_Test_Report.xlsx'):
        try:
            dirname = os.path.basename(os.path.dirname(report_path))
            if 'huggingface' in dirname:
                benchmark = 'huggingface'
            elif 'timm' in dirname:
                benchmark = 'timm_models'
            elif 'torchbench' in dirname:
                benchmark = 'torchbench'
            else:
                continue
            
            report_wb = openpyxl.load_workbook(report_path)
            
            for sheet_name in report_wb.sheetnames:
                if not sheet_name.endswith('_acc'):
                    continue
                
                parts = sheet_name.replace(f'{benchmark}_', '').replace('_acc', '').split('_')
                
                dtype = 'float32'
                amp = False
                phase = ''
                
                if 'amp' in parts:
                    amp = True
                    idx = parts.index('amp')
                    if idx + 1 < len(parts):
                        dtype = parts[idx + 1]
                else:
                    if len(parts) >= 1:
                        dtype = parts[0]
                
                if 'inf' in parts:
                    phase = 'inf'
                elif 'tra' in parts:
                    phase = 'tra'
                
                ws = report_wb[sheet_name]
                for row in range(3, ws.max_row + 1):
                    model_name = ws.cell(row, 2).value
                    status = ws.cell(row, 4).value
                    
                    if model_name and status:
                        key = (benchmark, dtype, amp, phase, model_name.lower())
                        e2e_model_status[key] = status
        except Exception as e:
            print(f"  Warning: Failed to parse {report_path}: {e}")
    
    print(f"  Found {len(e2e_model_status)} E2E model status entries")
    
    matched = 0
    for row in range(2, ws_e2e.max_row + 1):
        benchmark = ws_e2e.cell(row, 3).value
        model = ws_e2e.cell(row, 4).value
        dtype = ws_e2e.cell(row, 6).value
        amp = ws_e2e.cell(row, 7).value
        phase = ws_e2e.cell(row, 5).value
        
        if benchmark and model and dtype and phase:
            dtype_key = dtype.lower().replace(' ', '_')
            phase_key = phase.lower().replace(' ', '_').replace('inference', 'inf').replace('training', 'tra')
            model_key = model.lower()
            
            key = (benchmark, dtype_key, bool(amp), phase_key, model_key)
            if key in e2e_model_status:
                ws_e2e.cell(row, 13, e2e_model_status[key])
                matched += 1
            else:
                key = (benchmark, dtype_key, False, phase_key, model_key)
                if key in e2e_model_status:
                    ws_e2e.cell(row, 13, e2e_model_status[key])
                    matched += 1
    
    print(f"  Matched {matched} E2E test cases with accuracy status")


def parse_e2e_sheet_name(sheet_name, benchmark):
    """
    Parse E2E report sheet name to extract dtype, amp, and phase.
    
    Args:
        sheet_name: Sheet name like 'huggingface_float32_inf_acc' or 'timm_amp_bf16_tra_acc'
        benchmark: Benchmark identifier (huggingface, timm_models, torchbench)
    
    Returns:
        tuple: (dtype, amp, phase)
    """
    parts = sheet_name.replace(f'{benchmark}_', '').replace('_acc', '').split('_')
    
    dtype = 'float32'
    amp = False
    phase = ''
    
    if 'amp' in parts:
        amp = True
        idx = parts.index('amp')
        if idx + 1 < len(parts):
            dtype = parts[idx + 1]
    else:
        if len(parts) >= 1:
            dtype = parts[0]
    
    if 'inf' in parts:
        phase = 'inf'
    elif 'tra' in parts:
        phase = 'tra'
    
    return dtype, amp, phase


def normalize_key_value(value):
    """
    Normalize key values for matching (dtype, phase, model).
    
    Args:
        value: Value to normalize
    
    Returns:
        str: Normalized value
    """
    if value is None:
        return ''
    return str(value).lower().replace(' ', '_')


def build_e2e_status_mapping(base_dir):
    """
    Build mapping from E2E report to (benchmark, dtype, amp, phase, model) -> status.
    
    Args:
        base_dir: Base directory containing E2E CI artifacts
    
    Returns:
        dict: Mapping of (benchmark, dtype, amp, phase, model) -> status
    """
    import openpyxl
    
    e2e_model_status = {}
    
    for report_path in glob.glob(f'{base_dir}/*E2E*/Inductor_E2E_Test_Report.xlsx'):
        try:
            dirname = os.path.basename(os.path.dirname(report_path))
            if 'huggingface' in dirname:
                benchmark = 'huggingface'
            elif 'timm' in dirname:
                benchmark = 'timm_models'
            elif 'torchbench' in dirname:
                benchmark = 'torchbench'
            else:
                continue
            
            report_wb = openpyxl.load_workbook(report_path)
            
            for sheet_name in report_wb.sheetnames:
                if not sheet_name.endswith('_acc'):
                    continue
                
                dtype, amp, phase = parse_e2e_sheet_name(sheet_name, benchmark)
                
                ws = report_wb[sheet_name]
                for row in range(3, ws.max_row + 1):
                    model_name = ws.cell(row, 2).value
                    status = ws.cell(row, 4).value
                    
                    if model_name and status:
                        key = (benchmark, dtype, amp, phase, model_name.lower())
                        e2e_model_status[key] = status
        except Exception as e:
            print(f"  Warning: Failed to parse {report_path}: {e}")
    
    return e2e_model_status


def match_e2e_status(e2e_model_status, benchmark, dtype, amp, phase, model):
    """
    Match E2E test case to status from the mapping.
    
    Args:
        e2e_model_status: dict mapping (benchmark, dtype, amp, phase, model) -> status
        benchmark: Benchmark identifier
        dtype: Data type (float32, bfloat16, float16, etc.)
        amp: Whether AMP is enabled
        phase: 'inf' or 'tra'
        model: Model name
    
    Returns:
        str or None: Status if found, None otherwise
    """
    dtype_key = normalize_key_value(dtype)
    phase_key = normalize_key_value(phase)
    model_key = normalize_key_value(model)
    
    key = (benchmark, dtype_key, bool(amp), phase_key, model_key)
    if key in e2e_model_status:
        return e2e_model_status[key]
    
    key = (benchmark, dtype_key, False, phase_key, model_key)
    if key in e2e_model_status:
        return e2e_model_status[key]
    
    return None