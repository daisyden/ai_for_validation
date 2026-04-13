#!/usr/bin/env python3
"""
PASS 1: CI Result Matching for Test Cases

Reads test results from test_cases_all.xlsx 'stock' and 'torch-xpu-ops' sheets.
Matches test cases from the 'Test Cases' sheet and populates:
    Col 8: Error Message
    Col 9: Traceback
    Col 12: XPU Status
    Col 13: Stock Status
    Col 14: No Match Reason (No test class / No test case)

Usage:
    from pass1_ci_matching import run_pass1_ci_matching
    run_pass1_ci_matching(ws)
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULT_DIR = os.environ.get('RESULT_DIR', '/home/daisydeng/ai_for_validation/opencode/issue_triage/result')


def normalize_class_name(class_name):
    """Extract short class name from various formats."""
    if not class_name:
        return None
    basename = class_name.split('.')[-1] if '.' in class_name else class_name
    short = basename.replace('XPU', '').replace('Tests', '').replace('Test', '')
    if short and short != basename and short.strip():
        return short.strip()
    return basename


def load_test_cases_all():
    """Load test_cases_all.xlsx and build lookup maps for stock and xpu results."""
    test_cases_all_path = os.path.join(RESULT_DIR, 'test_cases_all.xlsx')

    if not os.path.exists(test_cases_all_path):
        return None, None, None

    wb_all = openpyxl.load_workbook(test_cases_all_path)

    stock_case_map = {}
    stock_short_class_map = {}
    stock_classes = set()

    if 'stock' in wb_all.sheetnames:
        ws_stock = wb_all['stock']
        for row in range(2, ws_stock.max_row + 1):
            test_file = ws_stock.cell(row, 1).value
            test_class = ws_stock.cell(row, 2).value
            test_case = ws_stock.cell(row, 3).value
            if test_class and test_case:
                short_class = normalize_class_name(test_class)
                stock_case_map[(test_class, test_case)] = {
                    'status': ws_stock.cell(row, 4).value or '',
                    'error_msg': ws_stock.cell(row, 5).value or '',
                    'traceback': ws_stock.cell(row, 6).value or ''
                }
                stock_classes.add(test_class)
                stock_classes.add(short_class)
                if short_class != test_class:
                    stock_short_class_map[(short_class, test_case)] = stock_case_map[(test_class, test_case)]
                else:
                    stock_short_class_map[(test_class, test_case)] = stock_case_map[(test_class, test_case)]

    xpu_case_map = {}
    xpu_short_class_map = {}
    xpu_classes = set()

    if 'torch-xpu-ops' in wb_all.sheetnames:
        ws_xpu = wb_all['torch-xpu-ops']
        for row in range(2, ws_xpu.max_row + 1):
            test_file = ws_xpu.cell(row, 1).value
            test_class = ws_xpu.cell(row, 2).value
            test_case = ws_xpu.cell(row, 3).value
            if test_class and test_case:
                short_class = normalize_class_name(test_class)
                xpu_case_map[(test_class, test_case)] = {
                    'status': ws_xpu.cell(row, 5).value or '',
                    'error_msg': ws_xpu.cell(row, 6).value or '',
                    'traceback': ws_xpu.cell(row, 7).value or ''
                }
                xpu_classes.add(test_class)
                xpu_classes.add(short_class)
                if short_class != test_class:
                    xpu_short_class_map[(short_class, test_case)] = xpu_case_map[(test_class, test_case)]
                else:
                    xpu_short_class_map[(short_class, test_case)] = xpu_case_map[(test_class, test_case)]

    return (stock_case_map, stock_short_class_map, stock_classes), (xpu_case_map, xpu_short_class_map, xpu_classes)


def run_pass1_ci_matching(ws):
    """
    PASS 1: Match CI results from test_cases_all.xlsx to Test Cases sheet.
    
    Reads test results from test_cases_all.xlsx 'stock' and 'torch-xpu-ops' sheets.
    Matches test cases from the 'Test Cases' sheet and populates columns.
    
    Updates:
        Col 8: Error Message
        Col 9: Traceback
        Col 12: XPU Status
        Col 13: Stock Status
        Col 14: No Match Reason
    
    Column mapping for Test Cases sheet:
        Col 1: Issue ID
        Col 2: Test Reproducer
        Col 3: Test Type
        Col 4: Test File
        Col 5: Origin Test File
        Col 6: Test Class
        Col 7: Test Case
        Col 8: Error Message
        Col 9: Traceback
        Col 12: XPU Status
        Col 13: Stock Status
        Col 14: No Match Reason
    
    Returns:
        dict: issues_needing_llm - map of issue_id -> test info for unmatched cases
    """
    stock_lookup, xpu_lookup = load_test_cases_all()
    if not stock_lookup or not xpu_lookup:
        return {}

    stock_case_map, stock_short_class_map, stock_classes = stock_lookup
    xpu_case_map, xpu_short_class_map, xpu_classes = xpu_lookup

    for col, name in zip([8, 9, 12, 13, 14], ["Error Message", "Traceback", "XPU Status", "Stock Status", "No Match Reason"]):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    issues_needing_llm = {}
    total = ws.max_row - 1
    found_count = 0
    not_found_count = 0

    for i, row_idx in enumerate(range(2, ws.max_row + 1), 1):
        test_file = ws.cell(row_idx, 4).value
        test_class = ws.cell(row_idx, 6).value
        test_case = ws.cell(row_idx, 7).value
        issue_id = ws.cell(row_idx, 1).value

        if not test_case:
            ws.cell(row_idx, 14, 'No test case')
            continue

        xpu_status = None
        xpu_error_msg = None
        xpu_traceback = None
        stock_status = None
        stock_error_msg = None
        stock_traceback = None

        short_class = normalize_class_name(test_class)

        if test_class and test_case:
            xpu_key = (test_class, test_case)
            if xpu_key in xpu_case_map:
                xpu_status = xpu_case_map[xpu_key]['status']
                xpu_error_msg = xpu_case_map[xpu_key]['error_msg']
                xpu_traceback = xpu_case_map[xpu_key]['traceback']
            elif xpu_short_class_map.get((short_class, test_case)):
                xpu_status = xpu_short_class_map[(short_class, test_case)]['status']
                xpu_error_msg = xpu_short_class_map[(short_class, test_case)]['error_msg']
                xpu_traceback = xpu_short_class_map[(short_class, test_case)]['traceback']

            stock_key = (test_class, test_case)
            if stock_key in stock_case_map:
                stock_status = stock_case_map[stock_key]['status']
                stock_error_msg = stock_case_map[stock_key]['error_msg']
                stock_traceback = stock_case_map[stock_key]['traceback']
            elif stock_short_class_map.get((short_class, test_case)):
                stock_status = stock_short_class_map[(short_class, test_case)]['status']
                stock_error_msg = stock_short_class_map[(short_class, test_case)]['error_msg']
                stock_traceback = stock_short_class_map[(short_class, test_case)]['traceback']

        ws.cell(row_idx, 12, xpu_status if xpu_status else '')
        ws.cell(row_idx, 13, stock_status if stock_status else '')

        if xpu_error_msg:
            ws.cell(row_idx, 8, xpu_error_msg[:3000] if xpu_error_msg else '')
        elif stock_error_msg:
            ws.cell(row_idx, 8, stock_error_msg[:3000] if stock_error_msg else '')

        if xpu_traceback:
            ws.cell(row_idx, 9, xpu_traceback[:3000] if xpu_traceback else '')
        elif stock_traceback and not ws.cell(row_idx, 9).value:
            ws.cell(row_idx, 9, stock_traceback[:3000] if stock_traceback else '')

        if xpu_status or stock_status:
            found_count += 1
            ws.cell(row_idx, 14, '')
        else:
            not_found_count += 1
            issues_needing_llm[issue_id] = {
                'test_file': test_file,
                'test_class': test_class,
                'test_case': test_case,
                'origin_test_file': ws.cell(row_idx, 5).value
            }
            if not test_case:
                ws.cell(row_idx, 14, 'No test case')
            elif not test_class:
                ws.cell(row_idx, 14, 'No test class')
            elif not xpu_status and not stock_status:
                ws.cell(row_idx, 14, 'No test case')

        if i % 500 == 0:
            print(f"    Progress: {i}/{total}")

    print(f"  PASS 1: {found_count} matched, {not_found_count} not found")
    return issues_needing_llm