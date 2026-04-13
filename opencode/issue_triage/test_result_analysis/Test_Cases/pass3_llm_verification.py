"""
PASS 3: LLM Analysis for Test Existence Verification

This module handles LLM verification for test cases where:
- PASS 1 could not find test cases in CI (No Match Reason set)
- XPU Status AND Stock Status are both blank or 'not found'
- LLM checks if test cases actually exist in the codebase

Logic:
    1. Collect issues where XPU Status IS (blank OR 'not found') AND
                    Stock Status IS (blank OR 'not found')
    2. Process only FIRST case per unique issue
    3. Query LLM to verify if test cases actually exist
    4. Apply results to columns: CUDA Case Exist, XPU Case Exist,
       case_existence_comments, can_enable_on_xpu
"""

import time
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def log(msg, print_also=True):
    """Log message to stdout."""
    if print_also:
        print(msg)


def ensure_pass3_headers(ws):
    """Ensure PASS 3 column headers exist."""
    ensure_headers(ws, [16, 17, 18, 19], ["CUDA Case Exist", "XPU Case Exist", "case_existence_comments", "can_enable_on_xpu"])


def ensure_headers(ws, col_indices, col_names):
    """Ensure column headers exist for given columns."""
    from openpyxl.styles import Font, PatternFill
    for col_idx, header_name in zip(col_indices, col_names):
        existing = ws.cell(row=1, column=col_idx).value
        if not existing:
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")


def normalize_class_name(class_name):
    """Extract short class name from fully qualified class name."""
    if not class_name:
        return ''
    class_name = str(class_name)
    for suffix in ['Test', 'XPU', 'Tests', 'CUDA', 'MPS']:
        if class_name.endswith(suffix):
            class_name = class_name[:-len(suffix)]
    return class_name.split('.')[-1] if '.' in class_name else class_name


def load_ci_lookup_maps(ci_data_file):
    """Load CI test case data into lookup maps."""
    import pandas as pd
    
    test_cases_all_xlsx = pd.ExcelFile(ci_data_file)
    stock_df = pd.read_excel(test_cases_all_xlsx, sheet_name='stock')
    xpu_df = pd.read_excel(test_cases_all_xlsx, sheet_name='torch-xpu-ops')
    
    stock_case_map = {}
    stock_short_class_map = {}
    for _, row in stock_df.iterrows():
        test_file = row['Test File']
        test_class = row['Test Class']
        test_case = row['Test Case']
        status = row['Status']
        error_msg = row.get('Error Message', '')
        
        key = (test_class, test_case)
        stock_case_map[key] = {
            'test_file': test_file,
            'status': str(status) if pd.notna(status) else '',
            'error_msg': str(error_msg) if pd.notna(error_msg) else ''
        }
        
        short_class = test_class.split('.')[-1] if pd.notna(test_class) else ''
        short_key = (short_class, test_case)
        if short_key not in stock_short_class_map:
            stock_short_class_map[short_key] = {
                'status': str(status) if pd.notna(status) else '',
                'error_msg': str(error_msg) if pd.notna(error_msg) else ''
            }
    
    xpu_case_map = {}
    xpu_short_class_map = {}
    for _, row in xpu_df.iterrows():
        test_file = row['Test File']
        test_class = row['Test Class']
        test_case = row['Test Case']
        status = row['Status']
        error_msg = row.get('Error Message', '')
        
        key = (test_class, test_case)
        xpu_case_map[key] = {
            'test_file': test_file,
            'status': str(status) if pd.notna(status) else '',
            'error_msg': str(error_msg) if pd.notna(error_msg) else ''
        }
        
        short_class = test_class.split('.')[-1] if pd.notna(test_class) else ''
        short_key = (short_class, test_case)
        if short_key not in xpu_short_class_map:
            xpu_short_class_map[short_key] = {
                'status': str(status) if pd.notna(status) else '',
                'error_msg': str(error_msg) if pd.notna(error_msg) else ''
            }
    
    return stock_case_map, stock_short_class_map, xpu_case_map, xpu_short_class_map


def get_ci_status(test_class, test_case, stock_case_map, stock_short_class_map, xpu_case_map, xpu_short_class_map):
    """Get CI status for a test case by checking both full and short class names."""
    xpu_status = None
    stock_status = None
    
    short_class = normalize_class_name(test_class)
    
    if test_class and test_case:
        xpu_key = (test_class, test_case)
        if xpu_key in xpu_case_map:
            xpu_status = xpu_case_map[xpu_key]['status']
        elif xpu_short_class_map.get((short_class, test_case)):
            xpu_status = xpu_short_class_map[(short_class, test_case)]['status']
        
        stock_key = (test_class, test_case)
        if stock_key in stock_case_map:
            stock_status = stock_case_map[stock_key]['status']
        elif stock_short_class_map.get((short_class, test_case)):
            stock_status = stock_short_class_map[(short_class, test_case)]['status']
    
    return xpu_status, stock_status


def collect_eligible_issues_for_pass3(ws, ci_data_file):
    """
    Collect unique issues eligible for PASS 3 LLM verification.
    
    Criteria:
        - XPU Status IS (blank OR 'not found')
        - AND Stock Status IS (blank OR 'not found')
        - Process only FIRST case per unique issue
        - INCLUDE cases with No Match Reason (LLM to verify)
    
    Returns:
        dict: {issue_id: {'test_file': ..., 'test_class': ..., 'test_case': ..., 'origin_test_file': ...}}
    """
    stock_case_map, stock_short_class_map, xpu_case_map, xpu_short_class_map = load_ci_lookup_maps(ci_data_file)
    
    issues_needing_llm = {}
    processed_issues = set()
    
    for row_idx in range(2, ws.max_row + 1):
        xpu_status = ws.cell(row_idx, 12).value
        stock_status = ws.cell(row_idx, 13).value
        issue_id = ws.cell(row_idx, 1).value
        
        # Skip if XPU or Stock status has a value (passed/failed/skipped)
        if xpu_status and stock_status:
            continue
        
        # Get CI status from lookup maps
        test_class = ws.cell(row_idx, 6).value
        test_case = ws.cell(row_idx, 7).value
        ci_xpu_status, ci_stock_status = get_ci_status(
            test_class, test_case,
            stock_case_map, stock_short_class_map,
            xpu_case_map, xpu_short_class_map
        )
        
        # Double-check: if CI has the status, update the sheet
        if ci_xpu_status and not xpu_status:
            xpu_status = ci_xpu_status
            ws.cell(row_idx, 12, xpu_status)
        if ci_stock_status and not stock_status:
            stock_status = ci_stock_status
            ws.cell(row_idx, 13, stock_status)
        
        # Re-evaluate eligibility
        xpu_nf = xpu_status == 'not found' if xpu_status else False
        stock_nf = stock_status == 'not found' if stock_status else False
        xpu_blank = not xpu_status
        stock_blank = not stock_status
        
        # PASS 3 criteria: BOTH XPU AND Stock are blank or 'not found'
        if (xpu_blank or xpu_nf) and (stock_blank or stock_nf):
            if issue_id not in processed_issues:
                processed_issues.add(issue_id)
                issues_needing_llm[issue_id] = {
                    'test_file': ws.cell(row_idx, 4).value,
                    'test_class': test_class,
                    'test_case': test_case,
                    'origin_test_file': ws.cell(row_idx, 5).value
                }
    
    return issues_needing_llm


def analyze_test_case_with_llm(test_file, test_class, test_case, origin_test_file=None):
    """
    Use Qwen3-32B via internal API to check CUDA and XPU test case existence.
    
    Returns dict with:
        - cuda_exists: 'Yes'/'No'
        - xpu_exists: 'Yes'/'No'
        - can_enable_on_xpu: 'Yes'/'No'
        - base_test_name
        - cuda_test_file
        - xpu_test_file
        - cuda_test_name
        - xpu_test_name
        - explanation
    """
    import requests
    import json
    import time
    import os
    import re
    
    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"
    
    pytorch_root = os.path.expanduser('~/pytorch')
    if not os.path.exists(pytorch_root):
        pytorch_root = os.path.expanduser('~/issue_traige/pytorch')
    
    prompt = f"""You are in the pytorch directory: {pytorch_root}

Check if the CUDA and XPU test exists in respective test files based on pytest parameterization.

Paths:
- PyTorch test files: {pytorch_root}/test/
- torch-xpu-ops test files: {pytorch_root}/third_party/torch-xpu-ops/test/xpu/

Test File: {test_file}
Origin Test File: {origin_test_file if origin_test_file else 'Not provided'}
Test Class: {test_class}
Test Case: {test_case}

Base test is the actual test function in the test file.

cuda_exists: replace the subfix of XPU or xpu in { test_class } or {test_case}, check whether cuda test is in PyTorch test files after parameterizaiton
xpu_exists: first check torch-xpu-ops test files {test_file} whether { test_class } or {test_case} exists after parameterization and the patching of XPUPatchForImport, then check whehter the  PyTorch test files have the test for xpu.

IMPORTANT: Determine if this test case can be enabled on XPU (can_enable_on_xpu):
- can_enable_on_xpu = Yes if test has hardcoded 'cuda' device or skip decorators (just skips)
- can_enable_on_xpu = No if test has cuda specific features such as cudnn, or created for some feature that not applicable to xpu. 
If the api for cuda but xpu also has the conterpart, it is can_enable_on_xpu = Yes, for example torch.cuda.is_avialble() has a conterpart of torch.xpu.is_available()

IMPORTANT: Explain WHY XPU test does not exist:
1. REMOVED/RENAMED: cuda test also not exists, the test could be removed or renamed, what is the expected name
2. SKIP DECORATORS: @onlyCUDA, @skipCUDAIfNoHipdnn, @skipIfXpu, @requires_xccl
3. PARAMETERIZATION: @dtypes, @parametrize_test generating tests
5. NOT APPLICABLE: CUDA/ROCm specific (hipdnn backend)
5. OTHER: Other reasons

How to enable xpu test:
If xpu can be enabled, explain how to enable it. For example remove what decorators? how to update op db? How to make interface general, please be specific to point out the changes required.
If the cuda test is renamed, provide the expected xpu test name. 

Return ONLY valid JSON:
{{
    "explanation": "detailed explanation why XPU test exists or not",
    "cuda_exists": "Yes/No",
    "xpu_exists": "Yes/No",
    "can_enable_on_xpu": "Yes/No",
    "how_to_enable": details explainion how to enable the test on xpu"
}}
"""
    
    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "You are a PyTorch test analysis assistant. Return ONLY valid JSON."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 2048
    }
    
    start_time = time.time()
    
    try:
        response = requests.post(
            LLM_ENDPOINT,
            headers=headers,
            json=payload,
            timeout=180
        )
        response.raise_for_status()
        result = response.json()
        elapsed = time.time() - start_time
        
        content = result.get('choices', [{}])[0].get('message', {}).get('content', '')
        json_match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
        if json_match:
            data = json.loads(json_match.group())
            data['elapsed_time'] = elapsed
            return data
        
        return {
            'explanation': f'Parse failed: {content[:100]}',
            'cuda_exists': 'Unknown',
            'xpu_exists': 'Unknown',
            'can_enable_on_xpu': 'Unknown',
            'elapsed_time': elapsed
        }
        
    except Exception as e:
        return {
            'explanation': f'Error: {str(e)[:100]}',
            'cuda_exists': 'Unknown',
            'xpu_exists': 'Unknown',
            'can_enable_on_xpu': 'Unknown',
            'elapsed_time': time.time() - start_time
        }


def pass3_llm_verification(ws, ci_data_file=None):
    """
    PASS 3: LLM Analysis for Test Existence Verification
    
    For issues where CI results were not found (both XPU and Stock statuses
    are blank or 'not found'), uses LLM to determine if test cases exist.
    
    Criteria:
        - XPU Status IS (blank OR 'not found')
        - AND Stock Status IS (blank OR 'not found')
        - Process only FIRST case per unique issue
        - INCLUDE cases with No Match Reason
    
    Updates:
        Col 16: CUDA Case Exist
        Col 17: XPU Case Exist
        Col 18: case_existence_comments
        Col 19: can_enable_on_xpu
    """
    
    if ci_data_file is None:
        SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
        ci_data_file = os.path.join(os.path.dirname(os.path.dirname(SCRIPT_DIR)), 'result', 'test_cases_all.xlsx')
    
    ensure_pass3_headers(ws)
    
    log("  [PASS 3/5] Running LLM analysis for test existence (CI verification)...")
    log("-" * 60)
    pass3_start_time = time.time()

    # Collect eligible issues
    log("    Collecting eligible issues...")
    issues_needing_llm = collect_eligible_issues_for_pass3(ws, ci_data_file)
    issue_count = len(issues_needing_llm)

    issues_list = sorted(issues_needing_llm.items())
    issue_count = len(issues_list)
    
    log(f"    Found {issue_count} unique issues needing LLM verification")
    log("-" * 60)
    
    if issue_count == 0:
        log("    No eligible issues for PASS 3")
        log(f"  PASS 3 complete: 0 results (0.0s)")
        return
    
    llm_results = {}
    
    # Process each unique issue
    for idx, (issue_id, info) in enumerate(issues_list, 1):
        start_time = time.time()
        try:
            test_file = info['test_file']
            test_class = info['test_class']
            test_case = info['test_case']
            origin_test_file = info['origin_test_file']
            
            log(f"[PASS3] [{idx:3d}/{issue_count}] Starting at {time.strftime('%H:%M:%S')} | Issue:{issue_id}")
            log(f"[PASS3]   File: {test_file}")
            log(f"[PASS3]   Class: {test_class}")
            log(f"[PASS3]   Case: {test_case}")
            
            result = analyze_test_case_with_llm(test_file, test_class, test_case, origin_test_file)
            elapsed = result.get('elapsed_time', 0)
            llm_results[issue_id] = result
            
            cuda = result.get('cuda_exists', '?')
            xpu = result.get('xpu_exists', '?')
            enable = result.get('can_enable_on_xpu', '?')
            explanation = result.get('explanation', '')[:80] if result.get('explanation') else 'N/A'
            
            log(f"[PASS3]   [{idx:3d}/{issue_count}] Completed at {time.strftime('%H:%M:%S')} | Time:{elapsed:.1f}s")
            log(f"[PASS3]   Result: CUDA={cuda}, XPU={xpu}, Enable={enable}")
            log(f"[PASS3]   Note: {explanation}...")
            log(f"[PASS3]   {'-' * 56}")
            
        except Exception as e:
            elapsed = time.time() - start_time
            log(f"[PASS3 ERROR] [{idx:3d}/{issue_count}] Issue:{issue_id} | Time:{elapsed:.1f}s | Error: {e}")
            llm_results[issue_id] = {'error': str(e)}
    
    elapsed = time.time() - pass3_start_time
    log("-" * 60)
    log(f"  PASS 3 complete: {len(llm_results)} LLM results ({elapsed:.1f}s)")
    
    log("  [APPLY] Writing LLM results to test cases...")
    applied = 0
    
    # Apply results to first case of each issue (expand to all cases if needed)
    processed_issues = set()
    for row_idx in range(2, ws.max_row + 1):
        issue_id = ws.cell(row_idx, 1).value
        
        # Only apply to first case of each issue
        if issue_id in processed_issues:
            continue
            
        if issue_id not in llm_results:
            continue
            
        no_match = ws.cell(row_idx, 14).value
        xpu_status = ws.cell(row_idx, 12).value
        stock_status = ws.cell(row_idx, 13).value
        
        # Re-apply eligibility check
        xpu_nf = xpu_status == 'not found' if xpu_status else False
        stock_nf = stock_status == 'not found' if stock_status else False
        xpu_blank = not xpu_status
        stock_blank = not stock_status
        
        if (xpu_blank or xpu_nf) and (stock_blank or stock_nf):
            result = llm_results[issue_id]
            processed_issues.add(issue_id)
            
            ws.cell(row_idx, 16, result.get('cuda_exists', 'Unknown'))  # Col 16: CUDA Case Exist
            ws.cell(row_idx, 17, result.get('xpu_exists', 'Unknown'))    # Col 17: XPU Case Exist
            
            parts = []
            explanation = result.get('explanation', '')
            if explanation and explanation != 'N/A':
                parts.append('explanation: ' + str(explanation))
            
            for key in ['base_test_name', 'cuda_test_file', 'xpu_test_file', 'cuda_test_name', 'xpu_test_name']:
                val = result.get(key)
                if val and val != 'N/A':
                    parts.append(f'{key}:{val}')
            
            comment = '\n'.join(parts) if parts else 'LLM analysis complete'
            ws.cell(row_idx, 18, comment)  # Col 18: case_existence_comments
            ws.cell(row_idx, 19, result.get('can_enable_on_xpu', 'Unknown'))  # Col 19: can_enable_on_xpu
            applied += 1
    
    log(f"  Applied LLM results to {applied} cases")


def run_pass3(excel_file, ci_data_file=None, save=True):
    """
    Run PASS 3 LLM verification on the Excel file.
    
    Args:
        excel_file: Path to torch_xpu_ops_issues.xlsx
        ci_data_file: Path to test_cases_all.xlsx (optional)
        save: Whether to save after processing
    """
    from openpyxl import load_workbook
    
    log(f"Loading: {excel_file}")
    wb = load_workbook(excel_file)
    ws = wb['Test Cases']
    
    pass3_llm_verification(ws, ci_data_file)
    
    if save:
        wb.save(excel_file)
        log(f"Saved: {excel_file}")
    
    return wb


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='PASS 3: LLM Verification')
    parser.add_argument('--excel', default='/home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx',
                        help='Excel file to process')
    parser.add_argument('--ci-data', default='/home/daisydeng/ai_for_validation/opencode/issue_triage/result/test_cases_all.xlsx',
                        help='CI test cases data file')
    parser.add_argument('--no-save', action='store_true', help='Do not save results')
    
    args = parser.parse_args()
    
    run_pass3(args.excel, args.ci_data, save=not args.no_save)
