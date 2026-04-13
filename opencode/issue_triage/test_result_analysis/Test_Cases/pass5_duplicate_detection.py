"""
PASS 5: Cross-Issue Duplicate Detection

Detects duplicate issues based on:
1. Test Class + Test Case (exact match)
2. Traceback similarity (normalized)

Updates:
    Col 20: duplicated_issue (comma-separated list of other issue IDs)
"""

import time
from collections import defaultdict


def log(msg, print_also=True):
    """Log message to stdout."""
    if print_also:
        print(msg)


def _normalize_traceback(traceback):
    """Normalize traceback for duplicate detection."""
    if not traceback:
        return None
    lines = traceback.strip().split('\n')
    if len(lines) < 2:
        return None
    core = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith('File "') or line.startswith('  File '):
            continue
        if line.startswith('---'):
            break
        core.append(line)
    norm = ' | '.join(core[:6])
    norm = norm.replace('"', "'").replace('  ', ' ')
    norm = norm.replace('test_serialization_xpu', 'test_serialization')
    norm = norm.replace('test_int64_upsample3d_xpu', 'test_int64_upsample3d')
    return norm


def ensure_headers(ws, col_indices, col_names):
    """Ensure column headers exist for given columns."""
    from openpyxl.styles import Font, PatternFill
    for col_idx, header_name in zip(col_indices, col_names):
        existing = ws.cell(row=1, column=col_idx).value
        if not existing:
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")


def pass5_duplicate_detection(ws):
    """
    Step 5: Cross-Issue Duplicate Detection

    Finds duplicates based on multiple criteria:
    1. Test Class + Test Case (exact match)
    2. Traceback similarity (normalized)

    Updates:
        Col 20: duplicated_issue (comma-separated list of other issue IDs)

    Returns:
        dict: traceback_dups for external reference if needed
    """
    ensure_headers(ws, [20], ["duplicated_issue"])

    log("  [PASS 5/5] Detecting cross-issue duplicates...")
    start_time = time.time()

    log("    [-] Building test_class+test_case index...")
    test_case_to_issues = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        issue_id = ws.cell(row, 1).value
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        if test_class and test_case:
            key = (test_class, test_case)
            test_case_to_issues[key].append((row, str(issue_id).strip()))

    log("    [-] Building traceback index (with noise filtering)...")
    traceback_to_rows = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        traceback = ws.cell(row, 9).value or ''
        issue_id = ws.cell(row, 1).value
        if 'AssertionError: Tensor-likes are not close!' in traceback:
            continue
        norm = _normalize_traceback(traceback)
        if norm:
            traceback_to_rows[norm].append((row, str(issue_id).strip(), traceback))

    log("    [-] Finding test_class+test_case duplicates...")
    test_case_dups = defaultdict(set)
    for key, entries in test_case_to_issues.items():
        if len(entries) > 1:
            issue_ids = [e[1] for e in entries]
            for row, issue_id in entries:
                other_issues = [i for i in issue_ids if i != issue_id]
                if other_issues:
                    test_case_dups[row].update(other_issues)

    log("    [-] Finding traceback similarity duplicates...")
    traceback_dups = defaultdict(set)
    for norm, entries in traceback_to_rows.items():
        if len(entries) > 1:
            issue_ids = [e[1] for e in entries]
            for row, issue_id, traceback in entries:
                other_issues = [i for i in issue_ids if i != issue_id]
                if other_issues:
                    traceback_dups[row].update(other_issues)

    log("    [-] Merging duplicate sources...")
    dup_count = 0
    for row in range(2, ws.max_row + 1):
        issue_id = ws.cell(row, 1).value
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value

        dup_set = set()
        if test_class and test_case:
            dup_set.update(test_case_dups.get(row, []))
        dup_set.update(traceback_dups.get(row, []))

        if dup_set:
            other_unique = sorted(list(set(dup_set)))
            ws.cell(row, 20, ','.join(other_unique))
            dup_count += 1

    elapsed = time.time() - start_time
    log(f"  PASS 5 complete: Marked {dup_count} duplicates ({elapsed:.1f}s)")

    log("  [CLEANUP] Fixing old boolean values...")
    old_values_fixed = 0
    for row in range(2, ws.max_row + 1):
        dup_issue = ws.cell(row, 22).value
        if dup_issue in ['True', 'False']:
            ws.cell(row, 22, '')
            old_values_fixed += 1
    if old_values_fixed > 0:
        log(f"  Cleared {old_values_fixed} old boolean values")

    return traceback_dups


if __name__ == '__main__':
    import argparse
    from openpyxl import load_workbook

    parser = argparse.ArgumentParser(description='PASS 5: Duplicate Detection')
    parser.add_argument('--excel', default='/home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx',
                        help='Excel file to process')
    parser.add_argument('--no-save', action='store_true', help='Do not save results')

    args = parser.parse_args()

    log(f"Loading: {args.excel}")
    wb = load_workbook(args.excel)
    ws = wb['Test Cases']

    pass5_duplicate_detection(ws)

    if not args.no_save:
        wb.save(args.excel)
        log(f"Saved: {args.excel}")