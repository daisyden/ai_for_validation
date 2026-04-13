#!/usr/bin/env python3
"""
Run Test Cases Processor - Step by Step

Usage:
    python3 run_processor_steps.py              # Run all steps
    python3 run_processor_steps.py --step 1      # Run only step 1 (PASS 1)
    python3 run_processor_steps.py --steps 1 2  # Run steps 1 and 2
    python3 run_processor_steps.py --steps 1-3    # Run steps 1, 2, 3
    python3 run_processor_steps.py --list       # List available steps
    python3 run_processor_steps.py --help       # Show help

Steps:
    1. PASS 1: Create test_cases_all.xlsx with stock and torch-xpu-ops sheets, match CI results
    2. PASS 2: Torch-ops extraction (pattern + LLM fallback)
    3. PASS 3: LLM analysis for test existence (CUDA/XPU)
    4. PASS 4: Dependency RAG (match ops to deps)
    5. PASS 5: Duplicate detection (cross-issue)

Example:
    # Run steps 1 and 2 only (no LLM calls - fast)
    python3 run_processor_steps.py --steps 1 2

    # Run full processor
    python3 run_processor_steps.py

    # Run specific steps
    python3 run_processor_steps.py --steps 1 2 3 4 5
"""

import os
import sys
import time
import argparse
import re
import glob
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = '/home/daisydeng'
RESULT_DIR = os.environ.get('RESULT_DIR', '/home/daisydeng/ai_for_validation/opencode/issue_triage/result')
LOG_FILE = os.path.join(RESULT_DIR, "pipeline.log")

sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, os.path.join(SCRIPT_DIR, '..'))

from test_cases_processor import (
    pass2_extract_torch_ops,
    pass3_llm_analysis_for_test_existence,
    pass4_dependency_rag,
    pass5_duplicate_detection,
    log as base_log
)

STEPS_DESC = {
    1: "PASS 1: Create test_cases_all.xlsx, collect stock & xpu CI results, match CI",
    2: "PASS 2: Torch-ops extraction (pattern + LLM fallback)",
    3: "PASS 3: LLM analysis for test existence (CUDA/XPU)",
    4: "PASS 4: Dependency RAG (match ops to deps)",
    5: "PASS 5: Duplicate detection (cross-issue)"
}

STEP_FUNCS = {
    2: pass2_extract_torch_ops,
    3: pass3_llm_analysis_for_test_existence,
    4: pass4_dependency_rag,
    5: pass5_duplicate_detection
}

llm_steps = {3}


def log(msg, print_also=True):
    """Log message to file and optionally print to console."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {msg}"
    try:
        with open(LOG_FILE, "a") as f:
            f.write(log_msg + "\n")
    except Exception:
        pass
    if print_also:
        print(log_msg)


def parse_ci_xml_content(content):
    """Parse pytest XML content and extract test cases with results."""
    try:
        root = ET.fromstring(content)
    except Exception:
        return []

    test_cases = []
    for testcase in root.findall('.//testcase'):
        name = testcase.get('name', '')
        classname = testcase.get('classname', '')
        file_path = testcase.get('file', '')

        failure = testcase.find('failure')
        skipped = testcase.find('skipped')

        status = 'passed'
        error_msg = ''
        traceback = ''

        if failure is not None:
            status = 'failed'
            msg = failure.text or failure.get('message', '') or ''
            error_msg, traceback = parse_failure_message(msg)
        elif skipped is not None:
            status = 'skipped'
            msg = skipped.text or skipped.get('message', '') or ''
            error_msg = msg[:500] if msg else 'skipped'

        test_cases.append({
            'test_file': file_path,
            'test_class': classname,
            'test_case': name,
            'status': status,
            'error_msg': error_msg,
            'traceback': traceback
        })

    return test_cases


def parse_failure_message(content):
    """Parse failure message to extract error_msg and traceback."""
    error_msg = ""
    traceback = ""

    if not content:
        return error_msg, traceback

    lines = content.split('\n')

    error_patterns = [
        (r'^RuntimeError', 'RuntimeError'),
        (r'^AssertionError', 'AssertionError'),
        (r'^ValueError', 'ValueError'),
        (r'^TypeError', 'TypeError'),
        (r'^IndexError', 'IndexError'),
        (r'^KeyError', 'KeyError'),
        (r'^ImportError', 'ImportError'),
        (r'^NotImplementedError', 'NotImplementedError'),
        (r'^AttributeError', 'AttributeError'),
        (r'^InductorError', 'InductorError'),
    ]

    error_line_idx = -1
    error_type = None
    last_error_msg = ""

    for idx, line in enumerate(lines):
        stripped = line.strip()
        for pattern, etype in error_patterns:
            if re.match(pattern, stripped):
                error_line_idx = idx
                error_type = etype
                clean_line = re.sub(r'^' + etype + r'[:\s]*', '', stripped)
                error_msg = clean_line[:200]
                break
        if error_line_idx >= 0:
            break
        for ep in [r'\braise\s+(RuntimeError|AssertionError|ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|AttributeError|InductorError)\s*[\(\'"]']:
            if re.search(ep, stripped):
                error_line_idx = idx
                match = re.search(r'raise\s+\w+\s*[\(\'"](.+?)[\'\"]?', stripped)
                if match:
                    last_error_msg = match.group(1).strip()[:200]

    traceback = ""
    if 'Traceback (most recent call last):' in content:
        tb_lines = []
        end_idx = error_line_idx if error_line_idx >= 0 else len(lines)
        for idx, line in enumerate(lines):
            if 'Traceback (most recent call last):' in line:
                for j in range(idx, end_idx + 1):
                    tb_lines.append(lines[j])
                break

        if tb_lines:
            traceback = '\n'.join(tb_lines)
        elif last_error_msg:
            for idx, line in enumerate(lines):
                stripped = line.strip()
                for ep in [r'\braise\s+(RuntimeError|AssertionError|ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|AttributeError|InductorError)\s*[\(\'"]']:
                    if re.search(ep, stripped):
                        traceback = '\n'.join(lines[idx:])
                        break
                if traceback:
                    break
    else:
        traceback = ""
        error_msg = content[:200]

    if last_error_msg and not error_msg:
        error_msg = last_error_msg

    return error_msg[:300] if error_msg else error_msg, traceback[:3000] if traceback else traceback


def collect_stock_test_cases():
    """Collect all test cases from stock PyTorch CI pytest XML files."""
    log("  Collecting stock CI test cases...")
    stock_base = '/home/daisydeng/issue_traige/ci_results/stock'

    stock_test_cases = []
    total_files = 0
    total_cases = 0

    for mount_point in glob.glob(f'{stock_base}/test-reports-runattempt1*.zip'):
        try:
            pytest_dir = os.path.join(mount_point, 'test-reports', 'python-pytest')
            if os.path.isdir(pytest_dir):
                for root, dirs, files in os.walk(pytest_dir):
                    for f in files:
                        if f.endswith('.xml'):
                            xml_path = os.path.join(root, f)
                            try:
                                with open(xml_path, 'r') as file:
                                    content = file.read()
                                test_cases = parse_ci_xml_content(content)
                                if test_cases:
                                    test_module = os.path.basename(root)
                                    for tc in test_cases:
                                        tc['source'] = f"stock:{test_module}"
                                    stock_test_cases.extend(test_cases)
                                    total_files += 1
                                    total_cases += len(test_cases)
                            except Exception:
                                pass
            else:
                with zipfile.ZipFile(mount_point, 'r') as zf:
                    for name in zf.namelist():
                        if name.endswith('.xml') and '/python-pytest/' in name:
                            try:
                                with zf.open(name) as f:
                                    content = f.read().decode('utf-8', errors='ignore')
                                test_cases = parse_ci_xml_content(content)
                                if test_cases:
                                    parts = name.split('/')
                                    test_module = parts[-2] if len(parts) >= 2 else name
                                    for tc in test_cases:
                                        tc['source'] = f"stock:{test_module}"
                                    stock_test_cases.extend(test_cases)
                                    total_files += 1
                                    total_cases += len(test_cases)
                            except Exception:
                                pass
        except Exception:
            continue

    log(f"  Stock CI: {total_cases} test cases from {total_files} XML files")
    return stock_test_cases


def collect_torch_xpu_ops_test_cases():
    """Collect all test cases from torch-xpu-ops CI pytest XML files."""
    log("  Collecting torch-xpu-ops CI test cases...")
    base_dir = '/home/daisydeng/issue_traige/ci_results/torch-xpu-ops'

    xpu_test_cases = []
    total_files = 0
    total_cases = 0

    for d in os.listdir(base_dir):
        if d.startswith('Inductor-XPU-UT-Data-'):
            match = re.match(r'Inductor-XPU-UT-Data-([a-f0-9]+)-.*-(\d+)-1$', d)
            if match:
                folder_path = os.path.join(base_dir, d, d)
                if not os.path.exists(folder_path):
                    continue
                for f in os.listdir(folder_path):
                    if f.endswith('.xml') and (f.startswith('op_ut_with_all') or f.startswith('op_ut_with_skip') or f == 'op_extended.xml'):
                        xml_path = os.path.join(folder_path, f)
                        try:
                            tree = ET.parse(xml_path)
                            root = tree.getroot()
                            count = len(root.findall('.//testcase'))
                            if count > 0:
                                prefix = f.replace('.xml', '')
                                for testcase in root.findall('.//testcase'):
                                    name = testcase.get('name', '')
                                    classname = testcase.get('classname', '')
                                    file_path = testcase.get('file', '')

                                    failure = testcase.find('failure')
                                    skipped = testcase.find('skipped')

                                    status = 'passed'
                                    error_msg = ''
                                    traceback = ''

                                    if failure is not None:
                                        status = 'failed'
                                        msg = failure.text or failure.get('message', '') or ''
                                        error_msg, traceback = parse_failure_message(msg)
                                    elif skipped is not None:
                                        status = 'skipped'
                                        msg = skipped.text or skipped.get('message', '') or ''
                                        error_msg = msg[:500] if msg else 'skipped'

                                    xpu_test_cases.append({
                                        'prefix': prefix,
                                        'test_file': file_path,
                                        'test_class': classname,
                                        'test_case': name,
                                        'status': status,
                                        'error_msg': error_msg,
                                        'traceback': traceback
                                    })
                                    total_cases += 1
                                total_files += 1
                        except Exception:
                            pass

    log(f"  XPU CI: {total_cases} test cases from {total_files} XML files")
    return xpu_test_cases


def create_test_cases_all_excel(stock_cases, xpu_cases):
    """Create test_cases_all.xlsx with stock and torch-xpu-ops sheets."""
    output_path = os.path.join(RESULT_DIR, 'test_cases_all.xlsx')
    wb = openpyxl.Workbook()

    ws_stock = wb.active
    ws_stock.title = 'stock'
    stock_headers = ['Test File', 'Test Class', 'Test Case', 'Status', 'Error Message', 'Traceback']
    for col, header in enumerate(stock_headers, 1):
        cell = ws_stock.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for row_idx, tc in enumerate(stock_cases, 2):
        ws_stock.cell(row=row_idx, column=1, value=tc.get('test_file', ''))
        ws_stock.cell(row=row_idx, column=2, value=tc.get('test_class', ''))
        ws_stock.cell(row=row_idx, column=3, value=tc.get('test_case', ''))
        ws_stock.cell(row=row_idx, column=4, value=tc.get('status', ''))
        ws_stock.cell(row=row_idx, column=5, value=tc.get('error_msg', ''))
        ws_stock.cell(row=row_idx, column=6, value=tc.get('traceback', '')[:3000] if tc.get('traceback') else '')

    ws_xpu = wb.create_sheet('torch-xpu-ops')
    xpu_headers = ['Test File', 'Test Class', 'Test Case', 'XML Prefix', 'Status', 'Error Message', 'Traceback']
    for col, header in enumerate(xpu_headers, 1):
        cell = ws_xpu.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for row_idx, tc in enumerate(xpu_cases, 2):
        ws_xpu.cell(row=row_idx, column=1, value=tc.get('test_file', ''))
        ws_xpu.cell(row=row_idx, column=2, value=tc.get('test_class', ''))
        ws_xpu.cell(row=row_idx, column=3, value=tc.get('test_case', ''))
        ws_xpu.cell(row=row_idx, column=4, value=tc.get('prefix', ''))
        ws_xpu.cell(row=row_idx, column=5, value=tc.get('status', ''))
        ws_xpu.cell(row=row_idx, column=6, value=tc.get('error_msg', ''))
        ws_xpu.cell(row=row_idx, column=7, value=tc.get('traceback', '')[:3000] if tc.get('traceback') else '')

    for ws in [ws_stock, ws_xpu]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(output_path)
    log(f"  Created: {output_path}")
    return output_path


def build_stock_status_map(stock_cases):
    """Build a map for fast lookup: (test_class, test_case) -> status, error_msg, traceback."""
    status_map = {}
    for tc in stock_cases:
        key = (tc.get('test_class', ''), tc.get('test_case', ''))
        status_map[key] = {
            'status': tc.get('status', ''),
            'error_msg': tc.get('error_msg', ''),
            'traceback': tc.get('traceback', '')
        }
    return status_map


def build_xpu_status_map(xpu_cases):
    """Build a map for fast lookup: (test_class, test_case, prefix) -> status, error_msg, traceback."""
    status_map = {}
    for tc in xpu_cases:
        key = (tc.get('test_class', ''), tc.get('test_case', ''), tc.get('prefix', ''))
        status_map[key] = {
            'status': tc.get('status', ''),
            'error_msg': tc.get('error_msg', ''),
            'traceback': tc.get('traceback', '')
        }
    return status_map


def normalize_class_name(class_name):
    """Extract short class name from various formats."""
    if not class_name:
        return None
    basename = class_name.split('.')[-1] if '.' in class_name else class_name
    short = basename.replace('XPU', '').replace('Tests', '').replace('Test', '')
    if short and short != basename and short.strip():
        return short.strip()
    return basename


def normalize_file_path(file_path):
    """Get basename of file path."""
    if not file_path:
        return None
    return file_path.split('/')[-1].replace('.py', '').replace('_xpu', '').replace('_cuda', '').strip()


def load_test_cases_all():
    """Load test_cases_all.xlsx and build lookup maps for stock and xpu results."""
    test_cases_all_path = os.path.join(RESULT_DIR, 'test_cases_all.xlsx')

    if not os.path.exists(test_cases_all_path):
        log(f"  Warning: {test_cases_all_path} not found, will collect from XML")
        return None, None, None, None

    wb_all = openpyxl.load_workbook(test_cases_all_path)

    stock_case_map = {}
    stock_short_class_map = {}
    stock_classes = set()

    if 'stock' in wb_all.sheetnames:
        ws_stock = wb_all['stock']
        for row in range(2, ws_stock.max_row + 1):
            test_file = ws_stock.cell(row, 1).value
            test_class = ws_stock.cell(row, 2).value
            test_case = ws_stock.cell(row, 3).value
            if test_class and test_case:
                short_class = normalize_class_name(test_class)
                stock_case_map[(test_class, test_case)] = {
                    'status': ws_stock.cell(row, 4).value or '',
                    'error_msg': ws_stock.cell(row, 5).value or '',
                    'traceback': ws_stock.cell(row, 6).value or ''
                }
                stock_classes.add(test_class)
                stock_classes.add(short_class)
                if short_class != test_class:
                    stock_short_class_map[(short_class, test_case)] = stock_case_map[(test_class, test_case)]
                else:
                    stock_short_class_map[(test_class, test_case)] = stock_case_map[(test_class, test_case)]

    xpu_case_map = {}
    xpu_short_class_map = {}
    xpu_classes = set()

    if 'torch-xpu-ops' in wb_all.sheetnames:
        ws_xpu = wb_all['torch-xpu-ops']
        for row in range(2, ws_xpu.max_row + 1):
            test_file = ws_xpu.cell(row, 1).value
            test_class = ws_xpu.cell(row, 2).value
            test_case = ws_xpu.cell(row, 3).value
            if test_class and test_case:
                short_class = normalize_class_name(test_class)
                xpu_case_map[(test_class, test_case)] = {
                    'status': ws_xpu.cell(row, 5).value or '',
                    'error_msg': ws_xpu.cell(row, 6).value or '',
                    'traceback': ws_xpu.cell(row, 7).value or ''
                }
                xpu_classes.add(test_class)
                xpu_classes.add(short_class)
                if short_class != test_class:
                    xpu_short_class_map[(short_class, test_case)] = xpu_case_map[(test_class, test_case)]
                else:
                    xpu_short_class_map[(short_class, test_case)] = xpu_case_map[(test_class, test_case)]

    return (stock_case_map, stock_short_class_map, stock_classes), (xpu_case_map, xpu_short_class_map, xpu_classes)


def pass1_match_ci_results(ws, output_path):
    """
    PASS 1: Create test_cases_all.xlsx, collect stock & xpu CI results, match CI.

    Reads test results from test_cases_all.xlsx 'stock' and 'torch-xpu-ops' sheets.

    Updates:
        Col 8: Error Message
        Col 9: Traceback
        Col 12: XPU Status
        Col 13: Stock Status

    Column mapping for Test Cases sheet:
        Col 1: Issue ID
        Col 2: Test Reproducer
        Col 3: Test Type
        Col 4: Test File
        Col 5: Origin Test File
        Col 6: Test Class
        Col 7: Test Case
        Col 8: Error Message
        Col 9: Traceback
        Col 10: torch-ops
        Col 11: dependency
        Col 12: XPU Status
        Col 13: Stock Status
        Col 16: CUDA Case Exist
        Col 17: XPU Case Exist
        Col 18: case_existence_comments
        Col 19: can_enable_on_xpu
        Col 20: duplicated_issue
    """
    from openpyxl.styles import Font, PatternFill

    for col, name in zip([8, 9, 12, 13, 14], ["Error Message", "Traceback", "XPU Status", "Stock Status", "No Match Reason"]):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    test_cases_all_path = os.path.join(RESULT_DIR, 'test_cases_all.xlsx')
    if not os.path.exists(test_cases_all_path):
        stock_cases = collect_stock_test_cases()
        xpu_cases = collect_torch_xpu_ops_test_cases()
        create_test_cases_all_excel(stock_cases, xpu_cases)

    stock_lookup, xpu_lookup = load_test_cases_all()
    if not stock_lookup or not xpu_lookup:
        log(f"  Warning: Could not load test_cases_all.xlsx")
        return {}
    stock_case_map, stock_short_class_map, stock_classes = stock_lookup
    xpu_case_map, xpu_short_class_map, xpu_classes = xpu_lookup

    issues_needing_llm = {}
    total = ws.max_row - 1
    found_count = 0
    not_found_count = 0

    log("  Matching CI results from test_cases_all.xlsx...")
    for i, row_idx in enumerate(range(2, ws.max_row + 1), 1):
        test_file = ws.cell(row_idx, 4).value
        test_class = ws.cell(row_idx, 6).value
        test_case = ws.cell(row_idx, 7).value
        issue_id = ws.cell(row_idx, 1).value

        if not test_case:
            continue

        xpu_status = None
        xpu_error_msg = None
        xpu_traceback = None
        stock_status = None
        stock_error_msg = None
        stock_traceback = None

        short_class = normalize_class_name(test_class)

        if test_class and test_case:
            xpu_key = (test_class, test_case)
            if xpu_key in xpu_case_map:
                xpu_status = xpu_case_map[xpu_key]['status']
                xpu_error_msg = xpu_case_map[xpu_key]['error_msg']
                xpu_traceback = xpu_case_map[xpu_key]['traceback']
            elif xpu_short_class_map.get((short_class, test_case)):
                xpu_status = xpu_short_class_map[(short_class, test_case)]['status']
                xpu_error_msg = xpu_short_class_map[(short_class, test_case)]['error_msg']
                xpu_traceback = xpu_short_class_map[(short_class, test_case)]['traceback']

            stock_key = (test_class, test_case)
            if stock_key in stock_case_map:
                stock_status = stock_case_map[stock_key]['status']
                stock_error_msg = stock_case_map[stock_key]['error_msg']
                stock_traceback = stock_case_map[stock_key]['traceback']
            elif stock_short_class_map.get((short_class, test_case)):
                stock_status = stock_short_class_map[(short_class, test_case)]['status']
                stock_error_msg = stock_short_class_map[(short_class, test_case)]['error_msg']
                stock_traceback = stock_short_class_map[(short_class, test_case)]['traceback']

        ws.cell(row_idx, 12, xpu_status if xpu_status else '')
        ws.cell(row_idx, 13, stock_status if stock_status else '')

        if xpu_error_msg:
            ws.cell(row_idx, 8, xpu_error_msg[:3000] if xpu_error_msg else '')
        elif stock_error_msg:
            ws.cell(row_idx, 8, stock_error_msg[:3000] if stock_error_msg else '')

        if xpu_traceback:
            ws.cell(row_idx, 9, xpu_traceback[:3000] if xpu_traceback else '')
        elif stock_traceback and not ws.cell(row_idx, 9).value:
            ws.cell(row_idx, 9, stock_traceback[:3000] if stock_traceback else '')

        if xpu_status or stock_status:
            found_count += 1
            ws.cell(row_idx, 14, '')
        else:
            not_found_count += 1
            issues_needing_llm[issue_id] = {
                'test_file': test_file,
                'test_class': test_class,
                'test_case': test_case,
                'origin_test_file': ws.cell(row_idx, 5).value
            }
            if not test_case:
                ws.cell(row_idx, 14, 'No test case')
            elif not test_class:
                ws.cell(row_idx, 14, 'No test class')
            elif not xpu_status and not stock_status:
                ws.cell(row_idx, 14, 'No test case')

        if i % 500 == 0:
            log(f"    Progress: {i}/{total}")

    log(f"  PASS 1 complete: {found_count} matched, {not_found_count} not found")
    return issues_needing_llm


def parse_steps(steps_arg):
    """Parse step arguments like '1', '1-3', '1 2 3' into sorted list."""
    steps = set()

    for arg in steps_arg:
        arg = arg.strip()
        if '-' in arg:
            try:
                start, end = map(int, arg.split('-'))
                steps.update(range(start, end + 1))
            except ValueError:
                pass
        elif arg.isdigit():
            steps.add(int(arg))

    return sorted(steps)


def print_step_info(steps):
    """Print information about steps being.run"""
    print("\n" + "=" * 60)
    print("Test Cases Processor - Step by Step Runner")
    print("=" * 60)

    print("\nSteps to run:")
    all_llm = all(s in llm_steps for s in steps)
    has_llm = any(s in llm_steps for s in steps)
    fast_mode = not has_llm

    for step in sorted(steps):
        desc = STEPS_DESC.get(step, f"Unknown step {step}")
        llm_marker = " [LLM - SLOW]" if step in llm_steps else ""
        print(f"  {step}. {desc}{llm_marker}")

    if fast_mode:
        print("\nMode: FAST (pattern-based, no LLM calls)")
    elif all_llm:
        print("\nMode: LLM ONLY (all steps require LLM)")
    else:
        print("\nMode: MIXED (some LLM calls required)")

    print(f"Total steps: {len(steps)}")
    print("=" * 60 + "\n")


def run_steps(steps_to_run, input_file=None, save=True):
    """
    Run specified steps of the test cases processor.

    Args:
        steps_to_run: list of step numbers to execute
        input_file: optional path to input Excel file
        save: whether to save results after each step

    Returns:
        tuple: (workbook, issues_needing_llm, issue_duplicated_map)
    """
    start_total = time.time()

    excel_file = input_file or os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx')

    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        return None, None, None

    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)

    if 'Test Cases' not in wb.sheetnames:
        print("ERROR: 'Test Cases' sheet not found in workbook")
        return None, None, None

    ws = wb['Test Cases']
    total_rows = ws.max_row - 1
    print(f"Total test cases: {total_rows}\n")

    issues_needing_llm = None
    issue_duplicated_map = None

    if 1 in steps_to_run:
        print_step_info([1])
        issues_needing_llm = pass1_match_ci_results(ws, os.path.join(RESULT_DIR, 'test_cases_all.xlsx'))
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    if 2 in steps_to_run:
        print_step_info([2])
        pass2_extract_torch_ops(ws)
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    if 3 in steps_to_run:
        print_step_info([3])
        if issues_needing_llm is None:
            print("Note: Running PASS 1 first to get issues needing LLM...")
            issues_needing_llm = pass1_match_ci_results(ws, os.path.join(RESULT_DIR, 'test_cases_all.xlsx'))
        pass3_llm_analysis_for_test_existence(ws, issues_needing_llm)
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    if 4 in steps_to_run:
        print_step_info([4])
        pass4_dependency_rag(ws)
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    if 5 in steps_to_run:
        print_step_info([5])
        issue_duplicated_map = pass5_duplicate_detection(ws)
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    elapsed_total = time.time() - start_total
    print("\n" + "=" * 60)
    print(f"Processing complete in {elapsed_total:.1f}s ({elapsed_total/60:.1f} min)")
    print("=" * 60)

    return wb, issues_needing_llm, issue_duplicated_map


def main():
    parser = argparse.ArgumentParser(
        description='Run Test Cases Processor - Step by Step',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )

    parser.add_argument(
        '--steps', nargs='+', default=None,
        help='Steps to run (e.g., "1", "1-3", "1 2 3", or "all")'
    )
    parser.add_argument(
        '--list', action='store_true',
        help='List all available steps with descriptions'
    )
    parser.add_argument(
        '--input', '-i', type=str, default=None,
        help='Input Excel file path (default: auto-detect from RESULT_DIR)'
    )
    parser.add_argument(
        '--no-save', action='store_true',
        help='Do not save results after each step (for debugging)'
    )
    parser.add_argument(
        '--fast', action='store_true',
        help='Run only fast steps (1, 2, 4, 5) - skip LLM dependent steps'
    )

    args = parser.parse_args()

    if args.list:
        print("\nAvailable Steps:")
        print("-" * 60)
        for step, desc in STEPS_DESC.items():
            llm = " [LLM - SLOW]" if step in llm_steps else ""
            print(f"  {step}. {desc}{llm}")
        print("-" * 60)
        return 0

    if args.fast:
        steps_to_run = [1, 2, 4, 5]
        print("Fast mode: Skipping step 3 (LLM analysis)")
    elif args.steps is None or (len(args.steps) == 1 and args.steps[0].lower() in ['all', 'a']):
        steps_to_run = [1, 2, 3, 4, 5]
        print("Running all steps (full processor)")
    else:
        steps_to_run = parse_steps(args.steps)
        if not steps_to_run:
            print("ERROR: No valid steps specified")
            print("Use --steps 1, --steps 1-3, or --steps 1 2 3")
            return 1

    os.makedirs(RESULT_DIR, exist_ok=True)

    save = not args.no_save
    run_steps(steps_to_run, input_file=args.input, save=save)

    return 0


if __name__ == '__main__':
    sys.exit(main())