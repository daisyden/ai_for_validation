#!/usr/bin/env python3
"""
Test Cases processor module.

Contains all logic for filling in fields of 'Test_Cases' sheet in torch_xpu_ops_issues.xlsx:
- CI results from torch-xpu-ops nightly and stock PyTorch XPU CI (PASS 1 moved to pass1_ci_matcher.py)
- Torch ops extraction from error messages and test names (uses torch-ops-extraction module)
- Test case existence analysis (CUDA/XPU)
- Duplicated issue detection
- Dependency analysis from torch ops

Usage:
    from test_result.Test_Cases.test_cases_processor import process_test_cases_sheet

    wb = openpyxl.load_workbook('torch_xpu_ops_issues.xlsx')
    process_test_cases_sheet(wb)
    wb.save('torch_xpu_ops_issues.xlsx')

Column mapping for Test Cases sheet:
    Col 1: Issue ID
    Col 2: Test Reproducer
    Col 3: Test Type
    Col 4: Test File
    Col 5: Origin Test File
    Col 6: Test Class
    Col 7: Test Case
    Col 8: Error Message
    Col 9: Traceback
    Col 10: torch-ops
    Col 11: dependency
    Col 12: XPU Status
    Col 13: Stock Status
    Col 16: CUDA Case Exist
    Col 17: XPU Case Exist
    Col 18: case_existence_comments
    Col 19: can_enable_on_xpu
    Col 20: duplicated_issue
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import re
import time
import csv
from difflib import SequenceMatcher

try:
    torch_ops_dir = os.path.join(os.path.dirname(__file__), '..', 'torch-ops-extraction')
    torch_ops_module = __import__('sys').path.insert(0, torch_ops_dir)
    from extract_torch_ops import (
        extract_torch_ops as torch_ops_extract_torch_ops,
        extract_torch_ops_with_llm as torch_ops_extract_torch_ops_with_llm,
    )
    TORCH_OPS_MODULE_AVAILABLE = True
    print("[TORCH_OPS] Using torch-ops-extraction module")
except ImportError as e:
    TORCH_OPS_MODULE_AVAILABLE = False
    print(f"[TORCH_OPS] Module not available, using inline functions: {e}")

ROOT_DIR = "/home/daisydeng"
RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")
LOG_FILE = os.path.join(RESULT_DIR, "pipeline.log")


def log(msg, print_also=True):
    """Log message to file and optionally print to console."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {msg}"
    try:
        with open(LOG_FILE, "a") as f:
            f.write(log_msg + "\n")
    except Exception:
        pass
    if print_also:
        print(log_msg)


def parse_failure_content(content):
    """Parse failure message to extract error_msg and traceback until error type."""
    error_msg = ""
    traceback = ""

    if not content:
        return error_msg, traceback

    lines = content.split('\n')

    error_patterns = [
        (r'^RuntimeError', 'RuntimeError'),
        (r'^AssertionError', 'AssertionError'),
        (r'^ValueError', 'ValueError'),
        (r'^TypeError', 'TypeError'),
        (r'^IndexError', 'IndexError'),
        (r'^KeyError', 'KeyError'),
        (r'^ImportError', 'ImportError'),
        (r'^NotImplementedError', 'NotImplementedError'),
        (r'^AttributeError', 'AttributeError'),
        (r'^InductorError', 'InductorError'),
    ]

    exception_raise_patterns = [
        r'\braise\s+(RuntimeError|AssertionError|ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|AttributeError|InductorError)\s*[\(\'"]',
    ]

    error_line_idx = -1
    error_type = None
    last_error_msg = ""

    for idx, line in enumerate(lines):
        stripped = line.strip()
        for pattern, etype in error_patterns:
            if re.match(pattern, stripped):
                error_line_idx = idx
                error_type = etype
                clean_line = re.sub(r'^' + etype + r'[:\s]*', '', stripped)
                error_msg = clean_line[:200]
                break
        if error_line_idx >= 0:
            break
        for ep in exception_raise_patterns:
            if re.search(ep, stripped):
                error_line_idx = idx
                match = re.search(r'raise\s+\w+\s*[\(\'"](.+?)[\'\"]?', stripped)
                if match:
                    last_error_msg = match.group(1).strip()[:200]

    traceback = ""
    if 'Traceback (most recent call last):' in content:
        tb_lines = []
        end_idx = error_line_idx if error_line_idx >= 0 else len(lines)
        for idx, line in enumerate(lines):
            if 'Traceback (most recent call last):' in line:
                for j in range(idx, end_idx + 1):
                    tb_lines.append(lines[j])
                break

        if tb_lines:
            traceback = '\n'.join(tb_lines)
        elif last_error_msg:
            for idx, line in enumerate(lines):
                stripped = line.strip()
                for ep in exception_raise_patterns:
                    if re.search(ep, stripped):
                        traceback = '\n'.join(lines[idx:])
                        break
                if traceback:
                    break
    else:
        traceback = ""
        error_msg = content[:200]

    if last_error_msg and not error_msg:
        error_msg = last_error_msg

    if error_msg and traceback and error_msg not in traceback:
        traceback += f"\n{error_msg}"

    return error_msg, traceback[:3000] if traceback else traceback


def extract_torch_ops_with_llm(test_file, test_case, error_msg, traceback):
    """LLM-based torch ops extraction using torch-ops-extraction module.
    Falls back to inline implementation if module unavailable.
    Returns list of torch ops and elapsed time.
    """
    if TORCH_OPS_MODULE_AVAILABLE:
        return torch_ops_extract_torch_ops_with_llm(test_file, test_case, error_msg, traceback)
    
    import requests
    import time as time_module

    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"

    context_parts = []
    if test_file:
        context_parts.append(f"Test file: {test_file}")
    if test_case:
        context_parts.append(f"Test case: {test_case}")
    if error_msg:
        error_sample = str(error_msg)[:1500] if len(str(error_msg)) > 1500 else str(error_msg)
        context_parts.append(f"Error message: {error_sample}")
    if traceback:
        traceback_sample = str(traceback)[:1000] if len(str(traceback)) > 1000 else str(traceback)
        context_parts.append(f"Traceback: {traceback_sample}")

    context = '\n'.join(context_parts)

    prompt = f"""You are a PyTorch expert. Extract the torch operations involved in this test failure.

Context:
{context}

Common torch ops to identify:
- aten ops: add, mm, bmm, matmul, conv2d, softmax, layernorm, gelu, dropout, linear, embedding, etc.
- torch ops: torch.add, torch.matmul, torch.nn.functional.*, torch.linalg.*, torch.fft.*, etc.
- aten.*_default ops: _flash_attention_forward, _scaled_mm, _convolution_forward, etc.

Return ONLY a JSON list of torch operation names, e.g.:
["aten.add", "aten.mm", "aten.conv2d"]

If no specific torch op can be identified, return empty list []."""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "You are a PyTorch operation analysis assistant. Return ONLY valid JSON list."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 1024
    }

    start_time = time_module.time()

    try:
        response = requests.post(
            LLM_ENDPOINT,
            headers=headers,
            json=payload,
            timeout=60
        )
        response.raise_for_status()
        result = response.json()
        elapsed = time_module.time() - start_time

        if result.get('choices') and len(result['choices']) > 0:
            content = result['choices'][0].get('message', {}).get('content', '')
            json_match = re.search(r'\[[^\]]*\]', content, re.DOTALL)
            if json_match:
                ops = json.loads(json_match.group())
                if isinstance(ops, list):
                    ops = [str(op) for op in ops if op]
                    log(f"    [LLM OPS] Extracted {len(ops)} ops in {elapsed:.1f}s")
                    return ops, elapsed

        return [], elapsed

    except Exception as e:
        elapsed = time_module.time() - start_time
        log(f"    [LLM OPS ERROR] {e} (elapsed: {elapsed:.1f}s)")
        return [], elapsed


def clean_op_inline(op):
    """Remove device and dtype suffixes from op name"""
    op = re.sub(r'_(xpu|cuda)_(float\d+|int\d+|complex\d+)$', '', op)
    op = re.sub(r'_(float\d+|int\d+|complex\d+)$', '', op)
    op = re.sub(r'_(xpu|cuda)$', '', op)
    return op


def extract_ops_from_test_name_inline(test_name):
    """Extract ops from test name using OpDB patterns"""
    if not test_name:
        return []

    test_name = str(test_name)

    match = re.search(r'torch_ops_aten__(\w+)', test_name)
    if match:
        return [f'aten._{clean_op_inline(match.group(1))}']

    match = re.search(r'__refs_(\w+)', test_name)
    if match:
        return [f'aten._{clean_op_inline(match.group(1))}']

    match = re.search(r'_nn_(\w+)', test_name)
    if match:
        return [f'nn.{clean_op_inline(match.group(1))}']

    match = re.search(r'_refs_(\w+)', test_name)
    if match:
        return [f'_{clean_op_inline(match.group(1))}']

    match = re.search(r'aten__(\w+)', test_name)
    if match:
        return [f'aten.{clean_op_inline(match.group(1))}']

    if 'fused_attention' in test_name or 'fused_kernel' in test_name:
        return ['aten.fused_attention']
    if 'sdpa' in test_name.lower() or 'sdp' in test_name.lower():
        return ['aten.scaled_dot_product_attention']
    if 'cudnn_attention' in test_name:
        return ['aten.cudnn_attention']
    if 'transformerencoder' in test_name:
        return ['torch.nn.TransformerEncoder']
    if 'transformer' in test_name:
        return ['torch.nn.Transformer']
    if 'flash_attention' in test_name or 'flash_atteention' in test_name:
        return ['aten.flash_attention']
    if 'mem_eff_attention' in test_name:
        return ['aten.memory_efficient_attention']

    match = re.search(r'vjp_linalg_(\w+)', test_name)
    if match:
        return [f'torch.linalg.{clean_op_inline(match.group(1))}']

    if 'csr_matvec' in test_name:
        return ['aten.csr_matvec']
    if 'sparse_csr' in test_name or 'SparseCSR' in test_name:
        return ['aten.sparse_csr']
    if 'to_sparse' in test_name:
        return ['aten.to_sparse']
    if 'sparse' in test_name.lower():
        return ['sparse_ops']

    if 'rms_norm_decomp' in test_name:
        return ['aten.rms_norm']
    if '_fft_' in test_name or 'fft_' in test_name:
        return ['torch.fft']

    if 'has_decomposition' in test_name:
        return ['decomp_ops']

    return []


def extract_from_error_or_traceback_inline(text):
    """Extract torch ops from error message or traceback"""
    if not text:
        return []

    text = str(text)
    found = []

    matches = re.findall(r'torch\.ops\.aten\.(\w+)\.default', text)
    for m in matches[:3]:
        found.append(f'torch.ops.aten.{m}.default')

    if not found:
        matches = re.findall(r'torch\.ops\.aten\.(\w+)', text)
        for m in matches[:3]:
            found.append(f'torch.ops.aten.{m}')

    if not found:
        matches = re.findall(r'aten::(\w+)', text)
        for m in matches[:3]:
            found.append(f'aten.{m}')

    return list(set(found))


def extract_from_test_case_name_inline(test_case):
    """Extract torch ops from test case name mapping"""
    if not test_case:
        return []

    test_case = str(test_case)
    name = test_case
    name = re.sub(r'^test_out_', '', name)
    name = re.sub(r'^test_quick_', '', name)
    name = re.sub(r'^test_comprehensive_', '', name)
    name = re.sub(r'^test_error_', '', name)
    name = re.sub(r'^test_noncontiguous_samples_', '', name)
    name = re.sub(r'^test_neg_view_', '', name)
    name = re.sub(r'_xpu.*$', '', name)
    name = re.sub(r'_cuda.*$', '', name)

    op_mappings = {
        'addmv': 'torch.addmv', 'addmm': 'torch.addmm', 'bmm': 'torch.bmm',
        'matmul': 'torch.matmul', 'dot': 'torch.dot', 'mm': 'torch.mm', 'mv': 'torch.mv',
        'conv2d': 'torch.nn.functional.conv2d', 'conv_transpose2d': 'torch.nn.functional.conv_transpose2d',
        'conv_transpose3d': 'torch.nn.functional.conv_transpose3d',
        'cross_entropy': 'torch.nn.functional.cross_entropy',
        'logaddexp': 'torch.logaddexp', 'histogram': 'torch.histogram',
        'linalg_tensorsolve': 'torch.linalg.tensorsolve',
        'baddbmm': 'aten.baddbmm', 'logspace': 'torch.logspace', 'linspace': 'torch.linspace',
        'arange': 'torch.arange', 'range': 'torch.range',
        'ones': 'torch.ones', 'zeros': 'torch.zeros', 'full': 'torch.full',
        'empty': 'torch.empty', 'rand': 'torch.rand', 'randn': 'torch.randn',
        'randint': 'torch.randint', 'tensor': 'torch.tensor', 'tensor_split': 'torch.tensor_split',
        'sum': 'torch.sum', 'mean': 'torch.mean', 'prod': 'torch.prod',
        'neg': 'torch.neg', 'abs': 'torch.abs',
        'exp': 'torch.exp', 'log': 'torch.log', 'sqrt': 'torch.sqrt',
        'sin': 'torch.sin', 'cos': 'torch.cos', 'tan': 'torch.tan',
        'tanh': 'torch.tanh', 'sigmoid': 'torch.sigmoid',
        'view': 'torch.view', 'reshape': 'torch.reshape', 'flatten': 'torch.flatten',
        'squeeze': 'torch.squeeze', 'unsqueeze': 'torch.unsqueeze',
        'transpose': 'torch.transpose', 'perm': 'torch.permute',
    }

    for key, op in op_mappings.items():
        if key in name:
            return [op]

    return []


def extract_torch_ops(test_file, test_case, error_msg, traceback, use_llm_fallback=True):
    """Extract torch ops using torch-ops-extraction module.
    Uses pattern-based extraction from module, which includes its own fallback.
    Falls back to inline pattern matching if module unavailable.
    Returns: (ops_list, llm_elapsed_time_or_None)
    """
    import time as time_module

    if TORCH_OPS_MODULE_AVAILABLE:
        found_ops, llm_elapsed = torch_ops_extract_torch_ops(test_file, test_case, error_msg, traceback, use_llm_fallback)
        return found_ops if found_ops else [], llm_elapsed

    found_ops = []

    if error_msg:
        extracted = extract_from_error_or_traceback_inline(error_msg)
        if extracted:
            return extracted, None

    if test_case:
        extracted = extract_ops_from_test_name_inline(test_case)
        if extracted:
            return extracted, None

    if test_case:
        extracted = extract_from_test_case_name_inline(test_case)
        if extracted:
            return extracted, None

    if traceback:
        extracted = extract_from_error_or_traceback_inline(traceback)
        if extracted:
            return extracted, None

    return found_ops if found_ops else [], None


def load_ops_dependency():
    """Load ops_dependency.csv into a list for RAG matching"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    rag_dir = os.path.join(script_dir, '..', 'rag')
    ops_dep_file = os.path.join(rag_dir, 'ops_dependency.csv')
    ops_dep_list = []
    if os.path.exists(ops_dep_file):
        with open(ops_dep_file, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                torch_op = row.get('torch_op', '').strip()
                dependency = row.get('dependency', '').strip() if row.get('dependency') else ''
                if torch_op and dependency and dependency.lower() != 'none':
                    ops_dep_list.append((torch_op, dependency))
    return ops_dep_list


def get_dependency_from_ops_rag(ops_list, ops_dep_list):
    """
    RAG-based dependency matching using ops_dependency.csv as knowledge base.
    Score each op against the dependency mapping with multiple strategies.
    Returns unique dependencies matched with high confidence.
    """
    if not ops_list or len(ops_list) == 0:
        return []

    SCORE_EXACT = 100
    SCORE_ATEN_PREFIX = 95
    SCORE_CONTAIN = 80
    SCORE_WORD_OVERLAP = 70
    SCORE_SEQMATCH = 50
    THRESHOLD = 50

    def score_op_match(op, csv_op):
        """Score how well op matches csv_op (higher is better)"""
        op_clean = op.lower().replace('aten::', '').replace('aten.', '').strip()
        csv_clean = csv_op.lower().replace('aten::', '').replace('aten.', '').strip()

        if op_clean == csv_clean:
            return SCORE_EXACT
        if op_clean in csv_clean or csv_clean in op_clean:
            return SCORE_CONTAIN

        op_words = set(re.split(r'[_\.\-]', op_clean))
        csv_words = set(re.split(r'[_\.\-]', csv_clean))
        common = op_words & csv_words
        if common:
            overlap = len(common) / max(len(op_words), len(csv_words))
            if overlap > 0.5:
                return SCORE_WORD_OVERLAP

        ratio = SequenceMatcher(None, op_clean, csv_clean).ratio()
        if ratio > 0.7:
            return int(ratio * SCORE_SEQMATCH)

        return 0

    matches = []
    for op in ops_list:
        op_stripped = op.strip()
        for csv_op, dependency in ops_dep_list:
            score = score_op_match(op_stripped, csv_op)
            if score >= THRESHOLD:
                matches.append((op_stripped, dependency, score))

    matches.sort(key=lambda x: -x[2])
    deps = set()
    for op, dependency, score in matches:
        deps.add(dependency)

    return sorted(list(deps))


def get_dependency_from_ops(ops_list, ops_dep_dict):
    """Get unique dependencies from a list of torch ops - exact matching"""
    deps = set()
    for op in ops_list:
        op = op.strip()
        if op in ops_dep_dict:
            deps.add(ops_dep_dict[op])
        op_clean = op.replace('aten::', '').replace('aten.', '').strip()
        if op_clean and op_clean in ops_dep_dict:
            deps.add(ops_dep_dict[op_clean])
    return sorted(list(deps))


def analyze_test_case_with_llm(test_file, test_class, test_case, origin_test_file):
    """
    Use LLM skills to check CUDA and XPU test case existence.
    Calls check-cuda-test-existence and check-xpu-test-existence skills via opencode.
    
    Note: This requires opencode to be available and pytorch skills to be loaded.
    """
    import subprocess
    import json
    
    if not test_file or not test_case:
        return {
            'cuda_exists': 'No',
            'xpu_exists': 'No',
            'explanation': 'No test file or test case',
            'cuda_decorators': [],
            'xpu_decorators': []
        }
    
    pytorch_root = os.path.expanduser('~/pytorch')
    if not os.path.exists(pytorch_root):
        pytorch_root = os.path.expanduser('~/issue_traige/pytorch')
    
    prompt = f"""You are in the pytorch directory: {pytorch_root}

Use the check-cuda-test-existence skill to check if the CUDA test exists in the original PyTorch test file.
Use the check-xpu-test-existence skill to check if the XPU test exists in torch-xpu-ops repo.

Paths:
- PyTorch test files: {pytorch_root}/test/
- torch-xpu-ops test files: {pytorch_root}/third_party/torch-xpu-ops/test/xpu/

Test File: {test_file}
Origin Test File: {origin_test_file if origin_test_file else 'Not provided'}
Test Class: {test_class}
Test Case: {test_case}

IMPORTANT: The base test name is NOT just removing '_xpu' suffix. The base test is the actual test function in the test file that can be parameterized to generate the XPU test case.

IMPORTANT: In the explanation, you MUST explain WHY the XPU test does not exist if cuda_exists is "No" or xpu_exists is "No". The reasons can be:
1. SKIP DECORATORS: Test has decorators like @onlyCUDA, @skipCUDAIfNoHipdnn, @skipIfXpu, @requires_xccl
2. PARAMETERIZATION: Test is generated from a parameterized base test
3. REMOVED/RENAMED: Test was removed or renamed in newer PyTorch versions
4. NOT APPLICABLE: Test is specific to CUDA/ROCm hardware
5. OTHER: Other reasons

For each check, provide:
1. Whether CUDA test exists (Yes/No)
2. Whether XPU test exists (Yes/N/A)
3. Key decorators found
4. Base test name
5. CUDA test file path if found
6. XPU test file path if found
7. CUDA test name found
8. XPU test name found
9. Detailed explanation

Return ONLY valid JSON format (no additional text):
{{
    "cuda_exists": "Yes/No",
    "xpu_exists": "Yes/No/N/A", 
    "cuda_decorators": ["decorator1", "decorator2"],
    "xpu_decorators": ["decorator1"],
    "base_test_name": "original_test_function_name",
    "cuda_test_file": "path/to/test_file.py",
    "xpu_test_file": "path/to/test_xpu.py", 
    "cuda_test_name": "test_name_found",
    "xpu_test_name": "test_name_found",
    "explanation": "detailed explanation"
}}
"""
    
    try:
        result = subprocess.run(
            ['opencode', 'run', '-m', 'opencode/gpt-5-nano', prompt],
            capture_output=True,
            text=True,
            timeout=180,
            cwd=pytorch_root
        )
        
        if result.returncode == 0:
            output = result.stdout.strip()
            try:
                json_match = re.search(r'\{[^{}]*\}', output, re.DOTALL)
                if json_match:
                    data = json.loads(json_match.group())
                    return data
                if output.startswith('{') and output.endswith('}'):
                    data = json.loads(output)
                    return data
            except (json.JSONDecodeError, AttributeError) as e:
                pass
            return {
                'cuda_exists': 'Unknown',
                'xpu_exists': 'Unknown',
                'cuda_decorators': [],
                'xpu_decorators': [],
                'explanation': f"Output: {output[:500]}"
            }
        else:
            return {
                'cuda_exists': 'Error',
                'xpu_exists': 'Error',
                'cuda_decorators': [],
                'xpu_decorators': [],
                'explanation': f"LLM error (rc={result.returncode}): {result.stderr[:300]}"
            }
    except subprocess.TimeoutExpired:
        return {
            'cuda_exists': 'Timeout',
            'xpu_exists': 'Timeout',
            'cuda_decorators': [],
            'xpu_decorators': [],
            'explanation': 'LLM analysis timed out after 240s'
        }
    except FileNotFoundError:
        return {
            'cuda_exists': 'Error',
            'xpu_exists': 'Error',
            'cuda_decorators': [],
            'xpu_decorators': [],
            'explanation': 'opencode command not found'
        }
    except Exception as e:
        return {
            'cuda_exists': 'Error',
            'xpu_exists': 'Error',
            'cuda_decorators': [],
            'xpu_decorators': [],
            'explanation': f"Failed to run LLM: {str(e)}"
        }


def analyze_test_case(test_file, test_case):
    """
    Analyze why test case exists or not (basic file existence check).
    Uses simple file-based analysis for fast detection.
    """
    result = {
        'cuda_file': None, 'cuda_exists': 'No',
        'xpu_file': None, 'xpu_exists': 'No',
        'explanation': '', 'expected_name': test_case
    }
    
    if not test_file:
        result['explanation'] = 'No test file'
        return result
    
    test_file = str(test_file)
    
    if '/' in test_file:
        cuda_rel = test_file.replace('torch-xpu-ops/test/xpu/', '').replace('_xpu', '')
        xpu_rel = test_file.replace('torch-xpu-ops/test/xpu/', '')
    elif '.' in test_file:
        parts = test_file.split('.')
        parts_len = len(parts)
        
        if parts_len == 4:
            cuda_rel = f"{parts[1]}/{parts[2]}.py"
            xpu_rel = f"{parts[1]}/{parts[2]}_xpu.py"
            result['expected_name'] = parts[3]
        elif parts_len == 3:
            if parts[0] == 'test':
                cuda_rel = f"{parts[1]}.py"
                xpu_rel = f"{parts[1]}_xpu.py"
            elif parts[0] == 'inductor':
                cuda_rel = f"inductor/{parts[1]}.py"
                xpu_rel = None
        elif parts_len == 2:
            cuda_rel = f"{parts[0].replace('_xpu', '')}.py"
            xpu_rel = f"{parts[0]}.py"
        else:
            return result
    else:
        return result
    
    pytorch_test_dir = os.path.expanduser('~/pytorch/test')
    torch_xpu_ops_dir = os.path.expanduser('~/pytorch/third_party/torch-xpu-ops/test/xpu')
    
    xpu_path = os.path.join(torch_xpu_ops_dir, xpu_rel) if xpu_rel else None
    xpu_content = None
    uses_xpu_patch = False
    
    if xpu_path and os.path.exists(xpu_path):
        with open(xpu_path, 'r') as f:
            xpu_content = f.read()
        uses_xpu_patch = 'XPUPatchForImport' in xpu_content
    
    cuda_path = os.path.join(pytorch_test_dir, cuda_rel) if cuda_rel else None
    cuda_content = None
    
    if cuda_path and os.path.exists(cuda_path):
        with open(cuda_path, 'r') as f:
            cuda_content = f.read()
    
    base_name = test_case.replace('_xpu', '') if test_case else None
    cuda_name = test_case.replace('_xpu', '_cuda') if test_case else None
    
    if cuda_content:
        result['cuda_file'] = cuda_rel
        result['cuda_exists'] = 'Yes'
        
        if test_case and test_case in cuda_content:
            result['explanation'] += 'CUDA: Found. '
        elif base_name and base_name in cuda_content:
            result['explanation'] += f"CUDA: Found (base: {base_name}). "
        elif cuda_name and cuda_name in cuda_content:
            result['explanation'] += f"CUDA: Found (cuda: {cuda_name}). "
        else:
            result['explanation'] += 'CUDA: Test not found in pytorch/test. '
            result['cuda_exists'] = 'No'
    else:
        result['cuda_file'] = f'Not found: {cuda_rel}'
        result['explanation'] += 'CUDA file missing. '
    
    uses_xpu_patch = xpu_content and 'XPUPatchForImport' in xpu_content
    
    if uses_xpu_patch:
        if cuda_content:
            if test_case in cuda_content:
                result['explanation'] += 'XPU: Found (imported from CUDA). '
            elif base_name and base_name in cuda_content:
                result['explanation'] += f"XPU: Found (imported from CUDA, base name). "
            elif cuda_name and cuda_name in cuda_content:
                result['explanation'] += f"XPU: Found (imported from CUDA, cuda name). "
            else:
                result['explanation'] += 'XPU: Imported but test not found in CUDA. '
        else:
            result['explanation'] += 'XPU: Uses XPUPatchForImport but CUDA file missing. '
    elif xpu_path and os.path.exists(xpu_path):
        result['xpu_file'] = xpu_rel
        result['xpu_exists'] = 'Yes'
        
        if test_case in xpu_content:
            result['explanation'] += 'XPU: Found in XPU file. '
        elif base_name and base_name in xpu_content:
            result['explanation'] += 'XPU: Found without _xpu. '
        else:
            result['explanation'] += 'XPU: Test not in file. '
    elif xpu_rel is None:
        result['xpu_file'] = 'N/A'
        result['explanation'] += 'XPU: Inductor tests use stock CI. '
    else:
        result['xpu_file'] = f'Not found: {xpu_rel}'
        result['explanation'] += 'XPU file missing. '
    
    return result


def analyze_test_case_with_llm_qwen(test_file, test_class, test_case, origin_test_file=None):
    """
    Use Qwen3-32B via internal API to check CUDA and XPU test case existence.
    Returns: dict with test existence info and measures elapsed time.
    Used for double "not found" cases analysis.
    """
    import requests
    import json
    import time
    
    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"
    
    pytorch_root = os.path.expanduser('~/pytorch')
    if not os.path.exists(pytorch_root):
        pytorch_root = os.path.expanduser('~/issue_traige/pytorch')
    
    prompt = f"""You are in the pytorch directory: {pytorch_root}

Check if the CUDA and XPU test exists in respective test files.

Paths:
- PyTorch test files: {pytorch_root}/test/
- torch-xpu-ops test files: {pytorch_root}/third_party/torch-xpu-ops/test/xpu/

Test File: {test_file}
Origin Test File: {origin_test_file if origin_test_file else 'Not provided'}
Test Class: {test_class}
Test Case: {test_case}

Base test is the actual test function in the test file.

IMPORTANT: Determine if this test case can be enabled on XPU (can_enable_on_xpu):
- can_enable_on_xpu = True if test has hardcoded 'cuda' device or skip decorators (just skips)
- can_enable_on_xpu = False if test is for HIPDNN/rocBlas/ROCm or uses @onlyCUDA, @skipIfXpu

IMPORTANT: Explain WHY XPU test does not exist if cuda_exists is "No" or xpu_exists is "No":
1. SKIP DECORATORS: @onlyCUDA, @skipCUDAIfNoHipdnn, @skipIfXpu, @requires_xccl
2. PARAMETERIZATION: @dtypes, @parametrize_test generating tests
3. REMOVED/RENAMED: Test removed/renamed in newer versions
4. NOT APPLICABLE: CUDA/ROCm specific (hipdnn backend)
5. OTHER: Other reasons

Return ONLY valid JSON:
{{
    "explanation": "detailed explanation why XPU test exists or not"
    "cuda_exists": "Yes/No",
    "xpu_exists": "Yes/No/N/A",
    "can_enable_on_xpu": "True/False",
    "cuda_decorators": ["decorator1"],
    "xpu_decorators": ["decorator1"],
    "base_test_name": "original_test_function_name",
    "cuda_test_file": "path/to/test_file.py",
    "xpu_test_file": "path/to/test_xpu.py",
    "cuda_test_name": "test_name_found",
    "xpu_test_name": "test_name_found",
}}
"""
    
    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "system", "content": "You are a PyTorch test analysis assistant. Return ONLY valid JSON."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 4096
    }
    
    start_time = time.time()
    
    try:
        response = requests.post(
            LLM_ENDPOINT,
            headers=headers,
            json=payload,
            timeout=180
        )
        response.raise_for_status()
        result = response.json()
        elapsed = time.time() - start_time
        
        if result.get('choices') and len(result['choices']) > 0:
            content = result['choices'][0].get('message', {}).get('content', '')
            json_match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
            if json_match:
                data = json.loads(json_match.group())
                data['elapsed_time'] = elapsed
                return data
        
        return {
            'explanation': f'API response parsing failed: {str(result)[:200]}',
            'cuda_exists': 'Unknown',
            'xpu_exists': 'Unknown',
            'elapsed_time': elapsed
        }
        
    except requests.exceptions.Timeout:
        elapsed = time.time() - start_time
        return {
            'explanation': 'Request timed out after 180s',
            'cuda_exists': 'Timeout',
            'xpu_exists': 'Timeout',
            'elapsed_time': elapsed
        }
    except Exception as e:
        elapsed = time.time() - start_time
        return {
            'explanation': f'API error: {str(e)}',
            'cuda_exists': 'Error',
            'xpu_exists': 'Error',
            'elapsed_time': elapsed
        }


def find_duplicated_issues(ws):
    """Find duplicated issues based on Test Class + Test Case or similar Traceback"""
    from collections import defaultdict
    
    class_case_index = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        issue_id = ws.cell(row, 1).value
        if test_class and test_case:
            key = (test_class, test_case)
            class_case_index[key].append((row, issue_id))
    
    traceback_index = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        traceback = ws.cell(row, 9).value or ''
        issue_id = ws.cell(row, 1).value
        
        if 'AssertionError: Tensor-likes are not close!' in traceback:
            continue
        
        norm = traceback.strip()
        if norm:
            traceback_index[norm].append((row, issue_id))
    
    class_case_duplicates = {}
    for key, rows in class_case_index.items():
        if len(rows) > 1:
            issue_ids = [rid for _, rid in rows]
            for row, rid in rows:
                other_issues = [i for i in issue_ids if i != rid]
                if other_issues:
                    class_case_duplicates[row] = other_issues
    
    traceback_duplicates = {}
    for key, rows in traceback_index.items():
        if len(rows) > 1:
            issue_ids = [rid for _, rid in rows]
            for row, rid in rows:
                other_issues = [i for i in issue_ids if i != rid]
                if other_issues:
                    if row in traceback_duplicates:
                        traceback_duplicates[row].extend(other_issues)
                    else:
                        traceback_duplicates[row] = other_issues
    
    merged_duplicates = {}
    for row in range(2, ws.max_row + 1):
        dup_set = set()
        if row in class_case_duplicates:
            dup_set.update(class_case_duplicates[row])
        if row in traceback_duplicates:
            dup_set.update(traceback_duplicates[row])
        if dup_set:
            merged_duplicates[row] = sorted(list(dup_set))
    
    return merged_duplicates


def ensure_headers(ws, col_indices, col_names):
    """Ensure column headers exist for given columns."""
    for col_idx, header_name in zip(col_indices, col_names):
        existing = ws.cell(row=1, column=col_idx).value
        if not existing:
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")


def pass2_extract_torch_ops(ws):
    """
    Step 2: Torch-ops extraction using torch-ops-extraction module.

    Uses pattern-based extraction first, with LLM fallback for unmatched cases.

    Updates:
        Col 10: torch-ops (comma-separated list)
    """
    from collections import defaultdict
    import time as time_module

    ensure_headers(ws, [10], ["torch-ops"])

    log("  [PASS 2/5] Extracting torch ops from test cases and error messages...")
    log("-" * 80)
    start_time = time.time()

    ops_extracted_count = 0
    llm_fallback_count = 0
    llm_total_time = 0
    pattern_matched_count = 0

    total_rows = ws.max_row - 1
    processed = 0

    for row in range(2, ws.max_row + 1):
        test_file = ws.cell(row, 4).value
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        error_msg = ws.cell(row, 12).value
        traceback = ws.cell(row, 9).value
        existing_ops = ws.cell(row, 10).value
        issue_id = ws.cell(row, 1).value

        processed += 1

        if existing_ops:
            continue

        extract_start = time_module.time()
        ops, llm_elapsed = extract_torch_ops(test_file, test_case, error_msg, traceback)
        extract_time = time_module.time() - extract_start

        if ops:
            is_llm = llm_elapsed is not None
            method = "LLM" if is_llm else "PATTERN"
            ws.cell(row, 10, ','.join(ops))
            ops_extracted_count += 1

            if is_llm:
                llm_fallback_count += 1
                llm_total_time += llm_elapsed
            else:
                pattern_matched_count += 1

            log(f"  [PASS2] Issue:{issue_id} | Class:{test_class} | Case:{test_case}")
            log(f"           Ops:{','.join(ops)} | Time:{extract_time:.3f}s | Method:{method}")

        if processed % 100 == 0:
            log(f"    Progress: {processed}/{total_rows}")

    elapsed = time.time() - start_time
    log("-" * 80)
    log(f"  PASS 2 complete: Extracted ops for {ops_extracted_count} cases ({elapsed:.1f}s)")
    log(f"    Pattern matched: {pattern_matched_count}")
    log(f"    LLM fallback: {llm_fallback_count}, avg time: {llm_total_time/llm_fallback_count:.2f}s" if llm_fallback_count > 0 else "    LLM fallback: 0")


def pass3_llm_analysis_for_test_existence(ws, issues_needing_llm):
    """
    PASS 3: LLM analysis for test existence (CUDA/XPU case existence).

    For issues where CI results were "not found", uses LLM to determine
    if CUDA/XPU test cases exist and can be enabled.

    Updates:
        Col 16: CUDA Case Exist
        Col 17: XPU Case Exist
        Col 18: case_existence_comments
        Col 19: can_enable_on_xpu
    """
    import time as time_module

    ensure_headers(ws, [16, 17, 18, 19], ["CUDA Case Exist", "XPU Case Exist", "case_existence_comments", "can_enable_on_xpu"])

    log("  [PASS 3/5] Running LLM analysis for test existence (CUDA/XPU)...")
    start_time = time_module.time()
    
    issue_count = len(issues_needing_llm)
    llm_results = {}
    
    for idx, (issue_id, info) in enumerate(issues_needing_llm.items(), 1):
        try:
            test_file = info['test_file']
            test_class = info['test_class']
            test_case = info['test_case']
            origin_test_file = info['origin_test_file']
            
            result = analyze_test_case_with_llm_qwen(test_file, test_class, test_case, origin_test_file)
            elapsed = result.get('elapsed_time', 0)
            llm_results[issue_id] = result
            log(f"    [LLM {idx}/{issue_count}] Issue {issue_id}: {elapsed:.1f}s")
        except Exception as e:
            log(f"    [LLM ERROR] Issue {issue_id}: {e}")
            llm_results[issue_id] = {'error': str(e)}
    
    elapsed = time_module.time() - start_time
    log(f"  PASS 3 complete: {len(llm_results)} LLM results ({elapsed:.1f}s)")
    
    log("  [APPLY] Writing LLM results to test cases...")
    applied = 0
    skipped = 0
    
    for row in range(2, ws.max_row + 1):
        issue_id = ws.cell(row, 1).value
        xpu_status = ws.cell(row, 12).value
        stock_status = ws.cell(row, 13).value
        ci_not_found = (xpu_status == 'not found' or not xpu_status) and (stock_status == 'not found' or stock_status == 'not in stock CI')
        
        if ci_not_found and issue_id in llm_results:
            result = llm_results[issue_id]
            
            if result.get('cuda_exists') or result.get('xpu_exists'):
                ws.cell(row, 16, result.get('cuda_exists', 'Unknown'))  # Col 16: CUDA Case Exist
                ws.cell(row, 17, result.get('xpu_exists', 'Unknown'))    # Col 17: XPU Case Exist
                
                parts = []
                explanation = result.get('explanation', '')
                if explanation:
                    parts.append('explanation: ' + explanation)
                
                for key in ['base_test_name', 'cuda_test_file', 'xpu_test_file', 'cuda_test_name', 'xpu_test_name']:
                    val = result.get(key)
                    if val:
                        parts.append(f'{key}:{val}')
                
                comment = '\n'.join(parts) if parts else 'Double not found - LLM analysis'
                ws.cell(row, 18, comment)  # Col 18: case_existence_comments
                ws.cell(row, 19, result.get('can_enable_on_xpu', 'Unknown'))  # Col 19: can_enable_on_xpu
                applied += 1
            else:
                skipped += 1
        else:
            skipped += 1
    
    log(f"  Applied to {applied} cases, skipped {skipped}")


def pass4_dependency_rag(ws):
    """
    Step 4: Dependency RAG - match ops to deps from ops_dependency.csv.

    Uses RAG to find dependencies for extracted torch ops.

    Updates:
        Col 11: dependency
    """
    ensure_headers(ws, [11], ["dependency"])

    log("  [PASS 4/5] Populating dependency using RAG...")
    start_time = time.time()
    
    ops_dep_list = load_ops_dependency()
    dep_count = 0
    
    for row in range(2, ws.max_row + 1):
        torch_ops_raw = ws.cell(row, 10).value
        if not torch_ops_raw:
            continue
        ops_list = str(torch_ops_raw).split(',')
        dependencies = get_dependency_from_ops_rag(ops_list, ops_dep_list)
        if dependencies:
            ws.cell(row, 11, ';'.join(dependencies))  # Col 11: dependency
            dep_count += 1
    
    elapsed = time.time() - start_time
    log(f"  PASS 4 complete: Set dependency_lib for {dep_count} cases ({elapsed:.1f}s)")


def process_test_cases_sheet(wb):
    """
    Process Test Cases sheet - fills all CI result and analysis columns.

    NOTE: PASS 1 has been moved to pass1_ci_matcher.py for better modularity.
    This function now starts from PASS 2.

    Steps (PASS 1 is now in pass1_ci_matcher.py):
        1. PASS 1: Match CI results from test_cases_all.xlsx (moved to pass1_ci_matcher.py)
        2. PASS 2: Torch-ops extraction (pattern + LLM fallback)
        3. PASS 3: LLM analysis for test existence (CUDA/XPU)
        4. PASS 4: Dependency RAG (match ops to deps)
        5. PASS 5: Duplicate detection (cross-issue)

    Columns added (10-13, 16-23 via PASS 1 + passes 2-5):
        - Col 8: Error Message (PASS 1)
        - Col 9: Traceback (PASS 1)
        - Col 10: torch-ops (PASS 2)
        - Col 11: dependency (PASS 4)
        - Col 12: XPU Status (PASS 1)
        - Col 13: Stock Status (PASS 1)
        - Col 14: No Match Reason (PASS 1)
        - Col 16: CUDA Case Exist (PASS 3)
        - Col 17: XPU Case Exist (PASS 3)
        - Col 18: case_existence_comments (PASS 3)
        - Col 19: can_enable_on_xpu (PASS 3)
        - Col 20: duplicated_issue (PASS 5)
    """
    import time as time_module
    
    ws = wb['Test Cases']
    
    log("Test Cases processor - PASS 1 moved to pass1_ci_matcher.py")
    log("Processing passes 2-5 only (PASS 1 must be run first via pass1_ci_matcher.py)")
    
    total = ws.max_row - 1
    log(f"\n[STEP] Processing {total} test cases (passes 2-5)...")
    log(f"  [START TIME] {time_module.strftime('%Y-%m-%d %H:%M:%S')}")
    
    pass2_extract_torch_ops(ws)
    
    pass3_llm_analysis_for_test_existence(ws, {})
    
    pass4_dependency_rag(ws)
    
    issue_duplicated_map = pass5_duplicate_detection(ws)
    
    log("Test Cases sheet processed successfully!")
    return ws, issue_duplicated_map