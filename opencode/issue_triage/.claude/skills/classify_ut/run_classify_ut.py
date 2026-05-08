#!/usr/bin/env python3
# pyright: reportMissingModuleSource=none

from __future__ import annotations

import argparse
import copy
from collections import Counter, OrderedDict
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook  # type: ignore[import-untyped]


HOME = Path.home()
DEFAULT_TARGET_XLSX = HOME / "opencode/classify/data/Non_inductor_ut_status_ww16_26.xlsx"
REFERENCE_XLSX = (
    HOME / "opencode/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx"
)
NOT_APPLIABLE_TXT = (
    HOME / "opencode/ai_for_validation/opencode/issue_triage/not_appliable.txt"
)
TARGET_SHEET = "Non-Inductor XPU Skip"
REFERENCE_SHEETS = ("Not Appliable", "Not Applicable")
DEEP_ANALYSIS_SKILL = (
    HOME
    / "opencode/ai_for_validation/opencode/issue_triage/.claude/skills/bug_scrub/"
    "analyze_ci_result/check_xpu_case_existence/SKILL.md"
)
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
BLUE_RGB = "FF0000FF"

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)
ET.register_namespace("mc", MC_NS)
ET.register_namespace("x14ac", X14AC_NS)


@dataclass(frozen=True)
class CellUpdate:
    value: object
    kind: str
    preferred_style: int | None = None


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def default_output_path(target_xlsx: Path) -> Path:
    return target_xlsx.with_name(target_xlsx.stem + ".agent" + target_xlsx.suffix)


def col_to_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


def letter_to_col(cell_ref: str) -> int:
    value = 0
    for char in cell_ref:
        if char.isdigit():
            break
        value = value * 26 + (ord(char.upper()) - ord("A") + 1)
    return value


def cell_ref(column: int, row: int) -> str:
    return f"{col_to_letter(column)}{row}"


def locate_or_allocate_column(headers: list[object], column_name: str) -> tuple[int, bool]:
    for idx, value in enumerate(headers, start=1):
        if value == column_name:
            return idx, False
    for idx, value in enumerate(headers, start=1):
        if value is None:
            headers[idx - 1] = column_name
            return idx, True
    headers.append(column_name)
    return len(headers), True


def is_real_case_row(row_values: tuple[object, ...] | list[object]) -> bool:
    return any(value is not None for value in row_values[:23])


def infer_xpu_class(classname_cuda: object, classname_xpu: object) -> object:
    if classname_xpu not in (None, ""):
        return classname_xpu
    if classname_cuda in (None, ""):
        return classname_xpu
    value = str(classname_cuda)
    if value.endswith("CUDA"):
        return value[:-4] + "XPU"
    return value.replace("CUDA", "XPU")


def infer_xpu_name(name_cuda: object, name_xpu: object) -> object:
    if isinstance(name_xpu, str) and "xpugraphs" in name_xpu:
        name_xpu = None
    if name_xpu not in (None, ""):
        return name_xpu
    if name_cuda in (None, ""):
        return name_xpu
    value = str(name_cuda)
    if value.endswith("_cuda"):
        return value[:-5] + "_xpu"
    return value


def infer_xpu_file(testfile_cuda: object, testfile_xpu: object) -> object:
    if testfile_xpu not in (None, ""):
        return testfile_xpu
    return testfile_cuda


def get_reference_sheet(workbook):
    for sheet_name in REFERENCE_SHEETS:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    raise KeyError(f"Missing reference sheet from {REFERENCE_SHEETS}")


def collect_not_applicable_items(target_xlsx: Path) -> list[str]:
    target_wb = load_workbook(target_xlsx, read_only=True)
    target_ws = target_wb[TARGET_SHEET]
    target_headers = [cell.value for cell in next(target_ws.iter_rows(min_row=1, max_row=1))]
    target_idx = {name: index for index, name in enumerate(target_headers) if name is not None}

    items: list[str] = []
    for row in target_ws.iter_rows(min_row=2, values_only=True):
        reason = normalize_text(row[target_idx["Reason"]])
        detail = normalize_text(row[target_idx["DetailReason"]])
        if reason.lower() == "not applicable" and detail:
            items.append(detail)

    reference_wb = load_workbook(REFERENCE_XLSX, read_only=True)
    reference_ws = get_reference_sheet(reference_wb)
    reference_headers = [
        cell.value for cell in next(reference_ws.iter_rows(min_row=1, max_row=1))
    ]
    reference_idx = {
        name: index for index, name in enumerate(reference_headers) if name is not None
    }
    operation_column = "Operation/API"
    if operation_column not in reference_idx:
        operation_column = "Torch Ops/API"

    for row in reference_ws.iter_rows(min_row=2, values_only=True):
        operation = normalize_text(row[reference_idx[operation_column]])
        if operation:
            items.append(operation)

    return list(OrderedDict.fromkeys(sorted(set(items), key=lambda item: item.lower())))


def write_not_appliable_file(items: list[str]) -> None:
    NOT_APPLIABLE_TXT.write_text("\n".join(items) + "\n")


def build_workbook_updates(
    target_xlsx: Path,
) -> tuple[dict[str, CellUpdate], int, Counter[str], list[tuple[int, str, str, str]], bool]:
    wb = load_workbook(target_xlsx, read_only=True, data_only=False, keep_links=False)
    ws = wb[TARGET_SHEET]

    headers: list[object] = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    explanation_col, explanation_added = locate_or_allocate_column(headers, "Exaplaination")
    reason_tbd_col, reason_tbd_added = locate_or_allocate_column(headers, "Reason TBD")
    idx = {name: index + 1 for index, name in enumerate(headers) if name is not None}

    updates: dict[str, CellUpdate] = {}
    dirty = explanation_added or reason_tbd_added
    updates[cell_ref(explanation_col, 1)] = CellUpdate("Exaplaination", "text")
    updates[cell_ref(reason_tbd_col, 1)] = CellUpdate("Reason TBD", "text")

    reason_counter: Counter[str] = Counter()
    pending_rows: list[tuple[int, str, str, str]] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values: list[object] = list(row)
        if not is_real_case_row(row_values):
            continue

        inferred_file = infer_xpu_file(row_values[idx["testfile_cuda"] - 1], row_values[idx["testfile_xpu"] - 1])
        inferred_class = infer_xpu_class(row_values[idx["classname_cuda"] - 1], row_values[idx["classname_xpu"] - 1])
        inferred_name = infer_xpu_name(row_values[idx["name_cuda"] - 1], row_values[idx["name_xpu"] - 1])

        for column_name, inferred_value in (
            ("testfile_xpu", inferred_file),
            ("classname_xpu", inferred_class),
            ("name_xpu", inferred_name),
        ):
            current_value = row_values[idx[column_name] - 1] if idx[column_name] - 1 < len(row_values) else None
            if inferred_value != current_value:
                updates[cell_ref(idx[column_name], row_idx)] = CellUpdate(inferred_value, "text")
                dirty = True

        reason = normalize_text(row_values[idx["Reason"] - 1])
        status_xpu = row_values[idx["status_xpu"] - 1]
        reason_counter[reason or "<blank>"] += 1

        reason_tbd_value = reason == ""
        updates[cell_ref(reason_tbd_col, row_idx)] = CellUpdate(reason_tbd_value, "bool")
        if row_values[reason_tbd_col - 1] != reason_tbd_value if reason_tbd_col - 1 < len(row_values) else True:
            dirty = True

        if status_xpu in (None, "") and reason_tbd_value:
            pending_rows.append(
                (
                    row_idx,
                    normalize_text(row_values[idx["testfile_cuda"] - 1]),
                    normalize_text(row_values[idx["classname_cuda"] - 1]),
                    normalize_text(row_values[idx["name_cuda"] - 1]),
                )
            )

    return updates, len(pending_rows), reason_counter, pending_rows, dirty


def resolve_sheet_xml_path(zf: ZipFile, sheet_name: str) -> str:
    ns = {"main": MAIN_NS, "pkgrel": PKG_REL_NS, "rel": REL_NS}
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rel_root.findall("pkgrel:Relationship", ns)
    }
    sheets_elem = workbook_root.find("main:sheets", ns)
    if sheets_elem is None:
        raise KeyError(f"No <sheets> element in workbook.xml; cannot find {sheet_name!r}")
    for sheet in sheets_elem:
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib[f"{{{REL_NS}}}id"]
        return f"xl/{rel_map[rel_id]}"
    raise KeyError(sheet_name)


def ensure_row(sheet_data, row_index: int):
    row_attr = str(row_index)
    for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
        if row.attrib.get("r") == row_attr:
            return row
        if int(row.attrib["r"]) > row_index:
            new_row = ET.Element(f"{{{MAIN_NS}}}row", {"r": row_attr})
            insert_at = list(sheet_data).index(row)
            sheet_data.insert(insert_at, new_row)
            return new_row
    new_row = ET.Element(f"{{{MAIN_NS}}}row", {"r": row_attr})
    sheet_data.append(new_row)
    return new_row


def ensure_cell(row, ref_name: str):
    for cell in row.findall(f"{{{MAIN_NS}}}c"):
        if cell.attrib.get("r") == ref_name:
            return cell, False
    new_cell = ET.Element(f"{{{MAIN_NS}}}c", {"r": ref_name})
    row.append(new_cell)
    sort_cells(row)
    return new_cell, True


def sort_cells(row) -> None:
    cells = [cell for cell in list(row) if cell.tag == f"{{{MAIN_NS}}}c"]
    others = [child for child in list(row) if child.tag != f"{{{MAIN_NS}}}c"]
    for child in list(row):
        row.remove(child)
    cells.sort(key=lambda cell: letter_to_col(cell.attrib["r"]))
    for child in cells + others:
        row.append(child)


def set_inline_text(cell, value: object) -> None:
    cell.attrib["t"] = "inlineStr"
    for child in list(cell):
        cell.remove(child)
    is_node = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
    text_node = ET.SubElement(is_node, f"{{{MAIN_NS}}}t")
    text_node.text = "" if value is None else str(value)


def set_bool(cell, value: bool) -> None:
    cell.attrib["t"] = "b"
    for child in list(cell):
        cell.remove(child)
    value_node = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
    value_node.text = "1" if value else "0"


def apply_updates_to_copy(output_xlsx: Path, updates: dict[str, CellUpdate]) -> None:
    patch_workbook_copy(output_xlsx, output_xlsx, updates)


def font_signature(font) -> tuple[tuple[str, tuple[tuple[str, str], ...], str | None], ...]:
    signature = []
    for child in list(font):
        signature.append(
            (
                child.tag,
                tuple(sorted(child.attrib.items())),
                child.text,
            )
        )
    return tuple(signature)


def clone_blue_font(font):
    new_font = copy.deepcopy(font)
    for child in list(new_font):
        if child.tag == f"{{{MAIN_NS}}}color":
            new_font.remove(child)
    color = ET.Element(f"{{{MAIN_NS}}}color", {"rgb": BLUE_RGB})
    inserted = False
    for index, child in enumerate(list(new_font)):
        if child.tag in {f"{{{MAIN_NS}}}sz", f"{{{MAIN_NS}}}b", f"{{{MAIN_NS}}}i", f"{{{MAIN_NS}}}u"}:
            continue
        new_font.insert(index, color)
        inserted = True
        break
    if not inserted:
        new_font.append(color)
    return new_font


def prepare_blue_styles(styles_root, base_style_ids: set[int]) -> dict[int, int]:
    fonts = styles_root.find(f"{{{MAIN_NS}}}fonts")
    cell_xfs = styles_root.find(f"{{{MAIN_NS}}}cellXfs")
    if fonts is None or cell_xfs is None:
        raise RuntimeError("Missing style tables")

    font_map = {font_signature(font): index for index, font in enumerate(list(fonts))}
    base_to_blue_font: dict[int, int] = {}
    blue_style_ids: dict[int, int] = {}
    xfs = list(cell_xfs)

    for base_style_id in sorted(base_style_ids):
        xf = xfs[base_style_id]
        base_font_id = int(xf.attrib.get("fontId", "0"))
        if base_font_id not in base_to_blue_font:
            blue_font = clone_blue_font(list(fonts)[base_font_id])
            signature = font_signature(blue_font)
            blue_font_id = font_map.get(signature)
            if blue_font_id is None:
                fonts.append(blue_font)
                blue_font_id = len(list(fonts)) - 1
                font_map[signature] = blue_font_id
            base_to_blue_font[base_font_id] = blue_font_id

        new_xf = copy.deepcopy(xf)
        new_xf.attrib["fontId"] = str(base_to_blue_font[base_font_id])
        new_xf.attrib["applyFont"] = "1"
        cell_xfs.append(new_xf)
        blue_style_ids[base_style_id] = len(list(cell_xfs)) - 1

    fonts.attrib["count"] = str(len(list(fonts)))
    cell_xfs.attrib["count"] = str(len(list(cell_xfs)))
    return blue_style_ids


def patch_workbook_copy(
    target_xlsx: Path,
    output_xlsx: Path,
    updates: dict[str, CellUpdate],
) -> None:
    with ZipFile(target_xlsx) as zf:
        sheet_xml_path = resolve_sheet_xml_path(zf, TARGET_SHEET)
        sheet_root = ET.fromstring(zf.read(sheet_xml_path))
        styles_root = ET.fromstring(zf.read("xl/styles.xml"))

    sheet_data = sheet_root.find(f"{{{MAIN_NS}}}sheetData")
    if sheet_data is None:
        raise RuntimeError("Missing sheetData")

    base_styles: set[int] = set()
    cell_objects = {}
    for ref_name, update in updates.items():
        row_index = int("".join(ch for ch in ref_name if ch.isdigit()))
        row = ensure_row(sheet_data, row_index)
        cell, created = ensure_cell(row, ref_name)
        style_id = int(cell.attrib.get("s", str(update.preferred_style or 0)))
        if created and update.preferred_style is not None:
            style_id = update.preferred_style
        base_styles.add(style_id)
        cell_objects[ref_name] = (cell, style_id, update)

    blue_style_ids = prepare_blue_styles(styles_root, base_styles)
    for cell, base_style_id, update in cell_objects.values():
        cell.attrib["s"] = str(blue_style_ids[base_style_id])
        if update.kind == "text":
            set_inline_text(cell, update.value)
        elif update.kind == "bool":
            set_bool(cell, bool(update.value))
        else:
            raise ValueError(update.kind)

    sheet_bytes = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
    styles_bytes = ET.tostring(styles_root, encoding="utf-8", xml_declaration=True)

    with NamedTemporaryFile(dir=output_xlsx.parent, suffix=output_xlsx.suffix, delete=False) as tmp:
        tmp_path = Path(tmp.name)

    try:
        with ZipFile(target_xlsx) as zin, ZipFile(tmp_path, "w") as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename == sheet_xml_path:
                    data = sheet_bytes
                elif info.filename == "xl/styles.xml":
                    data = styles_bytes
                zout.writestr(info, data)
        tmp_path.replace(output_xlsx)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET_XLSX)
    parser.add_argument("--output", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    target_xlsx = args.target.expanduser().resolve()
    output_xlsx = (
        args.output.expanduser().resolve() if args.output else default_output_path(target_xlsx)
    )

    if not target_xlsx.exists():
        raise FileNotFoundError(target_xlsx)
    if not REFERENCE_XLSX.exists():
        raise FileNotFoundError(REFERENCE_XLSX)

    items = collect_not_applicable_items(target_xlsx)
    write_not_appliable_file(items)
    updates, pending_count, reason_counter, pending_rows, dirty = build_workbook_updates(
        target_xlsx
    )
    patch_workbook_copy(target_xlsx, output_xlsx, updates)

    print(f"target workbook: {target_xlsx}")
    print(f"output workbook: {output_xlsx}")
    print(f"reference workbook: {REFERENCE_XLSX}")
    print(f"not appliable list: {NOT_APPLIABLE_TXT}")
    print(f"workbook_updated: {dirty}")
    print(f"pending_deep_analysis_rows: {pending_count}")
    print("reason counts:")
    for reason, count in reason_counter.most_common():
        print(f"  {reason}: {count}")
    print("not appliable items:")
    for item in items:
        print(f"  - {item}")

    if pending_rows:
        print("pending rows requiring deep case existence analysis:")
        for row_idx, test_file, test_class, test_name in pending_rows:
            print(f"  - row {row_idx}: {test_file} | {test_class} | {test_name}")
        print("run the documented deep analysis workflow before filling Reason/DetailReason/Exaplaination:")
        print(f"  {DEEP_ANALYSIS_SKILL}")


if __name__ == "__main__":
    main()
