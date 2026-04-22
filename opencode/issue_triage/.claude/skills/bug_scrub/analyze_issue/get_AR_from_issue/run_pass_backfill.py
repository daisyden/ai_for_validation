"""Backfill action_TBD / action_reason / action_Type for PASS-validation-status
Phase 4b issues whose verb was never emitted.

Rule (per-issue, apply first that matches):
  - any VERIFIED pr_analysis with merged=True or state in {MERGED,MERGED_PRESUMED}
      -> VERIFY_AND_CLOSE, verb: "Verify fix from merged PR <ref> and close"
  - else any VERIFIED with state=OPEN
      -> TRACK_PR, verb: "Track PR <ref> to merge"
  - else any VERIFIED with state=CLOSED (unmerged)
      -> RETRIAGE_PRS, verb: "PR <ref> closed unmerged; reassess fix path"
"""
import glob
import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[7]
EXCEL = str(REPO / "opencode/issue_triage/result/torch_xpu_ops_issues.xlsx")
RESULTS_GLOB = str(REPO / "agent_space/phase4b/wave*/result_*.json")


def pr_ref(pr: dict) -> str:
    repo = pr.get("repo", "")
    n = pr.get("pr_number")
    short = {"intel/torch-xpu-ops", "pytorch/pytorch"}
    return f"{repo}#{n}" if repo in short else f"{repo}#{n}"


def classify(pra: list) -> tuple[str, str, str]:
    """Return (action_Type, verb, reason)."""
    ver = [p for p in pra if p.get("verdict") == "VERIFIED"]
    merged = [p for p in ver if p.get("merged") or p.get("state") in ("MERGED", "MERGED_PRESUMED")]
    if merged:
        p = merged[0]
        return (
            "VERIFY_AND_CLOSE",
            f"Verify fix from merged PR {pr_ref(p)} and close",
            "Phase 4b backfill: verified merged PR from pr_analysis.",
        )
    opens = [p for p in ver if p.get("state") == "OPEN"]
    if opens:
        p = opens[0]
        return (
            "TRACK_PR",
            f"Track PR {pr_ref(p)} to merge",
            "Phase 4b backfill: verified open PR from pr_analysis.",
        )
    closed = [p for p in ver if p.get("state") == "CLOSED"]
    if closed:
        p = closed[0]
        return (
            "RETRIAGE_PRS",
            f"PR {pr_ref(p)} closed unmerged; reassess fix path",
            "Phase 4b backfill: verified PR closed without merging.",
        )
    raise RuntimeError("no VERIFIED PR but validation_status was PASS?")


def main() -> None:
    # Load Phase 4b results
    backfill: dict[int, tuple[str, str, str]] = {}
    for f in sorted(glob.glob(RESULTS_GLOB)):
        d = json.load(open(f))
        if d.get("validation_status") != "PASS":
            continue
        if d.get("action_TBD"):
            continue
        iid = d["issue_number"]
        backfill[iid] = classify(d.get("pr_analysis", []))

    print(f"Backfill candidates: {len(backfill)}")

    # Apply to Excel
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Issues"]
    hdr = [c.value for c in ws[1]]
    I_ID = hdr.index("Issue ID")
    I_TBD = hdr.index("action_TBD")
    I_RSN = hdr.index("action_reason")
    I_TYP = hdr.index("action_Type")

    applied = 0
    skipped_nonempty = []
    for row in ws.iter_rows(min_row=2):
        iid = row[I_ID].value
        if iid not in backfill:
            continue
        cur_tbd = row[I_TBD].value
        if cur_tbd and str(cur_tbd).strip() not in ("", "[]", "None"):
            skipped_nonempty.append((iid, cur_tbd))
            continue
        cat, verb, reason = backfill[iid]
        row[I_TBD].value = json.dumps([verb])
        row[I_RSN].value = json.dumps([reason])
        row[I_TYP].value = cat
        applied += 1
        print(f"  #{iid:<5} {cat:<16} {verb}")

    wb.save(EXCEL)
    print(f"\nApplied: {applied}")
    if skipped_nonempty:
        print(f"Skipped (already had action_TBD): {len(skipped_nonempty)}")
        for iid, v in skipped_nonempty:
            print(f"  #{iid}: {v!r}")


if __name__ == "__main__":
    main()
