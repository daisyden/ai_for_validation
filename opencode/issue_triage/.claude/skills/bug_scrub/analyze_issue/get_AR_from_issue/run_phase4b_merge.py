"""Merge Phase 4b wave results into Issues sheet.

Loads all 305 result_<id>.json under agent_space/phase4b/wave{1..7}/
and unions their action_TBD / action_reason / owner_transferred arrays
with any values already written by Phase 4a/4c.

Separator for Phase 4b additions is " | " (pipe) because action_reason
sentences contain commas. Existing comma-separated 4a/4c tokens are
preserved intact; dedupe uses substring membership.
"""
import json
import os
from pathlib import Path
from collections import defaultdict

import openpyxl

REPO   = Path(__file__).resolve().parents[7]
EXCEL  = REPO / "opencode/issue_triage/result/torch_xpu_ops_issues.xlsx"
# WAVES dir defaults to <repo>/agent_space/phase4b but can be overridden
# via PHASE4B_WAVES env var (e.g. when results live in a sibling clone).
WAVES  = Path(os.environ.get("PHASE4B_WAVES", REPO / "agent_space/phase4b"))

SEP = " | "

# ---- load all 305 results --------------------------------------------------
results = {}
wave_dirs = sorted(WAVES.glob("wave*"), key=lambda p: int(p.name.replace("wave", "")))
for wave_dir in wave_dirs:
    for f in sorted(wave_dir.glob("result_*.json")):
        with open(f) as fh:
            r = json.load(fh)
        iid = r["issue_number"]
        if iid in results:
            raise RuntimeError(f"duplicate issue {iid}")
        results[iid] = r

print(f"loaded {len(results)} Phase 4b results")

# ---- per-status counters ---------------------------------------------------
by_status = defaultdict(int)
for r in results.values():
    by_status[r.get("validation_status", "?")] += 1
print(f"status breakdown: {dict(by_status)}")

# ---- open Excel ------------------------------------------------------------
wb  = openpyxl.load_workbook(EXCEL)
ws  = wb["Issues"]
hdr = [c.value for c in ws[1]]

def col(name):
    return hdr.index(name)

I_ID  = col("Issue ID")
I_ACT = col("action_TBD")
I_RSN = col("action_reason")
I_OWN = col("owner_transferred")

def merge_cell(cell, additions):
    """Append each addition to cell (pipe-separated) unless already present."""
    cur = (cell.value or "").strip()
    for add in additions:
        if not add:
            continue
        add = add.strip()
        if not add:
            continue
        if cur and add in cur:
            continue
        cur = f"{cur}{SEP}{add}" if cur else add
    return cur or None

# ---- merge -----------------------------------------------------------------
updated = 0
missing = []
for row in ws.iter_rows(min_row=2):
    iid = row[I_ID].value
    if iid is None:
        continue
    r = results.get(iid)
    if r is None:
        continue
    acts = r.get("action_TBD") or []
    rsns = r.get("action_reason") or []
    owns = r.get("owner_transferred") or []
    # Coerce string -> single-element list (some agents emitted a bare string)
    if isinstance(acts, str): acts = [acts]
    if isinstance(rsns, str): rsns = [rsns]
    if isinstance(owns, str): owns = [owns]
    if not (acts or rsns or owns):
        continue
    if acts:
        row[I_ACT].value = merge_cell(row[I_ACT], acts)
    if rsns:
        row[I_RSN].value = merge_cell(row[I_RSN], rsns)
    if owns:
        row[I_OWN].value = merge_cell(row[I_OWN], owns)
    updated += 1

# Any result not matched to an Excel row?
excel_ids = {row[I_ID].value for row in ws.iter_rows(min_row=2)}
missing   = sorted(set(results) - excel_ids)

wb.save(EXCEL)
print(f"merged {updated} issues into Issues sheet")
if missing:
    print(f"WARN: {len(missing)} result IDs not in Excel: {missing[:10]}...")
else:
    print("all 305 result IDs mapped to Excel rows")
