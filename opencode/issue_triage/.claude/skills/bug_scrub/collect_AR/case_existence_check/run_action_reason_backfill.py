"""Phase 4c follow-up: populate action_reason for check_case_avaliablity issues.

For every row in the Issues sheet whose action_TBD contains
`check_case_avaliablity` and whose action_reason is blank, aggregate the
distinct non-empty `case_existence_comments` from the Test Cases sheet
and write them back into action_reason. A single distinct comment is
written as a plain string; multiple distinct comments are written as a
JSON array (matching the existing action_reason conventions).

Usage:
    python3 run_action_reason_backfill.py

Anchors the Excel path via __file__ so it runs from any CWD.
Backs up the workbook to ..._bk_before_action_reason_backfill.xlsx before writing.
"""
from __future__ import annotations

import json
import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[7]
EXCEL = REPO_ROOT / "opencode/issue_triage/result/torch_xpu_ops_issues.xlsx"
BACKUP = EXCEL.with_name(
    EXCEL.stem + "_bk_before_action_reason_backfill.xlsx"
)
TOKEN = "check_case_avaliablity"


def col_idx(ws, name: str) -> int:
    return [c.value for c in ws[1]].index(name)


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    issues = wb["Issues"]
    cases = wb["Test Cases"]

    ci_iid = col_idx(cases, "Issue ID")
    ci_cmnt = col_idx(cases, "case_existence_comments")

    # issue_id -> list of distinct non-empty comments (preserve insertion order)
    comments: dict[int, list[str]] = defaultdict(list)
    for row in cases.iter_rows(min_row=2, values_only=True):
        iid = row[ci_iid]
        c = row[ci_cmnt]
        if iid is None or c is None:
            continue
        s = str(c).strip()
        if not s or s.lower() == "none":
            continue
        if s not in comments[int(iid)]:
            comments[int(iid)].append(s)

    ii_iid = col_idx(issues, "Issue ID")
    ii_tbd = col_idx(issues, "action_TBD")
    ii_rsn = col_idx(issues, "action_reason")

    updated = 0
    skipped_no_blank = 0
    skipped_no_comments = 0
    for row in issues.iter_rows(min_row=2):
        iid_cell = row[ii_iid]
        if iid_cell.value is None:
            continue
        tbd = (row[ii_tbd].value or "").strip()
        if TOKEN not in tbd:
            continue
        existing = (row[ii_rsn].value or "")
        if str(existing).strip():
            skipped_no_blank += 1
            continue
        lst = comments.get(int(iid_cell.value), [])
        if not lst:
            skipped_no_comments += 1
            continue
        if len(lst) == 1:
            row[ii_rsn].value = lst[0]
        else:
            row[ii_rsn].value = json.dumps(lst, ensure_ascii=False)
        updated += 1

    print(
        f"check_case_avaliablity rows updated: {updated}\n"
        f"  skipped (action_reason already set): {skipped_no_blank}\n"
        f"  skipped (no Test Cases comments):    {skipped_no_comments}"
    )

    if updated == 0:
        print("no changes; skipping save")
        return

    shutil.copy(EXCEL, BACKUP)
    print(f"backed up to {BACKUP}")
    wb.save(EXCEL)
    print(f"wrote {EXCEL}")


if __name__ == "__main__":
    main()
