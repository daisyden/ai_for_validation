"""Convert result/bug_scrub.md → result/bug_scrub.html.

Phase 5b of the bug-scrub pipeline. Reads the markdown report produced by
generate_report, parses headings / paragraphs / tables, emits an
interactive single-file HTML report:

  - "Done" checkbox in every row of the §3 (Action required) and §4 (QA)
    tables. Per-issue checked-state persists in browser localStorage.
  - Sticky filter bar at the top of the page with five dropdowns
    (Assignee, Owner Transferred, Priority, Category, Dependency) plus
    a free-text search across Title / action_TBD. Filters apply to
    every table.
  - Fully self-contained: CSS / JS inlined, no CDN. Works offline.

Re-runs gen_bug_scrub_md.py first so the HTML always reflects the
current Excel workbook.
"""

from __future__ import annotations

import html
import re
import subprocess
import sys
from pathlib import Path

# ---- paths ---------------------------------------------------------------

THIS = Path(__file__).resolve()
REPO_ROOT = THIS.parents[7]
RESULT_DIR = REPO_ROOT / "opencode" / "issue_triage" / "result"
MD_PATH = RESULT_DIR / "bug_scrub.md"
HTML_PATH = RESULT_DIR / "bug_scrub.html"
XLSX_PATH = RESULT_DIR / "torch_xpu_ops_issues.xlsx"
GEN_MD_SCRIPT = (
    THIS.parents[1] / "generate_report" / "gen_bug_scrub_md.py"
)

# Section-prefix → semantics. Done checkbox is gated on this.
DONE_CHECKBOX_PREFIXES = ("3.", "4.")  # §3 Action required, §4 QA

# Header columns we treat as filter dimensions. Match is case-insensitive
# on the rendered header text, so "Owner Transferred" maps to a stable
# filter id "owner_transferred".
FILTER_COLUMNS = {
    "Assignee":          "assignee",
    "Owner":             "assignee",        # report renames Assignee → Owner
    "Owner Transferred": "owner_transferred",
    "Priority":          "priority",
    "Category":          "category",
    "Dependency":        "dependency",
    "Milestone":         "milestone",
}


# ---- markdown subset parser ---------------------------------------------
# We control the input shape (gen_bug_scrub_md.py), so this parser handles
# only the constructs that report actually emits. Anything unrecognized is
# emitted as a plain <p>.

INLINE_LINK = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
INLINE_CODE = re.compile(r"`([^`]+)`")
INLINE_BOLD = re.compile(r"\*\*([^*]+)\*\*")
INLINE_ITAL = re.compile(r"(?<!\*)\*([^*]+)\*(?!\*)")
INLINE_BR   = re.compile(r"<br\s*/?>", re.IGNORECASE)
ANCHOR_TAG  = re.compile(r'<a id="([^"]+)"></a>')
TABLE_SEP   = re.compile(r"^\|\s*[-:]+\s*(\|\s*[-:]+\s*)+\|?\s*$")


def render_inline(text: str) -> str:
    """Inline markdown → HTML. Order matters: escape first, then expand."""
    # extract <br> placeholders before escaping
    text = INLINE_BR.sub("\x00BR\x00", text)
    text = html.escape(text, quote=False)
    text = text.replace("\x00BR\x00", "<br>")
    text = INLINE_LINK.sub(
        lambda m: f'<a href="{html.escape(m.group(2), quote=True)}">{m.group(1)}</a>',
        text,
    )
    text = INLINE_CODE.sub(lambda m: f"<code>{m.group(1)}</code>", text)
    text = INLINE_BOLD.sub(lambda m: f"<strong>{m.group(1)}</strong>", text)
    text = INLINE_ITAL.sub(lambda m: f"<em>{m.group(1)}</em>", text)
    return text


def split_table_row(line: str) -> list[str]:
    """Split a markdown table row on `|`, honoring `\|` escape."""
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    # Protect escaped pipes during the split
    parts = line.replace(r"\|", "\x00PIPE\x00").split("|")
    return [p.replace("\x00PIPE\x00", "|").strip() for p in parts]


def slug_for_header(name: str) -> str:
    return FILTER_COLUMNS.get(name, "")


def extract_issue_id(cell_html: str) -> str:
    """Find the first /issues/<N> in a rendered cell (the Issue column)."""
    m = re.search(r"/issues/(\d+)", cell_html)
    return m.group(1) if m else ""


def load_issue_metadata(xlsx_path: Path) -> dict[str, dict[str, str]]:
    """Build {issue_id: {category, dependency, priority, assignee, owner_transferred}}.

    The markdown report's §3/§4 detail tables don't carry Category/Dependency
    columns, but the user wants to filter by them in the HTML report. We
    look these up from the Issues sheet and stamp them onto every row's
    data-* attributes regardless of which markdown table the row came from.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("WARN: openpyxl not available; Category/Dependency filters will be sparse",
              file=sys.stderr)
        return {}
    if not xlsx_path.exists():
        return {}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Issues" not in wb.sheetnames:
        return {}
    ws = wb["Issues"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(rows)]
    col = {h: i for i, h in enumerate(headers)}
    needed = ("Issue ID", "Category", "Dependency", "Priority", "Assignee", "owner_transferred", "Milestone")
    if "Issue ID" not in col:
        return {}

    def _get(r, key):
        i = col.get(key)
        if i is None or i >= len(r):
            return ""
        v = r[i]
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() == "none" else s

    out: dict[str, dict[str, str]] = {}
    for r in rows:
        iid = _get(r, "Issue ID")
        if not iid:
            continue
        # Issue IDs are bare integers in the workbook; markdown links use
        # /issues/<N> with no leading "#". Normalize to plain digits.
        iid = iid.lstrip("#").strip()
        out[iid] = {
            "category":          _get(r, "Category"),
            "dependency":        _get(r, "Dependency"),
            "priority":          _get(r, "Priority"),
            "assignee":          _get(r, "Assignee"),
            "owner_transferred": _get(r, "owner_transferred"),
            "milestone":         _get(r, "Milestone"),
        }
    wb.close()
    return out


# Loaded once, consulted by render_table for every row
ISSUE_META: dict[str, dict[str, str]] = {}


# ---- main parse → emit ---------------------------------------------------

def md_to_html(md_text: str) -> str:
    lines = md_text.split("\n")
    out: list[str] = []
    i = 0
    n = len(lines)
    pending_anchor: str | None = None
    current_section_path: list[str] = []  # ["3", "1", "1"] etc.

    def emit(s: str = ""):
        out.append(s)

    while i < n:
        line = lines[i]

        # standalone anchor (always followed by a heading or table)
        m = ANCHOR_TAG.fullmatch(line.strip())
        if m:
            pending_anchor = m.group(1)
            i += 1
            continue

        stripped = line.strip()

        # heading
        m = re.match(r"^(#{1,6})\s+(.+)$", stripped)
        if m:
            level = len(m.group(1))
            text = m.group(2).strip()
            anchor_attr = f' id="{pending_anchor}"' if pending_anchor else ""
            pending_anchor = None
            # track section path from heading text "## 3.1.1 Foo  ·  N issues"
            sec_match = re.match(r"^(\d+(?:\.\d+)*)", text)
            if sec_match:
                current_section_path = sec_match.group(1).split(".")
            emit(f'<h{level}{anchor_attr}>{render_inline(text)}</h{level}>')
            i += 1
            continue

        # blockquote / italic info line "_..._"
        if stripped.startswith("_") and stripped.endswith("_") and len(stripped) > 2:
            emit(f'<p class="meta">{render_inline(stripped)}</p>')
            i += 1
            continue

        # table — at least header + separator
        if stripped.startswith("|") and i + 1 < n and TABLE_SEP.match(lines[i + 1]):
            header_cells = split_table_row(lines[i])
            i += 2  # consume header + separator
            body_rows: list[list[str]] = []
            while i < n and lines[i].strip().startswith("|"):
                body_rows.append(split_table_row(lines[i]))
                i += 1

            section_id = ".".join(current_section_path) if current_section_path else ""
            show_done = any(
                section_id.startswith(p) or section_id == p.rstrip(".")
                for p in DONE_CHECKBOX_PREFIXES
            )
            emit(render_table(header_cells, body_rows, show_done, section_id))
            continue

        # bullet list — collect contiguous bullets
        if re.match(r"^\s*[-*]\s+", line):
            emit("<ul>")
            while i < n and re.match(r"^\s*[-*]\s+", lines[i]):
                item = re.sub(r"^\s*[-*]\s+", "", lines[i])
                emit(f"  <li>{render_inline(item)}</li>")
                i += 1
            emit("</ul>")
            continue

        # blank
        if not stripped:
            emit("")
            i += 1
            continue

        # plain paragraph — until blank line
        para_lines = [line]
        i += 1
        while i < n and lines[i].strip() and not lines[i].lstrip().startswith(("#", "|", "-", "*", "<a id")):
            para_lines.append(lines[i])
            i += 1
        emit(f"<p>{render_inline(' '.join(s.strip() for s in para_lines))}</p>")

    return "\n".join(out)


def render_table(headers: list[str], rows: list[list[str]], show_done: bool, section_id: str) -> str:
    """Emit a <table> with per-row data-* attributes for filtering and an
    optional Done checkbox column. The checkbox is only emitted when
    show_done is True (§3 / §4 sections)."""
    thead = ["<thead>", "<tr>"]
    if show_done:
        thead.append('<th class="done-col">Done</th>')
    for h in headers:
        thead.append(f"<th>{render_inline(h)}</th>")
    thead.append("</tr></thead>")

    # Map header position → filter slug
    col_to_slug: dict[int, str] = {}
    for idx_h, h in enumerate(headers):
        slug = slug_for_header(h.strip())
        if slug:
            col_to_slug[idx_h] = slug

    tbody = ["<tbody>"]
    for row in rows:
        # render cells first so we can extract issue id from rendered HTML
        rendered_cells = [render_inline(c) for c in row]
        issue_id = ""
        if rendered_cells:
            issue_id = extract_issue_id(rendered_cells[0])

        # build data-* attributes — only for per-issue rows. Stats tables
        # (§8) reuse "Category" as a header but rows have no Issue link,
        # so they must not contribute to filter dropdowns.
        emitted: dict[str, str] = {}
        if issue_id:
            # Workbook is authoritative for filter values: the markdown
            # cell may be truncated/wrapped (e.g. "Copilot / Ei…") which
            # would corrupt tokenization. Cells only fill in dimensions
            # the workbook doesn't have.
            if issue_id in ISSUE_META:
                for slug, val in ISSUE_META[issue_id].items():
                    if val:
                        emitted[slug] = val
            for idx_c, slug in col_to_slug.items():
                if slug in emitted or idx_c >= len(row):
                    continue
                raw = re.sub(r"<[^>]+>", " ", row[idx_c])
                raw = re.sub(r"\s+", " ", raw).strip()
                emitted[slug] = raw
        data_attrs = [
            f'data-{slug}="{html.escape(val, quote=True)}"'
            for slug, val in emitted.items()
        ]
        # full-row searchable text (Title + action_TBD) lowercased
        searchable = " ".join(re.sub(r"<[^>]+>", " ", c) for c in row[:5]).lower()
        data_attrs.append(f'data-search="{html.escape(searchable, quote=True)}"')
        if issue_id:
            data_attrs.append(f'data-issue="{issue_id}"')
        tr_attrs = " ".join(data_attrs)

        tbody.append(f"<tr {tr_attrs}>")
        if show_done:
            cb_id = f"done-{issue_id}" if issue_id else ""
            tbody.append(
                f'<td class="done-col"><input type="checkbox" class="ar-done" '
                f'data-issue="{issue_id}" id="{cb_id}"></td>'
            )
        for c in rendered_cells:
            tbody.append(f"<td>{c}</td>")
        tbody.append("</tr>")
    tbody.append("</tbody>")

    sec_attr = f' data-section="{section_id}"' if section_id else ""
    return (
        f'<table class="ar-table"{sec_attr}>\n'
        + "\n".join(thead) + "\n"
        + "\n".join(tbody) + "\n"
        + "</table>"
    )


# ---- CSS / JS (inlined) --------------------------------------------------

CSS = """
:root {
  --bg: #f8f9fa; --fg: #212529; --muted: #6c757d;
  --border: #dee2e6; --accent: #0066cc; --done-bg: #e9ecef; --done-fg: #adb5bd;
  --filter-bg: #ffffff; --hl: #fff3cd;
}
* { box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  font-size: 13px; line-height: 1.5; color: var(--fg); background: var(--bg);
  margin: 0; padding: 0; }
h1, h2, h3, h4 { line-height: 1.2; margin-top: 1.5em; }
h1 { font-size: 1.8em; } h2 { font-size: 1.4em; border-bottom: 2px solid var(--border); padding-bottom: .3em; }
h3 { font-size: 1.15em; } h4 { font-size: 1em; color: var(--muted); }
a { color: var(--accent); text-decoration: none; }
a:hover { text-decoration: underline; }
code { background: #f1f3f5; padding: 1px 4px; border-radius: 3px; font-size: .9em; }
.meta { color: var(--muted); font-size: .9em; margin: .3em 0; }
.content { max-width: 100%; padding: 1em 1.5em 4em; }

/* Sticky filter bar */
.filter-bar {
  position: sticky; top: 0; z-index: 100;
  background: var(--filter-bg); border-bottom: 1px solid var(--border);
  padding: .6em 1em; display: flex; flex-wrap: wrap; gap: .8em; align-items: center;
  box-shadow: 0 1px 3px rgba(0,0,0,.06);
}
.filter-bar > label { font-size: 12px; color: var(--muted); display: flex; flex-direction: column; gap: 2px; }
.filter-bar select, .filter-bar input[type=text] {
  font-size: 13px; padding: 3px 6px; border: 1px solid var(--border); border-radius: 3px;
  background: white; min-width: 140px;
}
.filter-bar input[type=text] { min-width: 240px; }
.filter-bar button {
  font-size: 12px; padding: 4px 10px; border: 1px solid var(--border); border-radius: 3px;
  background: white; cursor: pointer;
}
.filter-bar button:hover { background: #f1f3f5; }
.filter-bar .stats { margin-left: auto; color: var(--muted); font-size: 12px; }

/* Tables */
table.ar-table { border-collapse: collapse; width: 100%; margin: .5em 0 1em; font-size: 12px; }
table.ar-table th, table.ar-table td { border: 1px solid var(--border); padding: 4px 6px; vertical-align: top; text-align: left; }
table.ar-table thead th { background: #e7eaf0; position: sticky; top: 56px; z-index: 50; }
table.ar-table td.done-col, table.ar-table th.done-col { width: 40px; text-align: center; }
table.ar-table tr:nth-child(even) td { background: #fafbfc; }
table.ar-table tr.done td { background: var(--done-bg) !important; color: var(--done-fg); }
table.ar-table tr.done td a { color: var(--done-fg); }
table.ar-table tr.done td:not(.done-col) { text-decoration: line-through; }
table.ar-table tr.hidden { display: none; }
table.ar-table tr.hl td { background: var(--hl) !important; }

/* Whole sections (and TOC entries) collapsed when a filter empties them. */
.section-hidden { display: none !important; }

/* Lists */
ul { margin: .3em 0; padding-left: 1.5em; }

/* Multi-select dropdown */
.ms-dd { position: relative; display: inline-block; }
.ms-dd-btn {
  font-size: 13px; padding: 3px 24px 3px 6px; border: 1px solid var(--border);
  border-radius: 3px; background: white; cursor: pointer; min-width: 160px;
  text-align: left; position: relative; white-space: nowrap; overflow: hidden;
  text-overflow: ellipsis; max-width: 260px;
}
.ms-dd-btn::after {
  content: "\\25BE"; position: absolute; right: 6px; top: 50%;
  transform: translateY(-50%); color: var(--muted); font-size: 10px;
}
.ms-dd-panel {
  display: none; position: absolute; top: calc(100% + 2px); left: 0;
  background: white; border: 1px solid var(--border); border-radius: 3px;
  box-shadow: 0 2px 8px rgba(0,0,0,.12); z-index: 200;
  max-height: 320px; overflow-y: auto; min-width: 220px; padding: 4px 0;
}
.ms-dd.open .ms-dd-panel { display: block; }
.ms-dd-search {
  width: calc(100% - 12px); margin: 4px 6px; padding: 3px 5px;
  border: 1px solid var(--border); border-radius: 3px; font-size: 12px;
}
.ms-dd-item {
  display: flex; flex-direction: row; align-items: center; gap: 8px;
  padding: 4px 10px; cursor: pointer; font-size: 12px; user-select: none;
  color: var(--fg);
}
.ms-dd-item:hover { background: #f1f3f5; }
.ms-dd-item input[type=checkbox] {
  margin: 0; flex: 0 0 auto; width: 14px; height: 14px; accent-color: var(--accent);
}
.ms-dd-item > span:not(.count) {
  flex: 1 1 auto; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
.ms-dd-item .count {
  flex: 0 0 auto; color: var(--muted); font-size: 11px;
  background: #f1f3f5; padding: 0 6px; border-radius: 8px; min-width: 24px;
  text-align: center;
}
.ms-dd-actions {
  display: flex; gap: 4px; padding: 4px 6px; border-top: 1px solid var(--border);
  margin-top: 2px; position: sticky; bottom: 0; background: white;
}
.ms-dd-actions button {
  flex: 1; font-size: 11px; padding: 2px 4px; border: 1px solid var(--border);
  border-radius: 3px; background: white; cursor: pointer;
}
.ms-dd-actions button:hover { background: #f1f3f5; }
"""

JS = """
// ---- Done checkbox persistence ----
const STORAGE_KEY_PREFIX = 'bug_scrub_done_';
function loadDoneState() {
  document.querySelectorAll('input.ar-done').forEach(cb => {
    const id = cb.dataset.issue;
    if (!id) return;
    const v = localStorage.getItem(STORAGE_KEY_PREFIX + id);
    if (v === '1') {
      cb.checked = true;
      cb.closest('tr').classList.add('done');
    }
  });
}
function onDoneToggle(e) {
  const cb = e.target;
  const id = cb.dataset.issue;
  if (!id) return;
  const tr = cb.closest('tr');
  if (cb.checked) {
    localStorage.setItem(STORAGE_KEY_PREFIX + id, '1');
    tr.classList.add('done');
  } else {
    localStorage.removeItem(STORAGE_KEY_PREFIX + id);
    tr.classList.remove('done');
  }
}

// ---- Filter bar ----
// UI dimensions (what the user sees). owner_transferred is intentionally
// NOT in this list — selecting an Assignee implicitly matches rows whose
// owner_transferred token equals the selected assignee, so a separate
// dropdown would be redundant.
const FILTER_DIMS = ['assignee', 'priority', 'category', 'milestone', 'dependency'];
const FILTER_LABELS = {
  assignee: 'Assignee', priority: 'Priority', category: 'Category',
  milestone: 'Milestone', dependency: 'Dependency'
};

// Dimensions whose raw cell value may carry multiple comma-separated
// owners (e.g. "daisyden, CuiYifeng"). Selecting "daisyden" should match
// any row that contains that token.
const MULTI_VALUE_DIMS = new Set(['assignee', 'owner_transferred']);
const SYCL_KERNEL_GROUP = 'SYCL kernel';

// Map a raw cell value to the list of canonical filter tokens it
// contributes. For multi-owner cells we split on `,`; for Dependency we
// collapse every "SYCL kernel: <file>.cpp" into a single "SYCL kernel"
// group so users don't have to scroll through 50 file-specific options.
function tokensFor(dim, raw) {
  if (!raw) return [];
  if (dim === 'dependency') {
    return [/^SYCL kernel:/i.test(raw) ? SYCL_KERNEL_GROUP : raw];
  }
  if (MULTI_VALUE_DIMS.has(dim)) {
    return raw.split(',').map(s => s.trim()).filter(Boolean);
  }
  return [raw];
}

function collectValues(dim) {
  // {token -> count} so we can label grouped options like "SYCL kernel (50)".
  const counts = new Map();
  // For the Assignee dropdown, also surface owner_transferred tokens so a
  // user can pick a person who has only ever appeared via transfer.
  const sourceDims = (dim === 'assignee') ? ['assignee', 'owner_transferred'] : [dim];
  for (const sd of sourceDims) {
    document.querySelectorAll(`[data-${sd}]`).forEach(tr => {
      const raw = (tr.dataset[sd] || '').trim();
      for (const tok of tokensFor(sd, raw)) {
        counts.set(tok, (counts.get(tok) || 0) + 1);
      }
    });
  }
  return Array.from(counts.entries())
    .sort((a, b) => a[0].localeCompare(b[0], undefined, {numeric: true}));
}

// Selected-token state, per dimension.
const SELECTED = Object.fromEntries(FILTER_DIMS.map(d => [d, new Set()]));

function buildMultiSelect(dim, items) {
  const dd = document.createElement('div');
  dd.className = 'ms-dd';
  dd.dataset.dim = dim;

  const btn = document.createElement('button');
  btn.type = 'button';
  btn.className = 'ms-dd-btn';
  btn.textContent = '(all)';
  dd.appendChild(btn);

  const panel = document.createElement('div');
  panel.className = 'ms-dd-panel';

  const search = document.createElement('input');
  search.type = 'text';
  search.className = 'ms-dd-search';
  search.placeholder = 'filter...';
  panel.appendChild(search);

  const list = document.createElement('div');
  list.className = 'ms-dd-list';
  panel.appendChild(list);

  for (const [tok, count] of items) {
    const label = document.createElement('label');
    label.className = 'ms-dd-item';
    const cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.value = tok;
    cb.addEventListener('change', () => {
      if (cb.checked) SELECTED[dim].add(tok);
      else SELECTED[dim].delete(tok);
      updateButtonLabel(dd, dim);
      applyFilters();
    });
    const txt = document.createElement('span');
    const display = (dim === 'dependency' && tok === SYCL_KERNEL_GROUP)
      ? `${tok}` : tok;
    txt.textContent = display;
    const cnt = document.createElement('span');
    cnt.className = 'count';
    cnt.textContent = count;
    label.appendChild(cb);
    label.appendChild(txt);
    label.appendChild(cnt);
    list.appendChild(label);
  }

  search.addEventListener('input', () => {
    const q = search.value.trim().toLowerCase();
    list.querySelectorAll('.ms-dd-item').forEach(it => {
      const t = it.querySelector('span').textContent.toLowerCase();
      it.style.display = (!q || t.includes(q)) ? '' : 'none';
    });
  });

  const actions = document.createElement('div');
  actions.className = 'ms-dd-actions';
  const clr = document.createElement('button');
  clr.type = 'button';
  clr.textContent = 'Clear';
  clr.addEventListener('click', () => {
    SELECTED[dim].clear();
    list.querySelectorAll('input[type=checkbox]').forEach(cb => cb.checked = false);
    updateButtonLabel(dd, dim);
    applyFilters();
  });
  const close = document.createElement('button');
  close.type = 'button';
  close.textContent = 'Close';
  close.addEventListener('click', () => dd.classList.remove('open'));
  actions.appendChild(clr);
  actions.appendChild(close);
  panel.appendChild(actions);

  dd.appendChild(panel);

  btn.addEventListener('click', (e) => {
    e.stopPropagation();
    const wasOpen = dd.classList.contains('open');
    document.querySelectorAll('.ms-dd.open').forEach(d => d.classList.remove('open'));
    if (!wasOpen) dd.classList.add('open');
  });
  panel.addEventListener('click', (e) => e.stopPropagation());

  return dd;
}

function updateButtonLabel(dd, dim) {
  const btn = dd.querySelector('.ms-dd-btn');
  const sel = SELECTED[dim];
  if (sel.size === 0) {
    btn.textContent = '(all)';
    btn.title = '';
  } else if (sel.size === 1) {
    const v = Array.from(sel)[0];
    btn.textContent = v;
    btn.title = v;
  } else {
    const vs = Array.from(sel).join(', ');
    btn.textContent = `${sel.size} selected`;
    btn.title = vs;
  }
}

function buildFilterBar() {
  const bar = document.createElement('div');
  bar.className = 'filter-bar';

  for (const dim of FILTER_DIMS) {
    const wrap = document.createElement('label');
    wrap.textContent = FILTER_LABELS[dim];
    const dd = buildMultiSelect(dim, collectValues(dim));
    dd.id = 'filter-' + dim;
    wrap.appendChild(dd);
    bar.appendChild(wrap);
  }

  // free-text search
  const searchWrap = document.createElement('label');
  searchWrap.textContent = 'Search';
  const search = document.createElement('input');
  search.type = 'text'; search.id = 'filter-search';
  search.placeholder = 'title / action_TBD';
  search.addEventListener('input', applyFilters);
  searchWrap.appendChild(search);
  bar.appendChild(searchWrap);

  // hide-done toggle
  const hideWrap = document.createElement('label');
  hideWrap.style.flexDirection = 'row';
  hideWrap.style.alignItems = 'center';
  hideWrap.style.gap = '4px';
  const hideCb = document.createElement('input');
  hideCb.type = 'checkbox'; hideCb.id = 'filter-hide-done';
  hideCb.addEventListener('change', applyFilters);
  hideWrap.appendChild(hideCb);
  hideWrap.appendChild(document.createTextNode('Hide Done'));
  bar.appendChild(hideWrap);

  // reset button
  const reset = document.createElement('button');
  reset.textContent = 'Reset filters';
  reset.addEventListener('click', () => {
    for (const dim of FILTER_DIMS) {
      SELECTED[dim].clear();
      const dd = document.getElementById('filter-' + dim);
      dd.querySelectorAll('input[type=checkbox]').forEach(cb => cb.checked = false);
      updateButtonLabel(dd, dim);
    }
    document.getElementById('filter-search').value = '';
    document.getElementById('filter-hide-done').checked = false;
    applyFilters();
  });
  bar.appendChild(reset);

  // export-done button
  const exp = document.createElement('button');
  exp.textContent = 'Export Done IDs';
  exp.addEventListener('click', () => {
    const ids = [];
    for (let i = 0; i < localStorage.length; i++) {
      const k = localStorage.key(i);
      if (k.startsWith(STORAGE_KEY_PREFIX) && localStorage.getItem(k) === '1') {
        ids.push(k.slice(STORAGE_KEY_PREFIX.length));
      }
    }
    ids.sort((a, b) => Number(a) - Number(b));
    navigator.clipboard.writeText(ids.join(',')).then(
      () => alert(`Copied ${ids.length} done issue IDs to clipboard`),
      () => prompt('Done IDs (copy manually):', ids.join(','))
    );
  });
  bar.appendChild(exp);

  // stats
  const stats = document.createElement('span');
  stats.className = 'stats'; stats.id = 'filter-stats';
  bar.appendChild(stats);

  document.body.insertBefore(bar, document.body.firstChild);

  // dismiss any open dropdown on outside click
  document.addEventListener('click', () => {
    document.querySelectorAll('.ms-dd.open').forEach(d => d.classList.remove('open'));
  });
}

function rowMatchesDim(tr, dim, selectedSet) {
  if (selectedSet.size === 0) return true;
  // Assignee filter also matches owner_transferred tokens.
  const sourceDims = (dim === 'assignee') ? ['assignee', 'owner_transferred'] : [dim];
  for (const sd of sourceDims) {
    const tokens = tokensFor(sd, (tr.dataset[sd] || '').trim());
    for (const t of tokens) {
      if (selectedSet.has(t)) return true;
    }
  }
  return false;
}

function applyFilters() {
  const search = document.getElementById('filter-search').value.trim().toLowerCase();
  const hideDone = document.getElementById('filter-hide-done').checked;

  let total = 0, visible = 0;
  document.querySelectorAll('table.ar-table tbody tr').forEach(tr => {
    total++;
    let show = true;
    for (const dim of FILTER_DIMS) {
      if (!rowMatchesDim(tr, dim, SELECTED[dim])) { show = false; break; }
    }
    if (show && search) {
      show = (tr.dataset.search || '').includes(search);
    }
    if (show && hideDone && tr.classList.contains('done')) {
      show = false;
    }
    tr.classList.toggle('hidden', !show);
    if (show) visible++;
  });
  document.getElementById('filter-stats').textContent =
    `${visible} / ${total} rows`;

  hideEmptySections();
}

// Collapse blocks whose tables ended up with zero visible rows. A "block"
// is a <table.ar-table> together with the siblings immediately preceding
// it back to the previous table or section heading (h2/h3/h4) — i.e. the
// intro paragraph(s), the bullet that names the subsection, and the
// "back to index" link that go with that table. Then a parent <h2> (and
// its descendants up to the next <h2>) is hidden when every table under
// it is hidden. Finally TOC entries pointing to hidden anchors are
// hidden too.
function hideEmptySections() {
  const content = document.querySelector('.content');
  if (!content) return;

  // Reset previous hides on structure-level elements (rows already toggled above)
  content.querySelectorAll('.section-hidden').forEach(el => {
    el.classList.remove('section-hidden');
  });

  const STOP = el => el && (el.tagName === 'TABLE' ||
                            el.tagName === 'H2' || el.tagName === 'H3' || el.tagName === 'H4');

  // Pass 1: hide each empty table + its leading intro siblings.
  const tables = Array.from(content.querySelectorAll('table.ar-table'));
  for (const table of tables) {
    const visibleRows = table.querySelectorAll('tbody tr:not(.hidden)').length;
    if (visibleRows > 0) continue;
    table.classList.add('section-hidden');
    // Walk back over intro siblings up to (but not including) the previous
    // table or heading.
    let sib = table.previousElementSibling;
    while (sib && !STOP(sib)) {
      sib.classList.add('section-hidden');
      sib = sib.previousElementSibling;
    }
    // The triggering heading (h3/h4) immediately before the intro: if
    // we landed on one and it isn't an h2 (which may parent multiple
    // tables), hide it too.
    if (sib && (sib.tagName === 'H3' || sib.tagName === 'H4')) {
      sib.classList.add('section-hidden');
    }
  }

  // Pass 2: hide an <h2> (and everything until the next <h2>) when all
  // descendants in that span are already hidden.
  const kids = Array.from(content.children);
  let i = 0;
  while (i < kids.length) {
    if (kids[i].tagName !== 'H2') { i++; continue; }
    const start = i;
    let end = i + 1;
    while (end < kids.length && kids[end].tagName !== 'H2') end++;
    // span = [start, end)
    let anyVisible = false;
    for (let j = start + 1; j < end; j++) {
      if (!kids[j].classList.contains('section-hidden')) {
        // tables only matter for "is this section worth showing?"
        if (kids[j].tagName === 'TABLE') {
          anyVisible = true; break;
        }
      }
    }
    // If no table in the span is visible, hide the whole span.
    if (!anyVisible) {
      // But: only collapse the span if it contains at least one table to
      // begin with. Sections like §1 Summary and §2 Index have no
      // <table.ar-table> and should always remain visible.
      let hasTable = false;
      for (let j = start + 1; j < end; j++) {
        if (kids[j].tagName === 'TABLE' && kids[j].classList.contains('ar-table')) {
          hasTable = true; break;
        }
      }
      if (hasTable) {
        for (let j = start; j < end; j++) {
          kids[j].classList.add('section-hidden');
        }
      }
    }
    i = end;
  }

  // Pass 3: hide TOC entries (in §2 Index) whose target anchor is now
  // inside a hidden section.
  const tocLinks = content.querySelectorAll('ul li > a[href^="#sec-"]');
  tocLinks.forEach(a => {
    const target = document.getElementById(a.getAttribute('href').slice(1));
    const li = a.parentElement;
    if (!target) { li.classList.remove('section-hidden'); return; }
    // A target is "effectively hidden" if the target itself is hidden,
    // or every <table> following it in its h2-span is hidden.
    let hide = target.classList.contains('section-hidden');
    if (!hide) {
      // walk forward until next same-or-higher heading; if every table
      // in between is hidden AND there is at least one table, hide.
      const targetLevel = parseInt(target.tagName[1], 10);
      let sib = target.nextElementSibling;
      let sawTable = false, allHidden = true;
      while (sib) {
        if (/^H[1-6]$/.test(sib.tagName) && parseInt(sib.tagName[1], 10) <= targetLevel) break;
        if (sib.tagName === 'TABLE' && sib.classList.contains('ar-table')) {
          sawTable = true;
          if (!sib.classList.contains('section-hidden')) { allHidden = false; break; }
        }
        sib = sib.nextElementSibling;
      }
      hide = sawTable && allHidden;
    }
    li.classList.toggle('section-hidden', hide);
  });
}

// ---- init ----
document.addEventListener('DOMContentLoaded', () => {
  loadDoneState();
  document.querySelectorAll('input.ar-done').forEach(cb => {
    cb.addEventListener('change', onDoneToggle);
  });
  buildFilterBar();
  applyFilters();
});
"""


# ---- top-level ----------------------------------------------------------

def render_page(body_html: str, title: str) -> str:
    return (
        '<!doctype html>\n'
        '<html lang="en">\n<head>\n'
        '<meta charset="utf-8">\n'
        f'<title>{html.escape(title)}</title>\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        f'<style>{CSS}</style>\n'
        '</head>\n<body>\n'
        f'<div class="content">\n{body_html}\n</div>\n'
        f'<script>{JS}</script>\n'
        '</body>\n</html>\n'
    )


def main() -> int:
    # Phase 5b spec: re-run gen_bug_scrub_md.py first so HTML always
    # reflects the current workbook.
    if not GEN_MD_SCRIPT.exists():
        print(f"ERROR: missing {GEN_MD_SCRIPT}", file=sys.stderr)
        return 2
    print(f"running {GEN_MD_SCRIPT.name} ...")
    rc = subprocess.run([sys.executable, str(GEN_MD_SCRIPT)]).returncode
    if rc != 0:
        print(f"ERROR: gen_bug_scrub_md.py failed (rc={rc})", file=sys.stderr)
        return rc

    if not MD_PATH.exists():
        print(f"ERROR: missing {MD_PATH}", file=sys.stderr)
        return 3

    # Per-issue metadata for filter backfill (Category, Dependency, etc.)
    global ISSUE_META
    ISSUE_META = load_issue_metadata(XLSX_PATH)
    print(f"loaded metadata for {len(ISSUE_META)} issues from {XLSX_PATH.name}")

    md_text = MD_PATH.read_text(encoding="utf-8")
    body = md_to_html(md_text)
    page = render_page(body, "torch-xpu-ops bug scrub")
    HTML_PATH.write_text(page, encoding="utf-8")
    print(f"wrote {HTML_PATH} ({HTML_PATH.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
