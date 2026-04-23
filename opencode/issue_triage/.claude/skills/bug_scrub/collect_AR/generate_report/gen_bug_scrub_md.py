"""Generate bug_scrub.md from Issues sheet + action_Type classification.

Usage:
    # default full report
    python3 gen_bug_scrub_md.py

    # filtered report (e.g. UT-scoped)
    python3 gen_bug_scrub_md.py \\
        --issues-file ut_issues.txt \\
        --out opencode/issue_triage/result/bug_scrub_ut.md \\
        --title-suffix " — UT scope"
"""
from __future__ import annotations

import argparse
import re
import textwrap
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[7]

_p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
_p.add_argument("--issues-file", type=Path, default=None,
                help="Optional path to a file with whitespace/newline-"
                     "separated issue IDs. When given, only those issues "
                     "are included in the report.")
_p.add_argument("--out", type=Path, default=None,
                help="Output .md path (relative paths resolve against "
                     "the repo root). Defaults to "
                     "opencode/issue_triage/result/bug_scrub.md.")
_p.add_argument("--title-suffix", default="",
                help="String appended to the report's top-level heading.")
_args = _p.parse_args()

EXCEL = REPO_ROOT / "opencode/issue_triage/result/torch_xpu_ops_issues.xlsx"
if _args.out is None:
    OUT = REPO_ROOT / "opencode/issue_triage/result/bug_scrub.md"
elif _args.out.is_absolute():
    OUT = _args.out
else:
    OUT = REPO_ROOT / _args.out
REPO  = "intel/torch-xpu-ops"
TODAY = datetime(2026, 4, 21, tzinfo=timezone.utc)
RECENT_CUTOFF = TODAY - timedelta(days=7)

# display order for the two macro-groups (merged buckets — see MERGE below)
DEV_SECTIONS = ["NEED PR", "TRACK PR", "NEEDS_OWNER"]
QA_SECTIONS  = ["CLOSE or SKIP", "AWAIT_REPLY", "MONITOR", "CHECK_CASES"]

# Raw-atom → merged-bucket mapping. Atoms not listed map to themselves.
# The action_Type column in the Excel stays atomic; merging is display-only.
MERGE = {
    "NEED_ACTION":      "NEED PR",
    "ROOT_CAUSE":       "NEED PR",
    "IMPLEMENT":        "NEED PR",
    "TRACK_PR":         "TRACK PR",
    "RETRIAGE_PRS":     "TRACK PR",
    "CLOSE":            "CLOSE or SKIP",
    "VERIFY_AND_CLOSE": "CLOSE or SKIP",
    "SKIP":             "CLOSE or SKIP",
    "NOT_TARGET_CLOSE": "CLOSE or SKIP",
}

def merged(raw: str | None) -> str | None:
    if raw is None:
        return None
    return MERGE.get(raw, raw)

# priority ordering for "primary category" selection when a row has combos
PRIMARY_ORDER = [
    "CLOSE", "NOT_TARGET_CLOSE", "VERIFY_AND_CLOSE", "TRACK_PR",
    "IMPLEMENT", "RETRIAGE_PRS", "WAIT_EXTERNAL",
    "ROOT_CAUSE", "FILE_ISSUE", "MONITOR",
    "NEEDS_OWNER", "NEED_ACTION",
    "AWAIT_REPLY", "CHECK_CASES", "SKIP",
]

SECTION_TITLES = {
    # merged buckets
    "NEED PR":          "NEED PR — a PR must be produced or continued (no PR yet, or owner actively debugging root cause, or new code needed)",
    "TRACK PR":         "TRACK PR — a PR is identified; track it to merge, or re-evaluate if prior PRs are dead / unverified",
    "CLOSE or SKIP":    "CLOSE or SKIP — terminal QA action (close fixed, verify merged fix, skip not-target/wontfix, or label not_target and close)",
    # unmerged buckets
    "NEEDS_OWNER":      "NEEDS_OWNER — awaiting triage-lead to assign an owner",
    "AWAIT_REPLY":      "AWAIT_REPLY — open questions in thread; owner must respond",
    "MONITOR":          "MONITOR — long-running tracker / maintenance / scoping",
    "CHECK_CASES":      "CHECK_CASES — XPU test case missing in repo; QA must verify case existence before action",
    "UNCLASSIFIED":     "UNCLASSIFIED — Phase 4b produced no verdict; needs manual triage",
}

PRIO_RANK = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "": 9, None: 9}


# -------- load -------------------------------------------------------------
wb  = openpyxl.load_workbook(EXCEL)
ws  = wb["Issues"]
hdr = [c.value for c in ws[1]]
def col(n): return hdr.index(n)
C = {k: col(k) for k in [
    "Issue ID","Title","Status","Assignee","Reporter","Labels","Created Time",
    "Category","Priority","Root Cause","Fix Approach","action_TBD","action_reason",
    "owner_transferred","duplicated_issue","Dependency","action_Type",
]}

# Traceback index: {issue_id: [ {source, test_case, test_file, error, traceback}, ... ]}
tb_by_issue: dict[int, list[dict]] = defaultdict(list)
for sheet_name, src_label in [("Test Cases", "UT"), ("E2E Test Cases", "E2E")]:
    ws_tb = wb[sheet_name]
    h_tb = [c.value for c in ws_tb[1]]
    idx = {k: h_tb.index(k) for k in ("Issue ID", "Error Message", "Traceback")}
    tc_i = h_tb.index("Test Case") if "Test Case" in h_tb else None
    tf_i = h_tb.index("Test File") if "Test File" in h_tb else None
    mdl_i = h_tb.index("Model") if "Model" in h_tb else None
    bench_i = h_tb.index("Benchmark") if "Benchmark" in h_tb else None
    for r_tb in ws_tb.iter_rows(min_row=2, values_only=True):
        iid_v = r_tb[idx["Issue ID"]]
        tb_v = r_tb[idx["Traceback"]]
        if iid_v is None or not tb_v:
            continue
        try:
            iid_int = int(iid_v)
        except (TypeError, ValueError):
            continue
        name = ""
        if tc_i is not None and r_tb[tc_i]:
            name = str(r_tb[tc_i])
        elif mdl_i is not None and r_tb[mdl_i]:
            name = f"{r_tb[bench_i] or ''}/{r_tb[mdl_i]}".strip("/")
        tb_by_issue[iid_int].append({
            "source": src_label,
            "name": name,
            "file": str(r_tb[tf_i]) if tf_i is not None and r_tb[tf_i] else "",
            "error": str(r_tb[idx["Error Message"]]) if r_tb[idx["Error Message"]] else "",
            "traceback": str(tb_v),
        })

rows = [tuple(c.value for c in r) for r in ws.iter_rows(min_row=2)]
print(f"loaded {len(rows)} rows")

if _args.issues_file is not None:
    ifile = _args.issues_file
    if not ifile.is_absolute():
        # resolve relative to the script's directory first, then CWD
        cand = Path(__file__).resolve().parent / ifile
        ifile = cand if cand.exists() else ifile
    wanted = {int(tok.lstrip("#")) for tok in ifile.read_text().split()
              if tok.strip().lstrip("#").isdigit()}
    _iid = C["Issue ID"]
    rows = [r for r in rows if r[_iid] is not None and int(r[_iid]) in wanted]
    found_ids = {int(r[_iid]) for r in rows}
    missing = sorted(wanted - found_ids)
    print(f"filtered to {len(rows)} rows matching {len(wanted)} requested IDs"
          f"{' (missing: ' + ', '.join(str(m) for m in missing) + ')' if missing else ''}")


# -------- helpers ----------------------------------------------------------
def clean(v) -> str:
    if v is None: return ""
    s = str(v).strip()
    if s.lower() == "none": return ""
    return s

def owner(r) -> str:
    a = clean(r[C["Assignee"]])
    if a: return a
    o = clean(r[C["owner_transferred"]])
    return o

def esc(s: str, max_len: int = 0) -> str:
    s = s.replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s)
    if max_len and len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s

def fmt_list(v) -> str:
    """action_TBD / action_reason cells are JSON arrays; join with '; '."""
    s = clean(v)
    if not s:
        return ""
    if s.startswith("["):
        try:
            import json as _j
            items = _j.loads(s)
            if isinstance(items, list):
                return "; ".join(str(x) for x in items)
        except Exception:
            pass
    return s


def wrap_cell(s, width: int = 80) -> str:
    """Soft-wrap a cell to `width` chars per visual line using <br>.
    Escapes pipes, collapses whitespace, word-wraps at word boundaries."""
    s = clean(s).replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    return "<br>".join(
        textwrap.wrap(s, width=width, break_long_words=True,
                      break_on_hyphens=False)
    )


# Patterns for Fix Approach beautification.
# Match file paths with common source/config extensions.
_PATH_RE = re.compile(
    r"(?<![`\w/.])"                               # not preceded by backtick or path char
    r"([\w./-]+\.(?:py|cpp|cmake|h|hpp|cu|cuh|xml|md|rst|yaml|yml|json))"
    r"(?![\w/.])"                                 # not followed by path char
)
# Sentence boundary: period+space before an uppercase ASCII letter, OR "; ".
# Avoids splitting on "e.g.", "vs.", "i.e." because they end with a period+space+lowercase.
_SPLIT_RE = re.compile(r"(?:\.\s+(?=[A-Z])|;\s+)")


def format_fix_approach(s, width: int = 80) -> str:
    """Bulletize Fix Approach text and wrap paths / quoted identifiers in
    backticks for readability.

    Pipeline:
      1. Clean & normalise whitespace.
      2. Wrap `'…'` single-quoted tokens and file paths in backticks
         (skipping content already inside backticks).
      3. Split on sentence boundaries ('. ' before uppercase, '; ').
      4. Soft-wrap each bullet to `width`; join with '<br>• '.
    """
    s = clean(s).replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    # Pass 1: protect existing backtick spans so we don't double-wrap.
    spans: list[str] = []
    def _save(m):
        spans.append(m.group(0))
        return f"\x00{len(spans)-1}\x00"
    s = re.sub(r"`[^`]+`", _save, s)

    # Pass 2: wrap single-quoted identifiers → backticks.
    s = re.sub(r"'([^'\s][^']{0,120}?)'", r"`\1`", s)

    # Pass 3: wrap file paths in backticks.
    s = _PATH_RE.sub(r"`\1`", s)

    # Restore protected spans.
    s = re.sub(r"\x00(\d+)\x00", lambda m: spans[int(m.group(1))], s)

    # Split into bullets on sentence boundaries. Preserve trailing '.' on
    # each bullet except those produced by '; ' splits.
    bullets: list[str] = []
    buf = s
    while True:
        m = _SPLIT_RE.search(buf)
        if not m:
            if buf.strip():
                bullets.append(buf.strip())
            break
        head = buf[:m.start()].strip()
        sep = m.group(0)
        if head:
            # If the split was on '. ', restore the period.
            if sep.startswith("."):
                head = head + "."
            bullets.append(head)
        buf = buf[m.end():]

    if not bullets:
        return ""

    # Wrap each bullet; prefix '• '; join with '<br>'.
    lines: list[str] = []
    for b in bullets:
        wrapped = textwrap.wrap(b, width=width - 2,   # -2 for '• ' prefix
                                break_long_words=False,
                                break_on_hyphens=False)
        if not wrapped:
            continue
        lines.append("• " + wrapped[0])
        # Continuation lines of the same bullet: indent with 2 spaces so
        # they visually align under the bullet glyph.
        for cont in wrapped[1:]:
            lines.append("&nbsp;&nbsp;" + cont)
    return "<br>".join(lines)


DUP_TOKEN = re.compile(r"#?(\d+)")


# ---- per-issue detail files ----------------------------------------------
DETAILS_DIR = OUT.parent / "details"
DETAILS_REL = "details"

def _wrap_para(s: str, width: int = 100) -> str:
    """Wrap free-form prose for detail-file body (plain markdown, no <br>)."""
    s = clean(s).replace("\r", " ")
    s = re.sub(r"[ \t]+", " ", s)
    paras = [p.strip() for p in re.split(r"\n\s*\n", s) if p.strip()]
    out = []
    for p in paras:
        p_flat = re.sub(r"\s+", " ", p)
        out.append("\n".join(textwrap.wrap(p_flat, width=width,
                                           break_long_words=False,
                                           break_on_hyphens=False)))
    return "\n\n".join(out)

def _bullets(raw) -> str:
    """Render action_TBD / action_reason JSON list cells as markdown bullets."""
    s = clean(raw)
    if not s:
        return "_(none)_"
    items: list[str] = []
    if s.startswith("["):
        try:
            import json as _j
            parsed = _j.loads(s)
            if isinstance(parsed, list):
                items = [str(x) for x in parsed]
        except Exception:
            items = [s]
    if not items:
        items = [s]
    return "\n".join(f"- {it}" for it in items)

def _fix_approach_md(raw) -> str:
    """Render Fix Approach as bulleted markdown for detail files."""
    s = clean(raw)
    if not s:
        return "_(none)_"
    s = re.sub(r"\s+", " ", s).strip()
    parts = [p.strip() for p in _SPLIT_RE.split(s) if p.strip()]
    if not parts:
        return s
    out = []
    for i, p in enumerate(parts):
        if i < len(parts) - 1 and not p.endswith(".") and not p.endswith(":"):
            p = p + "."
        out.append(f"- {p}")
    return "\n".join(out)

def _preview(raw, max_chars: int = 100) -> str:
    """Short preview for Fix Approach table cell: first sentence or max_chars."""
    s = clean(raw).replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    m = _SPLIT_RE.search(s)
    if m and m.start() <= max_chars:
        head = s[:m.start()].rstrip()
        if m.group(0).startswith("."):
            head += "."
        return head
    if len(s) <= max_chars:
        return s
    return s[:max_chars - 1].rstrip() + "…"

def fix_approach_cell(r) -> str:
    """Truncated Fix Approach + link to per-issue detail file."""
    iid = r[C["Issue ID"]]
    preview = _preview(r[C["Fix Approach"]])
    link = f"[→ details]({DETAILS_REL}/{iid}.md)"
    return f"{preview}<br>{link}" if preview else link

def write_detail(r) -> None:
    iid = r[C["Issue ID"]]
    if iid is None:
        return
    title = clean(r[C["Title"]])
    buf: list[str] = []
    a = buf.append
    a(f"# Issue #{iid}: {title}")
    a("")
    a(f"- **GitHub**: https://github.com/{REPO}/issues/{iid}")
    for key, label in [
        ("Category", "Category"),
        ("Priority", "Priority"),
        ("Status", "Status"),
        ("Assignee", "Assignee"),
        ("owner_transferred", "owner_transferred"),
        ("Reporter", "Reporter"),
        ("Labels", "Labels"),
        ("Dependency", "Dependency"),
        ("action_Type", "action_Type"),
    ]:
        v = clean(r[C[key]])
        a(f"- **{label}**: {v if v else '_(blank)_'}")
    a("")
    a("## action_TBD")
    a("")
    a(_bullets(r[C["action_TBD"]]))
    a("")
    a("## action_reason")
    a("")
    a(_bullets(r[C["action_reason"]]))
    a("")
    a("## Root Cause")
    a("")
    rc = clean(r[C["Root Cause"]])
    a(_wrap_para(rc) if rc else "_(none)_")
    a("")
    a("## Fix Approach")
    a("")
    a(_fix_approach_md(r[C["Fix Approach"]]))
    a("")
    try:
        iid_int = int(iid)
    except (TypeError, ValueError):
        iid_int = None
    tbs = tb_by_issue.get(iid_int, []) if iid_int is not None else []
    if tbs:
        a(f"## Test Cases & Traceback ({len(tbs)})")
        a("")
        for i, e in enumerate(tbs, start=1):
            hdr_bits = [f"{e['source']}"]
            if e["name"]:
                hdr_bits.append(e["name"])
            a(f"### {i}. {' · '.join(hdr_bits)}")
            a("")
            if e["file"]:
                a(f"- **Test File**: `{e['file']}`")
            if e["error"]:
                err = e["error"].replace("\r", " ").strip()
                err = re.sub(r"\s+", " ", err)
                a(f"- **Error**: {err[:300] + ('…' if len(err) > 300 else '')}")
            a("")
            a("```")
            a(e["traceback"].rstrip())
            a("```")
            a("")
    (DETAILS_DIR / f"{iid}.md").write_text("\n".join(buf))





def parse_dup_ids(dup_cell, action_tbd) -> list[int]:
    """Extract duplicate issue IDs from duplicated_issue cell, with
    action_TBD 'duplicate of …' fallback. Forward refs only."""
    ids: set[int] = set()
    for tok in re.split(r"[,;\s]+", clean(dup_cell)):
        m = DUP_TOKEN.fullmatch(tok)
        if m:
            ids.add(int(m.group(1)))
    if not ids:
        s = clean(action_tbd).lower()
        idx = s.find("duplicate of")
        if idx >= 0:
            for m in DUP_TOKEN.finditer(s[idx:idx+200]):
                ids.add(int(m.group(1)))
    return sorted(ids)


BACK = '_[↑ Back to Index](#sec-2)_'

def issue_link(iid) -> str:
    return f"[#{iid}](https://github.com/{REPO}/issues/{iid})"

def primary(action_type: str) -> str | None:
    if not action_type: return None
    parts = action_type.split("+")
    for cat in PRIMARY_ORDER:
        if cat in parts:
            return cat
    return parts[0]

def prio_key(r):
    p = clean(r[C["Priority"]]) or None
    return PRIO_RANK.get(p, 9)

def parse_dt(s) -> datetime | None:
    s = clean(s)
    if not s: return None
    try:
        return datetime.fromisoformat(s.replace("Z","+00:00"))
    except Exception:
        return None


# -------- bucket rows into sections ---------------------------------------
by_section: dict[str, list] = defaultdict(list)
per_cat = Counter()         # multi-label raw-atom category counts
per_primary = Counter()     # primary raw-atom category counts
per_primary_merged = Counter()  # primary counts after MERGE remap
per_prio = Counter()
per_status = Counter()
per_category_col = Counter()
empty_action = 0
check_cases_ids: list = []
check_cases_rows: list = []   # all rows with CHECK_CASES in action_Type (any position)
unclassified_rows: list = []  # rows with empty action_Type

for r in rows:
    at = clean(r[C["action_Type"]])
    if at:
        parts = at.split("+")
        for c in parts:
            per_cat[c] += 1
        prim = primary(at)
        per_primary[prim] += 1
        mprim = merged(prim)
        per_primary_merged[mprim] += 1
        if mprim in DEV_SECTIONS + QA_SECTIONS:
            # Exclude issues labeled 'random' from the CLOSE-or-SKIP bucket —
            # they are flaky tests that happened to pass in one CI run, not
            # true fixes. Applies only when raw primary was CLOSE.
            if prim == "CLOSE" and "random" in clean(r[C["Labels"]]).lower():
                pass
            else:
                by_section[mprim].append(r)
        if "CHECK_CASES" in parts:
            check_cases_ids.append(r[C["Issue ID"]])
            check_cases_rows.append(r)
    else:
        empty_action += 1
        unclassified_rows.append(r)
    per_prio[clean(r[C["Priority"]]) or "(blank)"] += 1
    per_status[clean(r[C["Status"]]) or "(blank)"] += 1
    per_category_col[clean(r[C["Category"]]) or "(blank)"] += 1

# Duplicated: duplicated_issue non-empty OR action_TBD mentions "duplicate of"
dup_rows = [r for r in rows if clean(r[C["duplicated_issue"]]) or
            "duplicate of" in clean(r[C["action_TBD"]]).lower()]

# Dependency: non-blank AND not upstream-pytorch AND not SYCL kernel:* AND not CPU fallback
def dep_ok(r) -> bool:
    d = clean(r[C["Dependency"]]).lower()
    if not d: return False
    if d == "upstream-pytorch": return False
    if d.startswith("sycl kernel"): return False
    if d == "cpu fallback": return False
    return True
dep_rows = [r for r in rows if dep_ok(r)]

# Exclude terminal QA rows (CLOSE / VERIFY_AND_CLOSE / SKIP / NOT_TARGET_CLOSE
# — sections 4.1, 4.2, 4.4, 4.6) from §6 and §7.
TERMINAL_QA = {"CLOSE", "VERIFY_AND_CLOSE", "SKIP", "NOT_TARGET_CLOSE"}
def is_terminal(r) -> bool:
    at = clean(r[C["action_Type"]])
    return primary(at) in TERMINAL_QA if at else False
dep_rows    = [r for r in dep_rows    if not is_terminal(r)]

# Dependency sub-buckets: upstream-pytorch + CPU fallback (terminal-QA excluded)
upstream_rows = [r for r in rows
                 if clean(r[C["Dependency"]]).lower() == "upstream-pytorch"
                 and not is_terminal(r)]
cpu_fb_rows   = [r for r in rows
                 if clean(r[C["Dependency"]]).lower() == "cpu fallback"
                 and not is_terminal(r)]

# New <=7 days (exclude terminal-QA rows)
recent_rows = [r for r in rows
               if (dt := parse_dt(r[C["Created Time"]])) and dt >= RECENT_CUTOFF
               and not is_terminal(r)]


# -------- render -----------------------------------------------------------
def render_table(row_list) -> str:
    """Standard table: Issue | Title | Owner | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels"""
    if not row_list:
        return "_No issues._\n"
    head = "| Issue | Title | Owner | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            wrap_cell(r[C["Title"]], 50),
            esc(owner(r), 25),
            esc(fmt_list(r[C["action_TBD"]]), 100),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def slug(s: str) -> str:
    """GitHub-style anchor slug."""
    s = (s or "").lower().strip()
    s = s.replace("_", "-")
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    return s


def render_section_by_category(row_list, section_num: str, cat_prefix: str) -> list[str]:
    """Split rows by their `Category` column; emit a `#### cat_prefix.N <Cat>` sub-heading with a table per group.
    Returns list of markdown lines and list of (anchor, label) for TOC."""
    buckets: dict[str, list] = defaultdict(list)
    for r in row_list:
        buckets[clean(r[C["Category"]]) or "(blank)"].append(r)
    out: list[str] = []
    toc: list[tuple[str, str]] = []
    for idx, cat in enumerate(sorted(buckets), start=1):
        rows_c = buckets[cat]
        anchor = f"sec-{section_num.replace('.','-')}-{idx}-{slug(cat)}"
        out.append(f'<a id="{anchor}"></a>')
        out.append(f"#### {section_num}.{idx} {cat}  ·  {len(rows_c)} issues")
        out.append("")
        out.append(BACK)
        out.append("")
        out.append(render_table(rows_c))
        out.append("")
        toc.append((anchor, f"{section_num}.{idx} {cat} ({len(rows_c)})"))
    return out, toc


def render_dep_table(row_list) -> str:
    head = "| Issue | Dependency | Title | Owner | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), clean(r[C["Dependency"]]), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            esc(clean(r[C["Dependency"]]), 30),
            wrap_cell(r[C["Title"]], 50),
            esc(owner(r), 25),
            esc(fmt_list(r[C["action_TBD"]]), 100),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def render_recent(row_list) -> str:
    head = "| Issue | Created | Title | Owner | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(
        row_list,
        key=lambda r: (prio_key(r), -(parse_dt(r[C["Created Time"]]) or TODAY).timestamp(), str(r[C["Issue ID"]])),
    )
    for r in sorted_rows:
        dt = parse_dt(r[C["Created Time"]])
        created = dt.strftime("%Y-%m-%d") if dt else ""
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            created,
            wrap_cell(r[C["Title"]], 50),
            esc(owner(r), 25),
            esc(fmt_list(r[C["action_TBD"]]), 100),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def render_dup_table(row_list) -> str:
    """§5 table: adds a `Duplicates` column after `Issue` with clickable issue links
    parsed from `duplicated_issue` (fallback: 'duplicate of …' clause in action_TBD)."""
    if not row_list:
        return "_No issues._\n"
    head = "| Issue | Duplicates | Title | Owner | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        ids = parse_dup_ids(r[C["duplicated_issue"]], r[C["action_TBD"]])
        if ids:
            dup_cell = ", ".join(issue_link(i) for i in ids)
        else:
            dup_cell = esc(clean(r[C["duplicated_issue"]]), 30)
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            dup_cell,
            wrap_cell(r[C["Title"]], 50),
            esc(owner(r), 25),
            esc(fmt_list(r[C["action_TBD"]]), 100),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


# ---- write per-issue detail files ----------------------------------------
DETAILS_DIR.mkdir(parents=True, exist_ok=True)
for _r in rows:
    write_detail(_r)
print(f"wrote {len(rows)} detail files to {DETAILS_DIR}")


# ---- assemble report ------------------------------------------------------
lines: list[str] = []
def w(s=""): lines.append(s)

w(f"# XPU Ops Bug Scrub Report{_args.title_suffix}")
w()
w(f"- **Repository**: `{REPO}`")
w(f"- **Generated**: {TODAY.strftime('%Y-%m-%d')} (cutoff for Section 7: {RECENT_CUTOFF.strftime('%Y-%m-%d')})")
w(f"- **Total issues in workbook**: {len(rows)}")
w(f"- **Classified (non-empty `action_Type`)**: {len(rows) - empty_action}")
w(f"- **Empty `action_TBD` (no verdict)**: {empty_action}")
w()

# -- Section 1: Summary ----------------------------------------------------
w("## 1. Summary")
w()
w(f"This report groups the {len(rows)} tracked torch-xpu-ops issues into action "
  f"buckets derived from the `action_Type` classification column of the triage "
  f"workbook. Each issue appears in at most one Action-Required or QA section, "
  f"chosen by its highest-priority category. Cross-cutting slices (duplicated "
  f"issues, external dependency blockers, newly filed issues) are listed "
  f"separately for visibility.")
w()
w("**Headline counts (primary category):**")
w()
w("| Bucket | Categories | Issues |")
w("|---|---|---:|")
dev_total = sum(per_primary_merged[c] for c in DEV_SECTIONS)
qa_total  = sum(per_primary_merged[c] for c in QA_SECTIONS)
w(f"| Developer action required | {', '.join(DEV_SECTIONS)} | {dev_total} |")
w(f"| QA action required | {', '.join(QA_SECTIONS)} | {qa_total} |")
w(f"| Duplicated | — | {len(dup_rows)} |")
w(f"| External dependency (non-upstream-pytorch, non-SYCL-kernel) | — | {len(dep_rows)} |")
w(f"| Upstream-pytorch | — | {len(upstream_rows)} |")
w(f"| CPU fallback | — | {len(cpu_fb_rows)} |")
w(f"| Filed within last 7 days | — | {len(recent_rows)} |")
w()

# -- Section 2: Index ------------------------------------------------------
w('<a id="sec-2"></a>')
w("## 2. Index")
w()
w('- [3. Action required (Developer)](#sec-3)')
w('  - [3.0 UNCLASSIFIED](#sec-3-0-unclassified)')
for i, c in enumerate(DEV_SECTIONS, start=1):
    w(f'  - [3.{i} {c}](#sec-3-{i}-{slug(c)})')
w('- [4. QA](#sec-4)')
for i, c in enumerate(QA_SECTIONS, start=1):
    w(f'  - [4.{i} {c}](#sec-4-{i}-{slug(c)})')
w('- [5. Duplicated issues](#sec-5)')
w('- [6. Dependency (external blockers)](#sec-6)')
w('  - [6.1 Third Parties](#sec-6-1-third-parties)')
w('  - [6.2 upstream-pytorch](#sec-6-2-upstream-pytorch)')
w('  - [6.3 CPU fallback](#sec-6-3-cpu-fallback)')
w('- [7. New submitted issues (<7 days)](#sec-7)')
w('- [8. Statistics](#sec-8)')
w()

# -- Section 3: Developer --------------------------------------------------
w('<a id="sec-3"></a>')
w("## 3. Action required (Developer)")
w()
w(BACK)
w()
w("Issues in this section require developer work before they can progress. "
  "Each subsection is split by `Category` (existing taxonomy column); "
  "rows inside each category table are sorted by `Priority` (P0 → P3).")
w()

# §3.0 Unclassified — rows with empty action_Type (Phase 4b emitted no verb)
w('<a id="sec-3-0-unclassified"></a>')
w(f"- **UNCLASSIFIED**  ·  {len(unclassified_rows)} issues")
w()
w(BACK)
w()
w(f"**{SECTION_TITLES['UNCLASSIFIED']}**")
w()
w(render_table(unclassified_rows))
w()

dev_toc: list[tuple[str, str]] = []  # [(anchor, label), ...] for extended TOC
for i, cat in enumerate(DEV_SECTIONS, start=1):
    section_num = f"3.{i}"
    anchor = f"sec-3-{i}-{slug(cat)}"
    w(f'<a id="{anchor}"></a>')
    w(f"- **{cat}**  ·  {len(by_section[cat])} issues")
    w()
    w(f"**{SECTION_TITLES[cat]}**")
    w()
    dev_toc.append((anchor, f"{section_num} {cat}"))
    sub_lines, sub_toc = render_section_by_category(by_section[cat], section_num, cat)
    for line in sub_lines:
        w(line)
    # extend TOC with sub-category links
    dev_toc.extend([(a, "  " + lbl) for a, lbl in sub_toc])
w()

# -- Section 4: QA ---------------------------------------------------------
w('<a id="sec-4"></a>')
w("## 4. QA")
w()
w(BACK)
w()
w("Issues in this section are ready for QA action (close, verify, reply, etc.). "
  "Rows sorted by `Priority` (P0 → P3).")
w()
for i, cat in enumerate(QA_SECTIONS, start=1):
    section_num = f"4.{i}"
    anchor = f"sec-4-{i}-{slug(cat)}"
    # §4.7 CHECK_CASES shows ALL rows tagged CHECK_CASES (any position in
    # action_Type), not just those where CHECK_CASES is the primary bucket.
    bucket = check_cases_rows if cat == "CHECK_CASES" else by_section[cat]
    w(f'<a id="{anchor}"></a>')
    w(f"- **{cat}**  ·  {len(bucket)} issues")
    w()
    w(f"**{SECTION_TITLES[cat]}**")
    w()
    w(render_table(bucket))
    w()

# -- Section 5: Duplicated -------------------------------------------------
w('<a id="sec-5"></a>')
w("## 5. Duplicated issues")
w()
w(BACK)
w()
w(f"Rows where `duplicated_issue` is set or `action_TBD` contains "
  f"\"duplicate of\".  —  {len(dup_rows)} issues.")
w()
w(render_dup_table(dup_rows))
w()

# -- Section 6: Dependency -------------------------------------------------
w('<a id="sec-6"></a>')
w("## 6. Dependency (external blockers)")
w()
w(BACK)
w()
w("Issues with a non-blank `Dependency` value, excluding `upstream-pytorch`, "
  "`CPU fallback`, and `SYCL kernel:*` (in-repo kernel pointers). "
  "Terminal-QA rows (CLOSE / VERIFY_AND_CLOSE / SKIP / NOT_TARGET_CLOSE) are "
  f"also excluded.  —  {len(dep_rows)} issues.")
w()
w('<a id="sec-6-1-third-parties"></a>')
w("- **Third Parties**")
w()
w(BACK)
w()
w(render_dep_table(dep_rows))
w()

w('<a id="sec-6-2-upstream-pytorch"></a>')
w("- **upstream-pytorch**")
w()
w(BACK)
w()
w("Issues whose fix lives in `pytorch/pytorch` (Dynamo/Inductor, AOTAutograd, "
  "`_prims_common`, benchmark harness, test-list sync, etc.). Terminal-QA rows "
  f"excluded.  —  {len(upstream_rows)} issues.")
w()
w(render_dep_table(upstream_rows))
w()

w('<a id="sec-6-3-cpu-fallback"></a>')
w("- **CPU fallback**")
w()
w(BACK)
w()
w("Issues where the XPU operator is missing and a CPU fallback is registered "
  "in torch-xpu-ops. Terminal-QA rows excluded.  —  "
  f"{len(cpu_fb_rows)} issues.")
w()
w(render_dep_table(cpu_fb_rows))
w()

# -- Section 7: New <=7 days -----------------------------------------------
w('<a id="sec-7"></a>')
w("## 7. New submitted issues (<7 days)")
w()
w(BACK)
w()
w(f"Issues created on or after {RECENT_CUTOFF.strftime('%Y-%m-%d')}, "
  "excluding terminal-QA rows.  —  "
  f"{len(recent_rows)} issues.")
w()
w(render_recent(recent_rows))
w()

# -- Section 8: Statistics -------------------------------------------------
w('<a id="sec-8"></a>')
w("## 8. Statistics")
w()
w(BACK)
w()
w(f"- Total rows: **{len(rows)}**")
w(f"- Classified (non-empty `action_Type`): **{len(rows) - empty_action}**")
w(f"- Empty `action_TBD` (no verdict yet): **{empty_action}**")
w(f"- Issues flagged for test-case existence check (`CHECK_CASES`): **{len(check_cases_ids)}**")
w()

w("- **Primary action_Type distribution (exclusive — one bucket per issue)**")
w()
w(BACK)
w()
w("Merged buckets (as rendered in §3 and §4):")
w()
w("| Bucket | Issues |")
w("|---|---:|")
for c in DEV_SECTIONS + QA_SECTIONS:
    if per_primary_merged[c]:
        w(f"| {c} | {per_primary_merged[c]} |")
w()
w("Raw atoms (pre-merge, for reference):")
w()
w("| Category | Issues |")
w("|---|---:|")
_seen = set()
for c in PRIMARY_ORDER + ["WAIT_EXTERNAL","FILE_ISSUE","CHECK_CASES"]:
    if c in _seen:
        continue
    _seen.add(c)
    if per_primary[c]:
        w(f"| {c} | {per_primary[c]} |")
w()

w("- **action_Type distribution (multi-label — each category counted once per issue)**")
w()
w(BACK)
w()
w("| Category | Issues |")
w("|---|---:|")
for c in PRIMARY_ORDER:
    if per_cat[c]:
        w(f"| {c} | {per_cat[c]} |")
w()

w("- **Priority distribution**")
w()
w(BACK)
w()
w("| Priority | Issues |")
w("|---|---:|")
for p in ["P0","P1","P2","P3","(blank)"]:
    if per_prio.get(p):
        w(f"| {p} | {per_prio[p]} |")
w()

w("- **Status distribution**")
w()
w(BACK)
w()
w("| Status | Issues |")
w("|---|---:|")
for s, n in per_status.most_common():
    w(f"| {s} | {n} |")
w()

w("- **Category column distribution (top 20)**")
w()
w(BACK)
w()
w("| Category | Issues |")
w("|---|---:|")
for c, n in per_category_col.most_common(20):
    w(f"| {c} | {n} |")
w()

w("- **CHECK_CASES issue IDs**")
w()
w(BACK)
w()
w(f"{len(check_cases_ids)} issues flagged for `check_case_avaliablity` (missing "
  f"XPU test case in repo):")
w()
w("> " + ", ".join(f"#{i}" for i in sorted(check_cases_ids)))
w()

# ---- write ----------------------------------------------------------------
OUT.write_text("\n".join(lines))
print(f"wrote {OUT} ({OUT.stat().st_size} bytes, {len(lines)} lines)")
