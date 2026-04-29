"""Add action_Type column to Issues sheet.

Classifies each row's action_TBD into 17 leaf categories, joined by "+"
in a fixed priority order so the column is sortable/groupable.

Category taxonomy (see session notes):
  CLOSE, NOT_TARGET_CLOSE, VERIFY_AND_CLOSE, TRACK_PR, IMPLEMENT,
  RETRIAGE_PRS, WAIT_EXTERNAL,
  ROOT_CAUSE, FILE_ISSUE, MONITOR, NEEDS_OWNER, NEED_ACTION,
  AWAIT_REPLY, CHECK_CASES, SKIP
"""
import re
from collections import Counter
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[7]
EXCEL = str(REPO / "opencode/issue_triage/result/torch_xpu_ops_issues.xlsx")

PRIORITY = [
    "CLOSE", "NOT_TARGET_CLOSE", "VERIFY_AND_CLOSE", "TRACK_PR",
    "IMPLEMENT", "RETRIAGE_PRS", "WAIT_EXTERNAL",
    "ROOT_CAUSE", "FILE_ISSUE", "MONITOR",
    "NEEDS_OWNER", "NEED_ACTION",
    "AWAIT_REPLY", "CHECK_CASES", "SKIP",
]


def classify(s: str) -> list[str]:
    if not s:
        return []
    low = s.lower()
    cats: set[str] = set()

    # --- CLOSE ---
    if "close the fixed issue" in low:           cats.add("CLOSE")
    if "close as duplicate" in low:              cats.add("CLOSE")
    if "confirm acceptable gap and close" in low: cats.add("CLOSE")

    # --- NOT_TARGET_CLOSE (covers both full-close and label-only/partial) ---
    if "label not_target and close" in low:
        cats.add("NOT_TARGET_CLOSE")
    elif ("not_target" in low or "not target" in low) and "label" in low:
        cats.add("NOT_TARGET_CLOSE")

    # --- SKIP (4a) ---
    if "skip issue" in low:                      cats.add("SKIP")

    # --- VERIFY_AND_CLOSE ---
    if re.search(r"verify.*close", low):         cats.add("VERIFY_AND_CLOSE")
    if "verify pr" in low or "validate pr" in low: cats.add("VERIFY_AND_CLOSE")
    if "validate pytorch/" in low:               cats.add("VERIFY_AND_CLOSE")
    if "verify fix from merged pr" in low:       cats.add("VERIFY_AND_CLOSE")
    if "verify alignment landed and close" in low: cats.add("VERIFY_AND_CLOSE")
    if re.search(r"verify (ci display|op-perf|e2e|sycl cpp)", low): cats.add("VERIFY_AND_CLOSE")
    if "reporter verify" in low:                 cats.add("VERIFY_AND_CLOSE")
    if low.strip() in ("verify and close", "verify fix and close"): cats.add("VERIFY_AND_CLOSE")
    if "assignee verify fix and close" in low:   cats.add("VERIFY_AND_CLOSE")

    # --- RETRIAGE_PRS (only when a PR is actually cited) ---
    # The bare 'RETRIAGE_PRS' token (Phase 4b emitted bucket-name as verb
    # for issues with no PRs found) is intentionally NOT mapped here —
    # see NEED_ACTION below.
    if "re-evaluate" in low and "closed unmerged" in low: cats.add("RETRIAGE_PRS")
    if "closed unmerged; reassess fix path" in low:       cats.add("RETRIAGE_PRS")
    if "re-validate cross-referenced prs" in low:         cats.add("RETRIAGE_PRS")
    if "resolve unresolved review comments on pr" in low: cats.add("RETRIAGE_PRS")
    if "address ci failures on pr" in low:                cats.add("RETRIAGE_PRS")

    # --- TRACK_PR ---
    if re.search(r"track.*to merge", low):       cats.add("TRACK_PR")
    if "track open pr" in low:                   cats.add("TRACK_PR")
    if "track pytorch/pytorch#" in low:          cats.add("TRACK_PR")
    if "track intel/torch-xpu-ops#" in low:      cats.add("TRACK_PR")
    if "track pr to merge" in low:               cats.add("TRACK_PR")
    if re.search(r"\bwait for prs?\b", low):     cats.add("TRACK_PR")
    if re.search(r"^land (pr|pytorch/|intel/|remaining|follow-up)", low): cats.add("TRACK_PR")
    if " land pytorch/" in low or " land intel/" in low: cats.add("TRACK_PR")
    if "land follow-up" in low or "land remaining" in low: cats.add("TRACK_PR")
    if re.search(r"move pr .* (out of wip )?and land", low): cats.add("TRACK_PR")
    if re.search(r"push pytorch/pytorch#\d+ review", low):   cats.add("TRACK_PR")
    if re.search(r"track pr\s+\S+#\d+", low):                cats.add("TRACK_PR")
    # Phase 4b: "Re-merge main into <branch> ..." (interim re-base)
    if "re-merge main into" in low:                          cats.add("TRACK_PR")

    # --- WAIT_EXTERNAL ---
    if "wait for onednn" in low:                 cats.add("WAIT_EXTERNAL")
    if "wait for triton" in low:                 cats.add("WAIT_EXTERNAL")
    if "mfdnn-" in low or "mlsl-" in low:        cats.add("WAIT_EXTERNAL")
    if "track in jira" in low:                   cats.add("WAIT_EXTERNAL")
    if "target pt 2." in low:                    cats.add("WAIT_EXTERNAL")
    if "after oneapi dle" in low or "revisit after" in low or "dle 2026" in low:
        cats.add("WAIT_EXTERNAL")
    if "track upstream" in low:                  cats.add("WAIT_EXTERNAL")
    # Phase 4b external trackers
    if "track driver/oneapi fix" in low:         cats.add("WAIT_EXTERNAL")
    if re.search(r"track dependency .*#\d+", low): cats.add("WAIT_EXTERNAL")
    if "gsd-" in low or "preqs-" in low:         cats.add("WAIT_EXTERNAL")

    # --- IMPLEMENT ---
    if "file fix pr" in low:                     cats.add("IMPLEMENT")
    if "needs new pr" in low:                    cats.add("IMPLEMENT")
    if re.search(r"\bimplement\b", low):         cats.add("IMPLEMENT")
    if "open xpu implementation pr" in low:      cats.add("IMPLEMENT")
    if "add input-tensor-expansion" in low:      cats.add("IMPLEMENT")
    if "evaluate adding" in low:                 cats.add("IMPLEMENT")
    if "migrate" in low and "usages off" in low: cats.add("IMPLEMENT")
    if "fix the duplicate division" in low:      cats.add("IMPLEMENT")
    if re.search(r"owner @\S+.* to skip .* tests", low): cats.add("IMPLEMENT")
    # Phase 4b "File upstream PR to ..."
    if "file upstream pr" in low:                cats.add("IMPLEMENT")
    if "submit upstream pr" in low:              cats.add("IMPLEMENT")

    # --- FILE_ISSUE (was INV_FILE_UPSTREAM) ---
    if "file upstream issue" in low:             cats.add("FILE_ISSUE")
    if "file driver-team issue" in low:          cats.add("FILE_ISSUE")
    if "file separate upstream issue" in low:    cats.add("FILE_ISSUE")

    # --- AWAIT_REPLY ---
    if "respond to open requests" in low:        cats.add("AWAIT_REPLY")
    # Phase 4b "Address comment AR from <user>: ..." -> AWAIT_REPLY (waiting on a person)
    if "address comment ar from" in low:         cats.add("AWAIT_REPLY")

    # --- INVESTIGATE sub-categories ---
    # NEED_ACTION (was NO_EVIDENCE): Phase 4b "no PR + no decision"
    if "needs owner investigation" in low:       cats.add("NEED_ACTION")
    if "weak/closed cross-ref candidates" in low: cats.add("NEED_ACTION")
    # Phase 4b "No action — investigate further" (em dash or hyphen)
    if "no action" in low and "investigate further" in low: cats.add("NEED_ACTION")
    # Phase 4b emitted the bare bucket name "RETRIAGE_PRS" as the verb
    # when its 6-vector PR search came up empty. There is no PR to
    # re-triage, so this is really a "needs investigation from scratch"
    # state — classify as NEED_ACTION rather than RETRIAGE_PRS.
    if re.search(r"\bretriage_prs\b", low):      cats.add("NEED_ACTION")

    # NEEDS_OWNER: no owner assigned / Phase 3 triage stub
    if "needs investigation / owner assignment" in low: cats.add("NEEDS_OWNER")
    if "assign owner" in low:                    cats.add("NEEDS_OWNER")
    if "triage owner needed" in low:             cats.add("NEEDS_OWNER")

    # ROOT_CAUSE: owner/assignee actively working
    if "assignee continue" in low:               cats.add("ROOT_CAUSE")
    if "assignee investigate" in low:            cats.add("ROOT_CAUSE")
    if "assignee triage" in low:                 cats.add("ROOT_CAUSE")
    if "assignee re-baseline" in low:            cats.add("ROOT_CAUSE")
    if re.search(r"owner @\S+ to (investigate|triage|re-check)", low): cats.add("ROOT_CAUSE")

    # RETRIAGE_PRS (also catches standalone re-triage of cross-refs)
    # (rule already handled above in RETRIAGE_PRS block)

    # MONITOR
    if re.search(r"owner @\S+.*to (maintain|scope|expose)", low):
        cats.add("MONITOR")
    if re.search(r"owner @\S+(?: / @\S+)+.*to track", low):
        cats.add("MONITOR")
    if re.search(r"owner @\S+ to evaluate\b", low) and "evaluate adding" not in low:
        cats.add("MONITOR")
    # Phase 4b "Track release X.Y.Z through Phase ..."
    if re.search(r"track release \d", low):      cats.add("MONITOR")

    # --- CHECK_CASES ---
    if "check_case_avaliablity" in low:          cats.add("CHECK_CASES")

    return [c for c in PRIORITY if c in cats]


def main():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Issues"]
    hdr = [c.value for c in ws[1]]
    I_ACT = hdr.index("action_TBD")
    I_ASSIGN = hdr.index("Assignee")
    I_OWNER = hdr.index("owner_transferred")
    I_STATUS = hdr.index("Status")

    # add action_Type column if absent
    if "action_Type" in hdr:
        I_TYPE = hdr.index("action_Type")
    else:
        I_TYPE = ws.max_column
        ws.cell(row=1, column=I_TYPE + 1, value="action_Type")

    rows = list(ws.iter_rows(min_row=2))
    total = len(rows)
    empty = 0
    per_cat = Counter()
    combos = Counter()
    unclassified: list[tuple[int, str]] = []

    def is_blank(v) -> bool:
        if v is None: return True
        s = str(v).strip()
        return not s or s.lower() == "none"

    for r in rows:
        act = r[I_ACT].value
        cats = classify(act) if act else []

        # Row-level rule: NEEDS_OWNER applies whenever an OPEN issue has
        # neither an Assignee nor a non-empty owner_transferred,
        # regardless of action_TBD verbs. This catches the truly
        # unowned issues that the verb-based classifier never tags.
        # Exception: if the row already has a terminal verdict
        # (CLOSE / NOT_TARGET_CLOSE / SKIP / VERIFY_AND_CLOSE), the
        # next step is closing — it doesn't really need an owner first.
        TERMINAL = {"CLOSE", "NOT_TARGET_CLOSE", "SKIP", "VERIFY_AND_CLOSE"}
        status = (r[I_STATUS].value or "").strip().lower() if r[I_STATUS].value else ""
        if status != "closed" and not (set(cats) & TERMINAL):
            if is_blank(r[I_ASSIGN].value) and is_blank(r[I_OWNER].value):
                cats = list(cats) if cats else []
                if "NEEDS_OWNER" not in cats:
                    cats.append("NEEDS_OWNER")

        if not act and not cats:
            empty += 1
            r[I_TYPE].value = None
            continue
        if not cats:
            unclassified.append((r[0].value, act))
            r[I_TYPE].value = "UNCLASSIFIED"
            continue
        # Re-sort by PRIORITY since NEEDS_OWNER may have been appended above.
        cats = [c for c in PRIORITY if c in set(cats)]
        label = "+".join(cats)
        r[I_TYPE].value = label
        for c in cats:
            per_cat[c] += 1
        combos[label] += 1

    wb.save(EXCEL)

    print(f"Total rows:        {total}")
    print(f"Empty action_TBD:  {empty}")
    print(f"Unclassified:      {len(unclassified)}")
    print("\n=== Per-category issue counts (multi-label) ===")
    for c in PRIORITY:
        print(f"  {c:<20}{per_cat[c]:>5}")
    print(f"\n=== Combo signatures ({len(combos)} distinct) ===")
    for combo, n in combos.most_common():
        print(f"  {n:>4}  {combo}")
    if unclassified:
        print("\n=== UNCLASSIFIED ===")
        for iid, s in unclassified:
            print(f"  #{iid}: {s!r}")


if __name__ == "__main__":
    main()
