#!/usr/bin/env python3
"""
Create Not applicable Sheet
Extracts wontfix/not_target issues and creates a summary sheet
"""
import openpyxl
import re

def create_not_applicable_sheet(
    excel_path: str = '/home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx',
    output_path: str = None
):
    """Create Not applicable sheet from issues with wontfix/not_target labels"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Issues']
    
    not_appliable = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        issue_id = row[0]
        title = row[1] if row[1] else ''
        labels = row[5] if row[5] else ''
        
        if 'wontfix' in labels.lower() or 'not_target' in labels.lower():
            torch_ops = extract_torch_ops(title)
            not_appliable.append({
                'issue_id': issue_id,
                'title': title,
                'torch_ops': torch_ops,
                'labels': labels,
                'reason': 'wontfix/not_target'
            })
    
    if 'Not applicable' in wb.sheetnames:
        del wb['Not applicable']
    
    ws_na = wb.create_sheet('Not applicable')
    ws_na.append(['Issue ID', 'Title', 'Torch Ops/API', 'Labels', 'Reason'])
    
    for item in not_appliable:
        ws_na.append([
            item['issue_id'],
            item['title'],
            item['torch_ops'],
            item['labels'],
            item['reason']
        ])
    
    save_path = output_path or excel_path
    wb.save(save_path)
    
    return len(not_appliable)

def extract_torch_ops(title: str) -> str:
    """Extract torch ops/API references from issue title"""
    ops = []
    
    aten_matches = re.findall(r"aten::[\w_.]+", title)
    ops.extend(aten_matches)
    
    quoted_aten = re.findall(r"'aten::[\w_]+'", title)
    ops.extend([m.strip("'") for m in quoted_aten])
    
    torch_matches = re.findall(r'torch\.[\w.]+(?:\.[\w.]+)*', title)
    for m in torch_matches:
        if 'torch' in m and len(m) < 30 and not m.startswith('torch._'):
            ops.append(m)
    
    if 'TypedStorage' in title or 'TypedTensors' in title:
        ops.append('TypedStorage/TypedTensors')
    if 'cuda_monkeypatch' in title:
        ops.append('cuda_monkeypatch (CUDA specific)')
    if 'scaled_dot_product_attention' in title.lower():
        ops.append('scaled_dot_product_attention')
    
    unique_ops = list(dict.fromkeys(ops))
    
    return ', '.join(unique_ops) if unique_ops else title[:60].strip()

if __name__ == '__main__':
    import sys
    
    excel_path = sys.argv[1] if len(sys.argv) > 1 else \
        '/home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx'
    
    count = create_not_applicable_sheet(excel_path)
    print(f'Created "Not applicable" sheet with {count} entries')