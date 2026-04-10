#!/usr/bin/env python3
"""
Process Issues sheet to add owner_transfer, action_TBD, priority, category, and root cause analysis.
This script is extracted from update_test_results.py for modularity.

Usage:
    python process_issues_sheet.py

Input:
    - /home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx (with Issues, Test Cases, E2E Test Cases sheets)

Output:
    - Updated /home/daisydeng/ai_for_validation/opencode/issue_triage/result/torch_xpu_ops_issues.xlsx
"""

import openpyxl
import os
import re
import time

ROOT_DIR = "/home/daisydeng"
RESULT_DIR = os.environ.get("RESULT_DIR", "/home/daisydeng/ai_for_validation/opencode/issue_triage/result")


# ============================================================================
# VERSION/BRANCH HELPERS
# ============================================================================

def extract_version_info(issue_content):
    """Extract version information from issue content."""
    version_patterns = [
        r'2\.5\.0\s*rc\d*', r'2\.5\.0\s*rc', r'2\.5\.0\s*dev',
        r'2\.4\.1\s*rc\d*', r'2\.4\.1\s*rc', r'2\.4\.1\s*dev',
        r'2\.4\.0\s*rc\d*', r'2\.4\.0\s*dev', r'2\.4\.0\s*stable',
        r'2\.3\.1\s*dev', r'2\.3\.1\s*stable', r'2\.3\.0',
        r' nightly ', r' nightlybuild', r' build ',
    ]
    found = []
    issue_lower = issue_content.lower() if issue_content else ''
    for pattern in version_patterns:
        matches = re.findall(pattern, issue_lower)
        found.extend(matches)
    return ' '.join(set(found)) if found else ''


def is_public_branch(version_str):
    """Check if version string indicates a public branch vs internal branch."""
    public_patterns = [
        'nightly', 'dev', 'rc1', 'rc2', 'rc3', 'rc4', 'rc5', 'rc6', 'rc7', 'rc8',
        r'\d+\.\d+\.\d+\w*', r'2\.5\.0\w*', r'2\.4\.1\w*', r'2\.4\.0\w*',
    ]
    for pattern in public_patterns:
        if re.search(pattern, version_str, re.IGNORECASE):
            return True
    return 'internal' in version_str.lower() or ' privately' in version_str.lower()


def check_info_requested_to_reporter(issue_content):
    """Check if the issue is already asking reporter to provide info."""
    content = issue_content.lower() if issue_content else ''
    keywords = ['reproduce', 'provide the full', 'please provide', 'need more information', 
                'missing error', 'lon des', 'what is your', 'can you tell', 'tell us more',
                'create a bug report', 'fill in the template', 'more details', 'additional context',
                'help us understand', 'reproduce the issue', 'reproduce this', 'steps to reproduce',
                'any update', 'any progress', 'bug still exists', 'still failing',
                'were you able', 'were you able to', "didn't receive", 'no response',
                'awaiting response', 'pending feedback', 'need feedback',
                'waiting for reporter', 'need info', 'need additional']
    return any(kw in content for kw in keywords)


def check_info_requested_to_reporter_llm(issue_title, issue_summary, error_msg, traceback):
    """Use LLM to determine if more info needs to be requested from reporter."""
    import requests
    import json

    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"

    combined_text = f"{issue_title} {issue_summary}".strip()
    combined_text = combined_text[:500]

    if not combined_text and not error_msg:
        return "Ready to analyze"

    prompt = f"""You are analyzing an issue to determine if additional information is needed from the reporter.

Issue Title: {issue_title}
Issue Summary (truncated): {issue_summary[:200] if issue_summary else 'None'}
Error Message (truncated): {str(error_msg)[:500] if error_msg else 'None'}

Analyze the issue and determine the next step:
1. If the issue has sufficient error information and test details to proceed with analysis, return: "Ready to analyze"
2. If the issue is missing critical error information or has vague error descriptions, return: "Need more information"
3. If the issue lacks reproduction steps, return: "Need reproduce steps"
4. If the issue needs additional details like version info, environment, or performance data, return specific details about what's missing

Common issues:
- Missing error message or traceback
- Vague error description like "Test failed" without specifics
- Need version information (PyTorch version, XPU driver, etc.)
- Need environment details
- Need reproduction steps for non-UT failures
- Need performance baseline or comparison data

Be specific about what's missing. If the issue has error logs, stack trace, or test case details, it's ready to analyze.

YOUR ANSWER (return ONLY one of these, no explanation):
- "Ready to analyze"
- "Need more information"
- "Need reproduce steps"
- "Need specific details about: [what's missing]"

YOUR ANSWER:"""

    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0,
        "max_tokens": 50
    }

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    start_time = time.time()

    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time

        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            content = content.replace('[', ' <').replace(']', '> ')
            content = re.sub(r'<[^>]*>', '', content)
            content = re.sub(r'\s+', ' ', content).strip()
            return content
        return f"API Error: {response.status_code}"
    except Exception as e:
        return f"Error: {str(e)[:30]}"


# ============================================================================
# PRIORITY DETERMINATION LLM
# ============================================================================

def determine_priority_llm(title, summary, error_msg, test_module, labels_str, test_cases_info):
    """Use Qwen3-32B via internal API to determine priority of an issue."""
    import requests
    import json
    import time

    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"

    if not title:
        return "P2", "No title provided - defaulting to P2"

    tc_info_str = ""
    failed_tcs = []
    if test_cases_info:
        for tc in test_cases_info[:5]:
            tc_status = tc.get('status') or ''
            if tc_status not in ['passed', 'skipped', 'not found']:
                failed_tcs.append(tc.get('test_case', '')[:50] if tc.get('test_case') else 'Unknown TC')
        if failed_tcs:
            tc_info_str = f"Failed tests: {', '.join(failed_tcs)}"
    
    error_info = ""
    if error_msg and error_msg not in ['Test case not found', 'Skipped']:
        error_info = f"Error details: {str(error_msg)[:200]}"

    prompt = f"""You are analyzing a PyTorch XPU issue to determine its priority.

Issue Title: {title}
Issue Summary (truncated): {summary[:300] if summary else 'None'}
Test Module: {test_module}
Labels: {labels_str}

Test Case Info:
{tc_info_str}

{error_info}

Determine the priority (P0, P1, P2, or P3) based on:
- P0: Build crash, critical regression, or model-level impact
- P1: Many UT failures (5+), or E2E accuracy issues, or security fixes
- P2: Few UT failures (1-5), or performance issues without regression, or feature gaps
- P3: Minor issues, documentation updates, or enhancement requests

Rules:
- E2E issues: P1 for accuracy/functionality, P2 for performance
- UT issues: P1 if >5 failures or involves sycl/cUDA backend, P2 if 1-5 failures
- Build/crash: Always P0
- Regression: Check if issue mentions "was pass[ing]" or "previously pass[ing]"
- Upstream issues (ut_upstream/inductor labels): Use P2 unless critical

Return format: "P<0-3> - <one line reason>"
Example: "P0 - Critical regression in E2E test with Huggingface models failed accuracy check"
Example: "P2 - Few UT failures (2 test cases) involving aten matrix operations"

YOUR ANSWER:"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": LLM_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.0,
        "max_tokens": 100
    }

    start_time = time.time()

    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time

        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            content = content.replace('[', ' <').replace(']', '> ')
            content = re.sub(r'<[^>]*>', '', content)
            content = re.sub(r'has ATTR\b', '', content)
            content = re.sub(r'\s+', ' ', content).strip()
            
            print(f"    LLM took {elapsed:.1f}s")
            
            if content.startswith('P') or content.startswith('p'):
                return content, ''
            else:
                return "P2", f"LLM returned non-standard format: {content[:50]}"
        return f"API Error: {response.status_code}", ''

    except Exception as e:
        return f"Error: {str(e)[:30]}", ''


# ============================================================================
# CATEGORY DETERMINATION
# ============================================================================

def determine_category(title, summary, test_cases_str, traceback, test_module, labels):
    """Rule-based fallback category determination."""
    categories = [
        ('distributed', 'Distributed'),
        ('torchao', 'TorchAO'),
        ('torch.export', 'PT2E'),
        ('dynamo', 'PT2E'),
        ('fake_tensor', 'PT2E'),
        ('exportedprogram', 'PT2E'),
        ('flash_attention', 'Flash Attention/Transformer'),
        ('scaled_dot_product_attention', 'Flash Attention/Transformer'),
        ('sdpa', 'Flash Attention/Transformer'),
        ('sparse', 'Sparse'),
        ('torch.compile', 'Inductor/Compilation'),
        ('inductor', 'Inductor/Compilation'),
        ('triton', 'Inductor/Compilation'),
        ('cuda runtime', 'Torch Runtime'),
        ('cudamalloc', 'Torch Runtime'),
        ('cudamemcpy', 'Torch Runtime'),
        ('out of memory', 'Torch Runtime'),
        ('oom', 'Torch Runtime'),
        ('device kernel', 'Torch Runtime'),
        ('cuda', 'Torch Runtime'),
        ('xpu', 'Torch Runtime'),
        ('sycl', 'Torch Runtime'),
        ('aten::', 'Torch Operations'),
        ('native', 'Torch Operations'),
        ('operator', 'Torch Operations'),
        ('backend', 'Backend/Device'),
        ('device', 'Backend/Device'),
        ('illegal memory', 'Backend/Device'),
        ('dtypel', 'Dtype/Precision'),
        ('float16', 'Dtype/Precision'),
        ('bfloat16', 'Dtype/Precision'),
        ('precision', 'Dtype/Precision'),
        ('autocast', 'Dtype/Precision'),
        ('not implemented', 'Feature Not Supported'),
        ('unsupported', 'Feature Not Supported'),
        ('skipped', 'Skip/No Test Exists'),
        ('skip', 'Skip/No Test Exists'),
        ('todo', 'Skip/No Test Exists'),
        ('unittest.skip', 'Skip/No Test Exists'),
    ]
    
    keywords = ' '.join([title or '', summary or '', test_cases_str or '', traceback or '']).lower()
    
    for pattern, category in categories:
        if pattern in keywords:
            return category
            
    return ""


# ============================================================================
# CATEGORY DETERMINATION LLM
# ============================================================================

def determine_category_llm(title, summary, test_cases_info, test_module, labels):
    """Use Qwen3-32B via internal API to determine the category of an issue."""
    import requests
    import json
    import time
    import re

    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"

    if not title and not summary:
        return "Others", ''

    tc_info_str = ""
    if test_cases_info:
        for tc in test_cases_info:
            error_val = tc.get('error_msg') or ''
            tc_info_str += f"- Test: {tc.get('test_case', '')}, Error: {str(error_val)[:100]}\n"

    test_module_str = test_module or "Unknown"
    labels_str = labels or "None"

    prompt = f"""You are analyzing PyTorch XPU issue to determine its category.

Issue Title: {title}
Issue Summary: {summary}
Test Module: {test_module_str}
Labels: {labels_str}

Test Cases Info:
{tc_info_str}

Categorize this issue into ONE of:

    1. Distributed - Keywords: distributed, XCCL, NCCL, Gloo, ProcessGroup, DDP, FSDP, torch.distributed, collective communication, multi-node, timeout, rendezvous, init_method, reduce_scatter, all_gather

    2. TorchAO - Keywords: torchao, quantize_, int4_weight_only, int8_dynamic_activation, fp8, nf4, autoquant, quantization_config, Adam8bit, AdamW4bit, Lion8bit, PagedAdam, OptimizerWithQuantization, quantized_optimizer, ao/sparsity, apply_dynamic_quant, change_linear_weights_to_int8_packed, QuantizedLinear, from_float (AO context), packed weight, dequantization, int4, int8 (when paired with torchao)

    3. PT2E - Keywords: torch.export(), Dynamo, fake_tensor, ExportedProgram, AOT, torch._export, graph break, tracing, exported_program.run_decompositions

    4. Flash Attention/Transformer - Keywords: flash_attention, scaled_dot_product_attention, SDPA, attention mask, transformer layer, F.scaled_dot_product_attention, memory-efficient attention, FlexAttention

    5. Sparse - Keywords: sparse tensor, CSR, CSC, COO, torch.sparse, sparse matrix multiplication, sparse_mask, to_sparse(), sparse_coo_tensor

    6. Inductor/Compilation - Keywords: torch.compile(), Inductor, Triton, codegen, AOT Autograd, FX graph, torch._inductor, compilation cache, torch._dynamo

    7. Torch Runtime - Keywords: CUDA runtime, cudaMalloc, cudaMemcpy, out of memory (OOM), device kernel launch, stream synchronization, cudaStreamSynchronize, device-side assert, cudaError, illegal memory access, context management, cudaSetDevice, device initialization, cudaGetDevice, driver error, CUDA_VISIBLE_DEVICES, memory leak, allocation failure, device reset

    8. Torch Operations - Keywords: operator implementation, aten::, native::, custom op, register_operator, operator overloading, tensor operation dispatch, kernel selection, op signature mismatch, unsupported op on device, op not implemented for device, device-specific op behavior, backward pass operation, autograd op

    9. Dtype/Precision - Keywords: dtype mismatch, float16, bfloat16, float32, mixed precision, autocast, GradScaler, precision loss, NaN/inf due to dtype, to(dtype=...), torch.int8 (without torchao), legacy torch.quantization

    10. Feature Not Supported - Keywords: unimplemented operator, missing kernel, feature not available in this build, unsupported combination, "not implemented for"

    11. Skip/No Test Exists - Keywords: test skipped, @unittest.skip, missing test decorator, CI test gap, skipIfNoTorchAO

    12. Others - None of the above (only if truly uncategorizable)

Classification Rules:
    - Select exactly ONE category
    - For int4/int8/fp8 errors: TorchAO takes precedence over Dtype/Precision
    - For quantized optimizers (Adam8bit, Lion8bit): TorchAO takes precedence
    - For CUDA runtime errors (memory, sync, context): Choose Torch Runtime
    - For operator dispatch/kernel errors: Choose Torch Operations
    - For device-specific op implementation issues: Choose Torch Operations
    - Use Others only as a last resort

    Distinction between Torch Runtime and Torch Operations:
        - Torch Runtime = Errors related to device management, memory allocation, synchronization, driver/runtime API calls
        - Torch Operations = Errors related to specific operator execution, kernel selection, op dispatch, custom op registration

Return the category AND a brief reason for your classification:
- The reason is REQUIRED, not optional
- Keep it concise (max 80 characters)
- Explain why you chose this category based on the issue details

Format: "X - Category Name | reason"
Example: "6 - Inductor/Compilation | compilation error in torch.compile"

YOUR ANSWER (must include reason after the pipe symbol):"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 100
    }

    start_time = time.time()

    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        elapsed = time.time() - start_time

        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            
            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            
            # Remove all bracket patterns and their content
            content = content.replace('[', ' <').replace(']', '> ')
            
            # Clean up markers
            content = re.sub(r'<[^>]*>', '', content)
            
            # Remove common thinking tag leftovers
            content = re.sub(r'has ATTR\b', '', content)
            
            # Clean up whitespace
            content = re.sub(r'\s+', ' ', content).strip()

            # Parse category and reason - format: "X - Category Name | reason"
            if '|' in content:
                parts = content.split('|', 1)
                category_part = parts[0].strip()
                reason = parts[1].strip() if len(parts) > 1 else ''
                return category_part, reason
            else:
                match = re.search(r'(\d+)\s*[-–]\s*[\w\s/]+', content)
                if match:
                    category = f"{match.group(1)} - " + content[match.start():match.end()].split('-')[-1].strip()
                    return category, ''
                else:
                    return content.strip()[:50], ''

        return f"API Error: {response.status_code}", ''

    except Exception as e:
        return f"Error: {str(e)[:30]}", ''


# ============================================================================
# ROOT CAUSE ANALYSIS
# ============================================================================

def analyze_root_cause_llm(issue_id, issue_title, issue_summary, test_file, test_class, test_case, error_msg, traceback, test_module=None):
    """Use Qwen3-32B via internal API to identify root cause of test failures."""
    import requests
    import time
    import re

    LLM_ENDPOINT = "http://10.239.15.43/v1/chat/completions"
    LLM_API_KEY = os.environ.get("OPENCODE_API_KEY", "sk-xxxxxxxxxx")
    LLM_MODEL = "Qwen3-32B"

    full_error = f"{error_msg}\n\n{traceback}" if traceback else error_msg
    full_error = full_error[:1500] if full_error else ''

    prompt = f"""You are analyzing PyTorch XPU test failures to identify the root cause.

Issue ID: {issue_id}
Issue Title: {issue_title}
Issue Summary: {issue_summary}

Test Information:
- Test File: {test_file}
- Test Class: {test_class if test_class else 'Unknown'}
- Test Case: {test_case if test_case else 'Unknown'}
- Test Module: {test_module if test_module else 'Unknown'}

Error and Traceback:
{full_error}

Identify the root cause category (choose ONE):
1. **Backend/Device Issue** - XPU device initialization, driver compatibility, SYCL runtime issues, device not found, memory allocation failures
2. **Operator Not Implemented** - aten:: operator missing for XPU backend, internal、中翰思、运营商、XPU算子缺失
3. **Dtype/Precision Issue** - Data type mismatches, precision problems, float16/bfloat16 issues, NaN/inf values
4. **Compilation/Graph Issue** - torch.compile failures, Dynamo tracing errors, Inductor codegen issues
5. **Environment/Configuration** - Missing dependencies, incorrect setup, version mismatches
6. **Test Itself Issue** - Flaky test, incorrect test logic, missing test fixtures
7. **Regression** - Was working before, now failing (check if related to recent PyTorch changes)
8. **Feature Gap** - Functionality not yet supported on XPU
9. **Performance Issue** - Slow execution, memory inefficiency, missing optimizations
10. **Input/Parameter Issue** - Incorrect input shape, invalid parameters, type mismatches

Provide a concise root cause analysis (max 150 characters):
- "Operator not implemented: aten::xxx not supported for dtype X on XPU"
- "Backend error: XPU driver out of memory during kernel launch"
- "Compilation error: torch.compile failed with FusionGroup issue"
- "Dtype mismatch: float32 expected but bfloat16 provided"

IMPORTANT: If the issue lacks specific error information (no error_msg or traceback), return an empty string.

YOUR ANSWER (root cause only, no explanation):"""

    headers = {
        "Authorization": f"Bearer {LLM_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": LLM_MODEL,
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 80
    }

    try:
        response = requests.post(LLM_ENDPOINT, headers=headers, json=payload, timeout=180)
        
        if response.status_code == 200:
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            
            # Clean up thinking tags
            content = re.sub(r'\[TO\]', '', content)
            content = re.sub(r'\[/TO\]', '', content)
            content = re.sub(r'\[RESULT\]', '', content)
            content = re.sub(r'\[/RESULT\]', '', content)
            content = re.sub(r'\[/TD\]', '', content)
            content = content.replace('[', ' <').replace(']', '> ')
            content = re.sub(r'<[^>]*>', '', content)
            content = re.sub(r'^[\s\-–1-10\.]+', '', content).strip()
            content = re.sub(r'^[\s\.]+', '', content).strip()
            
            if not content or len(content) < 5:
                return ""
            
            if content.lower() in ['null', 'none', 'empty', 'n/a', 'na']:
                return ""
            
            # Check if content indicates missing info
            if any(phrase in content.lower() for phrase in ['missing', 'no error', 'cannot determine', 'insufficient', 'not provided']):
                return ""
            
            # Return first line only
            content = content.split('\n')[0].strip()
            return content[:150]
        else:
            return ""

    except Exception as e:
        return ""


def analyze_root_cause(issue_title, issue_summary, test_file, test_class, test_case, error_msg, traceback, test_module=None):
    """Rule-based root cause analysis with keyword matching."""
    root_causes = {
        'backend': ['xpu', 'sycl', 'intel', 'gpu', 'driver', 'illegal memory', 'device not found', 'out of memory', 'runtime error'],
        'operator': ['not implemented', 'unsupported', 'aten::', 'operator not found', 'kernel'],
        'dtype': ['dtype', 'float16', 'bfloat16', 'float32', 'int8', 'int4', 'precision', 'nan', 'inf', 'overflow'],
        'inductor': ['inductor', 'compile', 'dynamo', 'triton', '编译', 'codegen'],
        'performance': ['slow', 'performance', 'latency', 'memory', 'throughput'],
    }
    
    issue_text = ' '.join([issue_title or '', issue_summary or '', error_msg or '', traceback or '']).lower()
    test_text = ' '.join([test_file or '', test_class or '', test_case or '']).lower()
    
    for category, keywords in root_causes.items():
        for keyword in keywords:
            if keyword in issue_text or keyword in test_text:
                if category == 'backend':
                    return f"Backend/Device Issue - {keyword} related"
                elif category == 'operator':
                    return f"Operator Issue - {keyword} related"
                elif category == 'dtype':
                    return f"Dtype/Precision Issue - {keyword} related"
                elif category == 'inductor':
                    return f"Compilation Issue - {keyword} related"
                elif category == 'performance':
                    return f"Performance Issue - {keyword} related"
    
    return ""


# ============================================================================
# MAIN PROCESS FUNCTION
# ============================================================================

def process_issues_sheet(wb):
    """Process Issues sheet to add owner_transfer, action_TBD, priority, category, root cause."""
    ws_issues = wb['Issues']
    ws_test = wb['Test Cases']

    ws_issues.cell(1, 19, 'owner_transfer')
    ws_issues.cell(1, 20, 'action_TBD')
    ws_issues.cell(1, 21, 'duplicated_issue')
    ws_issues.cell(1, 22, 'priority')
    ws_issues.cell(1, 23, 'priority_reason')
    ws_issues.cell(1, 24, 'Category')
    ws_issues.cell(1, 25, 'Root Cause')
    ws_issues.cell(1, 26, 'category_reason')

    MAX_LLM_ROOT_CAUSE = 500
    MAX_LLM_CATEGORY = 500
    MAX_LLM_PRIORITY = 500
    llm_root_cause_count = 0
    llm_category_count = 0
    llm_priority_count = 0
    root_cause_cache = {}

    issue_test_results = {}
    for row in range(2, ws_test.max_row + 1):
        issue_id = ws_test.cell(row, 1).value
        xpu_status = ws_test.cell(row, 11).value
        stock_status = ws_test.cell(row, 16).value
        cuda_case_exist = ws_test.cell(row, 18).value
        dup_issue = ws_test.cell(row, 21).value

        if issue_id not in issue_test_results:
            issue_test_results[issue_id] = {
                'xpu_statuses': set(),
                'stock_statuses': set(),
                'cuda_case_not_exist': False,
                'duplicated_issues': set()
            }

        if xpu_status:
            issue_test_results[issue_id]['xpu_statuses'].add(xpu_status)
        if stock_status:
            issue_test_results[issue_id]['stock_statuses'].add(stock_status)
        if cuda_case_exist == 'No':
            issue_test_results[issue_id]['cuda_case_not_exist'] = True
        if dup_issue:
            for dup_id in str(dup_issue).split(','):
                issue_test_results[issue_id]['duplicated_issues'].add(dup_id.strip())

    ws_e2e = wb['E2E Test Cases']
    for row in range(2, ws_e2e.max_row + 1):
        issue_id = ws_e2e.cell(row, 1).value
        e2e_status = ws_e2e.cell(row, 13).value

        if issue_id not in issue_test_results:
            issue_test_results[issue_id] = {
                'xpu_statuses': set(),
                'stock_statuses': set(),
                'cuda_case_not_exist': False,
                'duplicated_issues': set(),
                'e2e_statuses': set()
            }

        if 'e2e_statuses' not in issue_test_results[issue_id]:
            issue_test_results[issue_id]['e2e_statuses'] = set()

        if e2e_status:
            issue_test_results[issue_id]['e2e_statuses'].add(e2e_status)

    issue_prs = {}
    issue_reporters = {}
    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        pr = ws_issues.cell(row, 15).value
        reporter = ws_issues.cell(row, 5).value
        assignee = ws_issues.cell(row, 4).value

        issue_prs[issue_id] = pr
        issue_reporters[issue_id] = reporter

    for row in range(2, ws_issues.max_row + 1):
        issue_id = ws_issues.cell(row, 1).value
        pr = ws_issues.cell(row, 15).value
        reporter = ws_issues.cell(row, 5).value
        assignee = ws_issues.cell(row, 4).value
        labels = ws_issues.cell(row, 6).value
        title_raw = str(ws_issues.cell(row, 2).value) if ws_issues.cell(row, 2).value else ''
        module = str(ws_issues.cell(row, 12).value) if ws_issues.cell(row, 12).value else ''
        summary_raw = str(ws_issues.cell(row, 10).value) if ws_issues.cell(row, 10).value else ''
        test_module_raw = ws_issues.cell(row, 13).value
        test_module = str(test_module_raw) if test_module_raw else ''
        traceback = ''
        for tr in range(2, ws_test.max_row + 1):
            if ws_test.cell(tr, 1).value == issue_id:
                traceback = str(ws_test.cell(tr, 9).value) if ws_test.cell(tr, 9).value else ''
                break

        test_cases_str = ''
        for tr in range(2, ws_test.max_row + 1):
            if ws_test.cell(tr, 1).value == issue_id:
                tc = str(ws_test.cell(tr, 7).value) if ws_test.cell(tr, 7).value else ''
                test_cases_str += ' ' + tc

        test_info = issue_test_results.get(issue_id, {
            'xpu_statuses': set(),
            'stock_statuses': set(),
            'cuda_case_not_exist': False,
            'duplicated_issues': set()
        })

        xpu_statuses = test_info['xpu_statuses']
        stock_statuses = test_info['stock_statuses']
        cuda_case_not_exist = test_info['cuda_case_not_exist']
        duplicated_issues = test_info['duplicated_issues']
        e2e_statuses = test_info.get('e2e_statuses', set())

        owner_transfer = ''
        action_tbd = ''

        labels_str = str(labels).lower() if labels else ''
        title_lower = title_raw.lower()
        summary_lower = summary_raw.lower() if summary_raw else ''

        is_not_target = ('not target' in labels_str or 'wont' in labels_str or "won't" in labels_str)

        if is_not_target:
            owner_transfer = reporter
            action_tbd = 'add to skiplist'

        is_random = 'random' in labels_str
        ut_passed = False
        if not is_random and xpu_statuses and stock_statuses:
            xpu_all_passed = (xpu_statuses == {'passed'})
            stock_all_passed = (stock_statuses == {'passed'})

            if xpu_all_passed or stock_all_passed:
                owner_transfer = reporter
                action_tbd = 'Close fixed issue'
                ut_passed = True

        e2e_all_passed = all(s == 'pass' for s in e2e_statuses) if e2e_statuses else False
        if not ut_passed and e2e_all_passed:
            owner_transfer = reporter
            action_tbd = 'Close fixed issue'

        pr_status = ws_issues.cell(row, 17).value
        pr_closed = pr_status in ['closed', 'merged']

        if not owner_transfer and pr_closed:
            has_failed = ('failed' in xpu_statuses) or ('failed' in stock_statuses)

            if not has_failed:
                owner_transfer = reporter
                action_tbd = 'Verify the issue'
            else:
                owner_transfer = assignee
                action_tbd = 'Revisit the PR as case failed'

        is_upstream = 'ut_upstream' in labels_str or 'inductor' in labels_str
        is_wontfix = 'wont ' in labels_str or ' wont ' in labels_str or 'wontfix' in labels_str or 'not target' in labels_str.replace('nottarget', '')
        is_not_target_upstream = is_not_target and is_upstream

        if not action_tbd:
            if is_not_target_upstream:
                action_tbd = 'Needs Upstream Skip PR (not_target + ut_upstream)'
                owner_transfer = assignee
            elif is_wontfix:
                action_tbd = 'Needs Skip PR (wontfix / not_target)'
                owner_transfer = assignee
            elif is_upstream:
                action_tbd = ''
                owner_transfer = assignee

        priority = 'P2'
        priority_reason = ''

        is_model_issue = ('model' in title_raw.lower() or 'model' in summary_raw.lower() or
                        'application' in title_raw.lower() or 'application' in summary_raw.lower() or
                        'huggingface' in title_raw.lower() or 'timm' in title_raw.lower() or 'torchbench' in title_raw.lower())
        is_e2e = test_module == 'e2e'
        is_ut = test_module == 'ut'

        is_regression = ('regression' in labels_str or 'regression' in title_raw.lower() or
                        'was pass' in summary_raw.lower() or 'previously pass' in summary_raw.lower() or
                        ('before' in summary_raw.lower() and 'now' in summary_raw.lower()))

        is_build_crash = ('build' in test_module.lower() or 'build' in title_raw.lower() or
                         'crash' in title_raw.lower() or 'segmentation' in title_raw.lower() or
                         'segfault' in title_raw.lower() or 'signal' in summary_raw.lower())

        failed_count = 0
        for tr in range(2, ws_test.max_row + 1):
            if ws_test.cell(tr, 1).value == issue_id:
                tc_status = ws_test.cell(tr, 11).value
                if tc_status in ['failed', 'error']:
                    failed_count += 1

        if is_build_crash:
            priority = 'P0'
            priority_reason = 'Build crash - critical blocking issue'
        elif is_model_issue and not ('test' in title_raw.lower() and 'case' in title_raw.lower()):
            priority = 'P0'
            priority_reason = 'Impacts real model/application'
        elif is_regression:
            priority = 'P0'
            priority_reason = 'Regression - passed before but failed now'
        elif is_e2e and ('accuracy' in title_raw.lower() or 'accuracy' in summary_raw.lower() or
                        'fail' in title_raw.lower() or 'fail' in summary_raw.lower()):
            priority = 'P1'
            priority_reason = 'E2E benchmark accuracy/functionality issue'
        elif is_e2e and ('performance' in title_raw.lower() or 'slow' in title_raw.lower() or
                        'latency' in title_raw.lower()):
            priority = 'P2'
            priority_reason = 'E2E benchmark performance issue'
        elif is_ut and failed_count > 20:
            priority = 'P1'
            priority_reason = f'UT with {failed_count} failed test cases'
        else:
            priority = 'P2'
            priority_reason = 'UT issue with few failures'

        if llm_priority_count < MAX_LLM_PRIORITY:
            test_cases_for_llm = []
            for tr in range(2, ws_test.max_row + 1):
                if ws_test.cell(tr, 1).value == issue_id:
                    tc_info = {
                        'test_case': ws_test.cell(tr, 7).value,
                        'error_msg': ws_test.cell(tr, 8).value,
                        'traceback': ws_test.cell(tr, 9).value
                    }
                    test_cases_for_llm.append(tc_info)
            llm_priority, llm_reason, _ = determine_priority_llm(
                title_raw, summary_raw, error_msg if 'error_msg' in dir() else '',
                test_module, str(labels) if labels else '',
                test_cases_for_llm
            )
            if llm_priority.startswith('P') and not llm_reason.startswith('API'):
                priority = llm_priority
                priority_reason = llm_reason
                llm_priority_count += 1
                print(f"  [LLM PRIORITY #{llm_priority_count}] Issue {issue_id}: {priority} - {priority_reason}")

        if duplicated_issues:
            ws_issues.cell(row, 21, ','.join(sorted(duplicated_issues)))

        if not action_tbd:
            version_info = extract_version_info(title_raw + ' ' + summary_raw)
            is_public = is_public_branch(version_info)

            has_xpu_status = bool(xpu_statuses)
            has_stock_status = bool(stock_statuses)
            has_e2e_status = bool(e2e_statuses)

            is_e2e_issue = test_module_raw == 'e2e'

            all_statuses_empty = not has_xpu_status and not has_stock_status and not has_e2e_status

            is_feature_request = (
                ('feature' in title_lower or 'feature' in summary_lower) or
                ('request' in title_lower or 'request' in summary_lower) or
                ('implement' in title_lower or 'implement' in summary_lower) or
                ('add support' in title_lower or 'add support' in summary_lower) or
                ('enable' in title_lower and 'test' not in title_lower and 'feature' in summary_lower) or
                (labels and 'enhancement' in str(labels).lower())
            )

            if is_e2e_issue:
                if is_e2e_issue and has_e2e_status and e2e_all_passed:
                    owner_transfer = reporter
                    action_tbd = 'Close fixed issue'
                else:
                    action_tbd = ''
                    owner_transfer = assignee
            elif is_public and all_statuses_empty and not is_feature_request:
                if e2e_all_passed:
                    owner_transfer = reporter
                    action_tbd = 'Close fixed issue'
                else:
                    action_tbd = ''
                    owner_transfer = assignee
            elif is_public and all_statuses_empty and not is_feature_request and not is_e2e_issue:
                action_tbd = 'Check case availability'
                owner_transfer = reporter

            if not action_tbd and not is_feature_request and not is_e2e_issue:
                issue_content = title_raw + ' ' + summary_raw

                is_bug_or_perf = any(kw in title_lower or kw in summary_lower for kw in [
                    'bug', 'fail', 'error', 'crash', 'assertion', 'exception',
                    'performance', 'slow', 'latency', 'timeout', 'regression',
                    'accuracy', 'wrong result', 'precision', 'dtype'
                ])

                has_already_requested = check_info_requested_to_reporter(issue_content)
                llm_error_msg = ''
                llm_traceback = ''
                test_file = ''
                test_case = ''
                for tr in range(2, ws_test.max_row + 1):
                    if ws_test.cell(tr, 1).value == issue_id:
                        llm_error_msg = str(ws_test.cell(tr, 8).value) if ws_test.cell(tr, 8).value else ''
                        llm_traceback = str(ws_test.cell(tr, 9).value) if ws_test.cell(tr, 9).value else ''
                        test_file = str(ws_test.cell(tr, 4).value) if ws_test.cell(tr, 4).value else ''
                        test_case = str(ws_test.cell(tr, 7).value) if ws_test.cell(tr, 7).value else ''
                        break
                llm_info_action = check_info_requested_to_reporter_llm(
                    title_raw, summary_raw, llm_error_msg, llm_traceback
                )
                if llm_info_action and llm_info_action not in ['Ready to analyze']:
                    has_already_requested = True
                    if 'reproduce' in llm_info_action.lower():
                        action_tbd = 'Need reproduce steps'
                    else:
                        action_tbd = f'LLM Suggestion: {llm_info_action}'
                    owner_transfer = reporter

                has_test_info = bool(test_file and test_case)

                if is_bug_or_perf:

                    if has_already_requested and 'reproduce' in llm_info_action.lower():
                        action_tbd = 'Need reproduce steps'
                        owner_transfer = reporter
                    elif has_test_info:
                        action_tbd = ''
                    elif has_already_requested:
                        action_tbd = 'Awaiting response from reporter'
                        owner_transfer = reporter
                    else:
                        action_tbd = ''
                else:
                    if has_already_requested:
                        action_tbd = 'Awaiting response from reporter'
                        owner_transfer = reporter
                    else:
                        info_needed = []
                        if 'accuracy' in title_lower or 'accuracy' in summary_lower:
                            info_needed.append('accuracy comparison data')
                        if 'performance' in title_lower or 'performance' in summary_lower:
                            info_needed.append('performance numbers/baseline')
                        if 'regression' in title_lower or 'regression' in summary_lower:
                            info_needed.append('previous good version info')

                        if info_needed:
                            info_str = ', '.join(info_needed)
                            action_tbd = f'Need more information - {info_str}'
                            owner_transfer = reporter

        if owner_transfer:
            ws_issues.cell(row, 19, owner_transfer)
        if action_tbd:
            ws_issues.cell(row, 20, action_tbd)
        old_format = 'Need reproduce step and more information'
        current_cell_value = ws_issues.cell(row, 20).value
        if current_cell_value == old_format:
            ws_issues.cell(row, 20, action_tbd if action_tbd else '')

        ws_issues.cell(row, 22, priority)
        ws_issues.cell(row, 23, priority_reason)

        category = determine_category(title_raw, summary_raw, test_cases_str, traceback, test_module, labels)
        category_reason = ''
        if llm_category_count < MAX_LLM_CATEGORY:
            test_cases_for_llm = []
            for tr in range(2, ws_test.max_row + 1):
                if ws_test.cell(tr, 1).value == issue_id:
                    tc_info = {
                        'test_case': ws_test.cell(tr, 7).value,
                        'test_file': ws_test.cell(tr, 4).value,
                        'error_msg': ws_test.cell(tr, 8).value,
                        'torch_ops': ws_test.cell(tr, 10).value
                    }
                    test_cases_for_llm.append(tc_info)
            category_llm, category_reason = determine_category_llm(
                title_raw, summary_raw, test_cases_for_llm, test_module, labels
            )
            if category_llm and not category_llm.startswith('API') and not category_llm.startswith('Error'):
                category = category_llm
                llm_category_count += 1
                print(f"  [LLM CATEGORY #{llm_category_count}] Issue {issue_id}: {category}")
        ws_issues.cell(row, 24, category)
        ws_issues.cell(row, 26, category_reason or '')

        current_action_tbd = ws_issues.cell(row, 20).value
        if not current_action_tbd:
            issue_test_file = ''
            issue_test_class = ''
            issue_test_case = ''
            issue_error_msg = ''
            issue_traceback = ''

            for tr in range(2, ws_test.max_row + 1):
                if ws_test.cell(tr, 1).value == issue_id:
                    if not issue_test_file:
                        issue_test_file = str(ws_test.cell(tr, 4).value) if ws_test.cell(tr, 4).value else ''
                    if not issue_test_class:
                        issue_test_class = str(ws_test.cell(tr, 6).value) if ws_test.cell(tr, 6).value else ''
                    issue_test_case = str(ws_test.cell(tr, 7).value) if ws_test.cell(tr, 7).value else ''
                    if not issue_error_msg:
                        issue_error_msg = str(ws_test.cell(tr, 8).value) if ws_test.cell(tr, 8).value else ''
                    if not issue_traceback:
                        issue_traceback = str(ws_test.cell(tr, 9).value) if ws_test.cell(tr, 9).value else ''
                    if issue_error_msg and issue_traceback:
                        break

            root_cause = analyze_root_cause(
                title_raw,
                summary_raw,
                issue_test_file,
                issue_test_class,
                issue_test_case,
                issue_error_msg,
                issue_traceback,
                test_module
            )

            if llm_root_cause_count < MAX_LLM_ROOT_CAUSE:
                print(f"  [LLM ROOT CAUSE #{llm_root_cause_count+1}] Issue {issue_id}: Calling LLM for root cause analysis...")
                llm_start = time.time()
                llm_root_cause = analyze_root_cause_llm(
                    issue_id,
                    title_raw,
                    summary_raw,
                    issue_test_file,
                    issue_test_class,
                    issue_test_case,
                    issue_error_msg,
                    issue_traceback,
                    test_module
                )
                llm_elapsed = time.time() - llm_start
                if llm_root_cause:
                    root_cause = llm_root_cause
                    llm_root_cause_count += 1
                    print(f"  [LLM ROOT CAUSE #{llm_root_cause_count}] Issue {issue_id} -> '{llm_root_cause}' ({llm_elapsed:.2f}s)")
                else:
                    print(f"  [LLM ROOT CAUSE #{llm_root_cause_count+1}] Issue {issue_id}: LLM returned empty")

            if root_cause:
                ws_issues.cell(row, 25, root_cause)

    print(f"Processed {ws_issues.max_row - 1} issues")
    return wb


def main():
    """Main function to run process_issues_sheet."""
    print("=" * 60)
    print("Processing Issues sheet...")
    print("=" * 60)

    input_file = os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx')

    if not os.path.exists(input_file):
        print(f"Error: Input file {input_file} not found.")
        return False

    wb = openpyxl.load_workbook(input_file)
    process_issues_sheet(wb)
    wb.save(input_file)
    print(f"Saved to {input_file}")
    return True


if __name__ == '__main__':
    main()