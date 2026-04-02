#!/usr/bin/env python3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

# Load data
with open("/home/daisydeng/issue_traige/data/torch_xpu_ops_issues.json") as f:
    issues = json.load(f)

with open("/home/daisydeng/issue_traige/data/torch_xpu_ops_comments.json") as f:
    comments = json.load(f)

# Load ops dependency
ops_dep = {}
with open("/home/daisydeng/issue_traige/doc/ops_dependency.csv") as f:
    next(f)
    for line in f:
        parts = line.strip().split(',')
        if len(parts) >= 2:
            op_name = parts[0].strip().lower()
            ops_dep[op_name] = parts[1].strip()

# Known test types
KNOWN_TEST_TYPES = ['op_ut', 'op_extend', 'e2e', 'benchmark', 'ut']

# Map test case to torch ops (common patterns)
TEST_CASE_OPS = {
    'test_alias': 'aten.alias',
    'test_retain_autograd': 'aten.retain_autograd',
    'test_cross_entropy': 'aten.cross_entropy',
    'test_nll_loss': 'aten.nll_loss',
    'test_mse_loss': 'aten.mse_loss',
    'test_l1_loss': 'aten.l1_loss',
    'test_layer_norm': 'aten.layer_norm',
    'test_rms_norm': 'aten.rms_norm',
    'test_group_norm': 'aten.group_norm',
    'test_batch_norm': 'aten.batch_norm',
    'test_linear': 'aten.linear',
    'test_conv': 'aten.conv',
    'test_conv2d': 'aten.conv2d',
    'test_conv3d': 'aten.conv3d',
    'test_matmul': 'aten.matmul',
    'test_mm': 'aten.mm',
    'test_bmm': 'aten.bmm',
    'test_addmm': 'aten.addmm',
    'test_dot': 'aten.dot',
    'test_vdot': 'aten.vdot',
    'test_embedding': 'aten.embedding',
    'test_softmax': 'aten.softmax',
    'test_gelu': 'aten.gelu',
    'test_relu': 'aten.relu',
    'test_scaled_dot_product': 'aten.scaled_dot_product_attention',
    'test_sdpa': 'aten.scaled_dot_product_attention',
    'test_flash_attention': 'aten._flash_attention_forward',
    'test_cudnn_attention': 'scaled_dot_product_attention',
    'test_cudnn': 'scaled_dot_product_attention',
    'test_index_add': 'aten.index_add',
    'test_index_copy': 'aten.index_copy',
    'test_gather': 'aten.gather',
    'test_scatter': 'aten.scatter',
    'test_topk': 'aten.topk',
    'test_sort': 'aten.sort',
    'test_argsort': 'aten.argsort',
    'test_chunk': 'aten.chunk',
    'test_split': 'aten.split',
    'test_stack': 'aten.stack',
    'test_cat': 'aten.cat',
    'test_flatten': 'aten.flatten',
    'test_reshape': 'aten.reshape',
    'test_view': 'aten.view',
    'test_transpose': 'aten.transpose',
    'test_permute': 'aten.permute',
    'test_contiguous': 'aten.contiguous',
    'test_clone': 'aten.clone',
    'test_copy': 'aten.copy',
    'test_bernoulli': 'aten.bernoulli',
    'test_normal': 'aten.normal',
    'test_randn': 'aten.randn',
    'test_randint': 'aten.randint',
    'test_rand': 'aten.rand',
    'test_zeros': 'aten.zeros',
    'test_ones': 'aten.ones',
    'test_full': 'aten.full',
    'test_arange': 'aten.arange',
    'test_linspace': 'aten.linspace',
    'test_logspace': 'aten.logspace',
    'test_empty': 'aten.empty',
    'test_eye': 'aten.eye',
    'test_scalar_tensor': 'aten.scalar_tensor',
    'test_tensor': 'aten.tensor',
    'test_as_tensor': 'aten.as_tensor',
    'test_from_numpy': 'aten.from_numpy',
    'test_sparse': 'aten.sparse',
    'test_coalesce': 'aten.coalesce',
    'test_indices': 'aten.indices',
    'test_values': 'aten.values',
    'test_to_dense': 'aten.to_dense',
    'test_to_sparse': 'aten.to_sparse',
    'test_pdist': 'aten.pdist',
    'test_cdist': 'aten.cdist',
    'test_triangular_solve': 'aten.triangular_solve',
    'test_cholesky': 'aten.cholesky',
    'test_qr': 'aten.linalg_qr',
    'test_solve': 'aten.linalg_solve',
    'test_inv': 'aten.linalg_inv',
    'test_det': 'aten.linalg_det',
    'test_slogdet': 'aten.linalg_slogdet',
    'test_eig': 'aten.linalg_eig',
    'test_eigh': 'aten.linalg_eigh',
    'test_svd': 'aten.linalg_svd',
    'test_lu': 'aten.linalg_lu',
    
    # Additional ops from traceback/errors - MORE SPECIFIC patterns first!
    'test_block_addmv': 'torch.addmv',  # More specific before test_block
    'test_block_addmm': 'torch.addmm',  # More specific before test_block
    'test_block_triangular_solve': 'aten.triangular_solve.X',
    'test_scaled_dot_product': 'scaled_dot_product_attention',
    'test_sdpa': 'scaled_dot_product_attention',
    'test_flash_attention': '_flash_attention_forward',
    'test_triton_bsr': 'triton_bsr',
    'test_triton_scaled': 'triton_scaled_dot_product_attention',
    'test_triton': 'triton',
    'test_block': 'torch.addmm',  # Generic block tests (fallback)
    'test_baddbmm': 'aten.baddbmm',
    'test_bmm': 'aten.bmm',
    'test_mm': 'aten.mm',
    'test_addmm': 'aten.addmm',
    'test_addmv': 'torch.addmv',
    'test_matmul': 'aten.matmul',
    'test_sparse': 'sparse',
    'test_triton_bsr_dense': 'triton_bsr_dense_addmm',
    'test_sampled_addmm': 'sampled_addmm',
    'test_triton_sampled': 'sampled_addmm',
    'test_conv': 'aten.conv',
    'test_conv2d': 'aten.conv2d',
    'test_conv_transpose': 'aten.conv_transpose',
    'test_norm': 'aten.layer_norm',  # generic norm tests
    
    # Direct aten/torch op mentions in traceback
    'index_add': 'aten.index_add',
    'index_copy': 'aten.index_copy',
    'layer_norm': 'aten.layer_norm',
    'rms_norm': 'aten.rms_norm',
    'cross_entropy': 'aten.cross_entropy',
    'embedding': 'aten.embedding',
    'linear': 'aten.linear',
    'conv2d': 'aten.conv2d',
    'matmul': 'aten.matmul',
    'bmm': 'aten.bmm',
    'mm': 'aten.mm',
    'addmm': 'aten.addmm',
    'dot': 'aten.dot',
    'vdot': 'aten.vdot',
    'gelu': 'aten.gelu',
    'relu': 'aten.relu',
    'softmax': 'aten.softmax',
    'batch_norm': 'aten.batch_norm',
    'group_norm': 'aten.group_norm',
    'gather': 'aten.gather',
    'scatter': 'aten.scatter',
    'topk': 'aten.topk',
    'sort': 'aten.sort',
    'cholesky': 'aten.cholesky',
    'qr': 'aten.linalg_qr',
    'svd': 'aten.linalg_svd',
    'solve': 'aten.linalg_solve',
    'inverse': 'aten.linalg_inv',
    'det': 'aten.linalg_det',
}

# Direct op name patterns (for traceback without aten. prefix)
DIRECT_OPS = [
    'scaled_dot_product_attention',
    '_flash_attention_forward',
    '_efficient_attention_forward',
    'index_add',
    'index_copy',
    'layer_norm',
    'rms_norm',
    'cross_entropy',
    'embedding',
    'linear',
    'matmul',
    'bmm',
    'mm',
    'addmm',
    'dot',
    'vdot',
    'gelu',
    'relu',
    'softmax',
    'batch_norm',
    'group_norm',
    'gather',
    'scatter',
    'topk',
    'sort',
    'cholesky',
    'qr',
    'svd',
    'solve',
]

# Infer torch ops from test case name, error message, and traceback
def infer_torch_ops_from_test_case(test_case, body, error_msg="", traceback=""):
    ops = []
    all_text = f"{test_case} {body} {error_msg} {traceback}".lower()
    full_text = f"{test_case} {body} {error_msg} {traceback}"
    
    if not test_case:
        # Try to infer from error/traceback only
        for direct_op in DIRECT_OPS:
            if direct_op.lower() in all_text:
                ops.append(direct_op)
                break
        return ops
    
    # Check test case name patterns first
    test_case_lower = test_case.lower()
    for pattern, op in TEST_CASE_OPS.items():
        if pattern in test_case_lower:
            ops.append(op)
            break
    
    # If no match from test case, check error message and traceback
    if not ops:
        for pattern, op in TEST_CASE_OPS.items():
            if pattern in all_text:
                ops.append(op)
                break
    
    # Also check for direct aten/torch op mentions in error/traceback
    if not ops:
        for direct_op in DIRECT_OPS:
            if direct_op.lower() in all_text:
                ops.append(direct_op)
                break
    
    # Check for aten.xxx pattern (including aten:: and aten.X patterns)
    # Matches: aten.index_add, aten::triangular_solve.X, aten::_scaled_dot_product_efficient_attention_backward
    aten_pattern = re.findall(r'aten::([^\s\'"]+)', full_text)
    for a_op in aten_pattern:
        # Clean up - remove leading underscores or trailing .X
        clean_op = a_op.strip()
        if clean_op.startswith('_'):
            full_op = f'aten::{clean_op}'
        elif '.' in clean_op:
            # Handle aten.triangular_solve.X or aten::triangular_solve.X
            full_op = f'aten.{clean_op}'
        else:
            full_op = f'aten.{clean_op}'
        if full_op not in ops:
            ops.append(full_op)
    
    # Also check standard aten.xxx pattern
    aten_pattern2 = re.findall(r'aten\.(\w+)', full_text)
    for a_op in aten_pattern2:
        full_op = f'aten.{a_op}'
        if full_op not in ops:
            ops.append(full_op)
    
    # Check for torch.xxx pattern  
    torch_pattern = re.findall(r'torch\.(\w+)\(', full_text)
    for t_op in torch_pattern:
        full_op = f'torch.{t_op}'
        if full_op not in ops:
            ops.append(full_op)
    
    return ops

def map_origin_test_file(test_file):
    """Map torch-xpu-ops test file to pytorch test file"""
    if not test_file:
        return ""
    
    match = re.search(r'test/xpu/(.+?)(?:_xpu)?\.py$', test_file)
    if match:
        return f"test/{match.group(1)}.py"
    
    if 'benchmarks/' in test_file:
        return test_file
    
    return test_file

def parse_test_cases_from_body(body):
    """Parse test cases from body, skip wrapped with ~~"""
    cases = []
    
    if 'Cases:' in body:
        cases_section = body.split('Cases:')[1]
        
        end_markers = ['\n###', '\nVersions', '\n```', '\n\n']
        min_end = len(cases_section)
        for marker in end_markers:
            idx = cases_section.find(marker)
            if idx > 0 and idx < min_end:
                min_end = idx
        cases_section = cases_section[:min_end]
        
        lines = cases_section.split('\n')
        
        for line in lines:
            line = line.strip()
            
            if not line:
                continue
            
            if line.startswith('###') or line.startswith('...'):
                continue
            
            if line.startswith('~~') and line.endswith('~~'):
                continue
            
            parts = line.split(',')
            if len(parts) < 3:
                continue
            
            test_type = parts[0].strip()
            if test_type not in KNOWN_TEST_TYPES:
                continue
            
            test_path = parts[1].strip()
            test_case = parts[2].strip()
            
            if not test_path or not test_case or len(test_case) < 3:
                continue
            
            if ' ' in test_case:
                continue
            
            if 'torch-xpu-ops' in test_path:
                path_parts = test_path.split('.')
                try:
                    txpo_idx = path_parts.index('torch-xpu-ops')
                    rel_parts = path_parts[txpo_idx+1:]
                    
                    test_class = ""
                    test_file_parts = []
                    for part in rel_parts:
                        if part.startswith('Test'):
                            test_class = part
                            break
                        test_file_parts.append(part)
                    
                    if test_file_parts:
                        test_file = 'torch-xpu-ops/' + '/'.join(test_file_parts) + '.py'
                        if not test_file.endswith('_xpu.py'):
                            test_file = test_file.replace('.py', '_xpu.py')
                    else:
                        test_file = ""
                    
                    origin_file = map_origin_test_file(test_file)
                except:
                    test_file = ""
                    test_class = ""
                    origin_file = ""
            else:
                test_file = test_path
                origin_file = ""
                test_class = ""
            
            cases.append({
                'test_type': test_type,
                'test_file': test_file,
                'origin_test_file': origin_file,
                'test_class': test_class,
                'test_case': test_case
            })
    
    # Format 2: e2e benchmark
    if 'benchmarks/dynamo/' in body:
        matches = re.findall(r'(python\s+benchmarks/dynamo/[^\s]+)', body)
        for match in matches:
            test_file = match.replace('python ', '').strip()
            cases.append({
                'test_type': 'e2e',
                'test_file': test_file,
                'origin_test_file': test_file,
                'test_class': '',
                'test_case': match.strip()
            })
    
    # Format 3: pytest command
    if 'pytest' in body:
        k_match = re.search(r'pytest[^-]*(-k\s+[^\s]+)?', body)
        if k_match and k_match.group(1):
            cases.append({
                'test_type': 'ut',
                'test_file': '',
                'origin_test_file': '',
                'test_class': '',
                'test_case': k_match.group(1).strip()
            })
    
    return cases

def extract_error_and_traceback(body):
    """Extract error message and full traceback"""
    error_msg = ""
    traceback = ""
    
    error_patterns = [
        r'(AssertionError[^\n]*)',
        r'(RuntimeError[^\n]*)',
        r'(ValueError[^\n]*)',
        r'(TypeError[^\n]*)',
        r'(IndexError[^\n]*)',
        r'(KeyError[^\n]*)',
        r'(ImportError[^\n]*)',
        r'(NotImplementedError[^\n]*)',
        r'(AttributeError[^\n]*)',
        r'(InductorError[^\n]*)',
    ]
    
    for pattern in error_patterns:
        match = re.search(pattern, body, re.IGNORECASE)
        if match:
            error_msg = match.group(1).strip()[:200]
            break
    
    if 'Traceback (most recent call last):' in body:
        start = body.find('Traceback (most recent call last):')
        end = body.find('\n###', start)
        if end == -1:
            end = body.find('\n\n', start)
        if end == -1:
            end = body.find('\nVersions', start)
        if end == -1:
            end = min(start + 2000, len(body))
        traceback = body[start:end].strip()
    
    elif re.search(r'_{5,}\s+Test\w+', body):
        match = re.search(r'_{5,}\s+Test\w+.*?(?=_{5,}|\n###|\nVersions|\Z)', body, re.DOTALL)
        if match:
            traceback = match.group(0).strip()[:2000]
    
    if not traceback:
        lines = body.split('\n')
        in_trace = False
        trace_lines = []
        for i, line in enumerate(lines):
            if 'File "' in line and '.py"' in line:
                in_trace = True
            if in_trace:
                trace_lines.append(line)
                if any(e in line for e in ['Error:', 'Exception:', 'raise ']) and i > 3:
                    break
        if trace_lines:
            traceback = '\n'.join(trace_lines)[:2000]
    
    return error_msg, traceback

def generate_summary(body, title, error_msg):
    """Generate a 1-2 sentence summary"""
    if 'Error' in title or 'Exception' in title:
        match = re.search(r'(Error|Exception):\s*(.+)', title)
        if match:
            return match.group(2).strip()[:150]
        else:
            return title[:150]
    else:
        return title[:150]

def classify_issue_type(body, title, labels):
    """Classify issue type with more precise rules"""
    text = f"{title} {body}".lower()
    
    # Check labels first
    for label in labels:
        ln = label.get('name', '').lower()
        if 'task' == ln or 'internal task' in ln:
            return 'internal task'
    
    # Check for explicit performance keywords
    performance_keywords = [
        'performance regression', 'performance dropped', 'performance issue',
        'latency', 'throughput', 'slow performance', 'performance slow',
        'execution time', 'runtime performance', 'performance fail'
    ]
    
    has_performance_keyword = any(k in text for k in performance_keywords)
    
    # Check for bug/error keywords
    bug_keywords = [
        'assertionerror', 'runtimeerror', 'valueerror', 'typeerror', 'indexerror',
        'keyerror', 'importerror', 'notimplementederror', 'attributeerror',
        'inductorerror', 'crash', 'fail', 'bug', 'error', 'not implemented',
        'not supported', 'missing', 'incorrect', 'wrong', 'unexpected'
    ]
    
    has_bug_keyword = any(k in text for k in bug_keywords)
    
    # Check for feature request
    feature_keywords = ['feature request', 'support for', 'implement', 'add support', 'need feature']
    has_feature_keyword = any(k in text for k in feature_keywords)
    
    # Determine type with priority
    if has_feature_keyword:
        return 'feature request'
    
    if has_performance_keyword:
        return 'performance issue'
    
    if has_bug_keyword:
        return 'functionality bug'
    
    return 'unknown'

def classify_test_module(body, title, labels):
    """Classify test module with more precise rules"""
    text = f"{title} {body}".lower()
    
    # Check for pytest or python test command on pytorch/test or test/xpu/test
    # or benchmark tests (e2e)
    pytest_patterns = [
        r'pytest\s+.*test[/._]',  # pytest on test file
        r'python\s+.*test[/._]',   # python test file
        r'test/test_',            # pytorch test folder
        r'test/xpu/test_',         # xpu test folder
        r'benchmarks/dynamo/',    # e2e benchmark
    ]
    
    has_test_pattern = False
    for pattern in pytest_patterns:
        if re.search(pattern, text):
            has_test_pattern = True
            break
    
    # Check for build related keywords - ONLY for source code/build process issues
    # These are specific build-related patterns, not general test failures
    build_patterns = [
        r'\[win\]\[build\]',
        r'build from source',
        r'compile from source', 
        r'source build',
        r'build script',
        r'BUILD_SEPARATE',
        r'BUILD_SHARED',
        r'cmake build',
        r'cmake error',
        r'cmake fail',
        r'setup\.py install',
        r'pip install -e \.',
        r'python setup\.py develop',
    ]
    
    has_build = any(re.search(p, text, re.IGNORECASE) for p in build_patterns)
    
    # Check for infrastructure - ONLY for CI/workflow infrastructure issues
    # Must be about the infrastructure itself, not just mentioning it in test failures
    infra_patterns = [
        r'workflow\s+(error|fail|issue|problem)',  # workflow error
        r'github\s+action\s+(error|fail|issue)',  # github action issue
        r'azure\s+pipeline\s+(error|fail)',      # azure pipeline issue
        r'ci\s+(runner|config|setup)\s+(error|fail)',  # CI runner/config issue
        r'runner\s+(error|fail|timeout)\s+in\s+ci',      # runner issue in CI
        r'checkout\s+(error|fail)\s+in\s+(workflow|ci)', # checkout action issue
        r'githubaction',                            # github action issue
    ]
    
    has_infra = any(re.search(p, text) for p in infra_patterns)
    
    # Also check labels for infrastructure
    for label in labels:
        ln = label.get('name', '').lower()
        if 'infrastructure' in ln and ('ci' in ln or 'workflow' in ln or 'action' in ln):
            has_infra = True
            break
    
    if has_build:
        return 'build'
    
    if has_infra:
        return 'infrastructure'
    
    if has_test_pattern:
        if 'benchmarks/dynamo/' in text or 'benchmark' in text:
            return 'e2e'
        return 'ut'
    
    # Default to ut for most test failures
    return 'ut'

def classify_module(body, title, labels):
    """Classify module based on content"""
    text = f"{title} {body}".lower()
    
    module_keywords = [
        ('distributed', ['distributed', 'device_mesh', 'ProcessGroup', 'FSDP', 'DDP', 'c10d', 'tensor parallel']),
        ('inductor', ['inductor', 'compile', 'triton', 'codegen', 'lowering', 'inductor error']),
        ('dynamo', ['dynamo', 'torch.compile', '_dynamo', 'dynamo error']),
        ('autograd', ['autograd', 'backward', 'grad', 'gradient']),
        ('aten_ops', ['aten::', 'torch.ops.aten']),
        ('low_precision', ['quantization', 'int8', 'fp8', 'int4', 'amp', 'bf16', 'fp16']),
        ('optimizer', ['optimizer', 'lr_scheduler', 'adam', 'sgd']),
        ('profiling', ['profiling', 'profile', 'benchmark']),
        ('fx', ['torch.fx', 'fx.', 'symbolic']),
        ('export', ['torch.export', 'exported']),
    ]
    
    for m, kw in module_keywords:
        if any(k in text for k in kw):
            return m
    
    # Check labels
    for label in labels:
        ln = label.get('name', '').lower()
        if 'module: distributed' in ln:
            return 'distributed'
        if 'module: inductor' in ln:
            return 'inductor'
        if 'module: ao' in ln:
            return 'AO'
        if 'module: ut' in ln:
            return 'aten_ops'
    
    return 'unknown'

def get_dependency_from_body(body):
    """Get dependency from body content"""
    text = body.lower()
    
    dep_keywords = [
        ('transformers', ['transformers', 'huggingface']),
        ('AO', ['torchao', 'torchao']),
        ('oneDNN', ['onednn', 'mkl-dnn']),
        ('oneCCL', ['oneccl', 'ccl']),
        ('oneMKL', ['onemkl', 'mkl']),
        ('driver', ['level-zero', 'libze', 'intel-opencl', 'driver']),
        ('Triton', ['triton', 'triton-xpu']),
        ('oneAPI', ['oneapi', 'dpcpp', 'icx']),
    ]
    
    for d, kw in dep_keywords:
        if any(k in text for k in kw):
            return d
    
    return 'None'

# Create Excel
wb = openpyxl.Workbook()

ws_issues = wb.active
ws_issues.title = "Issues"

headers = ["Issue ID", "Title", "Status", "Assignee", "Reporter", "Labels", 
           "Created Time", "Updated Time", "Milestone", "Summary", "Type", 
           "Module", "Test Module", "Dependency"]

for col, header in enumerate(headers, 1):
    cell = ws_issues.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

ws_cases = wb.create_sheet("Test Cases")

case_headers = ["Issue ID", "Test Reproducer", "Test Type", "Test File", 
                 "Origin Test File", "Test Class", "Test Case", 
                 "Error Message", "Traceback", "torch-ops", "dependency"]

for col, header in enumerate(case_headers, 1):
    cell = ws_cases.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

issue_row = 2
case_row = 2

for issue in issues:
    num = issue['number']
    title = issue['title']
    body = issue.get('body', '') or ''
    
    status = issue['state']
    assignee = ", ".join([a['login'] for a in issue.get('assignees', [])]) or "None"
    reporter = issue['user']['login']
    labels = issue.get('labels', [])
    label_str = ", ".join([l['name'] for l in labels])
    created = issue['created_at']
    updated = issue['updated_at']
    milestone = issue.get('milestone', {})
    milestone_name = milestone.get('title', 'None') if milestone else 'None'
    
    # Classify
    issue_type = classify_issue_type(body, title, labels)
    module = classify_module(body, title, labels)
    test_module = classify_test_module(body, title, labels)
    dependency = get_dependency_from_body(body)
    
    error_msg, traceback = extract_error_and_traceback(body)
    summary = generate_summary(body, title, error_msg)
    
    ws_issues.cell(row=issue_row, column=1, value=num)
    ws_issues.cell(row=issue_row, column=2, value=title)
    ws_issues.cell(row=issue_row, column=3, value=status)
    ws_issues.cell(row=issue_row, column=4, value=assignee)
    ws_issues.cell(row=issue_row, column=5, value=reporter)
    ws_issues.cell(row=issue_row, column=6, value=label_str)
    ws_issues.cell(row=issue_row, column=7, value=created)
    ws_issues.cell(row=issue_row, column=8, value=updated)
    ws_issues.cell(row=issue_row, column=9, value=milestone_name)
    ws_issues.cell(row=issue_row, column=10, value=summary)
    ws_issues.cell(row=issue_row, column=11, value=issue_type)
    ws_issues.cell(row=issue_row, column=12, value=module)
    ws_issues.cell(row=issue_row, column=13, value=test_module)
    ws_issues.cell(row=issue_row, column=14, value=dependency)
    
    test_cases = parse_test_cases_from_body(body)
    
    if test_cases:
        for tc in test_cases:
            # Infer torch ops from test case, error message, and traceback
            test_case = tc.get('test_case', '')
            torch_ops = infer_torch_ops_from_test_case(test_case, body, error_msg, traceback)
            
            if not torch_ops:
                torch_ops = ['unknown']
            
            case_dep = 'sycl'
            if torch_ops[0] != 'unknown':
                for op in torch_ops:
                    op_lower = op.lower().replace('aten.', '')
                    if op_lower in ops_dep:
                        case_dep = ops_dep[op_lower]
                        break
            
            ws_cases.cell(row=case_row, column=1, value=num)
            ws_cases.cell(row=case_row, column=2, value=title[:150])
            ws_cases.cell(row=case_row, column=3, value=tc.get('test_type', ''))
            ws_cases.cell(row=case_row, column=4, value=tc.get('test_file', ''))
            ws_cases.cell(row=case_row, column=5, value=tc.get('origin_test_file', ''))
            ws_cases.cell(row=case_row, column=6, value=tc.get('test_class', ''))
            ws_cases.cell(row=case_row, column=7, value=tc.get('test_case', ''))
            ws_cases.cell(row=case_row, column=8, value=error_msg)
            ws_cases.cell(row=case_row, column=9, value=traceback)
            ws_cases.cell(row=case_row, column=10, value=", ".join(torch_ops))
            ws_cases.cell(row=case_row, column=11, value=case_dep)
            case_row += 1
    else:
        if error_msg:
            ws_cases.cell(row=case_row, column=1, value=num)
            ws_cases.cell(row=case_row, column=2, value=title[:150])
            ws_cases.cell(row=case_row, column=8, value=error_msg)
            ws_cases.cell(row=case_row, column=9, value=traceback)
            ws_cases.cell(row=case_row, column=10, value="unknown")
            ws_cases.cell(row=case_row, column=11, value="sycl")
            case_row += 1
    
    issue_row += 1
    
    if issue_row % 50 == 0:
        print(f"Processed {issue_row-1} issues...")

print(f"\nTotal issues: {issue_row-2}")
print(f"Total test case rows: {case_row-2}")

for ws in [ws_issues, ws_cases]:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

output_path = "/home/daisydeng/issue_traige/data/torch_xpu_ops_issues.xlsx"
wb.save(output_path)
print(f"\nSaved to {output_path}")