#!/usr/bin/env python3
"""
Script to update torch_xpu_ops_issues.xlsx with test case results from:
1. torch-xpu-ops nightly CI (XML files in Inductor-XPU-UT-Data-*-op_ut-*)
2. Stock PyTorch XPU CI (test-reports from pytorch xpu CI artifacts)
3. Case existence analysis (checking if tests exist in pytorch/test and torch-xpu-ops/test/xpu)

Usage:
    python update_test_results.py

Input:
    - /home/daisydeng/issue_traige/data/torch_xpu_ops_issues.xlsx
    - /home/daisydeng/issue_traige/ci_results/torch-xpu-ops/
    - /home/daisydeng/issue_traige/ci_results/stock/
    - ~/pytorch/test/
    - ~/pytorch/third_party/torch-xpu-ops/test/xpu/

Output:
    - Updated /home/daisydeng/issue_traige/data/torch_xpu_ops_issues.xlsx
      with additional columns:
      - Column K: status in torch-xpu-ops nightly
      - Column L: comments in torch-xpu-ops nightly
      - Column M: Commit
      - Column N: Run_id
      - Column O: XML
      - Column P: status in stock CI
      - Column Q: comments in stock CI
      - Column R: cuda_case_exist
      - Column S: xpu_case_exist
      - Column T: case_existence_comments
"""

import openpyxl
import xml.etree.ElementTree as ET
import os
import re
import glob


def get_torch_xpu_ops_xml_files():
    """Get all XML files from torch-xpu-ops nightly artifacts"""
    base_dir = '/home/daisydeng/issue_traige/ci_results/torch-xpu-ops'
    
    ut_folders = []
    for d in os.listdir(base_dir):
        if d.startswith('Inductor-XPU-UT-Data-'):
            match = re.match(r'Inductor-XPU-UT-Data-([a-f0-9]+)-.*-(\d+)-1$', d)
            if match:
                ut_folders.append((d, match.group(1), match.group(2)))
    
    xml_files = {}
    for folder_name, commit, run_id in ut_folders:
        folder_path = os.path.join(base_dir, folder_name, folder_name)
        if not os.path.exists(folder_path):
            continue
        for f in os.listdir(folder_path):
            if f.endswith('.xml') and (f.startswith('op_ut_with_all') or f.startswith('op_ut_with_skip')):
                xml_path = os.path.join(folder_path, f)
                try:
                    tree = ET.parse(xml_path)
                    root = tree.getroot()
                    count = len(root.findall('.//testcase'))
                    if count > 0:
                        prefix = f.replace('.xml', '')
                        xml_files[prefix] = (xml_path, commit, run_id, count)
                except:
                    pass
    
    return xml_files


def get_stock_xml_files():
    """Get all XML files from stock PyTorch XPU CI"""
    stock_base = '/home/daisydeng/issue_traige/ci_results/stock'
    stock_xml_files = {}
    
    for zip_file in glob.glob(f'{stock_base}/test-reports-*.zip'):
        pytest_dir = os.path.join(zip_file, 'test-reports', 'python-pytest')
        if not os.path.exists(pytest_dir):
            continue
        for root, dirs, files in os.walk(pytest_dir):
            for f in files:
                if f.endswith('.xml'):
                    xml_path = os.path.join(root, f)
                    test_module = os.path.basename(root)
                    stock_xml_files[test_module] = xml_path
    
    return stock_xml_files


def convert_test_file_to_xml_prefix(test_file):
    """Convert test file to XML prefix for torch-xpu-ops"""
    if not test_file:
        return None, 'No test file'
    
    test_file = str(test_file)
    
    if '/' in test_file:
        test_file = test_file.replace('torch-xpu-ops/test/xpu/', '')
        test_file = test_file.replace('.py', '')
        return 'op_ut_with_all.' + test_file.replace('/', '.'), None
    
    parts = test_file.split('.')
    parts_len = len(parts)
    
    if parts_len == 2:
        module = parts[0]
        if '_xpu' in module:
            return 'op_ut_with_skip.' + module, None
        return 'op_ut_with_all.' + module, None
    
    if parts_len == 3:
        module = parts[0]
        test_module = parts[1]
        if module == 'test':
            return 'op_ut_with_all.test_' + test_module + '_xpu', None
        if module == 'inductor':
            return None, 'inductor not in torch-xpu-ops'
        return 'op_ut_with_all.' + module + '.' + test_module, None
    
    if parts_len == 4:
        module = parts[0]
        subdir = parts[1]
        test_module = parts[2]
        if module == 'test':
            if subdir == 'functorch':
                return f'op_ut_with_all.{subdir}.{test_module}_xpu', None
            return None, f'{subdir} not in torch-xpu-ops'
        if module == 'inductor':
            return None, 'inductor not in torch-xpu-ops'
    
    return None, 'unknown pattern'


def convert_to_stock_prefix(test_file):
    """Convert test file to stock test module name"""
    if not test_file:
        return None
    
    test_file = str(test_file)
    
    if '/' in test_file:
        test_file = test_file.replace('torch-xpu-ops/test/xpu/', '')
        test_file = test_file.replace('.py', '')
        return test_file.replace('/', '.')
    
    parts = test_file.split('.')
    parts_len = len(parts)
    
    if parts_len == 2:
        return parts[0]
    if parts_len == 3:
        module = parts[0]
        test_module = parts[1]
        if module == 'test':
            return test_module
        return f'{module}.{test_module}'
    if parts_len == 4:
        module = parts[0]
        subdir = parts[1]
        test_module = parts[2]
        if module == 'test':
            return f'{subdir}.{test_module}'
        if module == 'inductor':
            return f'{module}.{test_module}'
    
    return None


def find_best_xml_match(xml_prefix, xml_files):
    """Find XML file matching prefix"""
    if not xml_prefix:
        return None
    if xml_prefix in xml_files:
        return xml_files[xml_prefix]
    return None


def get_test_result(xml_path, test_case):
    """Get test case result from XML file"""
    if not xml_path:
        return None, None
    
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        
        for testcase in root.findall('.//testcase'):
            if testcase.get('name') == test_case:
                failure = testcase.find('failure')
                if failure is not None:
                    msg = failure.text or failure.get('message', '')
                    return 'failed', msg[:500] if msg else 'failed'
                skipped = testcase.find('skipped')
                if skipped is not None:
                    msg = skipped.text or skipped.get('message', '')
                    return 'skipped', msg[:500] if msg else 'skipped'
                return 'passed', ''
        
        return 'not found', 'Test case not found'
    except Exception as e:
        return 'error', str(e)


def analyze_test_case(test_file, test_case):
    """Analyze why test case exists or not"""
    pytorch_test_dir = os.path.expanduser('~/pytorch/test')
    torch_xpu_ops_dir = os.path.expanduser('~/pytorch/third_party/torch-xpu-ops/test/xpu')
    
    if not test_file:
        return {'cuda_file': 'N/A', 'cuda_exists': 'No', 'xpu_file': 'N/A', 'xpu_exists': 'No', 
                'explanation': 'No test file', 'expected_name': None}
    
    test_file = str(test_file)
    result = {
        'cuda_file': None, 'cuda_exists': 'No',
        'xpu_file': None, 'xpu_exists': 'No',
        'explanation': '', 'expected_name': test_case
    }
    
    # Parse file path
    if '/' in test_file:
        cuda_rel = test_file.replace('torch-xpu-ops/test/xpu/', '').replace('_xpu', '')
        xpu_rel = test_file.replace('torch-xpu-ops/test/xpu/', '')
    elif '.' in test_file:
        parts = test_file.split('.')
        parts_len = len(parts)
        
        if parts_len == 4:
            cuda_rel = f"{parts[1]}/{parts[2]}.py"
            xpu_rel = f"{parts[1]}/{parts[2]}_xpu.py"
            result['expected_name'] = parts[3]
        elif parts_len == 3:
            if parts[0] == 'test':
                cuda_rel = f"{parts[1]}.py"
                xpu_rel = f"{parts[1]}_xpu.py"
            elif parts[0] == 'inductor':
                cuda_rel = f"inductor/{parts[1]}.py"
                xpu_rel = None
        elif parts_len == 2:
            cuda_rel = f"{parts[0].replace('_xpu', '')}.py"
            xpu_rel = f"{parts[0]}.py"
        else:
            return result
    else:
        return result
    
    # Check CUDA file
    if cuda_rel:
        cuda_path = os.path.join(pytorch_test_dir, cuda_rel)
        if os.path.exists(cuda_path):
            result['cuda_file'] = cuda_rel
            result['cuda_exists'] = 'Yes'
            
            with open(cuda_path, 'r') as f:
                cuda_content = f.read()
            
            if test_case and test_case in cuda_content:
                result['explanation'] += 'CUDA: Found. '
            elif test_case and test_case.replace('_xpu', '') in cuda_content:
                result['explanation'] += f"CUDA: Uses _xpu suffix (expected: {test_case.replace('_xpu', '')}). "
            elif test_case and test_case.replace('_xpu', '_cuda') in cuda_content:
                result['explanation'] += f"CUDA: Uses _cuda suffix (expected: {test_case.replace('_xpu', '_cuda')}). "
            else:
                if '@parametrize' in cuda_content:
                    has_device_param = re.search(r"device.*?['\"]xpu['\"]", cuda_content, re.IGNORECASE)
                    if has_device_param:
                        result['explanation'] += 'CUDA: Parametrized with xpu device (should match). '
                    else:
                        base_name = test_case.replace('_xpu', '') if test_case else None
                        if base_name and base_name in cuda_content:
                            result['explanation'] += f"CUDA: Has test but needs XPU parametrize. "
                        else:
                            result['explanation'] += 'CUDA: Test not found (possibly removed or renamed). '
                else:
                    if 'XPUPatchForImport' in cuda_content or 'XPUPatchForImport' in str(test_file):
                        result['explanation'] += 'CUDA: Uses XPUPatchForImport pattern. '
                    else:
                        result['explanation'] += 'CUDA: Test not in file. '
        else:
            result['cuda_file'] = f'Not found: {cuda_rel}'
            result['explanation'] += f'CUDA file missing. '
    
    # Check XPU file
    if xpu_rel:
        xpu_path = os.path.join(torch_xpu_ops_dir, xpu_rel)
        if os.path.exists(xpu_path):
            result['xpu_file'] = xpu_rel
            result['xpu_exists'] = 'Yes'
            
            with open(xpu_path, 'r') as f:
                xpu_content = f.read()
            
            if test_case and test_case in xpu_content:
                result['explanation'] += 'XPU: Found. '
            elif test_case and test_case.replace('_xpu', '') in xpu_content:
                result['explanation'] += f"XPU: Found without _xpu. "
            else:
                if 'XPUPatchForImport' in xpu_content:
                    result['explanation'] += 'XPU: Uses XPUPatchForImport. '
                else:
                    result['explanation'] += 'XPU: Test not in file. '
        else:
            result['xpu_file'] = f'Not found: {xpu_rel}'
            if 'inductor' in test_file.lower():
                result['explanation'] += 'XPU: Inductor not in torch-xpu-ops. '
            else:
                result['explanation'] += 'XPU file missing. '
    elif xpu_rel is None:
        result['xpu_file'] = 'N/A'
        result['explanation'] += 'XPU: Inductor tests use stock CI. '
    
    return result


def main():
    # Load workbook
    wb = openpyxl.load_workbook('/home/daisydeng/issue_traige/data/torch_xpu_ops_issues.xlsx')
    ws = wb['Test Cases']
    
    # Add new columns
    ws.cell(1, 11, 'status in torch-xpu-ops nightly')
    ws.cell(1, 12, 'comments in torch-xpu-ops nightly')
    ws.cell(1, 13, 'Commit')
    ws.cell(1, 14, 'Run_id')
    ws.cell(1, 15, 'XML')
    ws.cell(1, 16, 'status in stock CI')
    ws.cell(1, 17, 'comments in stock CI')
    ws.cell(1, 18, 'cuda_case_exist')
    ws.cell(1, 19, 'xpu_case_exist')
    ws.cell(1, 20, 'case_existence_comments')
    
    # Get XML files
    print("Loading XML files...")
    xpu_xml_files = get_torch_xpu_ops_xml_files()
    print(f"  Found {len(xpu_xml_files)} torch-xpu-ops XML files")
    
    stock_xml_files = get_stock_xml_files()
    print(f"  Found {len(stock_xml_files)} stock XML files")
    
    # Process all rows
    total = ws.max_row - 1
    for i, row in enumerate(range(2, ws.max_row + 1), 1):
        test_file = ws.cell(row, 4).value
        test_class = ws.cell(row, 6).value
        test_case = ws.cell(row, 7).value
        
        # Get torch-xpu-ops nightly result
        xml_prefix, reason = convert_test_file_to_xml_prefix(test_file)
        if xml_prefix:
            matched = find_best_xml_match(xml_prefix, xpu_xml_files)
            if matched:
                xml_path, commit, run_id, _ = matched
                status, comment = get_test_result(xml_path, test_case)
                ws.cell(row, 11, status)
                ws.cell(row, 12, comment)
                ws.cell(row, 13, commit)
                ws.cell(row, 14, run_id)
                ws.cell(row, 15, os.path.basename(xml_path))
            else:
                ws.cell(row, 11, 'not found')
                ws.cell(row, 12, f'No XML: {xml_prefix}')
        else:
            ws.cell(row, 11, 'not found')
            ws.cell(row, 12, reason)
        
        # Get stock CI result
        stock_prefix = convert_to_stock_prefix(test_file)
        if stock_prefix and stock_prefix in stock_xml_files:
            stock_xml = stock_xml_files[stock_prefix]
            stock_status, stock_comment = get_test_result(stock_xml, test_case)
            ws.cell(row, 16, stock_status)
            ws.cell(row, 17, stock_comment)
        else:
            ws.cell(row, 16, 'not found')
            ws.cell(row, 17, 'Not in stock CI')
        
        # Get case existence info
        result = analyze_test_case(test_file, test_case)
        ws.cell(row, 18, result['cuda_exists'])
        ws.cell(row, 19, result['xpu_exists'])
        
        comment = f"CUDA file: {result['cuda_file']}. XPU file: {result['xpu_file']}. {result['explanation']}"
        ws.cell(row, 20, comment.strip())
        
        if i % 200 == 0:
            print(f"Processed {i}/{total}")
    
    print(f"Processed {total}/{total}")
    wb.save('/home/daisydeng/issue_traige/data/torch_xpu_ops_issues.xlsx')
    print("Saved!")


if __name__ == '__main__':
    main()